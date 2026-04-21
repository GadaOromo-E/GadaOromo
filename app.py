# -*- coding: utf-8 -*-
"""
Created on Sun Jan 11 16:32:35 2026
@author: ademo

Gadaa Dictionary - Flask + SQLite + PWA-ready

Includes:
- Dictionary search + translate pages
- Admin login + dashboard + approve/reject
- Public submission: words + phrases (manual + CSV/XLSX)
- Legacy /submit_file kept
- Admin bulk import (TXT/CSV/XLSX English-only) -> Google Translate -> pending
- Community audio upload + admin approve/reject
- In-page mic recording posts to: POST /api/submit-audio
- /learn, /support

PWA support:
- /manifest.webmanifest
- /service-worker.js (root scope)
- /offline

SEO / Google:
- /robots.txt
- /sitemap.xml
- ProxyFix for Render (correct https URLs)
- google verification file route

Recorder mode (password protected):
- Recorder password login (/recorder)
- Recorder dashboard (/recorder/dashboard) to quickly record unlimited words/phrases
- Recorder recordings are AUTO-APPROVED and can REPLACE existing approved audio
- Recorder can DELETE approved Oromo audio (no admin approval)
- Public recording stays the same: pending + admin approval
"""

import os
import re
import sqlite3
import logging
import csv
import json
import hashlib
import shutil
import time
import threading
import click
from uuid import uuid4
from difflib import get_close_matches
from io import StringIO, BytesIO
from datetime import datetime
from urllib.parse import quote, unquote
import unicodedata

import requests
from flask import (
    Flask, render_template, request, redirect, session,
    jsonify, send_from_directory, send_file, abort, make_response, Response, g
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.exceptions import HTTPException
from openpyxl import load_workbook
from services.translation_service import google_translate_batch as service_google_translate_batch
from services.translation_service import google_translate_text as service_google_translate_text
from services.translation_service import get_or_generate_translation as service_get_or_generate_translation
from services.tts_service import azure_synthesize_mp3
from services.tts_service import generate_and_store_tts as service_generate_and_store_tts
try:
    from azure.storage.blob import BlobServiceClient, ContentSettings
    from azure.core.exceptions import ResourceExistsError
except Exception:
    BlobServiceClient = None
    ContentSettings = None
    ResourceExistsError = Exception

import os
import shutil

APP_ROOT = os.path.dirname(os.path.abspath(__file__))
BOOTSTRAP_DB_SOURCE = os.path.join(APP_ROOT, "gadaoromo.db")
RAILWAY_DB_PATH = "/data/gadaoromo.db"



def _is_prod_runtime() -> bool:
    return (
        (os.environ.get("FLASK_ENV") == "production")
        or bool(os.environ.get("RENDER"))
        or bool(os.environ.get("RAILWAY_ENVIRONMENT"))
        or bool(os.environ.get("RAILWAY_PROJECT_ID"))
    )


def _safe_file_size(path: str):
    try:
        if os.path.isfile(path):
            return os.path.getsize(path)
    except Exception:
        pass
    return None


def _inspect_sqlite_db(path: str):
    """
    Inspect a SQLite DB file for bootstrap decisions.
    Returns structured diagnostics and a conservative keep/reseed decision.
    """
    diag = {
        "path": path,
        "exists": os.path.isfile(path),
        "size": _safe_file_size(path),
        "valid_sqlite": False,
        "tables": [],
        "words_count": None,
        "phrases_count": None,
        "approved_words_count": None,
        "approved_phrases_count": None,
        "keep": False,
        "reseed": True,
        "reason": "",
        "error": None,
    }

    if not diag["exists"]:
        diag["reason"] = "missing_file"
        return diag

    conn = None
    try:
        conn = sqlite3.connect(path)
        c = conn.cursor()
        c.execute("PRAGMA quick_check;")
        quick = str((c.fetchone() or [""])[0] or "").strip().lower()
        if quick != "ok":
            diag["reason"] = f"quick_check_failed:{quick or 'unknown'}"
            return diag
        diag["valid_sqlite"] = True

        c.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = sorted(
            str(r[0]).strip()
            for r in (c.fetchall() or [])
            if r and str(r[0] or "").strip()
        )
        diag["tables"] = tables

        required = {"words", "phrases"}
        if not required.issubset(set(tables)):
            diag["reason"] = "missing_required_tables"
            return diag

        c.execute("SELECT COUNT(*) FROM words")
        words_count = int((c.fetchone() or [0])[0] or 0)
        c.execute("SELECT COUNT(*) FROM phrases")
        phrases_count = int((c.fetchone() or [0])[0] or 0)
        diag["words_count"] = words_count
        diag["phrases_count"] = phrases_count

        try:
            c.execute("SELECT COUNT(*) FROM words WHERE status='approved'")
            diag["approved_words_count"] = int((c.fetchone() or [0])[0] or 0)
        except Exception:
            diag["approved_words_count"] = None
        try:
            c.execute("SELECT COUNT(*) FROM phrases WHERE status='approved'")
            diag["approved_phrases_count"] = int((c.fetchone() or [0])[0] or 0)
        except Exception:
            diag["approved_phrases_count"] = None

        # Conservative overwrite rule:
        # auto-reseed only when DB is clearly empty bootstrap state.
        if words_count <= 0 and phrases_count <= 0:
            diag["reason"] = "empty_base_tables"
            return diag

        # Keep if base tables contain data (even if approved counts are low/missing),
        # to avoid destructive overwrite of legitimate user data.
        diag["keep"] = True
        diag["reseed"] = False
        if (
            diag["approved_words_count"] is not None
            and diag["approved_phrases_count"] is not None
            and diag["approved_words_count"] <= 0
            and diag["approved_phrases_count"] <= 0
        ):
            diag["reason"] = "base_data_present_but_no_approved_rows"
        else:
            diag["reason"] = "base_data_present"
        return diag
    except Exception as e:
        diag["error"] = repr(e)
        diag["reason"] = "sqlite_open_or_query_failed"
        return diag
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


def _bootstrap_railway_db():
    src = BOOTSTRAP_DB_SOURCE
    dst = RAILWAY_DB_PATH

    src_diag = _inspect_sqlite_db(src)
    dst_diag_before = _inspect_sqlite_db(dst)
    print(f"[startup-db] source_diag={src_diag}")
    print(f"[startup-db] dest_diag_before={dst_diag_before}")

    copy_ran = False
    copy_error = None
    copy_reason = "kept_existing_destination_db"
    if dst_diag_before.get("reseed", True):
        copy_reason = f"reseed_destination_db reason={dst_diag_before.get('reason', 'unknown')}"
        if not src_diag.get("valid_sqlite", False):
            copy_error = f"source_not_usable reason={src_diag.get('reason', '')} error={src_diag.get('error', '')}"
        elif (int(src_diag.get("words_count") or 0) <= 0 and int(src_diag.get("phrases_count") or 0) <= 0):
            copy_error = "source_missing_base_content"
        else:
            try:
                os.makedirs(os.path.dirname(dst), exist_ok=True)
                shutil.copy2(src, dst)
                copy_ran = True
            except Exception as e:
                copy_error = repr(e)

    dst_diag_after = _inspect_sqlite_db(dst)
    print(
        f"[startup-db] copy_ran={copy_ran} copy_reason={copy_reason} copy_error={copy_error} "
        f"dest_diag_after={dst_diag_after}"
    )

    if _is_prod_runtime() and not dst_diag_after.get("exists", False):
        raise RuntimeError(
            f"Production DB bootstrap failed: {dst} is missing after startup bootstrap. "
            f"source_diag={src_diag} dest_diag_after={dst_diag_after} copy_error={copy_error}"
        )
    if _is_prod_runtime() and dst_diag_after.get("reseed", True):
        raise RuntimeError(
            f"Production DB bootstrap failed: destination DB is not usable after startup bootstrap. "
            f"dest_diag_after={dst_diag_after} source_diag={src_diag} copy_error={copy_error}"
        )


_bootstrap_railway_db()


# ------------------ APP SETUP ------------------

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev")

from datetime import timedelta

# True i produksjon/https (Render + Cloudflare). False lokalt pÃ¥ http.
IS_PROD = (
    (os.environ.get("FLASK_ENV") == "production")
    or bool(os.environ.get("RENDER"))
    or bool(os.environ.get("RAILWAY_ENVIRONMENT"))
    or bool(os.environ.get("RAILWAY_PROJECT_ID"))
)
IS_RAILWAY = bool(os.environ.get("RAILWAY_ENVIRONMENT")) or bool(os.environ.get("RAILWAY_PROJECT_ID"))

app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=IS_PROD,
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
)

# âœ… IMPORTANT for Render / reverse proxy: makes Flask understand HTTPS + correct host
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)

logging.basicConfig(level=logging.INFO)
app.logger.setLevel(logging.INFO)

@app.route("/health")
def health():
    return "ok", 200

def _pick_base_dir() -> str:
    """
    Pick persistent data root.
    Railway volume mount should be /data. Keep env override for portability.
    """
    explicit = (
        os.environ.get("PERSISTENT_DATA_DIR", "").strip()
        or os.environ.get("DATA_DIR", "").strip()
    )
    if explicit:
        return explicit
    if os.path.isdir("/data"):
        return "/data"
    if os.path.isdir("/var/data"):
        return "/var/data"
    return os.path.abspath(os.path.dirname(__file__))


# Base directory for uploads/db
BASE_DIR = _pick_base_dir()
BASE_DIR = os.path.abspath(BASE_DIR)

UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Upload limit (total request size)
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024

DEFAULT_DB = os.path.join(BASE_DIR, "gadaoromo.db")
DB_NAME = (os.environ.get("DB_PATH", "").strip() or DEFAULT_DB)
DB_DIR = os.path.dirname(DB_NAME)
if DB_DIR:
    os.makedirs(DB_DIR, exist_ok=True)
app.DB_NAME = DB_NAME
app.logger.info(
    "Startup DB resolved: db_name=%s abs=%s exists=%s size=%s bootstrap_source=%s bootstrap_source_exists=%s bootstrap_source_size=%s",
    DB_NAME,
    os.path.abspath(DB_NAME),
    os.path.isfile(DB_NAME),
    _safe_file_size(DB_NAME),
    BOOTSTRAP_DB_SOURCE,
    os.path.isfile(BOOTSTRAP_DB_SOURCE),
    _safe_file_size(BOOTSTRAP_DB_SOURCE),
)

if IS_RAILWAY and os.path.abspath(DB_NAME) != os.path.abspath(RAILWAY_DB_PATH):
    raise RuntimeError(
        f"Railway must use DB_PATH={RAILWAY_DB_PATH}, but resolved DB_NAME={DB_NAME}. "
        "Set DB_PATH=/data/gadaoromo.db or remove DB_PATH to use the default."
    )

if IS_RAILWAY:
    railway_db_diag = _inspect_sqlite_db(DB_NAME)
    app.logger.info("Railway resolved DB diagnostics: %s", railway_db_diag)
    if not railway_db_diag.get("exists", False):
        raise RuntimeError(
            f"Railway DB missing at startup: {DB_NAME}. "
            f"bootstrap_source={BOOTSTRAP_DB_SOURCE} source_exists={os.path.isfile(BOOTSTRAP_DB_SOURCE)}"
        )
    if railway_db_diag.get("reseed", True):
        raise RuntimeError(
            f"Railway DB unusable after bootstrap: {DB_NAME}. "
            f"diag={railway_db_diag}"
        )
app.logger.info(f"âœ… Using DB_NAME={DB_NAME}")

DEFAULT_REQUIRE_EXPLICIT_DB_PATH = (
    "0" if BASE_DIR in {"/data", "/var/data"} else ("1" if IS_PROD else "0")
)
REQUIRE_EXPLICIT_DB_PATH = (
    os.environ.get("REQUIRE_EXPLICIT_DB_PATH", DEFAULT_REQUIRE_EXPLICIT_DB_PATH).strip() == "1"
)
if IS_PROD and REQUIRE_EXPLICIT_DB_PATH and (not os.environ.get("DB_PATH", "").strip()):
    raise RuntimeError(
        "Production requires explicit DB_PATH to avoid source-of-truth drift. "
        "Set DB_PATH to your production database file path."
    )

APP_NAME = os.environ.get("APP_NAME", "Gadaa Dictionary")
APP_BUILD_TOKEN = (
    (os.environ.get("APP_BUILD_TOKEN") or "").strip()
    or (os.environ.get("RENDER_GIT_COMMIT") or "").strip()
    or (os.environ.get("RAILWAY_GIT_COMMIT_SHA") or "").strip()
    or (os.environ.get("RAILWAY_GIT_COMMIT") or "").strip()
    or "dev-local"
)
APP_BUILD_TOKEN = APP_BUILD_TOKEN[:16]
LEARN_TEMPLATE_VERSION = "learn_multilingual_table_v1"
AUDIO_JS_VERSION = (os.environ.get("AUDIO_JS_VERSION") or APP_BUILD_TOKEN).strip()
PWA_UI_JS_VERSION = (os.environ.get("PWA_UI_JS_VERSION") or APP_BUILD_TOKEN).strip()
SW_JS_VERSION = (os.environ.get("SW_JS_VERSION") or APP_BUILD_TOKEN).strip()
SW_CANONICAL_URL = f"/service-worker.js?v={SW_JS_VERSION}"

ADMIN_MANAGE_PASSWORD = (os.environ.get("ADMIN_MANAGE_PASSWORD") or "").strip()

# If you set WEBSITE_URL in Render env vars, we use it for sitemap/canonical.
WEBSITE_URL = os.environ.get("WEBSITE_URL", "").strip().rstrip("/")
API_URL = os.environ.get("API_URL", "").strip()

SUPPORT_MIN_NOK = int(os.environ.get("SUPPORT_MIN_NOK", "200"))

DONATE_URLS = {
    "custom": os.environ.get("STRIPE_DONATE_CUSTOM_URL", "").strip(),
}

APP_ROOT_DIR = os.path.abspath(os.path.dirname(__file__))
APP_RUNTIME = (
    "railway" if IS_RAILWAY
    else ("render" if bool(os.environ.get("RENDER")) else ("production" if IS_PROD else "local"))
)
PERSISTENT_DATA_CONFIGURED = bool(
    os.environ.get("PERSISTENT_DATA_DIR", "").strip()
    or os.environ.get("DATA_DIR", "").strip()
    or BASE_DIR in {"/data", "/var/data"}
)

def _safe_url(u: str) -> str:
    u = (u or "").strip()
    if u.startswith("https://") or u.startswith("http://"):
        return u
    return ""

DONATE_URLS = {k: _safe_url(v) for k, v in DONATE_URLS.items()}
TIMED_ROUTE_PREFIXES = ("/dashboard", "/learn", "/uploads/", "/admin")


def _should_time_route(path: str) -> bool:
    p = (path or "").strip()
    return bool(
        p == "/dashboard"
        or p == "/learn"
        or p.startswith("/uploads/")
        or p.startswith("/admin")
    )

@app.before_request
def force_primary_domain():
    if request.path.startswith("/.well-known/"):
        return None
    redirect_hosts = {
        h.strip().lower()
        for h in (os.environ.get("PRIMARY_REDIRECT_HOSTS", "gadaoromo.onrender.com") or "").split(",")
        if h.strip()
    }
    host = ((request.host or "").split(":")[0] or "").strip().lower()
    if host and (host in redirect_hosts):
        primary_base = (WEBSITE_URL or "https://gadaadictionary.com").rstrip("/")
        dest = request.full_path or request.path or "/"
        if dest.endswith("?"):
            dest = dest[:-1]
        return redirect(primary_base + dest, code=301)

    return None


@app.before_request
def mark_route_timing_start():
    try:
        req_path = request.path or ""
        if _should_time_route(req_path):
            g._route_timing_started = time.perf_counter()
    except Exception:
        pass

def _site_base_url() -> str:
    if WEBSITE_URL:
        return WEBSITE_URL.rstrip("/")
    try:
        return (request.url_root or "").rstrip("/")
    except Exception:
        return "https://gadaadictionary.com"

    
@app.context_processor
def inject_globals():
    return dict(
        APP_NAME=APP_NAME,
        SUPPORT_MIN_NOK=SUPPORT_MIN_NOK,
        DONATE_URLS=DONATE_URLS,
        WEBSITE_URL=WEBSITE_URL,
        API_URL=API_URL,
        APP_BUILD_TOKEN=APP_BUILD_TOKEN,
        LEARN_TEMPLATE_VERSION=LEARN_TEMPLATE_VERSION,
        AUDIO_JS_VERSION=AUDIO_JS_VERSION,
        PWA_UI_JS_VERSION=PWA_UI_JS_VERSION,
        SW_JS_VERSION=SW_JS_VERSION,
        SW_CANONICAL_URL=SW_CANONICAL_URL,
    )

@app.route("/debug-vars")
def debug_vars():
    return f"SUPPORT_MIN_NOK={SUPPORT_MIN_NOK}, donate_url_set={bool(DONATE_URLS.get('custom'))}"


@app.after_request
def add_security_headers(resp):
    resp.headers["X-Content-Type-Options"] = "nosniff"
    resp.headers["X-Frame-Options"] = "SAMEORIGIN"
    if resp.mimetype == "text/html":
        resp.headers.setdefault("Cache-Control", "no-cache")
        # Keep utility/private surfaces out of search indexing.
        noindex_prefixes = ("/admin", "/recorder", "/create_admin", "/api/", "/recorder/api/")
        noindex_exact = ("/offline", "/health")
        req_path = (request.path or "").strip()
        if req_path in noindex_exact or any(req_path.startswith(p) for p in noindex_prefixes):
            resp.headers["X-Robots-Tag"] = "noindex, nofollow"
    return resp


@app.after_request
def log_route_duration(resp):
    try:
        started = getattr(g, "_route_timing_started", None)
        if started is not None:
            elapsed_ms = (time.perf_counter() - float(started)) * 1000.0
            app.logger.info(
                "route_timing method=%s path=%s status=%s duration_ms=%.3f",
                request.method,
                request.path,
                int(getattr(resp, "status_code", 0) or 0),
                elapsed_ms,
            )
    except Exception:
        pass
    return resp

# ------------------ SEO: ROBOTS + SITEMAP ------------------

@app.route("/robots.txt")
def robots_txt():
    base = _site_base_url()
    lines = [
        "User-agent: *",
        "Allow: /",
        "",
        f"Sitemap: {base}/sitemap.xml",
        "",
    ]
    resp = make_response("\n".join(lines))
    resp.headers["Content-Type"] = "text/plain; charset=utf-8"
    resp.headers["Cache-Control"] = "public, max-age=3600"
    return resp

@app.route("/sitemap.xml")
def sitemap_xml():
    base = _site_base_url()
    urls = [
        ("/", "daily", "1.0"),
        ("/dictionary", "daily", "0.9"),
        ("/translate", "daily", "0.8"),
        ("/learn", "weekly", "0.8"),
    ]

    now = datetime.utcnow().strftime("%Y-%m-%d")

    xml_parts = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">',
    ]
    static_url_count = len(urls)
    emitted_word_urls = 0
    emitted_phrase_urls = 0
    fetched_word_rows = 0
    fetched_phrase_rows = 0
    for path, freq, prio in urls:
        loc = f"{base}{path}"
        xml_parts.append("<url>")
        xml_parts.append(f"<loc>{loc}</loc>")
        xml_parts.append(f"<lastmod>{now}</lastmod>")
        xml_parts.append(f"<changefreq>{freq}</changefreq>")
        xml_parts.append(f"<priority>{prio}</priority>")
        xml_parts.append("</url>")

    # --- WORD URLS ---
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("""
            SELECT DISTINCT w.id, TRIM(w.english) AS english
            FROM words w
            WHERE w.status='approved'
              AND w.english IS NOT NULL
              AND TRIM(w.english) != ''
              AND w.oromo IS NOT NULL
              AND TRIM(w.oromo) != ''
              AND (
                    EXISTS (
                        SELECT 1
                        FROM generated_translations gt
                        WHERE gt.word_id = w.id
                          AND gt.translated_text IS NOT NULL
                          AND TRIM(gt.translated_text) != ''
                    )
                    OR EXISTS (
                        SELECT 1
                        FROM audio a
                        WHERE a.entry_type='word'
                          AND a.entry_id=w.id
                          AND a.status='approved'
                          AND a.file_path IS NOT NULL
                          AND TRIM(a.file_path) != ''
                    )
              )
            ORDER BY english ASC
        """)
        rows = c.fetchall()
        fetched_word_rows = len(rows)
        if not rows:
            # Safe fallback: keep approved word pages discoverable even if enrichment tables are empty.
            c.execute("""
                SELECT DISTINCT w.id, TRIM(w.english) AS english
                FROM words w
                WHERE w.status='approved'
                  AND w.english IS NOT NULL
                  AND TRIM(w.english) != ''
                  AND w.oromo IS NOT NULL
                  AND TRIM(w.oromo) != ''
                ORDER BY english ASC
            """)
            rows = c.fetchall()
            fetched_word_rows = len(rows)

        for _wid, en in rows:
            try:
                url = f"{base}/word/{quote(en, safe='')}"
                xml_parts.append(f"""
    <url>
        <loc>{url}</loc>
        <lastmod>{now}</lastmod>
        <changefreq>weekly</changefreq>
        <priority>0.8</priority>
    </url>
""")
                emitted_word_urls += 1
            except Exception as e:
                print("skip bad word:", en, e)

        # --- PHRASE URLS ---
        c.execute("""
            SELECT DISTINCT p.id, TRIM(p.english) AS english
            FROM phrases p
            WHERE p.status='approved'
              AND p.english IS NOT NULL
              AND TRIM(p.english) != ''
              AND p.oromo IS NOT NULL
              AND TRIM(p.oromo) != ''
              AND (
                    EXISTS (
                        SELECT 1
                        FROM generated_phrase_translations gpt
                        WHERE gpt.phrase_id = p.id
                          AND gpt.translated_text IS NOT NULL
                          AND TRIM(gpt.translated_text) != ''
                    )
                    OR EXISTS (
                        SELECT 1
                        FROM audio a
                        WHERE a.entry_type='phrase'
                          AND a.entry_id=p.id
                          AND a.status='approved'
                          AND a.file_path IS NOT NULL
                          AND TRIM(a.file_path) != ''
                    )
              )
            ORDER BY english ASC
        """)
        phrase_rows = c.fetchall()
        fetched_phrase_rows = len(phrase_rows)
        if not phrase_rows:
            c.execute("""
                SELECT DISTINCT p.id, TRIM(p.english) AS english
                FROM phrases p
                WHERE p.status='approved'
                  AND p.english IS NOT NULL
                  AND TRIM(p.english) != ''
                  AND p.oromo IS NOT NULL
                  AND TRIM(p.oromo) != ''
                ORDER BY english ASC
            """)
            phrase_rows = c.fetchall()
            fetched_phrase_rows = len(phrase_rows)

        for _pid, en in phrase_rows:
            try:
                slug = make_phrase_slug(en or "")
                if not slug:
                    continue
                url = f"{base}/phrase/{quote(slug, safe='')}"
                xml_parts.append(f"""
    <url>
        <loc>{url}</loc>
        <lastmod>{now}</lastmod>
        <changefreq>weekly</changefreq>
        <priority>0.7</priority>
    </url>
""")
                emitted_phrase_urls += 1
            except Exception as e:
                print("skip bad phrase:", en, e)

    except Exception as e:
        print("sitemap content error:", e)
    finally:
        try:
            conn.close()
        except Exception:
            pass

    xml_parts.append("</urlset>")
    final_total_urls = static_url_count + emitted_word_urls + emitted_phrase_urls
    sitemap_log_line = (
        f"sitemap_xml db_path={DB_NAME} "
        f"fetched_word_rows={fetched_word_rows} "
        f"fetched_phrase_rows={fetched_phrase_rows} "
        f"emitted_word_urls={emitted_word_urls} "
        f"emitted_phrase_urls={emitted_phrase_urls} "
        f"final_total_urls={final_total_urls}"
    )
    app.logger.info(sitemap_log_line)
    logging.getLogger("gunicorn.error").info(sitemap_log_line)

    xml = "".join(xml_parts)
    return Response(xml, mimetype="application/xml")


@app.route("/.well-known/assetlinks.json")
def assetlinks():
    return send_from_directory(
        os.path.join(app.static_folder, ".well-known"),
        "assetlinks.json",
        mimetype="application/json",
    )

# ------------------ UPLOAD CONFIG (AUDIO) ------------------

# Backward-compatible flag name: true when using a mounted persistent disk path.
IS_RENDER_DISK = BASE_DIR in {"/var/data", "/data"}
IS_PERSISTENT_STORAGE = bool(PERSISTENT_DATA_CONFIGURED)
DEFAULT_UPLOAD_FOLDER = "/data/uploads" if IS_RAILWAY else os.path.join(BASE_DIR, "uploads")
UPLOAD_FOLDER = (
    os.environ.get("AUDIO_UPLOAD_DIR", "").strip()
    or os.environ.get("UPLOAD_FOLDER", "").strip()
    or DEFAULT_UPLOAD_FOLDER
)
UPLOAD_FOLDER = os.path.abspath(UPLOAD_FOLDER)
AUDIO_SOURCE_BASE_URL = (os.environ.get("AUDIO_SOURCE_BASE_URL") or "").strip().rstrip("/")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
# Static assets are always served from the app's code static folder.
# Do not derive from BASE_DIR; app static is code assets, not persistent storage.
STATIC_UPLOADS_FOLDER = os.path.join(app.static_folder, "uploads")
os.makedirs(STATIC_UPLOADS_FOLDER, exist_ok=True)

def _copy_static_uploads_to_persistent_startup():
    """
    Startup repair: copy any bundled audio files into persistent storage.
    Railway-safe behavior: skip existing files and never overwrite.
    """
    copied = 0
    scanned = 0
    skipped_existing = 0
    if not os.path.isdir(STATIC_UPLOADS_FOLDER):
        app.logger.info(
            "startup_audio_copy status=skipped reason=static_uploads_missing static_uploads=%s upload_folder=%s",
            STATIC_UPLOADS_FOLDER,
            UPLOAD_FOLDER,
        )
        return
    try:
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        for name in os.listdir(STATIC_UPLOADS_FOLDER):
            src_abs = os.path.join(STATIC_UPLOADS_FOLDER, name)
            if not os.path.isfile(src_abs):
                continue
            scanned += 1
            dst_abs = os.path.join(UPLOAD_FOLDER, os.path.basename(name))
            if os.path.isfile(dst_abs):
                skipped_existing += 1
                continue
            try:
                shutil.copy2(src_abs, dst_abs)
                copied += 1
            except Exception:
                app.logger.exception(
                    "startup_audio_copy file_copy_failed file=%s src=%s dst=%s",
                    name,
                    src_abs,
                    dst_abs,
                )
        app.logger.info(
            "startup_audio_copy status=done scanned=%s copied=%s skipped_existing=%s static_uploads=%s upload_folder=%s",
            scanned,
            copied,
            skipped_existing,
            STATIC_UPLOADS_FOLDER,
            UPLOAD_FOLDER,
        )
    except Exception:
        app.logger.exception(
            "startup_audio_copy status=failed static_uploads=%s upload_folder=%s",
            STATIC_UPLOADS_FOLDER,
            UPLOAD_FOLDER,
        )


def _try_regenerate_missing_tts_file(name: str) -> bool:
    """
    Attempt lazy regeneration for missing generated TTS assets using filename metadata.
    Returns True only when the expected file exists in UPLOAD_FOLDER after generation.
    """
    clean_name = os.path.basename(name or "")
    m = GENERATED_TTS_FILENAME_RE.match(clean_name)
    if not m:
        return False

    entry_type, entry_id_raw, lang_code, _, _ = m.groups()
    try:
        entry_id = int(entry_id_raw)
    except Exception:
        return False

    texts = _get_entry_texts_for_tts(entry_type, entry_id)
    text_value = normalize_text((texts or {}).get(lang_code, "") or "")
    if not text_value:
        app.logger.warning(
            "uploads_missing_tts_regen_skipped reason=missing_text file=%s entry_type=%s entry_id=%s lang=%s",
            clean_name,
            entry_type,
            entry_id,
            lang_code,
        )
        return False

    regen_url = normalize_text(
        _resolve_or_generate_tts_for_text(
            entry_type,
            entry_id,
            lang_code,
            text_value,
            allow_generate=True,
        )
        or ""
    )
    if not regen_url:
        app.logger.warning(
            "uploads_missing_tts_regen_failed reason=generator_returned_empty file=%s entry_type=%s entry_id=%s lang=%s",
            clean_name,
            entry_type,
            entry_id,
            lang_code,
        )
        return False

    regenerated_abs = os.path.join(UPLOAD_FOLDER, clean_name)
    if os.path.isfile(regenerated_abs):
        app.logger.info(
            "uploads_missing_tts_regen_done file=%s entry_type=%s entry_id=%s lang=%s url=%s",
            clean_name,
            entry_type,
            entry_id,
            lang_code,
            regen_url,
        )
        return True
    app.logger.warning(
        "uploads_missing_tts_regen_failed reason=file_not_materialized file=%s entry_type=%s entry_id=%s lang=%s url=%s",
        clean_name,
        entry_type,
        entry_id,
        lang_code,
        regen_url,
    )
    return False


_copy_static_uploads_to_persistent_startup()
app.logger.info(
    "Audio storage configured runtime=%s persistent_data_configured=%s is_render_disk=%s is_persistent_storage=%s upload_folder=%s static_uploads=%s",
    APP_RUNTIME,
    PERSISTENT_DATA_CONFIGURED,
    IS_RENDER_DISK,
    IS_PERSISTENT_STORAGE,
    UPLOAD_FOLDER,
    STATIC_UPLOADS_FOLDER,
)
app.logger.info(
    "Startup runtime context runtime=%s is_prod=%s is_railway=%s base_dir=%s app_root=%s db_path=%s db_exists=%s upload_exists=%s",
    APP_RUNTIME,
    IS_PROD,
    IS_RAILWAY,
    BASE_DIR,
    APP_ROOT_DIR,
    DB_NAME,
    os.path.isfile(DB_NAME),
    os.path.isdir(UPLOAD_FOLDER),
)

ALLOWED_AUDIO = {"mp3", "wav", "m4a", "webm", "ogg"}
MAX_AUDIO_MB = int(os.environ.get("MAX_AUDIO_MB", "15"))
MAX_UPLOAD_MB = int(os.environ.get("MAX_UPLOAD_MB", "100"))
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_MB * 1024 * 1024

# ------------------ PWA ROUTES ------------------

@app.route("/manifest.webmanifest")
def manifest():
    resp = make_response(
        send_from_directory(app.static_folder, "manifest.webmanifest")
    )
    resp.headers["Content-Type"] = "application/manifest+json"
    resp.headers["Cache-Control"] = "public, max-age=3600"
    return resp


@app.route("/service-worker.js")
def service_worker():
    resp = make_response(
        send_from_directory(app.static_folder, "service-worker.js")
    )
    resp.headers["Content-Type"] = "application/javascript"
    resp.headers["Service-Worker-Allowed"] = "/"
    resp.headers["Cache-Control"] = "no-cache"
    resp.headers["X-Gadaa-Build"] = APP_BUILD_TOKEN
    resp.headers["X-SW-JS-Version"] = SW_JS_VERSION
    resp.headers["X-SW-Canonical-URL"] = SW_CANONICAL_URL
    return resp

@app.route("/sw.js")
def service_worker_legacy():
    # Legacy compatibility for older clients that still request /sw.js.
    return service_worker()


@app.route("/offline")
def offline():
    return render_template("offline.html")

@app.route("/favicon.ico")
def favicon():
    return send_from_directory(
        os.path.join(app.root_path, "static", "icons"),
        "favicon.ico",
        mimetype="image/vnd.microsoft.icon",
    )

@app.errorhandler(404)
def not_found_page(_err):
    # Keep JSON/API callers from unexpectedly receiving HTML.
    if request.path.startswith("/api/") or request.path.startswith("/recorder/api/"):
        return jsonify({"ok": False, "error": "Not found"}), 404
    return render_template("404.html"), 404

# ------------------ GOOGLE VERIFICATION ------------------

@app.route("/googledba38dd4b1b65cfb.html")
def google_verification():
    resp = make_response("google-site-verification: googledba38dd4b1b65cfb.html")
    resp.headers["Content-Type"] = "text/html; charset=utf-8"
    resp.headers["Cache-Control"] = "public, max-age=3600"
    return resp


# ------------------ PUBLIC UPLOADS ROUTE (AUDIO) ------------------

@app.route("/uploads/<path:filename>")
def uploads(filename):
    safe_name = os.path.basename(filename)
    full_path = os.path.abspath(os.path.join(UPLOAD_FOLDER, safe_name))
    t0 = time.perf_counter()

    def _audio_mimetype_for_name(name: str) -> str:
        ext = (os.path.splitext(name or "")[1] or "").lower().strip(".")
        if ext == "mp3":
            return "audio/mpeg"
        if ext == "wav":
            return "audio/wav"
        if ext == "ogg":
            return "audio/ogg"
        if ext == "webm":
            return "audio/webm"
        if ext == "m4a":
            return "audio/mp4"
        return "application/octet-stream"

    def _serve_audio_abs(abs_path: str, source: str, send_start: float):
        mime = _audio_mimetype_for_name(safe_name)
        size = _safe_file_size(abs_path)
        app.logger.info(
            "uploads_timing_send_prepare file=%s source=%s abs_path=%s mime=%s size=%s elapsed_ms=%.3f",
            safe_name,
            source,
            abs_path,
            mime,
            size,
            (send_start - t0) * 1000.0,
        )
        response = send_file(abs_path, mimetype=mime, conditional=True)
        done = time.perf_counter()
        app.logger.info(
            "uploads_timing_done file=%s status=%s source=%s total_ms=%.3f send_file_ms=%.3f",
            safe_name,
            int(response.status_code or 200),
            source,
            (done - t0) * 1000.0,
            (done - send_start) * 1000.0,
        )
        return response

    app.logger.info(
        "uploads_timing_start file=%s requested=%s upload_path=%s",
        safe_name,
        filename,
        full_path,
    )

    t_exists_start = time.perf_counter()
    upload_exists = os.path.exists(full_path)
    t_exists_done = time.perf_counter()

    app.logger.info(
        "uploads_timing_exists file=%s upload_exists=%s exists_ms=%.3f elapsed_ms=%.3f",
        safe_name,
        bool(upload_exists),
        (t_exists_done - t_exists_start) * 1000.0,
        (t_exists_done - t0) * 1000.0,
    )
    app.logger.info(
        "uploads_timing_resolved file=%s resolved_upload=%s",
        safe_name,
        full_path,
    )

    if upload_exists:
        send_start = time.perf_counter()
        return _serve_audio_abs(full_path, "upload_folder", send_start)

    app.logger.warning(
        "uploads_timing_404 file=%s requested=%s reason=missing_in_upload_folder total_ms=%.3f",
        safe_name,
        filename,
        (time.perf_counter() - t0) * 1000.0,
    )
    return jsonify({"ok": False, "error": "Audio file not found", "file": safe_name}), 404


# ------------------ ADMIN IMPORT CONFIG ------------------

IMPORT_BATCH_SIZE = 100
IMPORT_MAX_WORDS = 200
MISSING_OROMO_KEY_SENTINEL = "__missing_oromo__"
LEARN_RECENT_PHRASE_LIMIT = max(
    10,
    int((os.environ.get("LEARN_RECENT_PHRASE_LIMIT") or "200").strip() or 200),
)


# ------------------ STOPWORDS ------------------

OROMO_STOP = {"fi", "kan", "inni", "isaan", "ani", "ati", "nu", "keessa", "irratti"}
EN_STOP = {"the", "is", "are", "to", "and", "of", "in", "on", "a", "an", "for", "with", "it", "this"}

# ------------------ TEXT NORMALIZATION ------------------

def normalize_text(text: str) -> str:
    t = (text or "").strip()
    t = t.replace("â€™", "'").replace("â€˜", "'").replace("`", "'")
    t = re.sub(r"\s+", " ", t).strip()
    return t


_INVALID_GENERATED_TEXT_VALUES = {
    "-",
    "—",
    "n/a",
    "na",
    "none",
    "null",
    "missing",
    "not generated yet",
}


def _is_meaningful_generated_text(text: str) -> bool:
    t = normalize_text(text or "")
    if not t:
        return False
    return t.casefold() not in _INVALID_GENERATED_TEXT_VALUES


def _is_missing_value(value: str) -> bool:
    return not normalize_text(value or "")


def normalize_tokens(text: str):
    t = normalize_text(text)
    return t.split() if t else []


def dedup_preserve_order(items):
    seen = set()
    out = []
    for x in items:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def make_search_key(text: str) -> str:
    t = (text or "").strip()
    t = t.replace("â€™", "'").replace("â€˜", "'").replace("`", "'")
    t = unicodedata.normalize("NFKC", t)
    t = t.casefold()
    t = re.sub(r"\s+", " ", t).strip()
    return t


# ------------------ COMMUNITY FILE PARSERS (NO GOOGLE) ------------------

def parse_csv_pairs_from_path(path: str):
    # prÃ¸v UTF-8 fÃ¸rst, fallback latin-1
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                out = []
                for row in reader:
                    en = normalize_text((row.get("english") or row.get("English") or "").strip())
                    om = normalize_text((row.get("oromo") or row.get("Oromo") or "").strip())
                    if en or om:
                        out.append((en, om))

            seen = set()
            final = []
            for en, om in out:
                if en and en not in seen:
                    seen.add(en)
                    final.append((en, om))
            return final

        except UnicodeDecodeError:
            continue

    raise ValueError("Could not decode CSV file.")

def parse_xlsx_pairs_from_path(path: str):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    seen = set()
    final = []

    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if not row:
            continue

        a = (row[0] if len(row) > 0 else "") or ""
        b = (row[1] if len(row) > 1 else "") or ""

        # skip header row
        if idx == 0 and str(a).strip().lower() in ("english", "en") and str(b).strip().lower() in ("oromo", "om"):
            continue

        en = normalize_text(str(a))
        om = normalize_text(str(b))

        if en and en not in seen:
            seen.add(en)
            final.append((en, om))

    return final


# ------------------ ADMIN IMPORT PARSERS (ENGLISH-ONLY) ------------------

def parse_txt_english_rows(file_bytes: bytes):
    text = file_bytes.decode("utf-8", errors="replace")
    return [line for line in text.splitlines()]


def parse_csv_english_rows(file_bytes: bytes):
    text = file_bytes.decode("utf-8", errors="replace")
    f = StringIO(text)
    reader = csv.DictReader(f)

    if not reader.fieldnames:
        return []

    english_key = None
    for k in reader.fieldnames:
        if (k or "").strip().lower() == "english":
            english_key = k
            break

    first_key = reader.fieldnames[0]

    words = []
    for row in reader:
        raw = row.get(english_key, "") if english_key else row.get(first_key, "")
        words.append(str(raw or ""))

    return words


def parse_xlsx_english_rows(file_bytes: bytes):
    wb = load_workbook(BytesIO(file_bytes))
    ws = wb.active

    words = []
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if not row:
            words.append("")
            continue
        a = (row[0] if len(row) > 0 else "") or ""

        if idx == 0 and str(a).strip().lower() in ("english", "en"):
            continue

        words.append(str(a))

    return words


def parse_csv_admin_import_rows(file_bytes: bytes):
    """
    Admin import parser that preserves optional Oromo when provided.
    Returns rows as tuples: (entry_type_hint, english_text, oromo_text)
    """
    text = file_bytes.decode("utf-8", errors="replace")
    f = StringIO(text)
    reader = csv.reader(f)
    raw_rows = [list(r or []) for r in reader]
    if not raw_rows:
        return []

    header = [normalize_text(str(x or "")) for x in (raw_rows[0] or [])]
    header_keys = [h.casefold() for h in header]

    english_idx = None
    oromo_idx = None
    for i, hk in enumerate(header_keys):
        if hk in ("english", "en") and english_idx is None:
            english_idx = i
        if hk in ("oromo", "om") and oromo_idx is None:
            oromo_idx = i

    has_named_header = (english_idx is not None) or (oromo_idx is not None)
    start_idx = 1 if has_named_header else 0
    if english_idx is None:
        english_idx = 0
    if oromo_idx is None:
        # Keep Oromo optional, but reliably pick column B when present.
        oromo_idx = 1

    rows = []
    for cells in raw_rows[start_idx:]:
        if not cells:
            rows.append(("", "", ""))
            continue
        en_raw = cells[english_idx] if english_idx < len(cells) else ""
        om_raw = cells[oromo_idx] if oromo_idx < len(cells) else ""
        en = normalize_text(str(en_raw or ""))
        om = normalize_text(str(om_raw or ""))
        entry_type_hint = "phrase" if om else ""
        rows.append((entry_type_hint, en, om))
    return rows


def parse_xlsx_admin_import_rows(file_bytes: bytes):
    """
    XLSX admin import parser with optional Oromo in second column.
    Returns rows as tuples: (entry_type_hint, english_text, oromo_text)
    """
    wb = load_workbook(BytesIO(file_bytes))
    ws = wb.active

    rows = []
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if not row:
            rows.append(("", "", ""))
            continue
        a = normalize_text(str((row[0] if len(row) > 0 else "") or ""))
        b = normalize_text(str((row[1] if len(row) > 1 else "") or ""))

        # Header row variants
        if idx == 0 and (a or "").casefold() in ("english", "en"):
            continue

        entry_type_hint = "phrase" if b else ""
        rows.append((entry_type_hint, a, b))
    return rows


def parse_txt_english(file_bytes: bytes):
    words = [normalize_text(x) for x in parse_txt_english_rows(file_bytes)]
    return dedup_preserve_order([w for w in words if w])


def parse_csv_english(file_bytes: bytes):
    words = [normalize_text(x) for x in parse_csv_english_rows(file_bytes)]
    return dedup_preserve_order([w for w in words if w])


def parse_xlsx_english(file_bytes: bytes):
    words = [normalize_text(x) for x in parse_xlsx_english_rows(file_bytes)]
    return dedup_preserve_order([w for w in words if w])


# ------------------ ADMIN + RECORDER HELPERS ------------------

def require_admin() -> bool:
    return "admin" in session


def _admin_id() -> int:
    try:
        return int(session.get("admin"))
    except Exception:
        return 0


# âœ… recorder session (password-based)
def require_recorder() -> bool:
    return bool(session.get("recorder") == 1)


def _words_table_counts():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    out = {
        "total_rows": 0,
        "non_empty_english_rows": 0,
        "approved_rows": 0,
        "approved_non_empty_english_rows": 0,
        "distinct_approved_non_empty_english_rows": 0,
    }

    c.execute("SELECT COUNT(*) FROM words")
    out["total_rows"] = int((c.fetchone() or [0])[0] or 0)

    c.execute("SELECT COUNT(*) FROM words WHERE english IS NOT NULL AND TRIM(english) != ''")
    out["non_empty_english_rows"] = int((c.fetchone() or [0])[0] or 0)

    c.execute("SELECT COUNT(*) FROM words WHERE status='approved'")
    out["approved_rows"] = int((c.fetchone() or [0])[0] or 0)

    c.execute("""
        SELECT COUNT(*)
        FROM words
        WHERE status='approved' AND english IS NOT NULL AND TRIM(english) != ''
    """)
    out["approved_non_empty_english_rows"] = int((c.fetchone() or [0])[0] or 0)

    c.execute("""
        SELECT COUNT(*)
        FROM (
            SELECT DISTINCT TRIM(english) AS english
            FROM words
            WHERE status='approved' AND english IS NOT NULL AND TRIM(english) != ''
        )
    """)
    out["distinct_approved_non_empty_english_rows"] = int((c.fetchone() or [0])[0] or 0)

    conn.close()
    return out


# ------------------ GOOGLE TRANSLATE (CLOUD v2) ------------------

def _get_google_key() -> str:
    # Prefer explicit translation key; keep GOOGLE_API_KEY as compatibility fallback.
    return (
        os.environ.get("GOOGLE_TRANSLATE_API_KEY", "").strip()
        or os.environ.get("GOOGLE_API_KEY", "").strip()
    )


def _google_key_source_name() -> str:
    if os.environ.get("GOOGLE_TRANSLATE_API_KEY", "").strip():
        return "GOOGLE_TRANSLATE_API_KEY"
    if os.environ.get("GOOGLE_API_KEY", "").strip():
        return "GOOGLE_API_KEY"
    return ""


def _log_runtime_repair_context(context: str):
    try:
        key_src = _google_key_source_name()
        app.logger.info(
            "%s runtime_context google_key_present=%s google_key_source=%s db_path=%s db_abs=%s upload_folder=%s render_disk_active=%s",
            context,
            bool(_get_google_key()),
            (key_src or "missing"),
            DB_NAME,
            os.path.abspath(DB_NAME),
            UPLOAD_FOLDER,
            IS_RENDER_DISK,
        )
    except Exception:
        app.logger.exception("%s runtime context logging failed", context)


def _provider_health_snapshot() -> dict:
    return {
        "google_configured": bool(_get_google_key()),
        "google_key_source": (_google_key_source_name() or "missing"),
        "azure_configured": bool(_get_azure_speech_key() and _get_azure_speech_region()),
        "azure_voice_map": {lc: (DEFAULT_AZURE_VOICES.get(lc) or "") for lc in ("en", "om", "am", "ar", "fr", "zh-CN")},
        "azure_retry": {
            "max_retries": int(AZURE_TTS_429_MAX_RETRIES),
            "base_backoff_ms": int(AZURE_TTS_429_BASE_BACKOFF_MS),
            "max_backoff_ms": int(AZURE_TTS_429_MAX_BACKOFF_MS),
            "jitter_ms": int(AZURE_TTS_429_JITTER_MS),
        },
        "tts_job_throttle": {
            "chunk_size": int(TTS_JOB_CHUNK_SIZE),
            "entry_delay_ms": int(TTS_JOB_ENTRY_DELAY_MS),
        },
    }


def google_translate_batch_v2(texts, target: str, source: str = "en"):
    api_key = _get_google_key()
    if not api_key:
        app.logger.error("GOOGLE_TRANSLATE_API_KEY is missing at runtime!")
        return []

    if not texts:
        return []

    try:
        translated = service_google_translate_batch(
            texts,
            target=target,
            source=source,
            api_key=api_key,
            timeout=30,
        )
        return [normalize_text(t or "") for t in translated]
    except Exception as e:
        app.logger.exception(f"Google Translate exception: {repr(e)}")
        return []

# ------------------ MULTILINGUAL TRANSLATION CONFIG ------------------

LANGUAGE_OPTIONS = {
    "om": {"label": "Oromo", "google_code": "om", "speech_code": "om-ET", "rtl": False},
    "en": {"label": "English", "google_code": "en", "speech_code": "en-US", "rtl": False},
    "am": {"label": "Amharic", "google_code": "am", "speech_code": "am-ET", "rtl": False},
    "ar": {"label": "Arabic", "google_code": "ar", "speech_code": "ar-SA", "rtl": True},
    "fr": {"label": "French", "google_code": "fr", "speech_code": "fr-FR", "rtl": False},
    "zh-CN": {"label": "Chinese", "google_code": "zh-CN", "speech_code": "zh-CN", "rtl": False},
}
EXTRA_GENERATED_LANGS = ("am", "ar", "zh-CN", "fr")
LEARN_TTS_LANGS = ("en", "am", "ar", "fr", "zh-CN", "om")
AZURE_TTS_PROVIDER = "azure_speech"
AZURE_BLOB_PROVIDER = "azure_blob"

# Env vars (production):
# - GOOGLE_TRANSLATE_API_KEY
# - AZURE_SPEECH_KEY
# - AZURE_SPEECH_REGION
# Optional per-language voice overrides:
# - AZURE_VOICE_EN, AZURE_VOICE_AM, AZURE_VOICE_AR, AZURE_VOICE_FR, AZURE_VOICE_ZH_CN
DEFAULT_AZURE_VOICES = {
    "en": os.environ.get("AZURE_VOICE_EN", "en-US-JennyNeural").strip(),
    "am": os.environ.get("AZURE_VOICE_AM", "am-ET-MekdesNeural").strip(),
    "ar": os.environ.get("AZURE_VOICE_AR", "ar-SA-ZariyahNeural").strip(),
    "fr": os.environ.get("AZURE_VOICE_FR", "fr-FR-DeniseNeural").strip(),
    "zh-CN": os.environ.get("AZURE_VOICE_ZH_CN", "zh-CN-XiaoxiaoNeural").strip(),
    # Optional Oromo voice (if Azure account provides one).
    "om": (os.environ.get("AZURE_VOICE_OM", "").strip() or os.environ.get("AZURE_VOICE_OROMO", "").strip()),
}
_phrase_tts_voice_map_logged = False
app.logger.info(
    "Startup Azure voice map provider=%s voices=%s azure_speech_key_set=%s azure_speech_region_set=%s",
    AZURE_TTS_PROVIDER,
    {lc: (DEFAULT_AZURE_VOICES.get(lc) or "") for lc in ("en", "om", "am", "ar", "fr", "zh-CN")},
    bool((os.environ.get("AZURE_SPEECH_KEY") or "").strip()),
    bool((os.environ.get("AZURE_SPEECH_REGION") or "").strip()),
)
if not (DEFAULT_AZURE_VOICES.get("om") or "").strip():
    app.logger.warning(
        "Azure Oromo TTS voice is not configured (AZURE_VOICE_OM/AZURE_VOICE_OROMO). "
        "Oromo phrase/word TTS generation will be skipped with reason=missing_voice."
    )

LEARN_TTS_LAZY_WARMUP = (os.environ.get("LEARN_TTS_LAZY_WARMUP", "0").strip() == "1")
try:
    _learn_tts_lazy_max_raw = int((os.environ.get("LEARN_TTS_LAZY_MAX_ENTRIES", "6") or "6").strip())
except Exception:
    _learn_tts_lazy_max_raw = 6
LEARN_TTS_LAZY_MAX_ENTRIES = max(0, min(_learn_tts_lazy_max_raw, 50))
TTS_GENERATE_ON_LOOKUP = (os.environ.get("TTS_GENERATE_ON_LOOKUP", "0").strip() == "1")
TTS_GENERATE_ON_IMPORT = (os.environ.get("TTS_GENERATE_ON_IMPORT", "1").strip() == "1")
TTS_JOB_CHUNK_SIZE = max(1, min(int((os.environ.get("TTS_JOB_CHUNK_SIZE") or "10").strip() or 10), 200))
TTS_JOB_ENTRY_DELAY_MS = max(0, min(int((os.environ.get("TTS_JOB_ENTRY_DELAY_MS") or "2000").strip() or 2000), 15000))
TTS_JOB_LANGUAGE_DELAY_MS = max(0, min(int((os.environ.get("TTS_JOB_LANGUAGE_DELAY_MS") or "700").strip() or 700), 5000))
AZURE_TTS_429_MAX_RETRIES = max(0, min(int((os.environ.get("AZURE_TTS_429_MAX_RETRIES") or "2").strip() or 2), 6))
AZURE_TTS_429_BASE_BACKOFF_MS = max(50, min(int((os.environ.get("AZURE_TTS_429_BASE_BACKOFF_MS") or "500").strip() or 500), 10000))
AZURE_TTS_429_MAX_BACKOFF_MS = max(AZURE_TTS_429_BASE_BACKOFF_MS, min(int((os.environ.get("AZURE_TTS_429_MAX_BACKOFF_MS") or "4000").strip() or 4000), 60000))
AZURE_TTS_429_JITTER_MS = max(0, min(int((os.environ.get("AZURE_TTS_429_JITTER_MS") or "250").strip() or 250), 5000))
app.logger.info(
    "Startup provider settings google_configured=%s azure_configured=%s azure_retry={max_retries:%s,base_backoff_ms:%s,max_backoff_ms:%s,jitter_ms:%s} tts_job_throttle={chunk_size:%s,entry_delay_ms:%s}",
    bool(_get_google_key()),
    bool((os.environ.get("AZURE_SPEECH_KEY") or "").strip() and (os.environ.get("AZURE_SPEECH_REGION") or "").strip()),
    int(AZURE_TTS_429_MAX_RETRIES),
    int(AZURE_TTS_429_BASE_BACKOFF_MS),
    int(AZURE_TTS_429_MAX_BACKOFF_MS),
    int(AZURE_TTS_429_JITTER_MS),
    int(TTS_JOB_CHUNK_SIZE),
    int(TTS_JOB_ENTRY_DELAY_MS),
)

AZURE_BLOB_CONNECTION_STRING = (os.environ.get("AZURE_BLOB_CONNECTION_STRING") or "").strip()
AZURE_BLOB_CONTAINER = (os.environ.get("AZURE_BLOB_CONTAINER") or "").strip()
AZURE_BLOB_PREFIX = (os.environ.get("AZURE_BLOB_PREFIX") or "tts").strip().strip("/")
REQUIRE_BLOB_FOR_GENERATED_TTS = (os.environ.get("REQUIRE_BLOB_FOR_GENERATED_TTS", "1" if IS_PROD else "0").strip() == "1")
GENERATED_TTS_FILENAME_RE = re.compile(
    r"^tts_(word|phrase)_(\d+)_([A-Za-z0-9-]+)_([0-9a-f]{12})_(.+)\.mp3$"
)
RELAXED_TTS_FILENAME_RE = re.compile(
    r"^tts_(word|phrase)_(\d+)_([A-Za-z0-9-]+)_(.+)\.(mp3|wav|ogg|webm|m4a)$",
    re.IGNORECASE,
)

_blob_client_cache = None
_blob_client_error_logged = False

def _is_supported_lang(lang_code: str) -> bool:
    return lang_code in LANGUAGE_OPTIONS


def _google_lang_code(lang_code: str) -> str:
    cfg = LANGUAGE_OPTIONS.get(lang_code, {})
    return cfg.get("google_code", lang_code)


def _speech_lang_code(lang_code: str) -> str:
    cfg = LANGUAGE_OPTIONS.get(lang_code, {})
    return cfg.get("speech_code", "en-US")


def _is_rtl_lang(lang_code: str) -> bool:
    cfg = LANGUAGE_OPTIONS.get(lang_code, {})
    return bool(cfg.get("rtl"))


def _upsert_pending_word_base(conn, english_text: str, oromo_text: str, status: str = "pending"):
    """
    Insert a pending base word or safely complete an existing partial row.
    Returns: (word_id, inserted_new, repaired_existing_partial)
    """
    en = normalize_text(english_text or "")
    om = normalize_text(oromo_text or "")
    en_key = make_search_key(_strip_edge_punct(en))
    om_key = make_search_key(_strip_edge_punct(om))

    if not en or not om or not en_key or not om_key:
        return None, False, False

    c = conn.cursor()
    c.execute("""
        SELECT id, english, oromo
        FROM words
        WHERE english_key=? OR oromo_key=? OR english=? OR oromo=?
        LIMIT 1
    """, (en_key, om_key, en, om))
    row = c.fetchone()

    if row:
        wid, existing_en, existing_om = row
        existing_en_norm = normalize_text(existing_en or "")
        existing_om_norm = normalize_text(existing_om or "")

        merged_en = existing_en_norm or en
        merged_om = existing_om_norm or om
        merged_en_key = make_search_key(_strip_edge_punct(merged_en))
        merged_om_key = make_search_key(_strip_edge_punct(merged_om))

        repaired = (merged_en != existing_en_norm) or (merged_om != existing_om_norm)
        if repaired:
            c.execute(
                "UPDATE words SET english=?, oromo=?, english_key=?, oromo_key=? WHERE id=?",
                (merged_en, merged_om, merged_en_key, merged_om_key, wid),
            )
        return wid, False, repaired

    safe_status = status if status in ("pending", "approved") else "pending"
    c.execute(
        "INSERT INTO words (english, oromo, english_key, oromo_key, status) VALUES (?, ?, ?, ?, ?)",
        (en, om, en_key, om_key, safe_status),
    )
    return c.lastrowid, True, False


def _cache_extra_translations_for_words(word_items):
    """
    Best-effort cache warmup for extra languages.
    word_items: list[(word_id, english_text)]
    """
    cached_count, _ = ensure_missing_generated_translations_for_words(
        word_items,
        langs=EXTRA_GENERATED_LANGS,
        chunk_size=IMPORT_BATCH_SIZE,
        log_context="cache_warmup",
    )
    return cached_count


def ensure_missing_generated_translations_for_words(
    word_items,
    langs=None,
    chunk_size: int = None,
    log_context: str = "generated_backfill",
    cancel_check=None,
):
    """
    Robust best-effort generated translation backfill.
    word_items: list[(word_id, english_text)]
    Returns: (saved_count, stats_by_lang)
    """
    if not word_items:
        return 0, {}

    langs = tuple(langs or EXTRA_GENERATED_LANGS)
    if not langs:
        return 0, {}

    safe_chunk = int(chunk_size or IMPORT_BATCH_SIZE or 50)
    if safe_chunk < 1:
        safe_chunk = 50

    unique_items = []
    seen_ids = set()
    for wid, en in word_items:
        wid_int = int(wid or 0)
        en_norm = normalize_text(en or "")
        if not wid_int or not en_norm or wid_int in seen_ids:
            continue
        seen_ids.add(wid_int)
        unique_items.append((wid_int, en_norm))

    if not unique_items:
        return 0, {}

    total_saved = 0
    stats = {}
    canceled = False

    for lang in langs:
        lang_stats = {
            "items_seen": len(unique_items),
            "attempted": len(unique_items),
            "already_cached": 0,
            "invalid_cached_treated_missing": 0,
            "source_changed_treated_missing": 0,
            "missing_before": 0,
            "saved": 0,
            "empty_results": 0,
            "batch_mismatch_fallback": 0,
            "provider_errors": 0,
            "google_empty_result": 0,
            "google_request_failures": 0,
            "google_provider_errors": 0,
            "failed_db_write": 0,
        }
        stats[lang] = lang_stats

        missing_pairs = []
        cache_skip_log_count = 0
        invalid_cache_log_count = 0
        processed_new_log_count = 0
        for wid, en in unique_items:
            if callable(cancel_check) and bool(cancel_check()):
                canceled = True
                break
            source_hash = _text_hash(en)
            cached = _get_cached_generated_translation(wid, lang)
            if cached:
                cached_text = normalize_text((cached or [""])[0] or "")
                cached_hash = normalize_text((cached or ["", "", ""])[2] or "")
                if _is_meaningful_generated_text(cached_text) and cached_hash and cached_hash == source_hash:
                    lang_stats["already_cached"] += 1
                    if cache_skip_log_count < 5:
                        app.logger.info(
                            "%s skipped_existing type=generated_translation entry_type=word entry_id=%s lang=%s value=%r",
                            log_context,
                            wid,
                            lang,
                            cached_text[:120],
                        )
                        cache_skip_log_count += 1
                    continue
                if not _is_meaningful_generated_text(cached_text):
                    lang_stats["invalid_cached_treated_missing"] += 1
                else:
                    lang_stats["source_changed_treated_missing"] += 1
                if invalid_cache_log_count < 5:
                    app.logger.info(
                        "generated cache invalid_or_stale; regenerating word_id=%s lang=%s value=%r cached_hash=%r source_hash=%r",
                        wid,
                        lang,
                        cached_text[:120],
                        cached_hash[:12],
                        source_hash[:12],
                    )
                    invalid_cache_log_count += 1
            missing_pairs.append((wid, en))
        if canceled:
            break

        lang_stats["missing_before"] = len(missing_pairs)
        if not missing_pairs:
            continue

        target_code = _google_lang_code(lang)

        for i in range(0, len(missing_pairs), safe_chunk):
            chunk_pairs = missing_pairs[i:i + safe_chunk]
            if not chunk_pairs:
                continue

            used_batch = False
            try:
                english_batch = [en for _wid, en in chunk_pairs]
                translated_list = google_translate_batch_v2(
                    english_batch,
                    target=target_code,
                    source="en",
                )
                if translated_list and len(translated_list) == len(chunk_pairs):
                    used_batch = True
                    for (wid, _en), translated in zip(chunk_pairs, translated_list):
                        if callable(cancel_check) and bool(cancel_check()):
                            canceled = True
                            break
                        translated_text = normalize_text(translated or "")
                        if not _is_meaningful_generated_text(translated_text):
                            lang_stats["empty_results"] += 1
                            continue
                        write_ok = _save_generated_translation(
                            wid,
                            lang,
                            translated_text,
                            source_text=_en,
                            provider="google_translate_v2",
                            tts_audio_url=None,
                        )
                        if write_ok:
                            lang_stats["saved"] += 1
                            total_saved += 1
                        else:
                            lang_stats["failed_db_write"] += 1
                        if processed_new_log_count < 5:
                            app.logger.info(
                                "%s processed_new type=generated_translation entry_type=word entry_id=%s lang=%s",
                                log_context,
                                wid,
                                lang,
                            )
                            processed_new_log_count += 1
                    if canceled:
                        break
                else:
                    lang_stats["batch_mismatch_fallback"] += len(chunk_pairs)
                    lang_stats["google_request_failures"] += len(chunk_pairs)
                    if not translated_list:
                        lang_stats["google_empty_result"] += len(chunk_pairs)
                        lang_stats["provider_errors"] += len(chunk_pairs)
                        lang_stats["google_provider_errors"] += len(chunk_pairs)
            except Exception:
                lang_stats["provider_errors"] += 1
                lang_stats["google_provider_errors"] += 1
                lang_stats["google_request_failures"] += len(chunk_pairs)
                app.logger.exception(
                    "%s batch_translate_failed entry_type=word lang=%s batch_size=%s",
                    log_context,
                    lang,
                    len(chunk_pairs),
                )

            if used_batch:
                continue

            # Per-item fallback avoids all-or-nothing loss when batch fails/mismatches.
            for wid, en in chunk_pairs:
                if callable(cancel_check) and bool(cancel_check()):
                    canceled = True
                    break
                try:
                    translated = google_translate_text_v2(en, target=target_code, source="en")
                    translated_text = normalize_text(translated or "")
                    if not _is_meaningful_generated_text(translated_text):
                        lang_stats["empty_results"] += 1
                        lang_stats["google_empty_result"] += 1
                        lang_stats["google_request_failures"] += 1
                        lang_stats["provider_errors"] += 1
                        lang_stats["google_provider_errors"] += 1
                        continue
                    write_ok = _save_generated_translation(
                        wid,
                        lang,
                        translated_text,
                        source_text=en,
                        provider="google_translate_v2",
                        tts_audio_url=None,
                    )
                    if write_ok:
                        lang_stats["saved"] += 1
                        total_saved += 1
                    else:
                        lang_stats["failed_db_write"] += 1
                    if processed_new_log_count < 5:
                        app.logger.info(
                            "%s processed_new type=generated_translation entry_type=word entry_id=%s lang=%s",
                            log_context,
                            wid,
                            lang,
                        )
                        processed_new_log_count += 1
                except Exception:
                    lang_stats["provider_errors"] += 1
                    lang_stats["google_provider_errors"] += 1
                    lang_stats["google_request_failures"] += 1
                    app.logger.exception(
                        "%s fallback_translate_failed entry_type=word entry_id=%s lang=%s",
                        log_context,
                        wid,
                        lang,
                    )
            if canceled:
                break
        if canceled:
            break

    try:
        compact = {
            lang: {
                "missing": st["missing_before"],
                "saved": st["saved"],
                "cached": st["already_cached"],
                "invalid_cached": st["invalid_cached_treated_missing"],
                "empty": st["empty_results"],
                "fallback": st["batch_mismatch_fallback"],
                "errors": st["provider_errors"],
                "google_empty_result": st.get("google_empty_result", 0),
                "google_request_failures": st.get("google_request_failures", 0),
                "google_provider_errors": st.get("google_provider_errors", 0),
            }
            for lang, st in stats.items()
        }
        app.logger.info(
            "generated backfill summary context=%s total_saved=%s details=%s",
            log_context,
            total_saved,
            compact,
        )
    except Exception:
        pass

    if canceled:
        stats["__meta__"] = {"canceled": True}

    return total_saved, stats


def google_translate_text_v2(text: str, target: str, source: str = "en") -> str:
    t = normalize_text(text)
    if not t:
        return ""
    api_key = _get_google_key()
    if not api_key:
        return ""
    try:
        out = service_google_translate_text(
            t,
            target=target,
            source=source,
            api_key=api_key,
            timeout=30,
        )
        return normalize_text(out or "")
    except Exception:
        return ""


# ------------------ AUDIO HELPERS ------------------

def allowed_audio(filename: str) -> bool:
    if not filename or "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[1].lower()
    return ext in ALLOWED_AUDIO


def _is_remote_audio_ref(file_path: str) -> bool:
    fp = (file_path or "").strip().lower()
    return fp.startswith("https://") or fp.startswith("http://")


def _canonical_tts_lang_code(lang_code: str) -> str:
    raw = normalize_text(lang_code or "")
    if not raw:
        return ""
    key = raw.replace("_", "-").strip().casefold()
    if key in ("en", "english", "en-us", "en-gb"):
        return "en"
    if key in ("om", "or", "oromo", "om-et"):
        return "om"
    if key in ("am", "amharic", "am-et"):
        return "am"
    if key in ("ar", "arabic", "ar-sa"):
        return "ar"
    if key in ("fr", "french", "fr-fr"):
        return "fr"
    if key in ("zh", "zh-cn", "zh-hans", "chinese", "chinese-simplified", "zh-sg"):
        return "zh-CN"
    return raw


def _normalized_audio_basename(ref: str) -> str:
    fp = (ref or "").replace("\\", "/").strip()
    if not fp:
        return ""

    while fp.startswith("./"):
        fp = fp[2:]
    while fp.startswith("/"):
        fp = fp[1:]

    if fp.startswith("/data/uploads"):
        fp = fp[len("/data/uploads"):]

    if fp.startswith("static/uploads/"):
        fp = fp[len("static/uploads/"):]

    if fp.startswith("uploads/"):
        fp = fp[len("uploads/"):]

    return os.path.basename(fp)


def _canonical_local_audio_ref(file_path: str) -> str:
    fp = normalize_text((file_path or "").replace("\\", "/"))
    if not fp:
        return ""
    if _is_remote_audio_ref(fp):
        return fp
    name = _normalized_audio_basename(fp)
    if not name:
        return ""
    return f"uploads/{name}"


def _tts_ref_matches_text_hash(file_path: str, text_value: str) -> bool:
    """
    Verify local generated TTS reference hash segment matches the provided text hash.
    Non-generated/local-unknown references return False.
    """
    fp = normalize_text((file_path or "").replace("\\", "/"))
    txt = normalize_text(text_value or "")
    if (not fp) or (not txt) or _is_remote_audio_ref(fp):
        return False
    name = os.path.basename(fp)
    m = GENERATED_TTS_FILENAME_RE.match(name)
    if not m:
        return False
    hash12 = normalize_text((m.groups()[3] if m.groups() else "") or "")
    return bool(hash12 and hash12 == _text_hash(txt)[:12])


def _parse_tts_semantic_parts(file_name: str):
    name = os.path.basename((file_name or "").replace("\\", "/").strip())
    if not name:
        return None
    m = RELAXED_TTS_FILENAME_RE.match(name)
    if not m:
        return None
    kind, entry_id_raw, lang_raw, _tail, _ext = m.groups()
    try:
        entry_id = int(entry_id_raw)
    except Exception:
        return None
    lang = _canonical_tts_lang_code(lang_raw or "")
    if not lang:
        return None
    return {
        "kind": normalize_text(kind or ""),
        "entry_id": entry_id,
        "lang": lang,
    }


def _collect_semantic_tts_candidates(requested_parts: dict):
    candidates = []
    if not requested_parts:
        return candidates
    dirs = [("upload", UPLOAD_FOLDER), ("static", STATIC_UPLOADS_FOLDER)]
    for source, folder in dirs:
        try:
            names = os.listdir(folder)
        except Exception:
            continue
        for name in names:
            abs_path = os.path.join(folder, name)
            if not os.path.isfile(abs_path):
                continue
            parts = _parse_tts_semantic_parts(name)
            if not parts:
                continue
            if parts.get("kind") != requested_parts.get("kind"):
                continue
            if int(parts.get("entry_id") or 0) != int(requested_parts.get("entry_id") or 0):
                continue
            if _canonical_tts_lang_code(parts.get("lang") or "") != _canonical_tts_lang_code(requested_parts.get("lang") or ""):
                continue
            ext = (os.path.splitext(name)[1] or "").lower().strip(".")
            try:
                mtime = float(os.path.getmtime(abs_path) or 0.0)
            except Exception:
                mtime = 0.0
            candidates.append(
                {
                    "source": source,
                    "folder": folder,
                    "name": name,
                    "path": abs_path,
                    "parts": parts,
                    "ext": ext,
                    "mtime": mtime,
                }
            )
    return candidates


def _rank_semantic_tts_candidates(candidates, requested_name: str):
    if not candidates:
        return None, "no_semantic_candidates"
    req_ext = (os.path.splitext(requested_name or "")[1] or "").lower().strip(".")

    def _score(c):
        source_score = 0 if c.get("source") == "upload" else 1
        ext_score = 0 if (req_ext and (c.get("ext") == req_ext)) else 1
        mtime_score = -float(c.get("mtime") or 0.0)
        name_score = normalize_text(c.get("name") or "")
        return (source_score, ext_score, mtime_score, name_score)

    ranked = sorted(candidates, key=_score)
    return ranked[0], (
        f"ranked_semantic_match source={ranked[0].get('source')} "
        f"req_ext={req_ext or 'none'} candidates={len(candidates)} "
        "rank=[source(upload-first),ext(match-first),mtime(desc),name(asc)]"
    )


def resolve_existing_audio_file(requested_ref: str, return_meta: bool = False):
    """
    Resolve local audio safely:
    A) exact basename in uploads/static
    B) normalized exact basename in uploads/static
    C) semantic TTS match by kind+entry_id+lang
    D) deterministic ranking when multiple semantic candidates exist
    """
    req_raw = normalize_text(requested_ref or "")
    req_name = os.path.basename((req_raw or "").replace("\\", "/"))
    normalized_name = _normalized_audio_basename(req_raw)
    out = {
        "requested_ref": req_raw,
        "requested_name": req_name,
        "normalized_name": normalized_name,
        "path": "",
        "strategy": "",
        "reason": "",
        "candidate_count": 0,
    }

    def _try_exact(name: str, strategy_label: str):
        if not name:
            return ""
        upload_abs = os.path.join(UPLOAD_FOLDER, name)
        if os.path.isfile(upload_abs):
            out["path"] = upload_abs
            out["strategy"] = strategy_label
            out["reason"] = "found_in_upload_folder"
            return upload_abs
        static_abs = os.path.join(STATIC_UPLOADS_FOLDER, name)
        if os.path.isfile(static_abs):
            out["path"] = static_abs
            out["strategy"] = strategy_label
            out["reason"] = "found_in_static_uploads"
            return static_abs
        return ""

    # Step A: exact basename match.
    if _try_exact(req_name, "exact_basename"):
        if return_meta:
            return out
        return out.get("path", "")

    # Step B: normalized exact basename.
    if normalized_name and normalized_name != req_name:
        if _try_exact(normalized_name, "normalized_basename"):
            if return_meta:
                return out
            return out.get("path", "")

    # Step C/D: semantic TTS match + deterministic ranking.
    requested_parts = _parse_tts_semantic_parts(req_name) or _parse_tts_semantic_parts(normalized_name)
    if requested_parts:
        candidates = _collect_semantic_tts_candidates(requested_parts)
        out["candidate_count"] = len(candidates)
        best, reason = _rank_semantic_tts_candidates(candidates, req_name or normalized_name)
        if best and best.get("path"):
            out["path"] = best.get("path", "")
            out["strategy"] = "semantic_ranked_tts"
            out["reason"] = reason
            if return_meta:
                return out
            return out.get("path", "")
        out["strategy"] = "semantic_ranked_tts"
        out["reason"] = reason
    else:
        out["strategy"] = "semantic_ranked_tts"
        out["reason"] = "requested_name_not_semantic_tts"

    if return_meta:
        return out
    return ""


def _has_usable_audio_ref(file_path: str) -> bool:
    if _is_remote_audio_ref(file_path):
        return True
    abs_path = _audio_abs_path(file_path)
    if not (abs_path and os.path.isfile(abs_path)):
        return False
    try:
        return bool(os.path.getsize(abs_path) > 0)
    except Exception:
        return False


def _azure_blob_enabled() -> bool:
    return bool(AZURE_BLOB_CONNECTION_STRING and AZURE_BLOB_CONTAINER and BlobServiceClient)


def _blob_key_for_file(file_name: str) -> str:
    if AZURE_BLOB_PREFIX:
        return f"{AZURE_BLOB_PREFIX}/{file_name}"
    return file_name


def _get_blob_container_client():
    global _blob_client_cache, _blob_client_error_logged
    if _blob_client_cache is not None:
        return _blob_client_cache
    if not _azure_blob_enabled():
        if (not _blob_client_error_logged) and (AZURE_BLOB_CONNECTION_STRING or AZURE_BLOB_CONTAINER):
            app.logger.warning(
                "Azure Blob TTS disabled: missing dependency or incomplete config "
                "(AZURE_BLOB_CONNECTION_STRING / AZURE_BLOB_CONTAINER)."
            )
            _blob_client_error_logged = True
        return None
    try:
        svc = BlobServiceClient.from_connection_string(AZURE_BLOB_CONNECTION_STRING)
        container = svc.get_container_client(AZURE_BLOB_CONTAINER)
        try:
            container.create_container()
        except Exception:
            pass
        _blob_client_cache = container
        return _blob_client_cache
    except Exception:
        if not _blob_client_error_logged:
            app.logger.exception("Failed to initialize Azure Blob client for TTS storage.")
            _blob_client_error_logged = True
        return None


def _upload_tts_bytes_to_blob(file_name: str, audio_bytes: bytes) -> str:
    container = _get_blob_container_client()
    if not container or (not audio_bytes):
        return ""
    blob_name = _blob_key_for_file(file_name)
    blob_client = container.get_blob_client(blob_name)
    content_settings = ContentSettings(content_type="audio/mpeg") if ContentSettings else None
    try:
        blob_client.upload_blob(audio_bytes, overwrite=False, content_settings=content_settings)
    except ResourceExistsError:
        pass
    except Exception:
        app.logger.exception("Failed to upload TTS audio to Azure Blob: %s", blob_name)
        return ""
    try:
        return blob_client.url or ""
    except Exception:
        return ""


def _maybe_promote_tts_to_persistent(name: str, src_abs: str):
    if (not IS_RENDER_DISK) or (not (name or "").startswith("tts_")):
        return
    _maybe_promote_audio_to_persistent(name, src_abs)


def _maybe_promote_audio_to_persistent(name: str, src_abs: str):
    if not IS_RENDER_DISK:
        return
    dst_abs = os.path.join(UPLOAD_FOLDER, os.path.basename(name or ""))
    if (not dst_abs) or os.path.isfile(dst_abs) or (not os.path.isfile(src_abs)):
        return
    try:
        shutil.copy2(src_abs, dst_abs)
        app.logger.info("Promoted audio file to persistent storage: %s", name)
    except Exception:
        app.logger.exception("Failed promoting audio file to persistent storage: %s", name)


def _public_audio_url(file_path: str) -> str:
    """
    DB may store a local relative path ('uploads/xyz.webm') or a full blob URL.
    Returns a browser-usable URL.
    """
    fp = (file_path or "").replace("\\", "/").strip()
    if not fp:
        return ""
    if _is_remote_audio_ref(fp):
        return fp
    name = _normalized_audio_basename(fp) or os.path.basename(fp)

    # Generated Azure TTS assets may exist in either uploads root or static/uploads,
    # depending on where the job ran. Resolve to whichever real file exists.
    if name.startswith("tts_"):
        uploads_abs = os.path.join(UPLOAD_FOLDER, name)
        static_abs = os.path.join(STATIC_UPLOADS_FOLDER, name)
        if os.path.isfile(static_abs):
            _maybe_promote_tts_to_persistent(name, static_abs)
        if os.path.isfile(uploads_abs):
            return "/uploads/" + name
        # Route through /uploads even for static-backed legacy files; the route
        # already falls back to static/uploads and works consistently on Render.
        return "/uploads/" + name

    if fp.startswith("uploads/") or fp.startswith("/uploads/") or fp.startswith("static/uploads/") or fp.startswith("/static/uploads/"):
        return "/uploads/" + name
    return "/uploads/" + name


def _normalize_cached_tts_url(tts_url: str) -> str:
    """
    Canonicalize local TTS paths to /uploads/<name> so serving goes through one
    route with Render-safe fallbacks.
    """
    u = normalize_text(tts_url or "")
    if not u:
        return ""
    if _is_remote_audio_ref(u):
        return u
    name = os.path.basename(u.replace("\\", "/"))
    if name.startswith("tts_"):
        # Ensure stale /static/uploads URLs and relative paths converge.
        if _has_usable_audio_ref(f"uploads/{name}"):
            return "/uploads/" + name
        return ""
    if u.startswith("/uploads/"):
        return u
    if u.startswith("uploads/"):
        return "/" + u
    return u


def _audio_abs_path(file_path: str) -> str:
    fp = (file_path or "").replace("\\", "/").strip()
    if not fp:
        return ""
    if _is_remote_audio_ref(fp):
        return ""
    resolved = resolve_existing_audio_file(fp, return_meta=True)
    resolved_path = normalize_text((resolved or {}).get("path", "") or "")
    if resolved_path and os.path.isfile(resolved_path):
        if (resolved.get("strategy") == "semantic_ranked_tts") and resolved.get("candidate_count", 0) > 1:
            app.logger.info(
                "audio_abs_semantic_resolution requested=%s resolved=%s reason=%s candidates=%s",
                fp,
                resolved_path,
                resolved.get("reason", ""),
                resolved.get("candidate_count", 0),
            )
        if os.path.abspath(os.path.dirname(resolved_path)) == os.path.abspath(STATIC_UPLOADS_FOLDER):
            _maybe_promote_audio_to_persistent(os.path.basename(resolved_path), resolved_path)
            promoted = os.path.join(UPLOAD_FOLDER, os.path.basename(resolved_path))
            if os.path.isfile(promoted):
                return promoted
        return resolved_path
    name = _normalized_audio_basename(fp)
    uploads_abs = os.path.join(UPLOAD_FOLDER, name)
    static_abs = os.path.join(STATIC_UPLOADS_FOLDER, name)
    if os.path.isfile(static_abs):
        _maybe_promote_audio_to_persistent(name, static_abs)
    if os.path.isfile(uploads_abs):
        return uploads_abs
    if os.path.isfile(static_abs):
        return static_abs
    return uploads_abs


def _log_db_context(where: str):
    try:
        app.logger.info(
            "DB context=%s db_name=%s abs_db=%s exists=%s",
            where,
            DB_NAME,
            os.path.abspath(DB_NAME),
            os.path.isfile(DB_NAME),
        )
    except Exception:
        app.logger.exception("DB context logging failed for %s", where)


def _sql_count(c, query: str, params=()):
    c.execute(query, params)
    return int((c.fetchone() or [0])[0] or 0)


def _collect_db_diagnostics():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    out = {
        "runtime": APP_RUNTIME,
        "base_dir": BASE_DIR,
        "persistent_data_configured": bool(PERSISTENT_DATA_CONFIGURED),
        "is_persistent_storage": bool(IS_PERSISTENT_STORAGE),
        "db_path": DB_NAME,
        "db_abs": os.path.abspath(DB_NAME),
        "db_exists": os.path.isfile(DB_NAME),
        "upload_folder": UPLOAD_FOLDER,
        "is_render_disk": bool(IS_RENDER_DISK),
        "voice_map": {lc: (DEFAULT_AZURE_VOICES.get(lc) or "") for lc in ("en", "om", "am", "ar", "fr", "zh-CN")},
        "require_explicit_db_path": bool(REQUIRE_EXPLICIT_DB_PATH),
        "require_blob_for_generated_tts": bool(REQUIRE_BLOB_FOR_GENERATED_TTS),
        "azure_blob_enabled": bool(_azure_blob_enabled()),
        "tables": {},
        "generated_tts_storage": {},
    }

    table_counts = {
        "words": "SELECT COUNT(*) FROM words",
        "phrases": "SELECT COUNT(*) FROM phrases",
        "generated_translations": "SELECT COUNT(*) FROM generated_translations",
        "generated_phrase_translations": "SELECT COUNT(*) FROM generated_phrase_translations",
        "generated_tts_audio": "SELECT COUNT(*) FROM generated_tts_audio",
        "audio": "SELECT COUNT(*) FROM audio",
        "post_import_jobs": "SELECT COUNT(*) FROM post_import_jobs",
    }
    for table_name, sql in table_counts.items():
        try:
            out["tables"][table_name] = _sql_count(c, sql)
        except Exception:
            out["tables"][table_name] = -1

    out["generated_tts_storage"] = {
        "blob_url_rows": _sql_count(
            c,
            "SELECT COUNT(*) FROM generated_tts_audio WHERE file_path LIKE 'https://%' OR file_path LIKE 'http://%'",
        ),
        "local_path_rows": _sql_count(
            c,
            "SELECT COUNT(*) FROM generated_tts_audio WHERE file_path LIKE 'uploads/%' OR file_path LIKE '/uploads/%'",
        ),
        "empty_rows": _sql_count(
            c,
            "SELECT COUNT(*) FROM generated_tts_audio WHERE file_path IS NULL OR TRIM(file_path)=''",
        ),
    }
    conn.close()
    return out


def get_approved_audio(entry_type: str, entry_id: int) -> dict:
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("""
        SELECT lang, file_path
        FROM audio
        WHERE status='approved' AND entry_type=? AND entry_id=?
        ORDER BY id DESC
    """, (entry_type, entry_id))
    rows = c.fetchall()
    conn.close()

    out = {}
    for lang, path in rows:
        if lang not in out:
            out[lang] = _public_audio_url(path)
    return out


def get_approved_oromo_audio_ids(entry_type: str) -> set:
    if entry_type not in ("word", "phrase"):
        return set()

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("""
        SELECT DISTINCT entry_id
        FROM audio
        WHERE status='approved'
          AND entry_type=?
          AND lang='oromo'
    """, (entry_type,))
    ids = {r[0] for r in c.fetchall()}
    conn.close()
    return ids


def delete_audio_for_entry(entry_type: str, entry_id: int):
    """
    Admin helper: deletes ALL audio rows + files for an entry.
    """
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute(
        "SELECT id, file_path FROM audio WHERE entry_type=? AND entry_id=?",
        (entry_type, entry_id)
    )
    rows = c.fetchall()

    c.execute(
        "DELETE FROM audio WHERE entry_type=? AND entry_id=?",
        (entry_type, entry_id)
    )
    conn.commit()
    conn.close()

    for _aid, fp in rows:
        abs_path = _audio_abs_path(fp)
        if abs_path and os.path.isfile(abs_path):
            try:
                os.remove(abs_path)
            except Exception:
                app.logger.exception(f"Could not delete audio file: {abs_path}")


def delete_audio_for_entry_lang(
    entry_type: str,
    entry_id: int,
    lang: str,
    statuses=("approved", "pending")
) -> int:
    """
    Deletes audio rows + files for a specific entry/lang, filtered by statuses.
    Returns number of rows deleted.
    """
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    q_marks = ",".join(["?"] * len(statuses))

    c.execute(f"""
        SELECT id, file_path
        FROM audio
        WHERE entry_type=? AND entry_id=? AND lang=? AND status IN ({q_marks})
    """, (entry_type, entry_id, lang, *statuses))
    rows = c.fetchall()

    c.execute(f"""
        DELETE FROM audio
        WHERE entry_type=? AND entry_id=? AND lang=? AND status IN ({q_marks})
    """, (entry_type, entry_id, lang, *statuses))
    conn.commit()
    conn.close()

    deleted = 0
    for _aid, fp in rows:
        abs_path = _audio_abs_path(fp)
        if abs_path and os.path.isfile(abs_path):
            try:
                os.remove(abs_path)
            except Exception:
                app.logger.exception(f"Could not delete audio file: {abs_path}")
        deleted += 1

    return deleted


# ------------------ DATABASE SETUP ------------------

def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    c.execute("""
    CREATE TABLE IF NOT EXISTS words (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        english TEXT,
        oromo TEXT,
        english_key TEXT,
        oromo_key TEXT,
        status TEXT
    )
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS phrases (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        english TEXT,
        oromo TEXT,
        english_key TEXT,
        oromo_key TEXT,
        status TEXT
    )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS admin (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT,
            password TEXT
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS search_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            query TEXT NOT NULL,
            direction TEXT,
            is_phrase INTEGER DEFAULT 0,
            is_exact INTEGER DEFAULT 0,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS search_counts (
            query TEXT PRIMARY KEY,
            total_count INTEGER DEFAULT 0,
            today_count INTEGER DEFAULT 0,
            week_count INTEGER DEFAULT 0,
            last_searched_at DATETIME
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS audio (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_type TEXT,
            entry_id INTEGER,
            lang TEXT,
            file_path TEXT,
            status TEXT
        )
    """)

    conn.commit()
    conn.close()
    
def _strip_edge_punct(s: str) -> str:
    return re.sub(r"^[\s\"'â€œâ€â€˜â€™`]+|[.!?,;:\s\"'â€œâ€â€˜â€™`]+$", "", s or "").strip()


def ensure_key_columns():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    def has_col(table, col):
        c.execute(f"PRAGMA table_info({table})")
        return any(r[1] == col for r in c.fetchall())

    for table in ("words", "phrases"):
        if not has_col(table, "english_key"):
            c.execute(f"ALTER TABLE {table} ADD COLUMN english_key TEXT")
        if not has_col(table, "oromo_key"):
            c.execute(f"ALTER TABLE {table} ADD COLUMN oromo_key TEXT")

    conn.commit()
    conn.close()

def backfill_keys():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    # words
    c.execute("SELECT id, english, oromo FROM words")
    for wid, en, om in c.fetchall():
        en_norm = normalize_text(en or "")
        om_norm = normalize_text(om or "")
        en_key = make_search_key(_strip_edge_punct(en_norm))
        om_key = make_search_key(_strip_edge_punct(om_norm))
        c.execute(
            "UPDATE words SET english_key=?, oromo_key=? WHERE id=?",
            (en_key, om_key, wid)
        )

    # phrases
    c.execute("SELECT id, english, oromo FROM phrases")
    for pid, en, om in c.fetchall():
        en_norm = normalize_text(en or "")
        om_norm = normalize_text(om or "")
        en_key = make_search_key(_strip_edge_punct(en_norm))
        om_key = make_search_key(_strip_edge_punct(om_norm))
        c.execute(
            "UPDATE phrases SET english_key=?, oromo_key=? WHERE id=?",
            (en_key, om_key, pid)
        )

    conn.commit()
    conn.close()


def ensure_key_indexes():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("CREATE INDEX IF NOT EXISTS idx_words_english_key ON words(english_key)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_words_oromo_key ON words(oromo_key)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_phrases_english_key ON phrases(english_key)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_phrases_oromo_key ON phrases(oromo_key)")
    conn.commit()
    conn.close()
    
def ensure_phrase_aliases_table():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    c.execute("""
    CREATE TABLE IF NOT EXISTS phrase_aliases (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        phrase_id INTEGER NOT NULL,
        english_alias_key TEXT,
        oromo_alias_key TEXT,
        source TEXT DEFAULT 'auto',
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(phrase_id) REFERENCES phrases(id) ON DELETE CASCADE
    )
    """)

    # Unique indexes (SQLite allows multiple NULLs; fine)
    c.execute("CREATE UNIQUE INDEX IF NOT EXISTS uq_alias_en ON phrase_aliases(english_alias_key)")
    c.execute("CREATE UNIQUE INDEX IF NOT EXISTS uq_alias_om ON phrase_aliases(oromo_alias_key)")

    c.execute("CREATE INDEX IF NOT EXISTS idx_alias_en ON phrase_aliases(english_alias_key)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_alias_om ON phrase_aliases(oromo_alias_key)")

    conn.commit()
    conn.close()


def ensure_generated_translations_table():
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()

        c.execute("""
        CREATE TABLE IF NOT EXISTS generated_translations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            word_id INTEGER NOT NULL,
            lang_code TEXT NOT NULL,
            translated_text TEXT NOT NULL,
            source_text_hash TEXT,
            provider TEXT NOT NULL DEFAULT 'google_translate_v2',
            tts_audio_url TEXT,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(word_id, lang_code)
        )
        """)

        c.execute("CREATE INDEX IF NOT EXISTS idx_generated_translations_word_id ON generated_translations(word_id)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_generated_translations_lang_code ON generated_translations(lang_code)")

        conn.commit()
        conn.close()
        return True
    except Exception as e:
        app.logger.exception(f"Failed to ensure generated_translations table: {repr(e)}")
        return False


def ensure_generated_phrase_translations_table():
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("""
        CREATE TABLE IF NOT EXISTS generated_phrase_translations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            phrase_id INTEGER NOT NULL,
            lang_code TEXT NOT NULL,
            translated_text TEXT NOT NULL,
            source_text_hash TEXT,
            provider TEXT NOT NULL DEFAULT 'google_translate_v2',
            tts_audio_url TEXT,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(phrase_id, lang_code)
        )
        """)
        c.execute("CREATE INDEX IF NOT EXISTS idx_generated_phrase_translations_phrase_id ON generated_phrase_translations(phrase_id)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_generated_phrase_translations_lang_code ON generated_phrase_translations(lang_code)")
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        app.logger.exception(f"Failed to ensure generated_phrase_translations table: {repr(e)}")
        return False


def ensure_generated_tts_audio_table():
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("""
        CREATE TABLE IF NOT EXISTS generated_tts_audio (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_type TEXT NOT NULL,
            entry_id INTEGER NOT NULL,
            lang_code TEXT NOT NULL,
            text_value TEXT NOT NULL,
            text_hash TEXT NOT NULL,
            voice_provider TEXT NOT NULL DEFAULT 'azure_speech',
            voice_name TEXT NOT NULL,
            file_path TEXT NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(entry_type, entry_id, lang_code, text_hash, voice_provider, voice_name)
        )
        """)
        c.execute("CREATE INDEX IF NOT EXISTS idx_generated_tts_audio_entry ON generated_tts_audio(entry_type, entry_id)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_generated_tts_audio_lang ON generated_tts_audio(lang_code)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_generated_tts_audio_hash ON generated_tts_audio(text_hash)")
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        app.logger.exception(f"Failed to ensure generated_tts_audio table: {repr(e)}")
        return False


def ensure_generated_translation_hash_columns():
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()

        c.execute("PRAGMA table_info(generated_translations)")
        gt_cols = {str((r or ["", ""])[1] or "").strip() for r in (c.fetchall() or [])}
        if "source_text_hash" not in gt_cols:
            c.execute("ALTER TABLE generated_translations ADD COLUMN source_text_hash TEXT")

        c.execute("PRAGMA table_info(generated_phrase_translations)")
        gpt_cols = {str((r or ["", ""])[1] or "").strip() for r in (c.fetchall() or [])}
        if "source_text_hash" not in gpt_cols:
            c.execute("ALTER TABLE generated_phrase_translations ADD COLUMN source_text_hash TEXT")

        conn.commit()
        conn.close()
        return True
    except Exception as e:
        app.logger.exception(f"Failed to ensure generated translation hash columns: {repr(e)}")
        return False


def ensure_post_import_jobs_table():
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute(
            """
            CREATE TABLE IF NOT EXISTS post_import_jobs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                status TEXT NOT NULL DEFAULT 'pending',
                job_type TEXT NOT NULL DEFAULT 'post_import',
                word_ids_json TEXT NOT NULL DEFAULT '[]',
                phrase_ids_json TEXT NOT NULL DEFAULT '[]',
                chunk_size INTEGER,
                import_summary_json TEXT NOT NULL DEFAULT '{}',
                options_json TEXT NOT NULL DEFAULT '{}',
                attempts INTEGER NOT NULL DEFAULT 0,
                runtime TEXT NOT NULL DEFAULT '',
                result_json TEXT NOT NULL DEFAULT '{}',
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                started_at DATETIME,
                finished_at DATETIME,
                last_error TEXT
            )
            """
        )
        c.execute("PRAGMA table_info(post_import_jobs)")
        cols = {str((r or ["", ""])[1] or "").strip() for r in (c.fetchall() or [])}
        if "job_type" not in cols:
            c.execute("ALTER TABLE post_import_jobs ADD COLUMN job_type TEXT NOT NULL DEFAULT 'post_import'")
        if "options_json" not in cols:
            c.execute("ALTER TABLE post_import_jobs ADD COLUMN options_json TEXT NOT NULL DEFAULT '{}'")
        if "result_json" not in cols:
            c.execute("ALTER TABLE post_import_jobs ADD COLUMN result_json TEXT NOT NULL DEFAULT '{}'")
        if "cancel_requested" not in cols:
            c.execute("ALTER TABLE post_import_jobs ADD COLUMN cancel_requested INTEGER NOT NULL DEFAULT 0")
        c.execute("CREATE INDEX IF NOT EXISTS idx_post_import_jobs_status_id ON post_import_jobs(status, id)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_post_import_jobs_status_type_id ON post_import_jobs(status, job_type, id)")
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        app.logger.exception(f"Failed to ensure post_import_jobs table: {repr(e)}")
        return False


# Run DB init + migrations at startup
init_db()
ensure_key_columns()
backfill_keys()
ensure_key_indexes()
_generated_table_ready = ensure_generated_translations_table()
_generated_phrase_table_ready = ensure_generated_phrase_translations_table()
_generated_tts_table_ready = ensure_generated_tts_audio_table()
_generated_translation_hash_columns_ready = ensure_generated_translation_hash_columns()
_post_import_jobs_table_ready = ensure_post_import_jobs_table()


def record_search(raw_query: str, direction: str, is_phrase: int, is_exact: int):
    q = normalize_text(raw_query)
    if not q:
        return
    conn = None
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()

        c.execute(
            "INSERT INTO search_logs (query, direction, is_phrase, is_exact) VALUES (?, ?, ?, ?)",
            (q, direction, int(is_phrase), int(is_exact))
        )

        c.execute("SELECT total_count FROM search_counts WHERE query=?", (q,))
        row = c.fetchone()

        if row:
            c.execute("""
                UPDATE search_counts
                SET total_count = total_count + 1,
                    today_count = today_count + 1,
                    week_count = week_count + 1,
                    last_searched_at = CURRENT_TIMESTAMP
                WHERE query=?
            """, (q,))
        else:
            c.execute("""
                INSERT INTO search_counts (query, total_count, today_count, week_count, last_searched_at)
                VALUES (?, 1, 1, 1, CURRENT_TIMESTAMP)
            """, (q,))

        conn.commit()
    except Exception as e:
        app.logger.exception(f"record_search failed: {repr(e)}")
    finally:
        if conn:
            conn.close()


def get_trending(limit=20):
    conn = None
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("""
            SELECT query, today_count, week_count, total_count
            FROM search_counts
            ORDER BY today_count DESC, week_count DESC, total_count DESC
            LIMIT ?
        """, (limit,))
        return c.fetchall()
    except Exception as e:
        app.logger.exception(f"get_trending failed: {repr(e)}")
        return []
    finally:
        if conn:
            conn.close()

# ------------------ SUGGESTIONS (KEY-BASED, CASE-INSENSITIVE) ------------------

def suggest_terms(term: str, direction: str, limit: int = 8):
    """
    Returns suggestions for the user's input.
    Uses *_key columns => case-insensitive + normalized.
    """
    raw = normalize_text(term)
    if not raw:
        return {"closest": [], "prefix": [], "partial": []}

    tkey = make_search_key(raw)
    if not tkey:
        return {"closest": [], "prefix": [], "partial": []}

    # which table columns to use
    key_col = "oromo_key" if direction == "om_en" else "english_key"
    text_col = "oromo" if direction == "om_en" else "english"

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    # prefix match
    c.execute(f"""
        SELECT {text_col}
        FROM words
        WHERE status='approved' AND {key_col} LIKE ?
        LIMIT ?
    """, (tkey + "%", limit))
    prefix = [r[0] for r in c.fetchall() if r and r[0]]

    # partial match
    c.execute(f"""
        SELECT {text_col}
        FROM words
        WHERE status='approved' AND {key_col} LIKE ?
        LIMIT ?
    """, ("%" + tkey + "%", limit))
    partial = [r[0] for r in c.fetchall() if r and r[0]]

    # candidates for "closest" (use last 3000 for speed)
    c.execute(f"""
        SELECT {text_col}
        FROM words
        WHERE status='approved'
        ORDER BY id DESC
        LIMIT 3000
    """)
    candidates = [r[0] for r in c.fetchall() if r and r[0]]
    conn.close()

    # closest match (run on normalized form to be fair)
    # map: normalized -> original
    norm_map = {}
    norm_list = []
    for cand in candidates:
        nk = make_search_key(cand)
        if nk and nk not in norm_map:
            norm_map[nk] = cand
            norm_list.append(nk)

    closest_norm = get_close_matches(tkey, norm_list, n=limit, cutoff=0.75)
    closest = [norm_map[n] for n in closest_norm if n in norm_map]

    def dedup(seq):
        seen = set()
        out = []
        for x in seq:
            if x not in seen:
                seen.add(x)
                out.append(x)
        return out

    return {"closest": dedup(closest), "prefix": dedup(prefix), "partial": dedup(partial)}


# ------------------ AUTO LANGUAGE DETECT (KEY-BASED, PUNCT-SAFE) ------------------

_WORD_RE = re.compile(r"[A-Za-z0-9']+")

def detect_direction_auto(text: str) -> str:
    """
    Decides whether user typed Oromo or English.
    Uses *_key columns for matching (case-insensitive).
    """
    t = normalize_text(text)
    if not t:
        return "en_om"

    # extract "word tokens" without punctuation
    tokens = _WORD_RE.findall(t)
    if not tokens:
        return "en_om"

    filtered = [w for w in tokens if w.casefold() not in EN_STOP and w.casefold() not in OROMO_STOP]
    if not filtered:
        filtered = tokens

    full_key = make_search_key(_strip_edge_punct(t))

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    or_score = 0
    en_score = 0

    # token scoring via keys
    for w in filtered:
        wk = make_search_key(w)

        c.execute("SELECT 1 FROM words WHERE status='approved' AND oromo_key=? LIMIT 1", (wk,))
        if c.fetchone():
            or_score += 1

        c.execute("SELECT 1 FROM words WHERE status='approved' AND english_key=? LIMIT 1", (wk,))
        if c.fetchone():
            en_score += 1

    # phrase scoring via keys (stronger weight)
    c.execute("SELECT 1 FROM phrases WHERE status='approved' AND oromo_key=? LIMIT 1", (full_key,))
    if c.fetchone():
        or_score += 4

    c.execute("SELECT 1 FROM phrases WHERE status='approved' AND english_key=? LIMIT 1", (full_key,))
    if c.fetchone():
        en_score += 4

    conn.close()

    if or_score > en_score + 0.5:
        return "om_en"
    if en_score > or_score + 0.5:
        return "en_om"
    return "en_om"

# ------------------ TRANSLATION LOGIC ------------------

def translate_text(text: str, direction: str = "om_en"):
    t = normalize_text(text)
    if not t:
        return "", 0, 0

    # Try exact phrase match using *_key (case-insensitive)
    t_key = make_search_key(_strip_edge_punct(t))

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    if t_key:
        if direction == "om_en":
            c.execute(
                "SELECT id, english FROM phrases WHERE status='approved' AND oromo_key=?",
                (t_key,)
            )
            row = c.fetchone()
            if row:
                conn.close()
                return row[1], 1, 1
        else:
            c.execute(
                "SELECT id, oromo FROM phrases WHERE status='approved' AND english_key=?",
                (t_key,)
            )
            row = c.fetchone()
            if row:
                conn.close()
                return row[1], 1, 1

    # Single-word exact match using *_key
    tokens = t.split()
    if len(tokens) == 1:
        wkey = make_search_key(_strip_edge_punct(tokens[0]))
        if wkey:
            if direction == "om_en":
                c.execute(
                    "SELECT id, english FROM words WHERE status='approved' AND oromo_key=?",
                    (wkey,)
                )
                row = c.fetchone()
                if row:
                    conn.close()
                    return row[1], 1, 0
            else:
                c.execute(
                    "SELECT id, oromo FROM words WHERE status='approved' AND english_key=?",
                    (wkey,)
                )
                row = c.fetchone()
                if row:
                    conn.close()
                    return row[1], 1, 0

    # Word-by-word fallback using *_key (case-insensitive)
    out = []
    for w in tokens:
        wk = make_search_key(_strip_edge_punct(w))
        if not wk:
            out.append(w)
            continue

        if direction == "om_en":
            c.execute("SELECT english FROM words WHERE status='approved' AND oromo_key=?", (wk,))
            r = c.fetchone()
            out.append(r[0] if r else w)
        else:
            c.execute("SELECT oromo FROM words WHERE status='approved' AND english_key=?", (wk,))
            r = c.fetchone()
            out.append(r[0] if r else w)

    conn.close()
    return " ".join(out), 0, 0


def _get_cached_generated_translation(word_id: int, lang_code: str):
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("""
            SELECT translated_text, tts_audio_url, source_text_hash
            FROM generated_translations
            WHERE word_id=? AND lang_code=?
              AND translated_text IS NOT NULL
              AND TRIM(translated_text) != ''
            LIMIT 1
        """, (word_id, lang_code))
        row = c.fetchone()
        conn.close()
        return row
    except Exception as e:
        if "no such table: generated_translations" in str(e).lower():
            ensure_generated_translations_table()
        app.logger.exception(f"generated_translations cache read failed: {repr(e)}")
        return None


def _get_cached_extra_translations_for_word(word_id: int, langs=None, log_hits: bool = False):
    """
    DB-first read for extra generated translations.
    Returns only meaningful cached values and never calls provider APIs.
    """
    out = {}
    wid = int(word_id or 0)
    if not wid:
        return out

    lang_list = tuple(langs or EXTRA_GENERATED_LANGS)
    for lang in lang_list:
        cached = _get_cached_generated_translation(wid, lang)
        translated = normalize_text((cached or [""])[0] or "")
        if not _is_meaningful_generated_text(translated):
            continue
        if log_hits:
            app.logger.info(
                "using cached translation word_id=%s lang=%s value=%r",
                wid,
                lang,
                translated[:120],
            )
        out[lang] = {
            "text": translated,
            "tts_audio_url": (cached or [None, None])[1] if cached else None,
        }

    return out


def _save_generated_translation(
    word_id: int,
    lang_code: str,
    translated_text: str,
    source_text: str = "",
    provider: str = "google_translate_v2",
    tts_audio_url: str = None
):
    txt = normalize_text(translated_text or "")
    src_hash = _text_hash(source_text or "")
    if not _is_meaningful_generated_text(txt):
        return False
    conn = None
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("""
            INSERT INTO generated_translations
            (word_id, lang_code, translated_text, source_text_hash, provider, tts_audio_url, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(word_id, lang_code) DO UPDATE SET
                translated_text=excluded.translated_text,
                source_text_hash=excluded.source_text_hash,
                provider=excluded.provider,
                tts_audio_url=excluded.tts_audio_url,
                updated_at=CURRENT_TIMESTAMP
        """, (word_id, lang_code, txt, src_hash, provider, tts_audio_url))
        conn.commit()
        c.execute(
            """
            SELECT translated_text, source_text_hash
            FROM generated_translations
            WHERE word_id=? AND lang_code=?
            LIMIT 1
            """,
            (int(word_id or 0), normalize_text(lang_code or "")),
        )
        check = c.fetchone()
        saved_txt = normalize_text((check or ["", ""])[0] or "")
        saved_hash = normalize_text((check or ["", ""])[1] or "")
        return bool(_is_meaningful_generated_text(saved_txt) and (not src_hash or src_hash == saved_hash))
    except Exception as e:
        err_lower = str(e).lower()
        if "no such column: source_text_hash" in err_lower:
            ensure_generated_translation_hash_columns()
        if "no such table: generated_translations" in err_lower:
            if ensure_generated_translations_table():
                try:
                    conn = sqlite3.connect(DB_NAME)
                    c = conn.cursor()
                    c.execute("""
                        INSERT INTO generated_translations
                        (word_id, lang_code, translated_text, source_text_hash, provider, tts_audio_url, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                        ON CONFLICT(word_id, lang_code) DO UPDATE SET
                            translated_text=excluded.translated_text,
                            source_text_hash=excluded.source_text_hash,
                            provider=excluded.provider,
                            tts_audio_url=excluded.tts_audio_url,
                            updated_at=CURRENT_TIMESTAMP
                    """, (word_id, lang_code, txt, src_hash, provider, tts_audio_url))
                    conn.commit()
                    c.execute(
                        """
                        SELECT translated_text, source_text_hash
                        FROM generated_translations
                        WHERE word_id=? AND lang_code=?
                        LIMIT 1
                        """,
                        (int(word_id or 0), normalize_text(lang_code or "")),
                    )
                    check = c.fetchone()
                    saved_txt = normalize_text((check or ["", ""])[0] or "")
                    saved_hash = normalize_text((check or ["", ""])[1] or "")
                    return bool(_is_meaningful_generated_text(saved_txt) and (not src_hash or src_hash == saved_hash))
                except Exception as e2:
                    app.logger.exception(f"generated_translations retry write failed: {repr(e2)}")
        app.logger.exception(f"generated_translations cache write failed: {repr(e)}")
        return False
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


def clear_generated_translations_for_word(word_id: int):
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("DELETE FROM generated_translations WHERE word_id=?", (word_id,))
        conn.commit()
        conn.close()
    except Exception as e:
        app.logger.exception(f"generated_translations cache clear failed: {repr(e)}")


def _get_generated_phrase_translation_row_raw(phrase_id: int, lang_code: str):
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("""
            SELECT translated_text, tts_audio_url, source_text_hash
            FROM generated_phrase_translations
            WHERE phrase_id=? AND lang_code=?
            LIMIT 1
        """, (phrase_id, lang_code))
        row = c.fetchone()
        conn.close()
        return row
    except Exception as e:
        err_lower = str(e).lower()
        if "no such column: source_text_hash" in err_lower:
            ensure_generated_translation_hash_columns()
        if "no such table: generated_phrase_translations" in err_lower:
            ensure_generated_phrase_translations_table()
        app.logger.exception(f"generated_phrase_translations cache read failed: {repr(e)}")
        return None


def _get_cached_generated_phrase_translation(phrase_id: int, lang_code: str):
    row = _get_generated_phrase_translation_row_raw(phrase_id, lang_code)
    txt = normalize_text((row or [""])[0] or "")
    if not _is_meaningful_generated_text(txt):
        return None
    return row


def _save_generated_phrase_translation(
    phrase_id: int,
    lang_code: str,
    translated_text: str,
    source_text: str = "",
    provider: str = "google_translate_v2",
    tts_audio_url: str = None
):
    txt = normalize_text(translated_text or "")
    src_hash = _text_hash(source_text or "")
    if not _is_meaningful_generated_text(txt):
        app.logger.warning(
            "generated_phrase_translations write skipped phrase_id=%s lang=%s reason=empty_text text_len=%s",
            int(phrase_id or 0),
            normalize_text(lang_code or ""),
            len(txt),
        )
        return False
    conn = None
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("""
            INSERT INTO generated_phrase_translations
            (phrase_id, lang_code, translated_text, source_text_hash, provider, tts_audio_url, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(phrase_id, lang_code) DO UPDATE SET
                translated_text=excluded.translated_text,
                source_text_hash=excluded.source_text_hash,
                provider=excluded.provider,
                tts_audio_url=excluded.tts_audio_url,
                updated_at=CURRENT_TIMESTAMP
        """, (phrase_id, lang_code, txt, src_hash, provider, tts_audio_url))
        conn.commit()
        c.execute(
            """
            SELECT translated_text, source_text_hash
            FROM generated_phrase_translations
            WHERE phrase_id=? AND lang_code=?
            LIMIT 1
            """,
            (int(phrase_id or 0), normalize_text(lang_code or "")),
        )
        check_row = c.fetchone()
        saved_text = normalize_text((check_row or ["", ""])[0] or "")
        saved_hash = normalize_text((check_row or ["", ""])[1] or "")
        ok = bool(_is_meaningful_generated_text(saved_text) and (not src_hash or src_hash == saved_hash))
        if not ok:
            app.logger.error(
                "generated_phrase_translations write verification failed phrase_id=%s lang=%s saved_text_len=%s",
                int(phrase_id or 0),
                normalize_text(lang_code or ""),
                len(saved_text),
            )
        return bool(ok)
    except Exception as e:
        if "no such table: generated_phrase_translations" in str(e).lower():
            ensure_generated_phrase_translations_table()
        app.logger.exception(
            "generated_phrase_translations cache write failed phrase_id=%s lang=%s error=%s",
            int(phrase_id or 0),
            normalize_text(lang_code or ""),
            repr(e),
        )
        return False
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


def ensure_missing_generated_translations_for_phrases(
    phrase_items,
    langs=None,
    chunk_size: int = None,
    overwrite_existing: bool = False,
    log_context: str = "generated_phrase_backfill",
    cancel_check=None,
):
    if not phrase_items:
        return 0, {}

    langs = tuple(langs or EXTRA_GENERATED_LANGS)
    safe_chunk = int(chunk_size or IMPORT_BATCH_SIZE or 50)
    if safe_chunk < 1:
        safe_chunk = 50

    unique_items = []
    seen_ids = set()
    for pid, en in phrase_items:
        pid_int = int(pid or 0)
        en_norm = normalize_text(en or "")
        if not pid_int or not en_norm or pid_int in seen_ids:
            continue
        seen_ids.add(pid_int)
        unique_items.append((pid_int, en_norm))

    if not unique_items:
        return 0, {}

    # Track base Oromo completeness so empty-string Oromo is treated as missing,
    # not as valid data.
    oromo_missing_ids = set()
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        id_marks = ",".join("?" for _ in unique_items)
        c.execute(
            f"""
            SELECT id, oromo
            FROM phrases
            WHERE id IN ({id_marks})
            """,
            tuple(pid for pid, _en in unique_items),
        )
        for pid, om in c.fetchall():
            pid_int = int(pid or 0)
            om_text = normalize_text(om or "")
            oromo_len = len(om_text)
            treated_missing = bool(oromo_len == 0)
            if treated_missing:
                oromo_missing_ids.add(pid_int)
            app.logger.info(
                "%s phrase_oromo_missing_check phrase_id=%s oromo_len=%s treated_as_missing=%s",
                log_context,
                pid_int,
                oromo_len,
                treated_missing,
            )
        conn.close()
    except Exception:
        app.logger.exception("%s phrase_oromo_missing_check failed", log_context)

    total_saved = 0
    stats = {}
    canceled = False
    phrase_debug = {
        int(pid): {"attempted": [], "saved": [], "skips": []}
        for pid, _en in unique_items
    }
    run_summary = {
        "attempted": int(len(unique_items) * len(langs)),
        "generated_new": 0,
        "skipped_existing": 0,
        "skipped_empty_result": 0,
        "failed_db_write": 0,
    }

    def _log_phrase_lang_status(
        phrase_id: int,
        lang_code: str,
        has_existing_translation: bool,
        cache_hit: bool,
        result_text: str,
        skip_reason: str,
    ):
        txt = normalize_text(result_text or "")
        app.logger.info(
            "%s phrase_translation_status phrase_id=%s lang=%s has_existing_translation=%s cache_hit=%s result_text_len=%s skip_reason=%s",
            log_context,
            int(phrase_id or 0),
            normalize_text(lang_code or ""),
            bool(has_existing_translation),
            bool(cache_hit),
            len(txt),
            (skip_reason or ""),
        )

    for lang in langs:
        st = {
            "items_seen": len(unique_items),
            "attempted": len(unique_items),
            "already_cached": 0,
            "invalid_cached_treated_missing": 0,
            "source_changed_treated_missing": 0,
            "missing_before": 0,
            "saved": 0,
            "empty_results": 0,
            "batch_mismatch_fallback": 0,
            "provider_errors": 0,
            "google_empty_result": 0,
            "google_request_failures": 0,
            "google_provider_errors": 0,
            "missing_api_key": 0,
            "request_failures": 0,
            "fallback_errors": 0,
            "failed_db_write": 0,
        }
        stats[lang] = st
        missing = []
        cache_meta = {}
        skipped_existing_log_count = 0
        processed_new_log_count = 0
        for pid, en in unique_items:
            if callable(cancel_check) and bool(cancel_check()):
                canceled = True
                break
            phrase_debug.setdefault(int(pid), {"attempted": [], "saved": [], "skips": []})
            phrase_debug[int(pid)]["attempted"].append(lang)
            cached_raw = _get_generated_phrase_translation_row_raw(pid, lang)
            cached_text = normalize_text((cached_raw or ["", "", ""])[0] or "")
            cached_hash = normalize_text((cached_raw or ["", "", ""])[2] or "")
            source_hash = _text_hash(en)
            has_existing_translation = bool(cached_raw is not None)
            cache_hit = bool(_is_meaningful_generated_text(cached_text) and cached_hash and cached_hash == source_hash)
            cache_meta[int(pid)] = {
                "has_existing_translation": has_existing_translation,
                "cache_hit": cache_hit,
                "cached_text": cached_text,
            }
            if not overwrite_existing:
                if cache_hit:
                    st["already_cached"] += 1
                    run_summary["skipped_existing"] += 1
                    phrase_debug[int(pid)]["skips"].append(f"{lang}:already_cached")
                    _log_phrase_lang_status(
                        phrase_id=pid,
                        lang_code=lang,
                        has_existing_translation=has_existing_translation,
                        cache_hit=cache_hit,
                        result_text=cached_text,
                        skip_reason="already_cached",
                    )
                    if skipped_existing_log_count < 5:
                        app.logger.info(
                            "%s skipped_existing type=generated_translation entry_type=phrase entry_id=%s lang=%s value=%r",
                            log_context,
                            pid,
                            lang,
                            cached_text[:120],
                        )
                        skipped_existing_log_count += 1
                    continue
                if has_existing_translation:
                    if _is_meaningful_generated_text(cached_text):
                        st["source_changed_treated_missing"] += 1
                    else:
                        st["invalid_cached_treated_missing"] += 1
                    app.logger.warning(
                        "%s phrase_translation_cache_row_invalid_or_stale phrase_id=%s lang=%s result_text_len=%s cached_hash=%s source_hash=%s treating_as_missing=1",
                        log_context,
                        pid,
                        lang,
                        len(cached_text),
                        cached_hash[:12],
                        source_hash[:12],
                    )
            missing.append((pid, en))
        if canceled:
            break
        # Missing means either no generated translation for this language OR empty Oromo base text.
        missing_translation_ids = {int(pid or 0) for pid, _en in missing if int(pid or 0) > 0}
        st["missing_before"] = len(missing_translation_ids.union(oromo_missing_ids))

        for i in range(0, len(missing), safe_chunk):
            pairs = missing[i:i + safe_chunk]
            if not pairs:
                continue
            texts = [en for _pid, en in pairs]
            api_key = _get_google_key()
            if not api_key:
                st["provider_errors"] += len(pairs)
                st["google_provider_errors"] += len(pairs)
                st["google_request_failures"] += len(pairs)
                st["missing_api_key"] += len(pairs)
                for pid, _en in pairs:
                    phrase_debug[int(pid)]["skips"].append(f"{lang}:missing_api_key")
                    meta = cache_meta.get(int(pid), {}) or {}
                    _log_phrase_lang_status(
                        phrase_id=pid,
                        lang_code=lang,
                        has_existing_translation=bool(meta.get("has_existing_translation", False)),
                        cache_hit=bool(meta.get("cache_hit", False)),
                        result_text="",
                        skip_reason="missing_api_key",
                    )
                app.logger.error(
                    "phrase translation skipped: missing Google API key context=%s lang=%s missing_pairs=%s",
                    log_context,
                    lang,
                    len(pairs),
                )
                continue
            try:
                translated = google_translate_batch_v2(
                    texts,
                    target=_google_lang_code(lang),
                    source="en",
                )
                if not translated:
                    # Do not drop the chunk on empty batch response.
                    # Fall through to per-item fallback translate below.
                    st["provider_errors"] += len(pairs)
                    st["google_provider_errors"] += len(pairs)
                    st["google_empty_result"] += len(pairs)
                    st["google_request_failures"] += len(pairs)
                    st["request_failures"] += len(pairs)
                    for pid, _en in pairs:
                        phrase_debug[int(pid)]["skips"].append(f"{lang}:empty_batch_response")
                        meta = cache_meta.get(int(pid), {}) or {}
                        _log_phrase_lang_status(
                            phrase_id=pid,
                            lang_code=lang,
                            has_existing_translation=bool(meta.get("has_existing_translation", False)),
                            cache_hit=bool(meta.get("cache_hit", False)),
                            result_text="",
                            skip_reason="empty_batch_response",
                        )
                    app.logger.error(
                        "phrase batch translate empty response context=%s lang=%s pairs=%s key_source=%s; falling back to per-item translate",
                        log_context,
                        lang,
                        len(pairs),
                        (_google_key_source_name() or "missing"),
                    )
                if len(translated) != len(pairs):
                    st["batch_mismatch_fallback"] += len(pairs)
                    for pid, _en in pairs:
                        phrase_debug[int(pid)]["skips"].append(f"{lang}:batch_mismatch_fallback")
                        meta = cache_meta.get(int(pid), {}) or {}
                        _log_phrase_lang_status(
                            phrase_id=pid,
                            lang_code=lang,
                            has_existing_translation=bool(meta.get("has_existing_translation", False)),
                            cache_hit=bool(meta.get("cache_hit", False)),
                            result_text="",
                            skip_reason="batch_mismatch_fallback",
                        )
                    translated = []
            except Exception:
                translated = []
                st["provider_errors"] += len(pairs)
                st["google_provider_errors"] += len(pairs)
                st["google_request_failures"] += len(pairs)
                st["request_failures"] += len(pairs)
                for pid, _en in pairs:
                    phrase_debug[int(pid)]["skips"].append(f"{lang}:batch_exception")
                    meta = cache_meta.get(int(pid), {}) or {}
                    _log_phrase_lang_status(
                        phrase_id=pid,
                        lang_code=lang,
                        has_existing_translation=bool(meta.get("has_existing_translation", False)),
                        cache_hit=bool(meta.get("cache_hit", False)),
                        result_text="",
                        skip_reason="batch_exception",
                    )
                app.logger.exception(
                    "phrase batch translate exception context=%s lang=%s pairs=%s",
                    log_context,
                    lang,
                    len(pairs),
                )
                # Fall through to per-item fallback translate below.

            if translated and len(translated) == len(pairs):
                for (pid, _en), out in zip(pairs, translated):
                    if callable(cancel_check) and bool(cancel_check()):
                        canceled = True
                        break
                    txt = normalize_text(out or "")
                    meta = cache_meta.get(int(pid), {}) or {}
                    app.logger.info(
                        "%s phrase_translation_raw_response phrase_id=%s lang=%s result_text=%r result_text_len=%s",
                        log_context,
                        pid,
                        lang,
                        txt[:300],
                        len(txt),
                    )
                    if not _is_meaningful_generated_text(txt):
                        st["empty_results"] += 1
                        st["google_empty_result"] += 1
                        run_summary["skipped_empty_result"] += 1
                        phrase_debug[int(pid)]["skips"].append(f"{lang}:empty_text")
                        _log_phrase_lang_status(
                            phrase_id=pid,
                            lang_code=lang,
                            has_existing_translation=bool(meta.get("has_existing_translation", False)),
                            cache_hit=bool(meta.get("cache_hit", False)),
                            result_text=txt,
                            skip_reason="empty_result_batch",
                        )
                        continue
                    try:
                        write_ok = _save_generated_phrase_translation(
                            pid, lang, txt, source_text=_en, provider="google_translate_v2", tts_audio_url=None
                        )
                        if write_ok:
                            st["saved"] += 1
                            total_saved += 1
                            run_summary["generated_new"] += 1
                            phrase_debug[int(pid)]["saved"].append(lang)
                            _log_phrase_lang_status(
                                phrase_id=pid,
                                lang_code=lang,
                                has_existing_translation=bool(meta.get("has_existing_translation", False)),
                                cache_hit=bool(meta.get("cache_hit", False)),
                                result_text=txt,
                                skip_reason="generated",
                            )
                        else:
                            st["failed_db_write"] += 1
                            run_summary["failed_db_write"] += 1
                            phrase_debug[int(pid)]["skips"].append(f"{lang}:write_failed")
                            _log_phrase_lang_status(
                                phrase_id=pid,
                                lang_code=lang,
                                has_existing_translation=bool(meta.get("has_existing_translation", False)),
                                cache_hit=bool(meta.get("cache_hit", False)),
                                result_text=txt,
                                skip_reason="db_write_failed",
                            )
                            continue
                        if processed_new_log_count < 5:
                            app.logger.info(
                                "%s processed_new type=generated_translation entry_type=phrase entry_id=%s lang=%s",
                                log_context,
                                pid,
                                lang,
                            )
                            processed_new_log_count += 1
                    except Exception:
                        st["provider_errors"] += 1
                        phrase_debug[int(pid)]["skips"].append(f"{lang}:save_exception")
                        _log_phrase_lang_status(
                            phrase_id=pid,
                            lang_code=lang,
                            has_existing_translation=bool(meta.get("has_existing_translation", False)),
                            cache_hit=bool(meta.get("cache_hit", False)),
                            result_text=txt,
                            skip_reason="save_exception",
                        )
                if canceled:
                    break
            else:
                for pid, en in pairs:
                    if callable(cancel_check) and bool(cancel_check()):
                        canceled = True
                        break
                    try:
                        meta = cache_meta.get(int(pid), {}) or {}
                        translated = google_translate_text_v2(
                            en,
                            target=_google_lang_code(lang),
                            source="en",
                        )
                        txt = normalize_text(translated or "")
                        app.logger.info(
                            "%s phrase_translation_raw_response phrase_id=%s lang=%s result_text=%r result_text_len=%s source=fallback",
                            log_context,
                            pid,
                            lang,
                            txt[:300],
                            len(txt),
                        )
                        if not _is_meaningful_generated_text(txt):
                            st["request_failures"] += 1
                            st["empty_results"] += 1
                            st["google_request_failures"] += 1
                            st["google_empty_result"] += 1
                            st["provider_errors"] += 1
                            st["google_provider_errors"] += 1
                            run_summary["skipped_empty_result"] += 1
                            phrase_debug[int(pid)]["skips"].append(f"{lang}:fallback_empty_text")
                            _log_phrase_lang_status(
                                phrase_id=pid,
                                lang_code=lang,
                                has_existing_translation=bool(meta.get("has_existing_translation", False)),
                                cache_hit=bool(meta.get("cache_hit", False)),
                                result_text=txt,
                                skip_reason="empty_result_fallback",
                            )
                            continue
                        write_ok = _save_generated_phrase_translation(
                            pid, lang, txt, source_text=en, provider="google_translate_v2", tts_audio_url=None
                        )
                        if write_ok:
                            st["saved"] += 1
                            total_saved += 1
                            run_summary["generated_new"] += 1
                            phrase_debug[int(pid)]["saved"].append(lang)
                            _log_phrase_lang_status(
                                phrase_id=pid,
                                lang_code=lang,
                                has_existing_translation=bool(meta.get("has_existing_translation", False)),
                                cache_hit=bool(meta.get("cache_hit", False)),
                                result_text=txt,
                                skip_reason="generated_fallback",
                            )
                        else:
                            st["failed_db_write"] += 1
                            run_summary["failed_db_write"] += 1
                            phrase_debug[int(pid)]["skips"].append(f"{lang}:fallback_write_failed")
                            _log_phrase_lang_status(
                                phrase_id=pid,
                                lang_code=lang,
                                has_existing_translation=bool(meta.get("has_existing_translation", False)),
                                cache_hit=bool(meta.get("cache_hit", False)),
                                result_text=txt,
                                skip_reason="db_write_failed_fallback",
                            )
                            continue
                        if processed_new_log_count < 5:
                            app.logger.info(
                                "%s processed_new type=generated_translation entry_type=phrase entry_id=%s lang=%s",
                                log_context,
                                pid,
                                lang,
                            )
                            processed_new_log_count += 1
                    except Exception:
                        st["provider_errors"] += 1
                        st["google_provider_errors"] += 1
                        st["google_request_failures"] += 1
                        st["fallback_errors"] += 1
                        phrase_debug[int(pid)]["skips"].append(f"{lang}:fallback_exception")
                        _log_phrase_lang_status(
                            phrase_id=pid,
                            lang_code=lang,
                            has_existing_translation=False,
                            cache_hit=False,
                            result_text="",
                            skip_reason="fallback_exception",
                        )
                        app.logger.exception(
                            "phrase fallback translate failed context=%s phrase_id=%s lang=%s",
                            log_context,
                            pid,
                            lang,
                        )
                if canceled:
                    break
            if canceled:
                break
        if canceled:
            break

    try:
        compact = {
            lang: {
                "attempted": st.get("attempted", 0),
                "missing": st["missing_before"],
                "saved": st["saved"],
                "cached": st["already_cached"],
                "invalid_cached": st["invalid_cached_treated_missing"],
                "empty": st["empty_results"],
                "fallback": st["batch_mismatch_fallback"],
                "errors": st["provider_errors"],
                "google_empty_result": st.get("google_empty_result", 0),
                "google_request_failures": st.get("google_request_failures", 0),
                "google_provider_errors": st.get("google_provider_errors", 0),
                "failed_db_write": st.get("failed_db_write", 0),
                "missing_api_key": st.get("missing_api_key", 0),
                "request_failures": st.get("request_failures", 0),
                "fallback_errors": st.get("fallback_errors", 0),
            }
            for lang, st in stats.items()
        }
        app.logger.info(
            "generated phrase backfill summary context=%s total_saved=%s details=%s",
            log_context,
            total_saved,
            compact,
        )
        app.logger.info(
            "%s phrase_generation_run_summary attempted=%s generated_new=%s skipped_existing=%s skipped_empty_result=%s failed_db_write=%s",
            log_context,
            int(run_summary.get("attempted", 0) or 0),
            int(run_summary.get("generated_new", 0) or 0),
            int(run_summary.get("skipped_existing", 0) or 0),
            int(run_summary.get("skipped_empty_result", 0) or 0),
            int(run_summary.get("failed_db_write", 0) or 0),
        )
    except Exception:
        pass

    if canceled:
        stats["__meta__"] = {"canceled": True}

    try:
        for pid, dbg in phrase_debug.items():
            saved_langs = sorted(set(dbg.get("saved", []) or []))
            attempted_langs = sorted(set(dbg.get("attempted", []) or []))
            skips = dbg.get("skips", []) or []
            app.logger.info(
                "%s phrase_translation_detail phrase_id=%s attempted=%s saved=%s skipped=%s",
                log_context,
                pid,
                attempted_langs,
                saved_langs,
                skips[:20],
            )
    except Exception:
        pass

    return total_saved, stats


def _fetch_approved_word_items(limit: int = 0):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    sql = """
        SELECT id, english
        FROM words
        WHERE status='approved' AND english IS NOT NULL AND TRIM(english) != ''
        ORDER BY id ASC
    """
    if limit and int(limit) > 0:
        sql += " LIMIT ?"
        c.execute(sql, (int(limit),))
    else:
        c.execute(sql)
    rows = [(int(r[0]), normalize_text(r[1] or "")) for r in c.fetchall()]
    conn.close()
    return [(wid, en) for wid, en in rows if wid and en]


def _fetch_approved_phrase_items(limit: int = 0):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    sql = """
        SELECT id, english
        FROM phrases
        WHERE status='approved' AND english IS NOT NULL AND TRIM(english) != ''
        ORDER BY id ASC
    """
    if limit and int(limit) > 0:
        sql += " LIMIT ?"
        c.execute(sql, (int(limit),))
    else:
        c.execute(sql)
    rows = [(int(r[0]), normalize_text(r[1] or "")) for r in c.fetchall()]
    conn.close()
    return [(pid, en) for pid, en in rows if pid and en]


def ensure_missing_oromo_for_entries(
    entry_type: str,
    items,
    chunk_size: int = None,
    log_context: str = "post_import_oromo",
    cancel_check=None,
):
    """
    Best-effort Oromo backfill for base entries that were imported without Oromo text.
    Updates base table rows only when Oromo is currently blank.
    """
    summary = {
        "items_seen": 0,
        "missing_before": 0,
        "updated": 0,
        "already_present": 0,
        "empty_results": 0,
        "provider_errors": 0,
        # Explicit Oromo-fill counters for phrase pipeline diagnostics.
        "scanned_missing_oromo": 0,
        "filled_oromo": 0,
        "skipped_existing_oromo": 0,
        "failed_oromo_fill": 0,
        "fallback_oromo_used": 0,
        "update_skipped": 0,
    }
    if entry_type not in ("word", "phrase") or not items:
        return summary

    safe_chunk = int(chunk_size or IMPORT_BATCH_SIZE or 50)
    if safe_chunk < 1:
        safe_chunk = 50

    table = "words" if entry_type == "word" else "phrases"
    unique_items = []
    seen_ids = set()
    for eid, en in items:
        eid_int = int(eid or 0)
        en_norm = normalize_text(en or "")
        if not eid_int or not en_norm or eid_int in seen_ids:
            continue
        seen_ids.add(eid_int)
        unique_items.append((eid_int, en_norm))

    summary["items_seen"] = len(unique_items)
    if not unique_items:
        return summary

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    missing_pairs = []
    missing_meta = {}
    filled_ids = set()
    failed_examples = []

    def _log_phrase_fill_status(phrase_id: int, english_text: str, current_oromo_len: int, filled_oromo: bool, reason: str):
        if entry_type != "phrase":
            return
        app.logger.info(
            "%s phrase_oromo_fill_status phrase_id=%s english=%r current_oromo_len=%s filled_oromo=%s reason=%s",
            log_context,
            int(phrase_id or 0),
            normalize_text(english_text or ""),
            int(current_oromo_len or 0),
            bool(filled_oromo),
            normalize_text(reason or ""),
        )

    for eid, en in unique_items:
        if callable(cancel_check) and bool(cancel_check()):
            summary["canceled"] = True
            break
        if entry_type == "phrase":
            c.execute(
                """
                SELECT oromo
                FROM phrases
                WHERE id=?
                  AND status='approved'
                LIMIT 1
                """,
                (eid,),
            )
        else:
            c.execute(
                f"SELECT oromo FROM {table} WHERE id=? AND status='approved' LIMIT 1",
                (eid,),
            )
        row = c.fetchone()
        existing_om = normalize_text((row or [""])[0] or "")
        current_oromo_len = len(existing_om)
        if existing_om:
            summary["already_present"] += 1
            summary["skipped_existing_oromo"] += 1
            if entry_type == "phrase":
                _log_phrase_fill_status(
                    phrase_id=eid,
                    english_text=en,
                    current_oromo_len=current_oromo_len,
                    filled_oromo=False,
                    reason="already_present",
                )
            continue
        missing_pairs.append((eid, en))
        missing_meta[int(eid)] = {
            "english": en,
            "current_oromo_len": current_oromo_len,
        }

    summary["missing_before"] = len(missing_pairs)
    summary["scanned_missing_oromo"] = len(missing_pairs)
    if bool(summary.get("canceled")):
        conn.close()
        return summary
    if not missing_pairs:
        conn.close()
        return summary

    def _deterministic_oromo_fallback(english_text: str) -> str:
        # Deterministic non-empty fallback to avoid silently unresolved blank Oromo.
        en = normalize_text(english_text or "")
        return normalize_text(f"[AUTO_OROMO_PENDING] {en}") if en else "[AUTO_OROMO_PENDING]"

    def _translate_oromo_single_with_retry(english_text: str) -> tuple[str, str]:
        """
        Returns (oromo_text, reason).
        reason in {"google_single", "google_single_retry", "google_single_empty_result", "provider_exception"}
        """
        en = normalize_text(english_text or "")
        if not en:
            return "", "google_single_empty_result"

        try:
            om_first = normalize_text(
                google_translate_text_v2(en, target=_google_lang_code("om"), source="en") or ""
            )
        except Exception:
            return "", "provider_exception"
        if om_first:
            return om_first, "google_single"

        try:
            om_second = normalize_text(
                google_translate_text_v2(en, target=_google_lang_code("om"), source="en") or ""
            )
        except Exception:
            return "", "provider_exception"
        if om_second:
            return om_second, "google_single_retry"
        return "", "google_single_empty_result"

    def _update_oromo_with_noop_recovery(eid: int, om_text: str, om_key: str) -> tuple[bool, str]:
        """
        Returns (updated, reason).
        reason in {"updated", "forced_after_noop", "update_skipped_became_nonblank",
                   "update_skipped_not_approved_or_missing", "update_skipped_still_blank"}.
        """
        c.execute(
            f"""
            UPDATE {table}
            SET oromo=?, oromo_key=?
            WHERE id=?
              AND (oromo IS NULL OR TRIM(oromo)='')
            """,
            (om_text, om_key, eid),
        )
        if int(c.rowcount or 0) > 0:
            return True, "updated"

        # Investigate no-op path.
        c.execute(
            f"SELECT status, oromo FROM {table} WHERE id=? LIMIT 1",
            (eid,),
        )
        cur = c.fetchone()
        if not cur:
            return False, "update_skipped_not_approved_or_missing"
        status_now = normalize_text((cur or ["", ""])[0] or "")
        oromo_now = normalize_text((cur or ["", ""])[1] or "")
        if status_now != "approved":
            return False, "update_skipped_not_approved_or_missing"
        if oromo_now:
            # Another worker/request already filled Oromo after scan.
            return False, "update_skipped_became_nonblank"

        # Still blank -> perform guarded forced update for approved row.
        c.execute(
            f"""
            UPDATE {table}
            SET oromo=?, oromo_key=?
            WHERE id=?
              AND status='approved'
            """,
            (om_text, om_key, eid),
        )
        if int(c.rowcount or 0) > 0:
            return True, "forced_after_noop"
        return False, "update_skipped_still_blank"

    for i in range(0, len(missing_pairs), safe_chunk):
        chunk = missing_pairs[i:i + safe_chunk]
        if not chunk:
            continue

        english_batch = [en for _eid, en in chunk]
        chunk_by_id = {int(pid): en for pid, en in chunk}
        translated_list = google_translate_batch_v2(
            english_batch,
            target=_google_lang_code("om"),
            source="en",
        )

        if translated_list and len(translated_list) == len(chunk):
            for (eid, _en), om_raw in zip(chunk, translated_list):
                if callable(cancel_check) and bool(cancel_check()):
                    summary["canceled"] = True
                    break
                om = normalize_text(om_raw or "")
                if not om:
                    # Batch empty -> single translation with one retry.
                    summary["empty_results"] += 1
                    meta = missing_meta.get(int(eid), {}) or {}
                    en_text = meta.get("english", chunk_by_id.get(int(eid), ""))
                    _log_phrase_fill_status(
                        phrase_id=eid,
                        english_text=en_text,
                        current_oromo_len=int(meta.get("current_oromo_len", 0) or 0),
                        filled_oromo=False,
                        reason="google_batch_empty_result",
                    )
                    om_retry, retry_reason = _translate_oromo_single_with_retry(en_text)
                    if not om_retry:
                        # Deterministic fallback marker so row is no longer blank/unresolved.
                        om_retry = _deterministic_oromo_fallback(en_text)
                        summary["fallback_oromo_used"] += 1
                        reason_tail = "fallback_marker_used"
                    else:
                        reason_tail = retry_reason
                    om = normalize_text(om_retry or "")
                om_key = make_search_key(_strip_edge_punct(om))
                updated, update_reason = _update_oromo_with_noop_recovery(eid, om, om_key)
                if updated:
                    summary["updated"] += 1
                    summary["filled_oromo"] += 1
                    filled_ids.add(int(eid))
                    if entry_type == "phrase":
                        meta = missing_meta.get(int(eid), {}) or {}
                        _log_phrase_fill_status(
                            phrase_id=eid,
                            english_text=meta.get("english", chunk_by_id.get(int(eid), "")),
                            current_oromo_len=int(meta.get("current_oromo_len", 0) or 0),
                            filled_oromo=True,
                            reason=("google_batch" if not om_raw else f"google_batch_{update_reason}"),
                        )
                elif entry_type == "phrase":
                    summary["failed_oromo_fill"] += 1
                    summary["update_skipped"] += 1
                    meta = missing_meta.get(int(eid), {}) or {}
                    _log_phrase_fill_status(
                        phrase_id=eid,
                        english_text=meta.get("english", chunk_by_id.get(int(eid), "")),
                        current_oromo_len=int(meta.get("current_oromo_len", 0) or 0),
                        filled_oromo=False,
                        reason=update_reason,
                    )
                    if len(failed_examples) < 10:
                        failed_examples.append((int(eid), normalize_text(update_reason or "")))
            if bool(summary.get("canceled")):
                break
        else:
            for eid, en in chunk:
                if callable(cancel_check) and bool(cancel_check()):
                    summary["canceled"] = True
                    break
                try:
                    om, single_reason = _translate_oromo_single_with_retry(en)
                    if not om:
                        summary["empty_results"] += 1
                        # Provider returned empty twice -> deterministic fallback.
                        om = _deterministic_oromo_fallback(en)
                        summary["fallback_oromo_used"] += 1
                        if entry_type == "phrase":
                            meta = missing_meta.get(int(eid), {}) or {}
                            _log_phrase_fill_status(
                                phrase_id=eid,
                                english_text=meta.get("english", en),
                                current_oromo_len=int(meta.get("current_oromo_len", 0) or 0),
                                filled_oromo=True,
                                reason=f"{single_reason}_fallback_marker_used",
                            )
                    om_key = make_search_key(_strip_edge_punct(om))
                    updated, update_reason = _update_oromo_with_noop_recovery(eid, om, om_key)
                    if updated:
                        summary["updated"] += 1
                        summary["filled_oromo"] += 1
                        filled_ids.add(int(eid))
                        if entry_type == "phrase":
                            meta = missing_meta.get(int(eid), {}) or {}
                            _log_phrase_fill_status(
                                phrase_id=eid,
                                english_text=meta.get("english", en),
                                current_oromo_len=int(meta.get("current_oromo_len", 0) or 0),
                                filled_oromo=True,
                                reason=f"{single_reason}_{update_reason}",
                            )
                    elif entry_type == "phrase":
                        summary["failed_oromo_fill"] += 1
                        summary["update_skipped"] += 1
                        meta = missing_meta.get(int(eid), {}) or {}
                        _log_phrase_fill_status(
                            phrase_id=eid,
                            english_text=meta.get("english", en),
                            current_oromo_len=int(meta.get("current_oromo_len", 0) or 0),
                            filled_oromo=False,
                            reason=update_reason,
                        )
                        if len(failed_examples) < 10:
                            failed_examples.append((int(eid), normalize_text(update_reason or "")))
                except Exception:
                    summary["provider_errors"] += 1
                    summary["failed_oromo_fill"] += 1
                    if entry_type == "phrase":
                        meta = missing_meta.get(int(eid), {}) or {}
                        _log_phrase_fill_status(
                            phrase_id=eid,
                            english_text=meta.get("english", en),
                            current_oromo_len=int(meta.get("current_oromo_len", 0) or 0),
                            filled_oromo=False,
                            reason="provider_exception",
                        )
                    if len(failed_examples) < 10:
                        failed_examples.append((int(eid), "provider_exception"))
            if bool(summary.get("canceled")):
                break
        if bool(summary.get("canceled")):
            break

    conn.commit()
    conn.close()

    # Guard against accidental under-counting in race/no-op edge cases.
    if summary["failed_oromo_fill"] < 0:
        summary["failed_oromo_fill"] = 0

    app.logger.info(
        "%s Oromo backfill summary entry_type=%s seen=%s missing=%s updated=%s already_present=%s empty=%s errors=%s scanned_missing_oromo=%s filled_oromo=%s skipped_existing_oromo=%s failed_oromo_fill=%s fallback_oromo_used=%s update_skipped=%s",
        log_context,
        entry_type,
        summary["items_seen"],
        summary["missing_before"],
        summary["updated"],
        summary["already_present"],
        summary["empty_results"],
        summary["provider_errors"],
        summary["scanned_missing_oromo"],
        summary["filled_oromo"],
        summary["skipped_existing_oromo"],
        summary["failed_oromo_fill"],
        summary["fallback_oromo_used"],
        summary["update_skipped"],
    )
    if entry_type == "phrase" and failed_examples:
        app.logger.warning(
            "%s Oromo backfill failed_examples entry_type=phrase samples=%s",
            log_context,
            failed_examples[:5],
        )
    return summary


def _fetch_phrase_items_missing_generated_translations(limit: int = 0):
    """
    Return approved phrase rows with non-empty English source text where either:
    - Oromo base text is missing (NULL/empty/whitespace), OR
    - at least one generated translation among am/ar/fr/zh-CN is missing.
    """
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    sql = """
        SELECT p.id, p.english, p.oromo
        FROM phrases p
        LEFT JOIN generated_phrase_translations g_am
          ON g_am.phrase_id = p.id
         AND g_am.lang_code = 'am'
         AND g_am.translated_text IS NOT NULL
         AND TRIM(g_am.translated_text) != ''
        LEFT JOIN generated_phrase_translations g_ar
          ON g_ar.phrase_id = p.id
         AND g_ar.lang_code = 'ar'
         AND g_ar.translated_text IS NOT NULL
         AND TRIM(g_ar.translated_text) != ''
        LEFT JOIN generated_phrase_translations g_fr
          ON g_fr.phrase_id = p.id
         AND g_fr.lang_code = 'fr'
         AND g_fr.translated_text IS NOT NULL
         AND TRIM(g_fr.translated_text) != ''
        LEFT JOIN generated_phrase_translations g_zh
          ON g_zh.phrase_id = p.id
         AND g_zh.lang_code = 'zh-CN'
         AND g_zh.translated_text IS NOT NULL
         AND TRIM(g_zh.translated_text) != ''
        WHERE p.status='approved'
          AND p.english IS NOT NULL
          AND TRIM(p.english) != ''
          AND (
            p.oromo IS NULL OR TRIM(p.oromo) = '' OR
            g_am.id IS NULL OR
            g_ar.id IS NULL OR
            g_fr.id IS NULL OR
            g_zh.id IS NULL
          )
        ORDER BY p.id ASC
    """
    if limit and int(limit) > 0:
        sql += " LIMIT ?"
        c.execute(sql, (int(limit),))
    else:
        c.execute(sql)
    rows = c.fetchall()
    conn.close()
    out = []
    for r in (rows or []):
        pid = int((r or [0, "", ""])[0] or 0)
        en = normalize_text((r or [0, "", ""])[1] or "")
        om = normalize_text((r or [0, "", ""])[2] or "")
        if not pid or not en:
            continue
        treated_missing = bool(not om)
        app.logger.info(
            "phrase_selector_missing_check phrase_id=%s oromo_len=%s treated_as_missing=%s",
            pid,
            len(om),
            treated_missing,
        )
        out.append((pid, en))
    return out


def run_phrase_translation_backfill(limit: int = 0, chunk_size: int = None):
    phrase_items = _fetch_phrase_items_missing_generated_translations(limit=limit)
    skipped_missing_text = 0
    valid_items = []
    for pid, en in phrase_items:
        if not pid:
            continue
        en_norm = normalize_text(en or "")
        if not en_norm:
            skipped_missing_text += 1
            continue
        valid_items.append((int(pid), en_norm))

    saved, stats = ensure_missing_generated_translations_for_phrases(
        valid_items,
        langs=EXTRA_GENERATED_LANGS,
        chunk_size=chunk_size or IMPORT_BATCH_SIZE,
        overwrite_existing=False,
        log_context="cli_backfill_phrase_translations",
    )
    return {
        "phrases_processed": len(valid_items),
        "translations_generated": int(saved or 0),
        "translations_cached": sum(int((st or {}).get("already_cached", 0) or 0) for st in (stats or {}).values()),
        "failures": sum(int((st or {}).get("provider_errors", 0) or 0) for st in (stats or {}).values()),
        "skipped_missing_text": int(skipped_missing_text),
        "stats": stats or {},
    }


def run_translation_backfill(entry_type: str = "all", entry_id: int = 0, overwrite_existing: bool = False, limit: int = 0):
    summary = {"words_saved": 0, "phrases_saved": 0, "word_stats": {}, "phrase_stats": {}}

    if entry_type in ("all", "word"):
        word_items = _fetch_approved_word_items(limit=limit)
        if entry_id:
            word_items = [(wid, en) for wid, en in word_items if int(wid) == int(entry_id)]
        if word_items:
            saved, stats = ensure_missing_generated_translations_for_words(
                word_items,
                langs=EXTRA_GENERATED_LANGS,
                chunk_size=IMPORT_BATCH_SIZE,
                log_context="cli_backfill_words",
            )
            if overwrite_existing:
                # Explicit overwrite path: re-run one by one.
                for wid, en in word_items:
                    for lang in EXTRA_GENERATED_LANGS:
                        tr = google_translate_text_v2(en, target=_google_lang_code(lang), source="en")
                        trn = normalize_text(tr or "")
                        if _is_meaningful_generated_text(trn):
                            _save_generated_translation(
                                wid,
                                lang,
                                trn,
                                source_text=en,
                                provider="google_translate_v2",
                                tts_audio_url=None,
                            )
                saved = len(word_items) * len(EXTRA_GENERATED_LANGS)
            summary["words_saved"] = int(saved or 0)
            summary["word_stats"] = stats or {}

    if entry_type in ("all", "phrase"):
        phrase_items = _fetch_approved_phrase_items(limit=limit)
        if entry_id:
            phrase_items = [(pid, en) for pid, en in phrase_items if int(pid) == int(entry_id)]
        if phrase_items:
            saved, stats = ensure_missing_generated_translations_for_phrases(
                phrase_items,
                langs=EXTRA_GENERATED_LANGS,
                chunk_size=IMPORT_BATCH_SIZE,
                overwrite_existing=overwrite_existing,
                log_context="cli_backfill_phrases",
            )
            summary["phrases_saved"] = int(saved or 0)
            summary["phrase_stats"] = stats or {}

    return summary


def _get_azure_speech_key() -> str:
    return (os.environ.get("AZURE_SPEECH_KEY") or "").strip()


def _get_azure_speech_region() -> str:
    return (os.environ.get("AZURE_SPEECH_REGION") or "").strip()


def _azure_voice_for_lang(lang_code: str) -> str:
    return (DEFAULT_AZURE_VOICES.get(lang_code) or "").strip()


def _text_hash(text: str) -> str:
    return hashlib.sha256((normalize_text(text or "")).encode("utf-8")).hexdigest()


def _generated_tts_file_name(entry_type: str, entry_id: int, lang_code: str, text_hash: str, voice_name: str) -> str:
    safe_voice = re.sub(r"[^a-zA-Z0-9_-]+", "-", (voice_name or "default")).strip("-")[:40] or "default"
    return f"tts_{entry_type}_{entry_id}_{lang_code}_{text_hash[:12]}_{safe_voice}.mp3"


def _resolve_generated_tts_row(entry_type: str, entry_id: int, lang_code: str, text: str, voice_name: str):
    th = _text_hash(text)
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("""
        SELECT id, file_path
        FROM generated_tts_audio
        WHERE entry_type=? AND entry_id=? AND lang_code=? AND text_hash=?
          AND voice_provider=? AND voice_name=?
        ORDER BY id DESC
        LIMIT 1
    """, (entry_type, int(entry_id), lang_code, th, AZURE_TTS_PROVIDER, voice_name))
    row = c.fetchone()
    if not row:
        # Reuse exact text-hash audio even if provider/voice metadata differs.
        c.execute(
            """
            SELECT id, file_path, text_hash
            FROM generated_tts_audio
            WHERE entry_type=? AND entry_id=? AND lang_code=?
              AND text_hash=?
              AND file_path IS NOT NULL
              AND TRIM(file_path) != ''
            ORDER BY id DESC
            LIMIT 1
            """,
            (entry_type, int(entry_id), lang_code, th),
        )
        row = c.fetchone()
        conn.close()
        if not row:
            return None
        file_path = (row[1] or "").strip()
        if not file_path or (not _has_usable_audio_ref(file_path)):
            return None
        return {
            "id": int(row[0]),
            "file_path": file_path,
            "url": _public_audio_url(file_path),
            "text_hash": (row[2] or ""),
        }
    conn.close()
    file_path = (row[1] or "").strip()
    if not file_path:
        return None
    if not _has_usable_audio_ref(file_path):
        return None
    return {"id": int(row[0]), "file_path": file_path, "url": _public_audio_url(file_path), "text_hash": th}


def _save_generated_tts_row(entry_type: str, entry_id: int, lang_code: str, text: str, voice_name: str, file_path: str):
    txt = normalize_text(text or "")
    fp_raw = normalize_text(file_path or "")
    fp = _canonical_local_audio_ref(fp_raw)
    if (not txt) or (not fp):
        return False
    th = _text_hash(txt)
    if _is_remote_audio_ref(fp):
        expected_fp = fp
    else:
        expected_fp = _canonical_local_audio_ref(fp)
        if not _has_usable_audio_ref(expected_fp):
            app.logger.warning(
                "save_generated_tts_row skipped_missing_file entry_type=%s entry_id=%s lang=%s expected_fp=%s",
                entry_type,
                entry_id,
                lang_code,
                expected_fp,
            )
            return False
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("""
        INSERT INTO generated_tts_audio
        (entry_type, entry_id, lang_code, text_value, text_hash, voice_provider, voice_name, file_path, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(entry_type, entry_id, lang_code, text_hash, voice_provider, voice_name) DO UPDATE SET
            file_path=excluded.file_path
    """, (entry_type, int(entry_id), lang_code, txt, th, AZURE_TTS_PROVIDER, voice_name, expected_fp))
    conn.commit()
    c.execute(
        """
        SELECT file_path
        FROM generated_tts_audio
        WHERE entry_type=? AND entry_id=? AND lang_code=? AND text_hash=?
          AND voice_provider=? AND voice_name=?
        LIMIT 1
        """,
        (entry_type, int(entry_id), lang_code, th, AZURE_TTS_PROVIDER, voice_name),
    )
    row = c.fetchone()
    conn.close()
    saved_fp_raw = normalize_text((row or [""])[0] or "")
    saved_fp = _canonical_local_audio_ref(saved_fp_raw)
    same_ref = bool(saved_fp and saved_fp == expected_fp)
    exists_ok = True if _is_remote_audio_ref(saved_fp) else bool(_has_usable_audio_ref(saved_fp))
    playback_url = _public_audio_url(saved_fp) if saved_fp else ""
    app.logger.info(
        "save_generated_tts_row verify entry_type=%s entry_id=%s lang=%s expected_fp=%s saved_fp=%s same_ref=%s exists_ok=%s playback_url=%s",
        entry_type,
        entry_id,
        lang_code,
        expected_fp,
        saved_fp,
        same_ref,
        exists_ok,
        playback_url,
    )
    return bool(same_ref and exists_ok)


def _save_tts_url_to_translation_cache(entry_type: str, entry_id: int, lang_code: str, tts_url: str):
    tts = normalize_text(tts_url or "")
    if not tts:
        return False
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    if entry_type == "word":
        c.execute("""
            UPDATE generated_translations
            SET tts_audio_url=?, updated_at=CURRENT_TIMESTAMP
            WHERE word_id=? AND lang_code=?
        """, (tts, int(entry_id), lang_code))
        c.execute(
            """
            SELECT tts_audio_url
            FROM generated_translations
            WHERE word_id=? AND lang_code=?
            LIMIT 1
            """,
            (int(entry_id), lang_code),
        )
    else:
        c.execute("""
            UPDATE generated_phrase_translations
            SET tts_audio_url=?, updated_at=CURRENT_TIMESTAMP
            WHERE phrase_id=? AND lang_code=?
        """, (tts, int(entry_id), lang_code))
        c.execute(
            """
            SELECT tts_audio_url
            FROM generated_phrase_translations
            WHERE phrase_id=? AND lang_code=?
            LIMIT 1
            """,
            (int(entry_id), lang_code),
        )
    row = c.fetchone()
    conn.commit()
    conn.close()
    return bool(normalize_text((row or [""])[0] or ""))


def _persist_generated_tts_audio(file_name: str, audio_bytes: bytes):
    """
    Persist generated TTS bytes and return (stored_ref, public_url).
    stored_ref is what we write to generated_tts_audio.file_path.
    """
    abs_path = os.path.join(UPLOAD_FOLDER, file_name)
    rel_path = f"uploads/{file_name}"
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    with open(abs_path, "wb") as f:
        f.write(audio_bytes)
    exists_after_write = bool(os.path.isfile(abs_path) and (os.path.getsize(abs_path) > 0))
    playback_url = _public_audio_url(rel_path)
    app.logger.info(
        "persist_generated_tts_audio file_name=%s abs_path=%s exists_after_write=%s file_size=%s db_file_path=%s playback_url=%s",
        file_name,
        abs_path,
        exists_after_write,
        _safe_file_size(abs_path),
        rel_path,
        playback_url,
    )
    if not exists_after_write:
        return "", ""
    return rel_path, playback_url


def _resolve_or_generate_tts_for_text(
    entry_type: str,
    entry_id: int,
    lang_code: str,
    text: str,
    allow_generate: bool = False,
):
    txt = normalize_text(text or "")
    if entry_type not in ("word", "phrase"):
        return ""
    if not txt:
        return ""

    voice_name = _azure_voice_for_lang(lang_code)
    if not voice_name:
        return ""

    cached = _resolve_generated_tts_row(entry_type, int(entry_id), lang_code, txt, voice_name)
    if cached and cached.get("url"):
        return cached["url"]
    if not allow_generate:
        return ""

    speech_key = _get_azure_speech_key()
    speech_region = _get_azure_speech_region()
    if not speech_key or not speech_region:
        return ""

    # Primary path: service-layer DB+Blob workflow.
    service_conn = None
    try:
        service_conn = sqlite3.connect(DB_NAME)
        stored_raw = normalize_text(
            service_generate_and_store_tts(
                db=service_conn,
                entry_type=entry_type,
                entry_id=int(entry_id or 0),
                lang_code=lang_code,
                text=txt,
                speech_key=speech_key,
                speech_region=speech_region,
                voice_name=voice_name,
                speech_lang=_speech_lang_code(lang_code),
                upload_dir=UPLOAD_FOLDER,
                output_filename=_generated_tts_file_name(entry_type, int(entry_id), lang_code, _text_hash(txt), voice_name),
                voice_provider=AZURE_TTS_PROVIDER,
            )
            or ""
        )
        stored = _canonical_local_audio_ref(stored_raw)
        if stored:
            public_url = _public_audio_url(stored)
            if entry_type == "phrase":
                abs_save_path = os.path.join(UPLOAD_FOLDER, os.path.basename(stored))
                exists_after_write = bool(os.path.isfile(abs_save_path))
                file_size = int(_safe_file_size(abs_save_path) or 0)
                app.logger.info(
                    "phrase_tts_write_check entry_id=%s lang=%s abs_save_path=%s exists_after_write=%s file_size=%s db_file_path=%s playback_url=%s",
                    entry_id,
                    lang_code,
                    abs_save_path,
                    exists_after_write,
                    file_size,
                    stored,
                    public_url,
                )
            if not _save_generated_tts_row(entry_type, int(entry_id), lang_code, txt, voice_name, stored):
                app.logger.warning(
                    "tts service save verification failed entry_type=%s entry_id=%s lang=%s stored_raw=%s stored=%s",
                    entry_type,
                    entry_id,
                    lang_code,
                    stored_raw,
                    stored,
                )
                return ""
            app.logger.info(
                "tts service persisted entry_type=%s entry_id=%s lang=%s db_file_path=%s playback_url=%s",
                entry_type,
                entry_id,
                lang_code,
                stored,
                public_url,
            )
            if lang_code in EXTRA_GENERATED_LANGS:
                _save_tts_url_to_translation_cache(entry_type, int(entry_id), lang_code, public_url)
            return public_url
    except Exception as e:
        app.logger.exception(
            "service generate_and_store_tts failed entry_type=%s entry_id=%s lang=%s error=%s",
            entry_type,
            entry_id,
            lang_code,
            repr(e),
        )
    finally:
        if service_conn is not None:
            try:
                service_conn.close()
            except Exception:
                pass

    # Fallback path: legacy in-app synthesis + persistence.
    audio_bytes, error = azure_synthesize_mp3(
        text=txt,
        speech_key=speech_key,
        speech_region=speech_region,
        voice_name=voice_name,
        speech_lang=_speech_lang_code(lang_code),
    )
    if error or not audio_bytes:
        app.logger.warning(
            "TTS generation failed entry_type=%s entry_id=%s lang=%s error=%s",
            entry_type,
            entry_id,
            lang_code,
            error,
        )
        return ""

    text_hash = _text_hash(txt)
    file_name = _generated_tts_file_name(entry_type, int(entry_id), lang_code, text_hash, voice_name)
    try:
        stored_ref, public_url = _persist_generated_tts_audio(file_name, audio_bytes)
        if not stored_ref:
            return ""
        if entry_type == "phrase":
            abs_save_path = os.path.join(UPLOAD_FOLDER, file_name)
            exists_after_write = bool(os.path.isfile(abs_save_path))
            file_size = int(_safe_file_size(abs_save_path) or 0)
            app.logger.info(
                "phrase_tts_write_check entry_id=%s lang=%s abs_save_path=%s exists_after_write=%s file_size=%s db_file_path=%s playback_url=%s",
                entry_id,
                lang_code,
                abs_save_path,
                exists_after_write,
                file_size,
                stored_ref,
                public_url,
            )
        if not _save_generated_tts_row(entry_type, int(entry_id), lang_code, txt, voice_name, stored_ref):
            return ""
        app.logger.info(
            "tts fallback persisted entry_type=%s entry_id=%s lang=%s generated_filename=%s abs_save_path=%s db_file_path=%s playback_url=%s",
            entry_type,
            entry_id,
            lang_code,
            file_name,
            os.path.join(UPLOAD_FOLDER, file_name),
            stored_ref,
            public_url,
        )
        if lang_code in EXTRA_GENERATED_LANGS:
            _save_tts_url_to_translation_cache(entry_type, int(entry_id), lang_code, public_url)
        return public_url
    except Exception as e:
        app.logger.exception(f"Failed to persist TTS audio for {entry_type}:{entry_id}:{lang_code}: {repr(e)}")
        return ""


def _get_saved_generated_tts_audio(entry_type: str, entry_id: int, langs=None, text_by_lang: dict = None):
    if entry_type not in ("word", "phrase"):
        return {}, {"audio_rows_found": 0, "audio_urls_attached": 0}
    requested = {_canonical_tts_lang_code(lc) for lc in (langs or ("en", "om"))}
    requested.discard("")
    if not requested:
        return {}, {"audio_rows_found": 0, "audio_urls_attached": 0}

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute(
        """
        SELECT lang_code, file_path, text_hash
        FROM generated_tts_audio
        WHERE entry_type=? AND entry_id=?
        ORDER BY id DESC
        """,
        (entry_type, int(entry_id)),
    )
    rows = c.fetchall()
    conn.close()

    expected_hash_by_lang = {}
    for lc, txt in (text_by_lang or {}).items():
        canonical = _canonical_tts_lang_code(lc or "")
        normalized_txt = normalize_text(txt or "")
        if canonical and normalized_txt:
            expected_hash_by_lang[canonical] = _text_hash(normalized_txt)

    out = {}
    found_rows = len(rows)
    for lang_code, file_path, text_hash in rows:
        lang = _canonical_tts_lang_code(lang_code or "")
        if lang not in requested:
            continue
        expected_hash = expected_hash_by_lang.get(lang, "")
        if expected_hash and normalize_text(text_hash or "") != expected_hash:
            continue
        key = "english" if lang == "en" else ("oromo" if lang == "om" else lang)
        if key in out:
            continue
        if not _has_usable_audio_ref(file_path or ""):
            continue
        out[key] = _public_audio_url(file_path or "")

    return out, {"audio_rows_found": found_rows, "audio_urls_attached": len(out)}


def _get_saved_audio_for_entry(
    entry_type: str,
    entry_id: int,
    english_text: str = "",
    oromo_text: str = "",
    allow_generate: bool = False,
    return_meta: bool = False,
):
    out = get_approved_audio(entry_type, int(entry_id)) or {}
    gen_audio, gen_meta = _get_saved_generated_tts_audio(
        entry_type,
        int(entry_id),
        langs=("en", "om"),
        text_by_lang={
            "en": normalize_text(english_text or ""),
            "om": normalize_text(oromo_text or ""),
        },
    )
    if gen_audio.get("english") and not out.get("english"):
        out["english"] = gen_audio.get("english", "")
    if gen_audio.get("oromo") and not out.get("oromo"):
        out["oromo"] = gen_audio.get("oromo", "")

    if english_text and not out.get("english"):
        en_url = _resolve_or_generate_tts_for_text(
            entry_type,
            int(entry_id),
            "en",
            english_text,
            allow_generate=allow_generate,
        )
        if en_url:
            out["english"] = en_url
    if oromo_text and not out.get("oromo"):
        om_url = _resolve_or_generate_tts_for_text(
            entry_type,
            int(entry_id),
            "om",
            oromo_text,
            allow_generate=allow_generate,
        )
        if om_url:
            out["oromo"] = om_url
    if return_meta:
        return out, {
            "audio_rows_found": int(gen_meta.get("audio_rows_found", 0) or 0),
            "audio_urls_attached": len([u for u in out.values() if normalize_text(u or "")]),
        }
    return out


def _get_entry_texts_for_tts(entry_type: str, entry_id: int):
    if entry_type not in ("word", "phrase"):
        return {}
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    table = "words" if entry_type == "word" else "phrases"
    c.execute(
        f"SELECT english, oromo FROM {table} WHERE id=? AND status='approved' LIMIT 1",
        (int(entry_id),),
    )
    row = c.fetchone()
    texts = {}
    en = normalize_text((row or ["", ""])[0] or "")
    om = normalize_text((row or ["", ""])[1] or "")
    if en:
        texts["en"] = en
    if om:
        texts["om"] = om

    if entry_type == "word":
        c.execute(
            """
            SELECT lang_code, translated_text
            FROM generated_translations
            WHERE word_id=?
              AND lang_code IN (?, ?, ?, ?)
              AND translated_text IS NOT NULL
              AND TRIM(translated_text) != ''
            """,
            (int(entry_id), "am", "ar", "fr", "zh-CN"),
        )
    else:
        c.execute(
            """
            SELECT lang_code, translated_text
            FROM generated_phrase_translations
            WHERE phrase_id=?
              AND lang_code IN (?, ?, ?, ?)
              AND translated_text IS NOT NULL
              AND TRIM(translated_text) != ''
            """,
            (int(entry_id), "am", "ar", "fr", "zh-CN"),
        )
    for lang, txt in c.fetchall():
        norm = normalize_text(txt or "")
        if _is_meaningful_generated_text(norm):
            texts[lang] = norm
    conn.close()
    return texts


def _is_tts_rate_limited_error(error_text: str) -> bool:
    msg = normalize_text(error_text or "").lower()
    if not msg:
        return False
    return ("429" in msg) or ("too many requests" in msg) or ("throttl" in msg) or ("rate limit" in msg)


def generate_tts_for_entry(
    entry_type: str,
    entry_id: int,
    force_regenerate: bool = False,
    langs=None,
    per_language_delay_ms: int = 0,
    stop_on_429: bool = False,
):
    global _phrase_tts_voice_map_logged
    result = {
        "generated": 0,
        "cached": 0,
        "failed": 0,
        "skipped_missing_text": 0,
        "skipped_missing_voice": 0,
        "by_language": {},
        "rate_limited": False,
    }
    texts = _get_entry_texts_for_tts(entry_type, entry_id)
    speech_key = _get_azure_speech_key()
    speech_region = _get_azure_speech_region()
    if not speech_key or not speech_region:
        app.logger.warning("Azure Speech credentials missing; skipping TTS generation.")
        return result
    if entry_type == "phrase" and (not _phrase_tts_voice_map_logged):
        _phrase_tts_voice_map_logged = True
        app.logger.info(
            "phrase_tts_voice_map provider=%s voices=%s",
            AZURE_TTS_PROVIDER,
            {lc: (DEFAULT_AZURE_VOICES.get(lc) or "") for lc in ("en", "om", "am", "ar", "fr", "zh-CN")},
        )

    selected_langs = tuple(langs or LEARN_TTS_LANGS)
    lang_delay_s = max(0.0, float(int(per_language_delay_ms or 0)) / 1000.0)

    for idx, lang in enumerate(selected_langs):
        if lang_delay_s > 0 and idx > 0:
            time.sleep(lang_delay_s)
        text = normalize_text(texts.get(lang, "") or "")
        if not text:
            result["skipped_missing_text"] += 1
            result["by_language"][lang] = "missing_text"
            app.logger.info(
                "tts lang_skip entry_type=%s entry_id=%s lang=%s reason=missing_text",
                entry_type,
                entry_id,
                lang,
            )
            continue

        voice_name = _azure_voice_for_lang(lang)
        if not voice_name:
            result["skipped_missing_voice"] += 1
            result["by_language"][lang] = "missing_voice"
            app.logger.info(
                "tts lang_skip entry_type=%s entry_id=%s lang=%s reason=missing_voice",
                entry_type,
                entry_id,
                lang,
            )
            continue
        cached = None if force_regenerate else _resolve_generated_tts_row(entry_type, entry_id, lang, text, voice_name)
        if cached:
            result["cached"] += 1
            result["by_language"][lang] = "already_exists"
            app.logger.info(
                "tts lang_skip entry_type=%s entry_id=%s lang=%s reason=already_exists",
                entry_type,
                entry_id,
                lang,
            )
            if lang in EXTRA_GENERATED_LANGS:
                _save_tts_url_to_translation_cache(entry_type, entry_id, lang, cached["url"])
            continue

        service_error_text = ""
        skip_direct_fallback = False
        service_conn = None
        try:
            service_conn = sqlite3.connect(DB_NAME)
            stored_raw = normalize_text(
                service_generate_and_store_tts(
                    db=service_conn,
                    entry_type=entry_type,
                    entry_id=int(entry_id or 0),
                    lang_code=lang,
                    text=text,
                    speech_key=speech_key,
                    speech_region=speech_region,
                    voice_name=voice_name,
                    speech_lang=_speech_lang_code(lang),
                    upload_dir=UPLOAD_FOLDER,
                    output_filename=_generated_tts_file_name(entry_type, int(entry_id), lang, _text_hash(text), voice_name),
                    voice_provider=AZURE_TTS_PROVIDER,
                )
                or ""
            )
            stored = _canonical_local_audio_ref(stored_raw)
            if stored:
                public_url = _public_audio_url(stored)
                if entry_type == "phrase":
                    abs_save_path = os.path.join(UPLOAD_FOLDER, os.path.basename(stored))
                    exists_after_write = bool(os.path.isfile(abs_save_path))
                    file_size = int(_safe_file_size(abs_save_path) or 0)
                    app.logger.info(
                        "phrase_tts_write_check entry_id=%s lang=%s abs_save_path=%s exists_after_write=%s file_size=%s db_file_path=%s playback_url=%s",
                        entry_id,
                        lang,
                        abs_save_path,
                        exists_after_write,
                        file_size,
                        stored,
                        public_url,
                    )
                if not _save_generated_tts_row(entry_type, entry_id, lang, text, voice_name, stored):
                    app.logger.warning(
                        "tts service save verification failed entry_type=%s entry_id=%s lang=%s stored_raw=%s stored=%s",
                        entry_type,
                        entry_id,
                        lang,
                        stored_raw,
                        stored,
                    )
                    result["failed"] += 1
                    result["by_language"][lang] = "failed_persistence_check"
                    continue
                verified = _resolve_generated_tts_row(entry_type, entry_id, lang, text, voice_name)
                if not (verified and verified.get("url")):
                    app.logger.warning(
                        "tts service persistence verification failed entry_type=%s entry_id=%s lang=%s stored=%s",
                        entry_type,
                        entry_id,
                        lang,
                        stored,
                    )
                    result["failed"] += 1
                    result["by_language"][lang] = "failed_persistence_check"
                    continue
                if lang in EXTRA_GENERATED_LANGS:
                    _save_tts_url_to_translation_cache(entry_type, entry_id, lang, public_url)
                result["generated"] += 1
                result["by_language"][lang] = "generated"
                app.logger.info(
                    "tts lang_done entry_type=%s entry_id=%s lang=%s result=generated generated_filename=%s abs_save_path=%s db_file_path=%s playback_url=%s",
                    entry_type,
                    entry_id,
                    lang,
                    os.path.basename(stored),
                    os.path.join(UPLOAD_FOLDER, os.path.basename(stored)),
                    stored,
                    public_url,
                )
                continue
            app.logger.error(
                "service generate_and_store_tts returned empty entry_type=%s entry_id=%s lang=%s upload_folder=%s render_disk_active=%s",
                entry_type,
                entry_id,
                lang,
                UPLOAD_FOLDER,
                IS_RENDER_DISK,
            )
        except Exception as e:
            service_error_text = repr(e)
            app.logger.exception(
                "service generate_and_store_tts failed entry_type=%s entry_id=%s lang=%s error=%s",
                entry_type,
                entry_id,
                lang,
                service_error_text,
            )
            if _is_tts_rate_limited_error(service_error_text):
                result["failed"] += 1
                result["rate_limited"] = True
                result["by_language"][lang] = "failed_429"
                skip_direct_fallback = True
        finally:
            if service_conn is not None:
                try:
                    service_conn.close()
                except Exception:
                    pass

        if skip_direct_fallback:
            if stop_on_429:
                break
            continue

        audio_bytes, error = azure_synthesize_mp3(
            text=text,
            speech_key=speech_key,
            speech_region=speech_region,
            voice_name=voice_name,
            speech_lang=_speech_lang_code(lang),
        )
        if error or not audio_bytes:
            result["failed"] += 1
            if _is_tts_rate_limited_error(error or ""):
                result["rate_limited"] = True
                result["by_language"][lang] = "failed_429"
                if stop_on_429:
                    break
            else:
                result["by_language"][lang] = "failed"
            app.logger.error(
                "TTS generation failed entry_type=%s entry_id=%s lang=%s error=%s speech_key_present=%s speech_region_present=%s upload_folder=%s",
                entry_type,
                entry_id,
                lang,
                error,
                bool(speech_key),
                bool(speech_region),
                UPLOAD_FOLDER,
            )
            continue

        text_hash = _text_hash(text)
        file_name = _generated_tts_file_name(entry_type, entry_id, lang, text_hash, voice_name)
        try:
            stored_ref, public_url = _persist_generated_tts_audio(file_name, audio_bytes)
            if not stored_ref:
                result["failed"] += 1
                continue
            if entry_type == "phrase":
                abs_save_path = os.path.join(UPLOAD_FOLDER, file_name)
                exists_after_write = bool(os.path.isfile(abs_save_path))
                file_size = int(_safe_file_size(abs_save_path) or 0)
                app.logger.info(
                    "phrase_tts_write_check entry_id=%s lang=%s abs_save_path=%s exists_after_write=%s file_size=%s db_file_path=%s playback_url=%s",
                    entry_id,
                    lang,
                    abs_save_path,
                    exists_after_write,
                    file_size,
                    stored_ref,
                    public_url,
                )
            if not _save_generated_tts_row(entry_type, entry_id, lang, text, voice_name, stored_ref):
                result["failed"] += 1
                result["by_language"][lang] = "failed_persistence_check"
                continue
            if lang in EXTRA_GENERATED_LANGS:
                _save_tts_url_to_translation_cache(entry_type, entry_id, lang, public_url)
            result["generated"] += 1
            result["by_language"][lang] = "generated"
            app.logger.info(
                "tts lang_done entry_type=%s entry_id=%s lang=%s result=generated generated_filename=%s abs_save_path=%s db_file_path=%s playback_url=%s",
                entry_type,
                entry_id,
                lang,
                file_name,
                os.path.join(UPLOAD_FOLDER, file_name),
                stored_ref,
                public_url,
            )
        except Exception as e:
            result["failed"] += 1
            result["by_language"][lang] = "failed"
            app.logger.exception(f"Failed to persist TTS audio for {entry_type}:{entry_id}:{lang}: {repr(e)}")

    return result


def run_tts_backfill(
    entry_type: str = "all",
    entry_id: int = 0,
    force_regenerate: bool = False,
    limit: int = 0,
    cancel_check=None,
):
    summary = {
        "processed_items": 0,
        "generated": 0,
        "cached": 0,
        "failed": 0,
        "skipped_missing_text": 0,
        "skipped_missing_voice": 0,
    }
    plans = []
    if entry_type in ("all", "word"):
        items = _fetch_approved_word_items(limit=limit)
        if entry_id:
            items = [(wid, en) for wid, en in items if int(wid) == int(entry_id)]
        plans.extend([("word", wid) for wid, _en in items])
    if entry_type in ("all", "phrase"):
        items = _fetch_approved_phrase_items(limit=limit)
        if entry_id:
            items = [(pid, en) for pid, en in items if int(pid) == int(entry_id)]
        plans.extend([("phrase", pid) for pid, _en in items])

    for etype, eid in plans:
        if callable(cancel_check) and bool(cancel_check()):
            summary["canceled"] = True
            break
        row = generate_tts_for_entry(etype, eid, force_regenerate=force_regenerate)
        summary["processed_items"] += 1
        for key in ("generated", "cached", "failed", "skipped_missing_text", "skipped_missing_voice"):
            summary[key] += int(row.get(key, 0) or 0)
    return summary


def run_generated_tts_blob_migration(limit: int = 0, chunk_size: int = 100, dry_run: bool = False):
    """
    One-time storage migration:
    - read generated_tts_audio rows
    - upload existing local files to Azure Blob
    - update generated_tts_audio.file_path to blob URL
    No TTS regeneration is performed.
    """
    summary = {
        "rows_scanned": 0,
        "rows_migrated": 0,
        "rows_migration_candidates": 0,
        "rows_missing_file": 0,
        "rows_already_blob_backed": 0,
        "failures": 0,
        "dry_run": bool(dry_run),
    }

    if not _azure_blob_enabled():
        app.logger.error(
            "migrate-tts-to-blob skipped: Azure Blob not configured "
            "(AZURE_BLOB_CONNECTION_STRING / AZURE_BLOB_CONTAINER)."
        )
        return summary

    safe_chunk = max(1, min(int(chunk_size or 100), 1000))
    safe_limit = max(0, int(limit or 0))

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    sql = """
        SELECT id, entry_type, entry_id, lang_code, file_path
        FROM generated_tts_audio
        WHERE file_path IS NOT NULL AND TRIM(file_path) != ''
        ORDER BY id ASC
    """
    params = ()
    if safe_limit > 0:
        sql += " LIMIT ?"
        params = (safe_limit,)
    c.execute(sql, params)

    while True:
        rows = c.fetchmany(safe_chunk)
        if not rows:
            break

        for row_id, entry_type, entry_id, lang_code, file_path in rows:
            summary["rows_scanned"] += 1
            fp = (file_path or "").strip()

            if _is_remote_audio_ref(fp):
                summary["rows_already_blob_backed"] += 1
                continue

            abs_path = _audio_abs_path(fp)
            if not abs_path or (not os.path.isfile(abs_path)):
                summary["rows_missing_file"] += 1
                app.logger.warning(
                    "migrate-tts-to-blob missing local file row_id=%s entry=%s:%s lang=%s file_path=%s",
                    row_id,
                    entry_type,
                    entry_id,
                    lang_code,
                    fp,
                )
                continue

            summary["rows_migration_candidates"] += 1
            if dry_run:
                continue

            try:
                with open(abs_path, "rb") as fh:
                    audio_bytes = fh.read()
                if not audio_bytes:
                    summary["failures"] += 1
                    app.logger.warning(
                        "migrate-tts-to-blob empty local file row_id=%s abs_path=%s",
                        row_id,
                        abs_path,
                    )
                    continue

                blob_url = _upload_tts_bytes_to_blob(os.path.basename(abs_path), audio_bytes)
                if not blob_url:
                    summary["failures"] += 1
                    continue

                c.execute(
                    "UPDATE generated_tts_audio SET file_path=? WHERE id=?",
                    (blob_url, int(row_id)),
                )

                # Keep translation cache URLs aligned for direct dictionary/translate lookups.
                if lang_code in EXTRA_GENERATED_LANGS:
                    if entry_type == "word":
                        c.execute(
                            """
                            UPDATE generated_translations
                            SET tts_audio_url=?, updated_at=CURRENT_TIMESTAMP
                            WHERE word_id=? AND lang_code=?
                            """,
                            (blob_url, int(entry_id or 0), lang_code),
                        )
                    elif entry_type == "phrase":
                        c.execute(
                            """
                            UPDATE generated_phrase_translations
                            SET tts_audio_url=?, updated_at=CURRENT_TIMESTAMP
                            WHERE phrase_id=? AND lang_code=?
                            """,
                            (blob_url, int(entry_id or 0), lang_code),
                        )

                summary["rows_migrated"] += 1
            except Exception:
                summary["failures"] += 1
                app.logger.exception(
                    "migrate-tts-to-blob failed row_id=%s entry=%s:%s lang=%s",
                    row_id,
                    entry_type,
                    entry_id,
                    lang_code,
                )

        conn.commit()

    conn.close()
    return summary


def run_backfill_existing_audio_linkage(
    limit: int = 0,
    dry_run: bool = False,
    source_dirs=None,
    promote_to_uploads: bool = True,
    cancel_check=None,
):
    """
    DB/file linkage backfill only.
    - Registers existing tts_*.mp3 files into generated_tts_audio.
    - Syncs generated translation caches with usable persisted audio URLs.
    - Never calls Azure Speech, never regenerates audio bytes.
    """
    summary = {
        "files_scanned": 0,
        "files_promoted": 0,
        "files_already_in_uploads": 0,
        "files_promotion_failed": 0,
        "rows_linked": 0,
        "rows_already_present": 0,
        "rows_skipped_missing_text": 0,
        "rows_skipped_hash_mismatch": 0,
        "rows_skipped_missing_file": 0,
        "cache_rows_scanned": 0,
        "cache_rows_linked": 0,
        "dry_run": bool(dry_run),
    }

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    c.execute("SELECT id, english, oromo FROM words WHERE status='approved'")
    words = {int(wid): {"en": normalize_text(en or ""), "om": normalize_text(om or "")} for wid, en, om in c.fetchall()}
    c.execute("SELECT id, english, oromo FROM phrases WHERE status='approved'")
    phrases = {int(pid): {"en": normalize_text(en or ""), "om": normalize_text(om or "")} for pid, en, om in c.fetchall()}

    word_tr = {}
    c.execute(
        """
        SELECT word_id, lang_code, translated_text
        FROM generated_translations
        WHERE lang_code IN ('am', 'ar', 'fr', 'zh-CN')
          AND translated_text IS NOT NULL
          AND TRIM(translated_text) != ''
        """
    )
    for wid, lang, txt in c.fetchall():
        wid_int = int(wid or 0)
        norm = normalize_text(txt or "")
        if wid_int and norm:
            word_tr.setdefault(wid_int, {})[lang] = norm

    phrase_tr = {}
    c.execute(
        """
        SELECT phrase_id, lang_code, translated_text
        FROM generated_phrase_translations
        WHERE lang_code IN ('am', 'ar', 'fr', 'zh-CN')
          AND translated_text IS NOT NULL
          AND TRIM(translated_text) != ''
        """
    )
    for pid, lang, txt in c.fetchall():
        pid_int = int(pid or 0)
        norm = normalize_text(txt or "")
        if pid_int and norm:
            phrase_tr.setdefault(pid_int, {})[lang] = norm

    c.execute(
        """
        SELECT entry_type, entry_id, lang_code, text_hash, voice_provider, voice_name
        FROM generated_tts_audio
        """
    )
    existing_keys = {
        (
            entry_type,
            int(entry_id or 0),
            lang_code,
            text_hash,
            voice_provider,
            voice_name,
        )
        for entry_type, entry_id, lang_code, text_hash, voice_provider, voice_name in c.fetchall()
    }
    conn.close()

    folders = [UPLOAD_FOLDER, STATIC_UPLOADS_FOLDER]
    if source_dirs:
        for d in source_dirs:
            dn = normalize_text(d or "")
            if dn and dn not in folders:
                folders.append(dn)

    candidate_names = set()
    for folder in folders:
        if not os.path.isdir(folder):
            continue
        try:
            for name in os.listdir(folder):
                if name.startswith("tts_") and name.endswith(".mp3"):
                    candidate_names.add(name)
        except Exception:
            app.logger.exception("audio linkage backfill failed listing folder: %s", folder)

    names = sorted(candidate_names)
    if limit and int(limit) > 0:
        names = names[: int(limit)]

    for name in names:
        if callable(cancel_check) and bool(cancel_check()):
            summary["canceled"] = True
            break
        summary["files_scanned"] += 1
        m = GENERATED_TTS_FILENAME_RE.match(name)
        if not m:
            continue
        entry_type, entry_id_raw, lang_code, hash12, voice_name = m.groups()
        entry_id = int(entry_id_raw or 0)
        base_map = words if entry_type == "word" else phrases
        tr_map = word_tr if entry_type == "word" else phrase_tr
        text_value = ""
        if lang_code in ("en", "om"):
            text_value = normalize_text((base_map.get(entry_id, {}) or {}).get(lang_code, "") or "")
        elif lang_code in EXTRA_GENERATED_LANGS:
            text_value = normalize_text((tr_map.get(entry_id, {}) or {}).get(lang_code, "") or "")
        if not text_value:
            summary["rows_skipped_missing_text"] += 1
            continue

        full_hash = _text_hash(text_value)
        if not full_hash.startswith(hash12):
            summary["rows_skipped_hash_mismatch"] += 1
            continue

        uploads_abs = os.path.join(UPLOAD_FOLDER, name)
        source_abs = ""
        if os.path.isfile(uploads_abs):
            source_abs = uploads_abs
            summary["files_already_in_uploads"] += 1
        else:
            for folder in folders:
                cand = os.path.join(folder, name)
                if os.path.isfile(cand):
                    source_abs = cand
                    break

        if not source_abs:
            summary["rows_skipped_missing_file"] += 1
            continue

        if (os.path.abspath(source_abs) != os.path.abspath(uploads_abs)) and promote_to_uploads:
            if dry_run:
                summary["files_promoted"] += 1
            else:
                try:
                    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
                    shutil.copy2(source_abs, uploads_abs)
                    summary["files_promoted"] += 1
                    source_abs = uploads_abs
                except Exception:
                    summary["files_promotion_failed"] += 1
                    app.logger.exception("audio linkage promote failed for %s from %s", name, source_abs)
                    summary["rows_skipped_missing_file"] += 1
                    continue

        if not os.path.isfile(uploads_abs):
            # Keep generated_tts_audio paths stable for live serving (/uploads/...).
            summary["rows_skipped_missing_file"] += 1
            continue
        file_ref = f"uploads/{name}"

        key = (entry_type, int(entry_id), lang_code, full_hash, AZURE_TTS_PROVIDER, voice_name)
        if key in existing_keys:
            summary["rows_already_present"] += 1
            continue

        if not dry_run:
            _save_generated_tts_row(entry_type, int(entry_id), lang_code, text_value, voice_name, file_ref)
            if lang_code in EXTRA_GENERATED_LANGS:
                _save_tts_url_to_translation_cache(entry_type, int(entry_id), lang_code, _public_audio_url(file_ref))
        existing_keys.add(key)
        summary["rows_linked"] += 1

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute(
        """
        SELECT word_id, lang_code, translated_text, tts_audio_url
        FROM generated_translations
        WHERE lang_code IN ('am', 'ar', 'fr', 'zh-CN')
          AND translated_text IS NOT NULL
          AND TRIM(translated_text) != ''
          AND tts_audio_url IS NOT NULL
          AND TRIM(tts_audio_url) != ''
        """
    )
    cache_rows_word = [("word", int(wid or 0), lang, normalize_text(txt or ""), normalize_text(url or "")) for wid, lang, txt, url in c.fetchall()]
    c.execute(
        """
        SELECT phrase_id, lang_code, translated_text, tts_audio_url
        FROM generated_phrase_translations
        WHERE lang_code IN ('am', 'ar', 'fr', 'zh-CN')
          AND translated_text IS NOT NULL
          AND TRIM(translated_text) != ''
          AND tts_audio_url IS NOT NULL
          AND TRIM(tts_audio_url) != ''
        """
    )
    cache_rows_phrase = [("phrase", int(pid or 0), lang, normalize_text(txt or ""), normalize_text(url or "")) for pid, lang, txt, url in c.fetchall()]
    conn.close()

    for entry_type, entry_id, lang_code, translated_text, tts_url in (cache_rows_word + cache_rows_phrase):
        if callable(cancel_check) and bool(cancel_check()):
            summary["canceled"] = True
            break
        summary["cache_rows_scanned"] += 1
        if (not entry_id) or (not translated_text):
            continue
        normalized_url = _normalize_cached_tts_url(tts_url)
        if not normalized_url:
            continue
        if not _has_usable_audio_ref(normalized_url):
            continue
        storage_ref = normalized_url if _is_remote_audio_ref(normalized_url) else f"uploads/{os.path.basename(normalized_url)}"
        voice_name = _azure_voice_for_lang(lang_code) or "unknown"
        if not dry_run:
            _save_generated_tts_row(entry_type, int(entry_id), lang_code, translated_text, voice_name, storage_ref)
            _save_tts_url_to_translation_cache(entry_type, int(entry_id), lang_code, _public_audio_url(storage_ref))
        summary["cache_rows_linked"] += 1

    return summary


def ensure_missing_tts_for_words(
    word_items,
    force_regenerate: bool = False,
    chunk_size: int = None,
    per_entry_delay_ms: int = 0,
    log_context: str = "word_tts_backfill",
    cancel_check=None,
):
    summary = {
        "words_seen": 0,
        "processed_new": 0,
        "skipped_existing": 0,
        "generated": 0,
        "cached": 0,
        "failed": 0,
        "skipped_missing_text": 0,
        "skipped_missing_voice": 0,
    }
    if not word_items:
        return summary

    safe_chunk = int(chunk_size or IMPORT_BATCH_SIZE or 50)
    if safe_chunk < 1:
        safe_chunk = 50
    delay_s = max(0.0, float(int(per_entry_delay_ms or 0)) / 1000.0 )
    unique_items = []
    seen = set()
    for wid, en in word_items:
        wid_int = int(wid or 0)
        en_norm = normalize_text(en or "")
        if not wid_int or not en_norm or wid_int in seen:
            continue
        seen.add(wid_int)
        unique_items.append((wid_int, en_norm))

    if not unique_items:
        return summary

    for i in range(0, len(unique_items), safe_chunk):
        chunk = unique_items[i:i + safe_chunk]
        for wid, _en in chunk:
            if callable(cancel_check) and bool(cancel_check()):
                summary["canceled"] = True
                return summary
            summary["words_seen"] += 1
            if not force_regenerate:
                texts = _get_entry_texts_for_tts("word", wid)
                missing_langs = []
                for lang in LEARN_TTS_LANGS:
                    text = normalize_text((texts or {}).get(lang, "") or "")
                    if not text:
                        continue
                    voice_name = _azure_voice_for_lang(lang)
                    if not voice_name:
                        continue
                    if not _resolve_generated_tts_row("word", wid, lang, text, voice_name):
                        missing_langs.append(lang)
                if not missing_langs:
                    summary["skipped_existing"] += 1
                    app.logger.info(
                        "%s skipped_existing type=tts entry_type=word entry_id=%s",
                        log_context,
                        wid,
                    )
                    continue
            summary["processed_new"] += 1
            app.logger.info(
                "%s processed_new type=tts entry_type=word entry_id=%s",
                log_context,
                wid,
            )
            row = generate_tts_for_entry("word", wid, force_regenerate=force_regenerate)
            for key in ("generated", "cached", "failed", "skipped_missing_text", "skipped_missing_voice"):
                summary[key] += int(row.get(key, 0) or 0)
            if delay_s > 0:
                time.sleep(delay_s)
        app.logger.info(
            "%s chunk_done start=%s size=%s processed_new=%s skipped_existing=%s generated=%s cached=%s failed=%s",
            log_context,
            i,
            len(chunk),
            summary["processed_new"],
            summary["skipped_existing"],
            summary["generated"],
            summary["cached"],
            summary["failed"],
        )
    return summary


def ensure_missing_tts_for_phrases(
    phrase_items,
    force_regenerate: bool = False,
    chunk_size: int = None,
    per_entry_delay_ms: int = 0,
    log_context: str = "phrase_tts_backfill",
    cancel_check=None,
):
    summary = {
        "phrases_seen": 0,
        "processed_new": 0,
        "skipped_existing": 0,
        "generated": 0,
        "cached": 0,
        "failed": 0,
        "skipped_missing_text": 0,
        "skipped_missing_voice": 0,
        "generated_by_language": {lang: 0 for lang in LEARN_TTS_LANGS},
    }
    if not phrase_items:
        return summary

    safe_chunk = int(chunk_size or IMPORT_BATCH_SIZE or 50)
    if safe_chunk < 1:
        safe_chunk = 50
    delay_s = max(0.0, float(int(per_entry_delay_ms or 0)) / 1000.0 )
    unique_items = []
    seen = set()
    for pid, en in phrase_items:
        pid_int = int(pid or 0)
        en_norm = normalize_text(en or "")
        if not pid_int or not en_norm or pid_int in seen:
            continue
        seen.add(pid_int)
        unique_items.append((pid_int, en_norm))

    if not unique_items:
        return summary

    for i in range(0, len(unique_items), safe_chunk):
        chunk = unique_items[i:i + safe_chunk]
        for pid, _en in chunk:
            if callable(cancel_check) and bool(cancel_check()):
                summary["canceled"] = True
                return summary
            summary["phrases_seen"] += 1
            if not force_regenerate:
                texts = _get_entry_texts_for_tts("phrase", pid)
                missing_langs = []
                lang_status = {}
                for lang in LEARN_TTS_LANGS:
                    text = normalize_text((texts or {}).get(lang, "") or "")
                    if not text:
                        lang_status[lang] = "missing_text"
                        app.logger.info(
                            "%s phrase_tts_lang_status entry_id=%s lang=%s reason=missing_text",
                            log_context,
                            pid,
                            lang,
                        )
                        continue
                    voice_name = _azure_voice_for_lang(lang)
                    if not voice_name:
                        lang_status[lang] = "missing_voice"
                        app.logger.info(
                            "%s phrase_tts_lang_status entry_id=%s lang=%s reason=missing_voice",
                            log_context,
                            pid,
                            lang,
                        )
                        continue
                    if not _resolve_generated_tts_row("phrase", pid, lang, text, voice_name):
                        missing_langs.append(lang)
                        lang_status[lang] = "missing_audio"
                        app.logger.info(
                            "%s phrase_tts_lang_status entry_id=%s lang=%s reason=missing_audio",
                            log_context,
                            pid,
                            lang,
                        )
                    else:
                        lang_status[lang] = "already_exists"
                        app.logger.info(
                            "%s phrase_tts_lang_status entry_id=%s lang=%s reason=already_exists",
                            log_context,
                            pid,
                            lang,
                        )
                if not missing_langs:
                    summary["skipped_existing"] += 1
                    app.logger.info(
                        "%s skipped_existing type=tts entry_type=phrase entry_id=%s lang_status=%s",
                        log_context,
                        pid,
                        lang_status,
                    )
                    continue
            summary["processed_new"] += 1
            app.logger.info(
                "%s processed_new type=tts entry_type=phrase entry_id=%s",
                log_context,
                pid,
            )
            row = generate_tts_for_entry("phrase", pid, force_regenerate=force_regenerate)
            for key in ("generated", "cached", "failed", "skipped_missing_text", "skipped_missing_voice"):
                summary[key] += int(row.get(key, 0) or 0)
            by_lang = (row or {}).get("by_language", {}) or {}
            for lang, state in by_lang.items():
                app.logger.info(
                    "%s phrase_tts_lang_status entry_id=%s lang=%s reason=%s",
                    log_context,
                    pid,
                    lang,
                    state,
                )
                if state == "generated":
                    summary["generated_by_language"][lang] = int(summary["generated_by_language"].get(lang, 0) or 0) + 1
            if delay_s > 0:
                time.sleep(delay_s)
        app.logger.info(
            "%s chunk_done start=%s size=%s processed_new=%s skipped_existing=%s generated=%s cached=%s failed=%s",
            log_context,
            i,
            len(chunk),
            summary["processed_new"],
            summary["skipped_existing"],
            summary["generated"],
            summary["cached"],
            summary["failed"],
        )
    return summary


def _new_pipeline_summary(entry_type: str):
    return {
        "entry_type": entry_type,
        "inserted": 0,
        "updated": 0,
        "imported_with_oromo": 0,
        "imported_missing_oromo": 0,
        "scanned_missing_oromo": 0,
        "filled_oromo": 0,
        "failed_oromo_fill": 0,
        "translations_attempted": 0,
        "translations_generated": 0,
        "translations_skipped_existing": 0,
        "translations_failed": 0,
        "audio_attempted": 0,
        "audio_generated": 0,
        "audio_skipped_existing": 0,
        "audio_missing_voice": 0,
        "audio_failed": 0,
        "learn_ready_phrases_with_audio": 0,
        "learn_visible_phrase_count": 0,
    }


def _count_phrases_with_any_audio(phrase_ids):
    ids = [int(x) for x in (phrase_ids or []) if int(x or 0) > 0]
    if not ids:
        return 0
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    marks = ",".join("?" for _ in ids)

    c.execute(
        f"""
        SELECT DISTINCT entry_id, file_path
        FROM generated_tts_audio
        WHERE entry_type='phrase'
          AND entry_id IN ({marks})
          AND file_path IS NOT NULL
          AND TRIM(file_path) != ''
        """,
        tuple(ids),
    )
    generated_rows = c.fetchall()

    c.execute(
        f"""
        SELECT DISTINCT entry_id, file_path
        FROM audio
        WHERE status='approved'
          AND entry_type='phrase'
          AND entry_id IN ({marks})
          AND file_path IS NOT NULL
          AND TRIM(file_path) != ''
        """,
        tuple(ids),
    )
    approved_rows = c.fetchall()
    conn.close()

    ready_ids = set()
    for eid, fp in (generated_rows or []):
        if _has_usable_audio_ref(fp or ""):
            ready_ids.add(int(eid or 0))
    for eid, fp in (approved_rows or []):
        if _has_usable_audio_ref(fp or ""):
            ready_ids.add(int(eid or 0))
    return len([x for x in ready_ids if x > 0])


def _count_recent_visible_learn_phrases(limit: int):
    safe_limit = int(limit or 0)
    if safe_limit <= 0:
        return 0
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute(
        """
        SELECT COUNT(*)
        FROM (
            SELECT id
            FROM phrases
            WHERE status='approved'
              AND english IS NOT NULL
              AND TRIM(english) != ''
            ORDER BY id DESC
            LIMIT ?
        ) recent
        """,
        (safe_limit,),
    )
    out = int((c.fetchone() or [0])[0] or 0)
    conn.close()
    return out


def _run_post_import_pipeline(
    word_ids,
    phrase_ids,
    chunk_size: int = None,
    import_summary: dict = None,
    cancel_check=None,
):
    safe_chunk = int(chunk_size or IMPORT_BATCH_SIZE or 50)
    if safe_chunk < 1:
        safe_chunk = 50

    word_ids = [int(x) for x in (word_ids or []) if int(x or 0) > 0]
    phrase_ids = [int(x) for x in (phrase_ids or []) if int(x or 0) > 0]
    import_summary = import_summary or {}
    word_import_summary = dict((import_summary or {}).get("word", {}) or {})
    phrase_import_summary = dict((import_summary or {}).get("phrase", {}) or {})
    words_summary = _new_pipeline_summary("word")
    phrases_summary = _new_pipeline_summary("phrase")
    for k in ("inserted", "updated", "imported_with_oromo", "imported_missing_oromo"):
        words_summary[k] = int(word_import_summary.get(k, 0) or 0)
        phrases_summary[k] = int(phrase_import_summary.get(k, 0) or 0)
    app.logger.info(
        "post_import_pipeline started runtime=%s db_path=%s upload_folder=%s word_ids=%s phrase_ids=%s import_summary=%s",
        APP_RUNTIME,
        DB_NAME,
        UPLOAD_FOLDER,
        len(word_ids),
        len(phrase_ids),
        {"word": words_summary, "phrase": phrases_summary},
    )
    app.logger.info("post_import_pipeline provider_health=%s", _provider_health_snapshot())

    pipeline_ok = True
    word_tr_stats = {}
    phrase_tr_stats = {}
    conn = None
    cursor = None
    try:
        # Background thread must use its own SQLite connection and cursor.
        conn = sqlite3.connect(app.DB_NAME)
        cursor = conn.cursor()

        def _fetch_items_by_ids(table_name, ids):
            if not ids:
                return []

            ordered_ids = []
            seen_ids = set()
            for raw_id in ids:
                rid = int(raw_id or 0)
                if rid > 0 and rid not in seen_ids:
                    ordered_ids.append(rid)
                    seen_ids.add(rid)

            if not ordered_ids:
                return []

            rows_by_id = {}
            for i in range(0, len(ordered_ids), safe_chunk):
                chunk_ids = ordered_ids[i:i + safe_chunk]
                placeholders = ",".join("?" for _ in chunk_ids)
                cursor.execute(
                    f"""
                    SELECT id, english
                    FROM {table_name}
                    WHERE status='approved'
                      AND id IN ({placeholders})
                      AND english IS NOT NULL
                      AND TRIM(english) != ''
                    """,
                    tuple(chunk_ids),
                )
                for rid, english in cursor.fetchall():
                    rid_int = int(rid or 0)
                    english_norm = normalize_text(english or "")
                    if rid_int and english_norm:
                        rows_by_id[rid_int] = english_norm

            return [(rid, rows_by_id[rid]) for rid in ordered_ids if rid in rows_by_id]

        # Re-fetch from DB inside this thread; do not use caller-owned row objects.
        word_items = _fetch_items_by_ids("words", word_ids)
        phrase_items = _fetch_items_by_ids("phrases", phrase_ids)
        if callable(cancel_check) and bool(cancel_check()):
            return {
                "ok": False,
                "canceled": True,
                "completed_with_provider_failures": False,
                "provider_failure_summary": {},
                "words_summary": words_summary,
                "phrases_summary": phrases_summary,
            }

        app.logger.info(
            "post_import_pipeline voice_map provider=%s voices=%s",
            AZURE_TTS_PROVIDER,
            {lc: (DEFAULT_AZURE_VOICES.get(lc) or "") for lc in ("en", "om", "am", "ar", "fr", "zh-CN")},
        )

        # Fill Oromo base field first for imported entries.
        if word_items:
            word_oromo_summary = ensure_missing_oromo_for_entries(
                "word",
                word_items,
                chunk_size=safe_chunk,
                log_context="post_import_pipeline",
                cancel_check=cancel_check,
            )
            words_summary["scanned_missing_oromo"] = int(word_oromo_summary.get("scanned_missing_oromo", 0) or 0)
            words_summary["filled_oromo"] = int(word_oromo_summary.get("filled_oromo", 0) or 0)
            words_summary["failed_oromo_fill"] = int(word_oromo_summary.get("failed_oromo_fill", 0) or 0)
            if bool(word_oromo_summary.get("canceled")):
                return {
                    "ok": False,
                    "canceled": True,
                    "words_summary": words_summary,
                    "phrases_summary": phrases_summary,
                }
        if phrase_items:
            phrase_oromo_summary = ensure_missing_oromo_for_entries(
                "phrase",
                phrase_items,
                chunk_size=safe_chunk,
                log_context="post_import_pipeline",
                cancel_check=cancel_check,
            )
            phrases_summary["scanned_missing_oromo"] = int(phrase_oromo_summary.get("scanned_missing_oromo", 0) or 0)
            phrases_summary["filled_oromo"] = int(phrase_oromo_summary.get("filled_oromo", 0) or 0)
            phrases_summary["failed_oromo_fill"] = int(phrase_oromo_summary.get("failed_oromo_fill", 0) or 0)
            if bool(phrase_oromo_summary.get("canceled")):
                return {
                    "ok": False,
                    "canceled": True,
                    "words_summary": words_summary,
                    "phrases_summary": phrases_summary,
                }

        if word_items:
            words_saved, word_tr_stats = ensure_missing_generated_translations_for_words(
                word_items,
                langs=EXTRA_GENERATED_LANGS,
                chunk_size=safe_chunk,
                log_context="post_import_pipeline_words",
                cancel_check=cancel_check,
            )
            words_summary["translations_generated"] = int(words_saved or 0)
            words_summary["translations_attempted"] = int(sum(int((st or {}).get("attempted", 0) or 0) for st in (word_tr_stats or {}).values()))
            words_summary["translations_skipped_existing"] = int(sum(int((st or {}).get("already_cached", 0) or 0) for st in (word_tr_stats or {}).values()))
            words_summary["translations_failed"] = int(
                sum(
                    int((st or {}).get("provider_errors", 0) or 0) +
                    int((st or {}).get("failed_db_write", 0) or 0)
                    for st in (word_tr_stats or {}).values()
                )
            )
            if bool(((word_tr_stats or {}).get("__meta__") or {}).get("canceled")):
                return {
                    "ok": False,
                    "canceled": True,
                    "words_summary": words_summary,
                    "phrases_summary": phrases_summary,
                }
        if phrase_items:
            phrases_saved, phrase_tr_stats = ensure_missing_generated_translations_for_phrases(
                phrase_items,
                langs=EXTRA_GENERATED_LANGS,
                chunk_size=safe_chunk,
                log_context="post_import_pipeline_phrases",
                cancel_check=cancel_check,
            )
            phrases_summary["translations_generated"] = int(phrases_saved or 0)
            phrases_summary["translations_attempted"] = int(sum(int((st or {}).get("attempted", 0) or 0) for st in (phrase_tr_stats or {}).values()))
            phrases_summary["translations_skipped_existing"] = int(sum(int((st or {}).get("already_cached", 0) or 0) for st in (phrase_tr_stats or {}).values()))
            phrases_summary["translations_failed"] = int(
                sum(
                    int((st or {}).get("provider_errors", 0) or 0) +
                    int((st or {}).get("failed_db_write", 0) or 0)
                    for st in (phrase_tr_stats or {}).values()
                )
            )
            if bool(((phrase_tr_stats or {}).get("__meta__") or {}).get("canceled")):
                return {
                    "ok": False,
                    "canceled": True,
                    "words_summary": words_summary,
                    "phrases_summary": phrases_summary,
                }

        tts_chunk = max(1, min(int(safe_chunk), int(TTS_JOB_CHUNK_SIZE)))
        words_tts = ensure_missing_tts_for_words(
            word_items,
            force_regenerate=False,
            chunk_size=tts_chunk,
            per_entry_delay_ms=TTS_JOB_ENTRY_DELAY_MS,
            log_context="post_import_pipeline_tts_words",
            cancel_check=cancel_check,
        ) if word_items else {}
        phrases_tts = ensure_missing_tts_for_phrases(
            phrase_items,
            force_regenerate=False,
            chunk_size=tts_chunk,
            per_entry_delay_ms=TTS_JOB_ENTRY_DELAY_MS,
            log_context="post_import_pipeline_tts_phrases",
            cancel_check=cancel_check,
        ) if phrase_items else {}
        words_summary["audio_generated"] = int((words_tts or {}).get("generated", 0) or 0)
        words_summary["audio_skipped_existing"] = int((words_tts or {}).get("skipped_existing", 0) or 0)
        words_summary["audio_missing_voice"] = int((words_tts or {}).get("skipped_missing_voice", 0) or 0)
        words_summary["audio_failed"] = int((words_tts or {}).get("failed", 0) or 0)
        words_summary["audio_attempted"] = int(
            words_summary["audio_generated"] +
            int((words_tts or {}).get("cached", 0) or 0) +
            words_summary["audio_missing_voice"] +
            int((words_tts or {}).get("skipped_missing_text", 0) or 0) +
            words_summary["audio_failed"]
        )

        phrases_summary["audio_generated"] = int((phrases_tts or {}).get("generated", 0) or 0)
        phrases_summary["audio_skipped_existing"] = int((phrases_tts or {}).get("skipped_existing", 0) or 0)
        phrases_summary["audio_missing_voice"] = int((phrases_tts or {}).get("skipped_missing_voice", 0) or 0)
        phrases_summary["audio_failed"] = int((phrases_tts or {}).get("failed", 0) or 0)
        phrases_summary["audio_attempted"] = int(
            phrases_summary["audio_generated"] +
            int((phrases_tts or {}).get("cached", 0) or 0) +
            phrases_summary["audio_missing_voice"] +
            int((phrases_tts or {}).get("skipped_missing_text", 0) or 0) +
            phrases_summary["audio_failed"]
        )
        phrases_summary["learn_ready_phrases_with_audio"] = _count_phrases_with_any_audio(phrase_ids)
        phrases_summary["learn_visible_phrase_count"] = _count_recent_visible_learn_phrases(LEARN_RECENT_PHRASE_LIMIT)
        if bool((words_tts or {}).get("canceled")) or bool((phrases_tts or {}).get("canceled")):
            return {
                "ok": False,
                "canceled": True,
                "words_summary": words_summary,
                "phrases_summary": phrases_summary,
            }

        words_google_provider_errors = int(sum(int((st or {}).get("google_provider_errors", (st or {}).get("provider_errors", 0)) or 0) for st in (word_tr_stats or {}).values()))
        phrases_google_provider_errors = int(sum(int((st or {}).get("google_provider_errors", (st or {}).get("provider_errors", 0)) or 0) for st in (phrase_tr_stats or {}).values()))
        words_google_empty_result = int(sum(int((st or {}).get("google_empty_result", (st or {}).get("empty_results", 0)) or 0) for st in (word_tr_stats or {}).values()))
        phrases_google_empty_result = int(sum(int((st or {}).get("google_empty_result", (st or {}).get("empty_results", 0)) or 0) for st in (phrase_tr_stats or {}).values()))
        words_google_request_failures = int(sum(int((st or {}).get("google_request_failures", (st or {}).get("request_failures", 0)) or 0) for st in (word_tr_stats or {}).values()))
        phrases_google_request_failures = int(sum(int((st or {}).get("google_request_failures", (st or {}).get("request_failures", 0)) or 0) for st in (phrase_tr_stats or {}).values()))
        provider_failure_summary = {
            "google_provider_errors": int(words_google_provider_errors + phrases_google_provider_errors),
            "google_empty_result": int(words_google_empty_result + phrases_google_empty_result),
            "google_request_failures": int(words_google_request_failures + phrases_google_request_failures),
            "azure_tts_failures": int(words_summary.get("audio_failed", 0) + phrases_summary.get("audio_failed", 0)),
        }
        completed_with_provider_failures = bool(
            int(provider_failure_summary.get("google_provider_errors", 0) or 0) > 0
            or int(provider_failure_summary.get("azure_tts_failures", 0) or 0) > 0
        )
        if completed_with_provider_failures:
            app.logger.warning(
                "post_import_pipeline completed_with_provider_failures provider_failure_summary=%s words=%s phrases=%s",
                provider_failure_summary,
                words_summary,
                phrases_summary,
            )
        result_payload = {
            "ok": True,
            "completed_with_provider_failures": completed_with_provider_failures,
            "provider_failure_summary": provider_failure_summary,
            "words_summary": words_summary,
            "phrases_summary": phrases_summary,
        }

        app.logger.info("post_import_pipeline summary words=%s", words_summary)
        app.logger.info("post_import_pipeline summary phrases=%s", phrases_summary)
        conn.commit()
    except Exception:
        pipeline_ok = False
        result_payload = {
            "ok": False,
            "completed_with_provider_failures": False,
            "provider_failure_summary": {},
            "words_summary": words_summary,
            "phrases_summary": phrases_summary,
        }
        app.logger.exception("Post-import pipeline failed with unhandled exception")
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
    finally:
        if cursor is not None:
            try:
                cursor.close()
            except Exception:
                pass
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass
    if not pipeline_ok:
        return result_payload
    if "result_payload" not in locals():
        result_payload = {
            "ok": True,
            "completed_with_provider_failures": False,
            "provider_failure_summary": {},
            "words_summary": words_summary,
            "phrases_summary": phrases_summary,
        }
    return result_payload


def _fetch_approved_items_for_audio_regen(entry_type: str, limit: int = 0):
    if entry_type not in ("word", "phrase"):
        return []
    table = "words" if entry_type == "word" else "phrases"
    safe_limit = int(limit or 0)
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    sql = (
        f"""
        SELECT id, english
        FROM {table}
        WHERE status='approved'
          AND english IS NOT NULL
          AND TRIM(english) != ''
        ORDER BY id DESC
        """
    )
    params = []
    if safe_limit > 0:
        sql += " LIMIT ?"
        params.append(safe_limit)
    c.execute(sql, tuple(params))
    rows = [(int(eid or 0), normalize_text(en or "")) for eid, en in c.fetchall()]
    conn.close()
    return [(eid, en) for eid, en in rows if eid > 0 and en]


def _scan_missing_audio_targets(entry_type: str, items, cancel_check=None):
    scanned_entries = 0
    audio_attempted = 0
    audio_skipped_existing = 0
    audio_missing_voice = 0
    target_items = []
    for entry_id, english_text in (items or []):
        if callable(cancel_check) and bool(cancel_check()):
            return {
                "target_items": target_items,
                "scanned_entries": int(scanned_entries),
                "audio_attempted": int(audio_attempted),
                "audio_skipped_existing": int(audio_skipped_existing),
                "audio_missing_voice": int(audio_missing_voice),
                "canceled": True,
            }
        eid = int(entry_id or 0)
        if eid <= 0:
            continue
        scanned_entries += 1
        texts = _get_entry_texts_for_tts(entry_type, eid)
        has_missing_for_entry = False
        for lang in LEARN_TTS_LANGS:
            txt = normalize_text((texts or {}).get(lang, "") or "")
            if not txt:
                continue
            voice_name = _azure_voice_for_lang(lang)
            if not voice_name:
                audio_missing_voice += 1
                continue
            if _resolve_generated_tts_row(entry_type, eid, lang, txt, voice_name):
                audio_skipped_existing += 1
                continue
            audio_attempted += 1
            has_missing_for_entry = True
        if has_missing_for_entry:
            target_items.append((eid, normalize_text(english_text or "")))
    return {
        "target_items": target_items,
        "scanned_entries": int(scanned_entries),
        "audio_attempted": int(audio_attempted),
        "audio_skipped_existing": int(audio_skipped_existing),
        "audio_missing_voice": int(audio_missing_voice),
        "canceled": False,
    }

def _run_admin_audio_regen_job(entry_type: str, limit: int = 0, chunk_size: int = None, cancel_check=None):
    et = normalize_text(entry_type or "").lower()
    if et not in ("word", "phrase"):
        return False, {"error": "invalid_entry_type"}

    # Strict hard cap for admin-run batch size to prevent overload.
    safe_limit = max(1, min(int(limit or 100), 100000))
    items = _fetch_approved_items_for_audio_regen(et, limit=safe_limit)
    scan = _scan_missing_audio_targets(et, items, cancel_check=cancel_check)
    target_items = list((scan.get("target_items", []) or [])[:safe_limit])
    if bool(scan.get("canceled")):
        return False, {
            "entry_type": et,
            "scanned_entries": int(scan.get("scanned_entries", 0) or 0),
            "audio_attempted": int(scan.get("audio_attempted", 0) or 0),
            "audio_generated": 0,
            "audio_skipped_existing": int(scan.get("audio_skipped_existing", 0) or 0),
            "audio_missing_voice": int(scan.get("audio_missing_voice", 0) or 0),
            "audio_failed": 0,
            "processed_entries": 0,
            "target_entries": int(len(target_items)),
            "sample_failed_ids": [],
            "canceled": True,
        }

    safe_chunk = int(chunk_size or IMPORT_BATCH_SIZE or 50)
    if safe_chunk < 1:
        safe_chunk = 50
    safe_chunk = min(int(safe_chunk), int(TTS_JOB_CHUNK_SIZE))

    delay_s = max(0.0, float(int(TTS_JOB_ENTRY_DELAY_MS or 0)) / 1000.0)
    per_lang_delay_ms = int(TTS_JOB_LANGUAGE_DELAY_MS or 0)

    app.logger.info(
        "admin_audio_regen_job provider_health=%s tts_job_chunk_size=%s tts_job_entry_delay_ms=%s tts_job_language_delay_ms=%s limit=%s",
        _provider_health_snapshot(),
        int(safe_chunk),
        int(TTS_JOB_ENTRY_DELAY_MS),
        int(per_lang_delay_ms),
        int(safe_limit),
    )

    generated = 0
    failed = 0
    sample_failed_ids = []
    processed_entries = 0

    for entry_id, _en in target_items:
        if callable(cancel_check) and bool(cancel_check()):
            summary = {
                "entry_type": et,
                "scanned_entries": int(scan.get("scanned_entries", 0) or 0),
                "audio_attempted": int(scan.get("audio_attempted", 0) or 0),
                "audio_generated": int(generated),
                "audio_skipped_existing": int(scan.get("audio_skipped_existing", 0) or 0),
                "audio_missing_voice": int(scan.get("audio_missing_voice", 0) or 0),
                "audio_failed": int(failed),
                "processed_entries": int(processed_entries),
                "target_entries": int(len(target_items)),
                "sample_failed_ids": sample_failed_ids,
                "canceled": True,
            }
            return False, summary
        row = generate_tts_for_entry(
            et,
            int(entry_id),
            force_regenerate=False,
            langs=LEARN_TTS_LANGS,
            per_language_delay_ms=per_lang_delay_ms,
            stop_on_429=True,
        )
        processed_entries += 1
        generated += int((row or {}).get("generated", 0) or 0)

        row_failed = int((row or {}).get("failed", 0) or 0)
        failed += row_failed

        if row_failed > 0 and len(sample_failed_ids) < 20:
            sample_failed_ids.append(int(entry_id))

        if processed_entries % max(1, safe_chunk) == 0:
            app.logger.info(
                "admin_audio_regen_job_progress entry_type=%s processed_entries=%s/%s generated=%s failed=%s",
                et,
                processed_entries,
                len(target_items),
                generated,
                failed,
            )

        if delay_s > 0:
            time.sleep(delay_s)

    summary = {
        "entry_type": et,
        "scanned_entries": int(scan.get("scanned_entries", 0) or 0),
        "audio_attempted": int(scan.get("audio_attempted", 0) or 0),
        "audio_generated": int(generated),
        "audio_skipped_existing": int(scan.get("audio_skipped_existing", 0) or 0),
        "audio_missing_voice": int(scan.get("audio_missing_voice", 0) or 0),
        "audio_failed": int(failed),
        "processed_entries": int(processed_entries),
        "target_entries": int(len(target_items)),
        "sample_failed_ids": sample_failed_ids,
    }

    summary["completed_with_provider_failures"] = bool(
        int(summary.get("target_entries", 0) or 0) > 0
        and int(summary.get("audio_generated", 0) or 0) <= 0
        and int(summary.get("audio_failed", 0) or 0) > 0
    )

    app.logger.info("admin_audio_regen_job_summary %s", summary)
    return True, summary

def _run_admin_repair_generated_job(
    entry_type: str,
    max_items: int = 5000,
    chunk_size: int = None,
    cancel_check=None,
):
    et = normalize_text(entry_type or "").lower()
    if et not in ("word", "phrase"):
        return False, {"error": "invalid_entry_type"}
    safe_limit = max(1, min(int(max_items or 5000), 50000))
    safe_chunk = max(1, int(chunk_size or IMPORT_BATCH_SIZE or 100))

    if et == "word":
        items = _fetch_approved_word_items(limit=safe_limit)
        saved, stats = ensure_missing_generated_translations_for_words(
            items,
            langs=EXTRA_GENERATED_LANGS,
            chunk_size=safe_chunk,
            log_context="admin_repair_generated_queue",
            cancel_check=cancel_check,
        )
        summary = {
            "entry_type": et,
            "scanned_entries": int(len(items)),
            "translations_generated": int(saved or 0),
            "stats_by_lang": stats or {},
            "canceled": bool(((stats or {}).get("__meta__") or {}).get("canceled")),
        }
        return True, summary

    phrase_items = _fetch_approved_phrase_items(limit=safe_limit)
    saved, stats = ensure_missing_generated_translations_for_phrases(
        phrase_items,
        langs=EXTRA_GENERATED_LANGS,
        chunk_size=safe_chunk,
        overwrite_existing=False,
        log_context="admin_repair_generated_phrases_queue",
        cancel_check=cancel_check,
    )
    summary = {
        "entry_type": et,
        "scanned_entries": int(len(phrase_items)),
        "translations_generated": int(saved or 0),
        "stats_by_lang": stats or {},
        "canceled": bool(((stats or {}).get("__meta__") or {}).get("canceled")),
    }
    return True, summary


POST_IMPORT_QUEUE_POLL_SECONDS = max(1, int((os.environ.get("POST_IMPORT_QUEUE_POLL_SECONDS") or "2").strip() or 2))
POST_IMPORT_QUEUE_MAX_IDLE_POLLS = max(1, int((os.environ.get("POST_IMPORT_QUEUE_MAX_IDLE_POLLS") or "3").strip() or 3))
POST_IMPORT_WORKER_START_ON_BOOT = ((os.environ.get("POST_IMPORT_WORKER_START_ON_BOOT") or "1").strip() == "1")
POST_IMPORT_REQUEUE_STALE_MINUTES = max(1, int((os.environ.get("POST_IMPORT_REQUEUE_STALE_MINUTES") or "30").strip() or 30))
_post_import_worker_lock = threading.Lock()
_post_import_worker_started = False


def _enqueue_post_import_job(
    word_ids,
    phrase_ids,
    chunk_size: int = None,
    import_summary: dict = None,
    job_type: str = "post_import",
    options: dict = None,
):
    word_ids = [int(x) for x in (word_ids or []) if int(x or 0) > 0]
    phrase_ids = [int(x) for x in (phrase_ids or []) if int(x or 0) > 0]
    normalized_job_type = normalize_text(job_type or "post_import") or "post_import"
    if normalized_job_type == "post_import" and (not word_ids and not phrase_ids):
        return 0

    conn = None
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute(
            """
            INSERT INTO post_import_jobs
            (status, job_type, word_ids_json, phrase_ids_json, chunk_size, import_summary_json, options_json, runtime)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "pending",
                normalized_job_type,
                json.dumps(word_ids, separators=(",", ":")),
                json.dumps(phrase_ids, separators=(",", ":")),
                int(chunk_size or 0) if chunk_size else None,
                json.dumps(import_summary or {}, separators=(",", ":")),
                json.dumps(options or {}, separators=(",", ":")),
                APP_RUNTIME,
            ),
        )
        conn.commit()
        job_id = int(c.lastrowid or 0)
        app.logger.info(
            "post_import_pipeline_job_enqueued job_id=%s job_type=%s word_ids=%s phrase_ids=%s chunk_size=%s runtime=%s options=%s",
            job_id,
            normalized_job_type,
            len(word_ids),
            len(phrase_ids),
            int(chunk_size or 0),
            APP_RUNTIME,
            (options or {}),
        )
        return job_id
    except Exception:
        app.logger.exception("Failed to enqueue post-import pipeline job")
        return 0
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


def _claim_next_post_import_job():
    conn = None
    try:
        conn = sqlite3.connect(DB_NAME, timeout=30)
        c = conn.cursor()
        c.execute("BEGIN IMMEDIATE")
        c.execute(
            """
            SELECT id, job_type, word_ids_json, phrase_ids_json, chunk_size, import_summary_json, options_json, attempts
            FROM post_import_jobs
            WHERE status='pending'
            ORDER BY id ASC
            LIMIT 1
            """
        )
        row = c.fetchone()
        if not row:
            conn.commit()
            return None
        job_id = int((row or [0])[0] or 0)
        c.execute(
            """
            UPDATE post_import_jobs
            SET status='running', started_at=CURRENT_TIMESTAMP, attempts=attempts + 1, last_error=NULL
            WHERE id=? AND status='pending'
            """,
            (job_id,),
        )
        if int(c.rowcount or 0) <= 0:
            conn.commit()
            return None
        conn.commit()
        return {
            "id": job_id,
            "job_type": normalize_text((row[1] or "") or "post_import") or "post_import",
            "word_ids_json": row[2] or "[]",
            "phrase_ids_json": row[3] or "[]",
            "chunk_size": row[4],
            "import_summary_json": row[5] or "{}",
            "options_json": row[6] or "{}",
            "attempts_before": int((row[7] or 0)),
        }
    except Exception:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        app.logger.exception("Failed to claim post-import pipeline job")
        return None
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


def _is_post_import_job_cancel_requested(job_id: int) -> bool:
    conn = None
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute(
            "SELECT cancel_requested FROM post_import_jobs WHERE id=? LIMIT 1",
            (int(job_id or 0),),
        )
        row = c.fetchone()
        return bool(int((row or [0])[0] or 0))
    except Exception:
        app.logger.exception("Failed to read post-import job cancel flag job_id=%s", int(job_id or 0))
        return False
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


def _complete_post_import_job(
    job_id: int,
    ok: bool = None,
    error_text: str = "",
    result_payload: dict = None,
    status: str = "",
):
    normalized_status = normalize_text(status or "").lower()
    if normalized_status not in ("done", "failed", "canceled"):
        normalized_status = "done" if bool(ok) else "failed"
    conn = None
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute(
            """
            UPDATE post_import_jobs
            SET status=?, finished_at=CURRENT_TIMESTAMP, last_error=?, result_json=?
            WHERE id=?
            """,
            (
                normalized_status,
                normalize_text(error_text or "")[:2000],
                json.dumps(result_payload or {}, separators=(",", ":")),
                int(job_id or 0),
            ),
        )
        conn.commit()
        app.logger.info(
            "post_import_worker job_%s job_id=%s error=%s result=%s",
            normalized_status,
            int(job_id or 0),
            normalize_text(error_text or ""),
            (result_payload or {}),
        )
    except Exception:
        app.logger.exception("Failed to finalize post-import job id=%s", int(job_id or 0))
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


def _requeue_stale_post_import_jobs():
    conn = None
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute(
            f"""
            UPDATE post_import_jobs
            SET status='pending',
                started_at=NULL,
                finished_at=NULL,
                last_error=COALESCE(last_error, '') || CASE WHEN COALESCE(last_error, '')='' THEN '' ELSE ' | ' END || 'requeued_stale_running_job'
            WHERE status='running'
              AND (
                    started_at IS NULL
                    OR started_at < datetime('now', '-{int(POST_IMPORT_REQUEUE_STALE_MINUTES)} minutes')
              )
            """
        )
        moved = int(c.rowcount or 0)
        conn.commit()
        if moved > 0:
            app.logger.warning(
                "post_import_worker requeued_stale_running_jobs count=%s stale_minutes=%s",
                moved,
                int(POST_IMPORT_REQUEUE_STALE_MINUTES),
            )
    except Exception:
        app.logger.exception("Failed to requeue stale post-import jobs")
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


def _post_import_worker_loop():
    global _post_import_worker_started
    idle_polls = 0
    app.logger.info(
        "post_import_worker started runtime=%s poll_seconds=%s max_idle_polls=%s db_path=%s",
        APP_RUNTIME,
        POST_IMPORT_QUEUE_POLL_SECONDS,
        POST_IMPORT_QUEUE_MAX_IDLE_POLLS,
        DB_NAME,
    )
    try:
        _requeue_stale_post_import_jobs()
        while True:
            job = _claim_next_post_import_job()
            if not job:
                idle_polls += 1
                if idle_polls >= POST_IMPORT_QUEUE_MAX_IDLE_POLLS:
                    break
                time.sleep(POST_IMPORT_QUEUE_POLL_SECONDS)
                continue
            idle_polls = 0
            job_id = int((job or {}).get("id") or 0)
            try:
                job_type = normalize_text((job or {}).get("job_type") or "post_import") or "post_import"
                word_ids = json.loads((job or {}).get("word_ids_json") or "[]")
                phrase_ids = json.loads((job or {}).get("phrase_ids_json") or "[]")
                import_summary = json.loads((job or {}).get("import_summary_json") or "{}")
                options = json.loads((job or {}).get("options_json") or "{}")
                chunk_size = int((job or {}).get("chunk_size") or 0) or None
            except Exception:
                app.logger.exception("Invalid post-import job payload job_id=%s", job_id)
                _complete_post_import_job(job_id, ok=False, error_text="invalid_job_payload")
                continue

            app.logger.info(
                "post_import_worker processing job_id=%s job_type=%s attempts_before=%s word_ids=%s phrase_ids=%s options=%s",
                job_id,
                job_type,
                int((job or {}).get("attempts_before") or 0),
                len(word_ids or []),
                len(phrase_ids or []),
                options,
            )
            app.logger.info("post_import_worker provider_health=%s", _provider_health_snapshot())
            cancel_check = lambda jid=job_id: _is_post_import_job_cancel_requested(jid)
            if bool(cancel_check()):
                _complete_post_import_job(
                    job_id,
                    status="canceled",
                    error_text="cancel_requested",
                    result_payload={
                        "ok": False,
                        "canceled": True,
                        "job_id": int(job_id or 0),
                        "job_type": job_type,
                    },
                )
                continue
            try:
                if job_type in ("admin_audio_regen_word", "admin_audio_regen_phrase"):
                    target_entry_type = "word" if job_type.endswith("_word") else "phrase"
                    limit = int((options or {}).get("limit") or 0)
                    run_ok, summary = _run_admin_audio_regen_job(
                        target_entry_type,
                        limit=limit,
                        chunk_size=chunk_size,
                        cancel_check=cancel_check,
                    )
                    if run_ok:
                        app.logger.info(
                            "post_import_worker admin_audio_regen_done job_id=%s job_type=%s summary=%s",
                            job_id,
                            job_type,
                            summary,
                        )
                    else:
                        app.logger.error(
                            "post_import_worker admin_audio_regen_failed job_id=%s job_type=%s summary=%s",
                            job_id,
                            job_type,
                            summary,
                        )
                elif job_type == "admin_import_missing_phrase_audio":
                    run_ok = True
                    summary = import_missing_phrase_audio_from_source(
                        limit=int((options or {}).get("limit") or 100),
                        offset=int((options or {}).get("offset") or 0),
                        entry_id=int((options or {}).get("entry_id") or 0),
                        dry_run=bool((options or {}).get("dry_run")),
                        cancel_check=cancel_check,
                    )
                    if not bool((summary or {}).get("ok", True)):
                        run_ok = False
                elif job_type in ("admin_repair_generated_word", "admin_repair_generated_phrase"):
                    target_entry_type = "word" if job_type.endswith("_word") else "phrase"
                    run_ok, summary = _run_admin_repair_generated_job(
                        target_entry_type,
                        max_items=int((options or {}).get("max_items") or 5000),
                        chunk_size=chunk_size,
                        cancel_check=cancel_check,
                    )
                elif job_type in ("admin_repair_missing_audio_import_only", "admin_repair_missing_audio_full_repair"):
                    mode = "full_repair" if job_type.endswith("_full_repair") else "import_only"
                    summary = run_repair_missing_audio(
                        limit=int((options or {}).get("limit") or 5000),
                        generate_missing=(mode == "full_repair"),
                        source_dirs=[STATIC_UPLOADS_FOLDER],
                        cancel_check=cancel_check,
                    )
                    summary["mode"] = mode
                    run_ok = True
                else:
                    pipeline_result = _run_post_import_pipeline(
                        word_ids,
                        phrase_ids,
                        chunk_size=chunk_size,
                        import_summary=import_summary,
                        cancel_check=cancel_check,
                    )
                    summary = pipeline_result if isinstance(pipeline_result, dict) else {}
                    run_ok = bool((pipeline_result or {}).get("ok", False)) if isinstance(pipeline_result, dict) else bool(pipeline_result)
                if bool((summary or {}).get("canceled")) or bool(cancel_check()):
                    canceled_summary = dict(summary or {})
                    canceled_summary["ok"] = False
                    canceled_summary["canceled"] = True
                    _complete_post_import_job(
                        job_id,
                        status="canceled",
                        error_text="cancel_requested",
                        result_payload=canceled_summary,
                    )
                    continue
                if run_ok:
                    warn_text = ""
                    if isinstance(summary, dict) and bool(summary.get("completed_with_provider_failures")):
                        warn_text = "completed_with_provider_failures"
                    _complete_post_import_job(job_id, ok=True, error_text=warn_text, result_payload=summary)
                else:
                    _complete_post_import_job(job_id, ok=False, error_text="pipeline_returned_failure", result_payload=summary)
            except Exception as e:
                app.logger.exception("post_import_worker job failed job_id=%s", job_id)
                _complete_post_import_job(job_id, ok=False, error_text=repr(e))
    finally:
        with _post_import_worker_lock:
            _post_import_worker_started = False
        app.logger.info("post_import_worker stopped runtime=%s db_path=%s", APP_RUNTIME, DB_NAME)


def _start_post_import_worker_if_needed():
    global _post_import_worker_started
    with _post_import_worker_lock:
        if _post_import_worker_started:
            return True
        try:
            worker = threading.Thread(
                target=_post_import_worker_loop,
                daemon=True,
                name="post_import_worker",
            )
            worker.start()
            _post_import_worker_started = True
            return True
        except Exception:
            app.logger.exception("Failed to start post-import worker thread")
            _post_import_worker_started = False
            return False


def trigger_post_import_pipeline_async(word_ids, phrase_ids, chunk_size: int = None, import_summary: dict = None):
    job_id = _enqueue_post_import_job(word_ids, phrase_ids, chunk_size=chunk_size, import_summary=import_summary)
    if not int(job_id or 0):
        return False
    if not _start_post_import_worker_if_needed():
        app.logger.warning("post_import_worker_not_started job_id=%s", int(job_id or 0))
    return True


def trigger_admin_audio_regen_async(entry_type: str, limit: int = 0, chunk_size: int = None):
    et = normalize_text(entry_type or "").lower()
    if et not in ("word", "phrase"):
        return False, 0
    safe_limit = max(0, int(limit or 0))
    job_type = f"admin_audio_regen_{et}"
    job_id = _enqueue_post_import_job(
        word_ids=[],
        phrase_ids=[],
        chunk_size=chunk_size,
        import_summary={},
        job_type=job_type,
        options={"entry_type": et, "limit": safe_limit},
    )
    if not int(job_id or 0):
        return False, 0
    if not _start_post_import_worker_if_needed():
        app.logger.warning("post_import_worker_not_started job_id=%s job_type=%s", int(job_id or 0), job_type)
    return True, int(job_id or 0)


def trigger_admin_import_missing_phrase_audio_async(limit: int = 100, offset: int = 0, entry_id: int = 0, dry_run: bool = False):
    job_type = "admin_import_missing_phrase_audio"
    job_id = _enqueue_post_import_job(
        word_ids=[],
        phrase_ids=[],
        chunk_size=None,
        import_summary={},
        job_type=job_type,
        options={
            "limit": max(1, min(int(limit or 100), 2000)),
            "offset": max(0, int(offset or 0)),
            "entry_id": max(0, int(entry_id or 0)),
            "dry_run": bool(dry_run),
        },
    )
    if not int(job_id or 0):
        return False, 0
    if not _start_post_import_worker_if_needed():
        app.logger.warning("post_import_worker_not_started job_id=%s job_type=%s", int(job_id or 0), job_type)
    return True, int(job_id or 0)


def trigger_admin_repair_generated_async(entry_type: str, max_items: int = 5000, chunk_size: int = None):
    et = normalize_text(entry_type or "").lower()
    if et not in ("word", "phrase"):
        return False, 0
    job_type = f"admin_repair_generated_{et}"
    job_id = _enqueue_post_import_job(
        word_ids=[],
        phrase_ids=[],
        chunk_size=chunk_size,
        import_summary={},
        job_type=job_type,
        options={"entry_type": et, "max_items": max(1, min(int(max_items or 5000), 50000))},
    )
    if not int(job_id or 0):
        return False, 0
    if not _start_post_import_worker_if_needed():
        app.logger.warning("post_import_worker_not_started job_id=%s job_type=%s", int(job_id or 0), job_type)
    return True, int(job_id or 0)


def trigger_admin_repair_missing_audio_async(mode: str, limit: int = 5000):
    selected = normalize_text(mode or "import_only")
    if selected not in ("import_only", "full_repair"):
        selected = "import_only"
    job_type = f"admin_repair_missing_audio_{selected}"
    job_id = _enqueue_post_import_job(
        word_ids=[],
        phrase_ids=[],
        chunk_size=None,
        import_summary={},
        job_type=job_type,
        options={"mode": selected, "limit": max(1, min(int(limit or 5000), 50000))},
    )
    if not int(job_id or 0):
        return False, 0
    if not _start_post_import_worker_if_needed():
        app.logger.warning("post_import_worker_not_started job_id=%s job_type=%s", int(job_id or 0), job_type)
    return True, int(job_id or 0)


if POST_IMPORT_WORKER_START_ON_BOOT:
    _start_post_import_worker_if_needed()


def run_word_audio_backfill(limit: int = 0, chunk_size: int = None, force_regenerate: bool = False):
    out = {
        "words_total": 0,
        "translation_saved": 0,
        "translation_stats": {},
        "tts": {
            "words_seen": 0,
            "generated": 0,
            "cached": 0,
            "failed": 0,
            "skipped_missing_text": 0,
            "skipped_missing_voice": 0,
        },
    }
    words = _fetch_approved_word_items(limit=limit)
    out["words_total"] = len(words)
    if not words:
        return out

    safe_chunk = int(chunk_size or IMPORT_BATCH_SIZE or 100)
    if safe_chunk < 1:
        safe_chunk = 100

    merged_translation_stats = {lang: {} for lang in EXTRA_GENERATED_LANGS}
    total_translation_saved = 0
    tts_totals = out["tts"]

    for i in range(0, len(words), safe_chunk):
        chunk = words[i:i + safe_chunk]
        saved_count, tr_stats = ensure_missing_generated_translations_for_words(
            chunk,
            langs=EXTRA_GENERATED_LANGS,
            chunk_size=safe_chunk,
            log_context="word_audio_backfill",
        )
        total_translation_saved += int(saved_count or 0)
        _merge_generated_stats(merged_translation_stats, tr_stats)

        tts_stats = ensure_missing_tts_for_words(
            chunk,
            force_regenerate=force_regenerate,
            chunk_size=safe_chunk,
            log_context="word_audio_backfill_tts",
        )
        for key in ("words_seen", "generated", "cached", "failed", "skipped_missing_text", "skipped_missing_voice"):
            tts_totals[key] += int(tts_stats.get(key, 0) or 0)

        app.logger.info(
            "word_audio_backfill progress processed=%s/%s translation_saved=%s tts_generated=%s tts_cached=%s tts_failed=%s",
            min(i + len(chunk), len(words)),
            len(words),
            total_translation_saved,
            tts_totals["generated"],
            tts_totals["cached"],
            tts_totals["failed"],
        )

    out["translation_saved"] = total_translation_saved
    out["translation_stats"] = merged_translation_stats
    return out


def _get_word_by_key(source_lang: str, key_text: str):
    if source_lang not in ("om", "en"):
        return None
    col = "oromo_key" if source_lang == "om" else "english_key"
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute(f"""
            SELECT id, english, oromo
            FROM words
            WHERE status='approved' AND {col}=?
            LIMIT 1
        """, (key_text,))
        row = c.fetchone()
        conn.close()
        return row
    except Exception as e:
        app.logger.exception(f"base word lookup failed ({source_lang}): {repr(e)}")
        return None


def _get_word_by_any_key(key_text: str):
    """Legacy-compatible base lookup: match either English or Oromo key."""
    if not key_text:
        return None
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("""
            SELECT id, english, oromo
            FROM words
            WHERE status='approved' AND (english_key=? OR oromo_key=?)
            LIMIT 1
        """, (key_text, key_text))
        row = c.fetchone()
        conn.close()
        return row
    except Exception as e:
        app.logger.exception(f"base word fallback lookup failed: {repr(e)}")
        return None


def _auto_translate_from_english(english_text: str, target_lang: str) -> str:
    if target_lang == "en":
        return normalize_text(english_text)
    return google_translate_text_v2(
        english_text,
        target=_google_lang_code(target_lang),
        source="en"
    )


def _get_or_generate_word_translation(
    word_id: int,
    english_text: str,
    target_lang: str,
    allow_tts_generation: bool = False,
):
    cached = _get_cached_generated_translation(word_id, target_lang)
    if cached and _is_meaningful_generated_text((cached or [""])[0]):
        app.logger.info("using cached translation word_id=%s lang=%s", word_id, target_lang)
        translated_cached = normalize_text(cached[0] or "")
        tts_cached_raw = _normalize_cached_tts_url((cached or [None, ""])[1] or "")
        tts_cached = tts_cached_raw
        if tts_cached and (not _tts_ref_matches_text_hash(tts_cached, translated_cached)):
            app.logger.warning(
                "cached_translation_tts_hash_mismatch word_id=%s lang=%s tts_ref=%s translated_hash=%s",
                int(word_id or 0),
                normalize_text(target_lang or ""),
                tts_cached,
                _text_hash(translated_cached)[:12],
            )
            tts_cached = ""
        if not tts_cached:
            tts_cached = _resolve_or_generate_tts_for_text(
                "word",
                int(word_id),
                target_lang,
                translated_cached,
                allow_generate=allow_tts_generation,
            )
            if tts_cached:
                _save_tts_url_to_translation_cache("word", int(word_id), target_lang, tts_cached)
        return translated_cached, tts_cached, True

    translated = ""
    conn = None
    try:
        conn = sqlite3.connect(DB_NAME)
        translated = normalize_text(
            service_get_or_generate_translation(
                db=conn,
                entry_type="word",
                entry_id=int(word_id or 0),
                source_text=english_text,
                target_lang=target_lang,
                api_key=_get_google_key(),
            )
            or ""
        )
    except Exception as e:
        app.logger.exception(f"service get_or_generate_translation failed word_id={word_id} lang={target_lang}: {repr(e)}")
        translated = _auto_translate_from_english(english_text, target_lang)
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass
    if not translated:
        app.logger.info(
            "skipping generation due to API failure word_id=%s lang=%s",
            word_id,
            target_lang,
        )
        return "", None, False

    # Ensure local cache table has the value even if service call failed over to direct provider path.
    _save_generated_translation(
        word_id,
        target_lang,
        translated,
        source_text=english_text,
        provider="google_translate_v2",
        tts_audio_url=None,
    )
    tts_url = _resolve_or_generate_tts_for_text(
        "word",
        int(word_id),
        target_lang,
        translated,
        allow_generate=allow_tts_generation,
    )
    if tts_url:
        _save_tts_url_to_translation_cache("word", int(word_id), target_lang, tts_url)
    return translated, tts_url, False


def safe_translate_multilingual(text: str, source_lang: str, target_lang: str):
    try:
        return translate_multilingual(text, source_lang, target_lang)
    except Exception as e:
        app.logger.exception(f"translate_multilingual failed: {repr(e)}")
        return {
            "text": "",
            "is_exact": 0,
            "is_phrase": 0,
            "is_auto_translation": (target_lang not in ("om", "en") or source_lang not in ("om", "en")),
            "tts_audio_url": None
        }


def _dictionary_lookup_result(query_text: str, source_lang: str, target_lang: str):
    q = normalize_text(query_text)
    if not q:
        return None, None, None, False

    base_row = None
    pivot_english = None

    if source_lang in ("om", "en"):
        wkey = make_search_key(_strip_edge_punct(q))
        if wkey:
            base_row = _get_word_by_key(source_lang, wkey)
            # Keep old dictionary behavior: if selected source misses, still try either side.
            if not base_row:
                base_row = _get_word_by_any_key(wkey)
    else:
        pivot_english = google_translate_text_v2(
            q,
            target="en",
            source=_google_lang_code(source_lang)
        )
        if pivot_english:
            wkey = make_search_key(_strip_edge_punct(pivot_english))
            if wkey:
                base_row = _get_word_by_key("en", wkey)

    if base_row:
        wid, en, om = base_row
        target_text = ""
        tts_audio_url = None
        is_auto = False
        auto_unavailable = False

        if target_lang == "en":
            target_text = en
        elif target_lang == "om":
            target_text = om
        else:
            try:
                target_text, tts_audio_url, _ = _get_or_generate_word_translation(
                    wid,
                    en,
                    target_lang,
                    allow_tts_generation=False,
                )
            except Exception as e:
                app.logger.exception(f"dictionary auto translation failed: {repr(e)}")
                target_text = ""
                tts_audio_url = None

            if not target_text:
                # Retry direct provider path before final fallback.
                tr_retry = safe_translate_multilingual(en, "en", target_lang)
                target_text = (tr_retry or {}).get("text", "") or ""
                if target_text:
                    try:
                        _save_generated_translation(
                            wid,
                            target_lang,
                            target_text,
                            source_text=en,
                            provider="google_translate_v2", tts_audio_url=None
                        )
                        tts_audio_url = _resolve_or_generate_tts_for_text(
                            "word",
                            int(wid),
                            target_lang,
                            target_text,
                            allow_generate=False,
                        )
                        if tts_audio_url:
                            _save_tts_url_to_translation_cache("word", int(wid), target_lang, tts_audio_url)
                    except Exception:
                        pass

            if not target_text:
                # Final fallback keeps page functional if provider is down.
                target_text = en
                auto_unavailable = True
            is_auto = True

        if source_lang == "en":
            source_text = en
        elif source_lang == "om":
            source_text = om
        else:
            source_text = q

        result = {
            "source_text": source_text,
            "target_text": target_text,
            "english": en,
            "oromo": om,
            "is_auto_translation": is_auto,
            "auto_unavailable": auto_unavailable,
            "tts_audio_url": tts_audio_url
        }
        return result, wid, None, True

    # Base entry not found: do not return generated result as primary dictionary output.
    return None, None, None, False


def get_or_generate_extra_translations(word_id: int, english_text: str):
    """
    Fetch or generate all configured extra-language translations for a base word.
    Failures per language are isolated so one provider/cache error does not break the page.
    """
    out = {}
    if not word_id:
        return out

    en_text = normalize_text(english_text or "")
    if not en_text:
        en_text = _get_word_english_by_id(word_id)
    if not en_text:
        app.logger.info("extra translation skipped: missing english text for word_id=%s", word_id)
        return out

    # Display is always DB-first so public routes remain API-independent.
    out = _get_cached_extra_translations_for_word(word_id, EXTRA_GENERATED_LANGS, log_hits=True)

    missing_for_display = [lang for lang in EXTRA_GENERATED_LANGS if lang not in out]

    saved_count = 0
    backfill_stats = {}
    if missing_for_display:
        try:
            saved_count, backfill_stats = ensure_missing_generated_translations_for_words(
                [(word_id, en_text)],
                langs=missing_for_display,
                chunk_size=len(EXTRA_GENERATED_LANGS),
                log_context="lookup_single_word",
            )
        except Exception as e:
            app.logger.exception(f"lookup generated backfill bootstrap failed for word_id={word_id}: {repr(e)}")
            backfill_stats = {}

    missing_langs = []
    cached_langs = []
    generated_langs = []
    failed_langs = []
    for lang in EXTRA_GENERATED_LANGS:
        st = backfill_stats.get(lang, {})
        missing_before = int(st.get("missing_before", 0) or 0)
        already_cached = int(st.get("already_cached", 0) or 0)
        saved = int(st.get("saved", 0) or 0)
        provider_errors = int(st.get("provider_errors", 0) or 0)

        if missing_before > 0:
            missing_langs.append(lang)
        if already_cached > 0:
            cached_langs.append(lang)
        if saved > 0:
            generated_langs.append(lang)
        if (missing_before > 0 and saved == 0) or provider_errors > 0:
            failed_langs.append(lang)

    app.logger.info(
        "lookup backfill word_id=%s missing=%s generated=%s cached=%s failed=%s saved_count=%s",
        word_id,
        missing_langs,
        generated_langs,
        cached_langs,
        failed_langs,
        saved_count,
    )

    if missing_for_display:
        refreshed = _get_cached_extra_translations_for_word(word_id, missing_for_display, log_hits=False)
        out.update(refreshed)

        still_missing = [lang for lang in missing_for_display if lang not in out]
        if still_missing:
            app.logger.info(
                "skipping generation due to API failure word_id=%s langs=%s",
                word_id,
                still_missing,
            )

    app.logger.info(
        "lookup backfill final word_id=%s available_langs=%s",
        word_id,
        sorted(out.keys()),
    )

    return out


def translate_multilingual(text: str, source_lang: str, target_lang: str):
    t = normalize_text(text)
    if not t:
        return {"text": "", "is_exact": 0, "is_phrase": 0, "is_auto_translation": False, "tts_audio_url": None}

    if source_lang == target_lang:
        return {"text": t, "is_exact": 1, "is_phrase": 0, "is_auto_translation": False, "tts_audio_url": None}

    if source_lang == "om":
        english_text, is_exact, is_phrase = translate_multipart_text(t, "om_en")
        english_text = normalize_text(english_text)
        if target_lang == "en":
            return {"text": english_text, "is_exact": is_exact, "is_phrase": is_phrase, "is_auto_translation": False, "tts_audio_url": None}

        tokens = t.split()
        if len(tokens) == 1:
            row = _get_word_by_key("om", make_search_key(_strip_edge_punct(tokens[0])))
            if row:
                translated, tts_url, _ = _get_or_generate_word_translation(
                    row[0],
                    row[1],
                    target_lang,
                    allow_tts_generation=False,
                )
                if translated:
                    return {"text": translated, "is_exact": 1, "is_phrase": 0, "is_auto_translation": True, "tts_audio_url": tts_url}

        translated = _auto_translate_from_english(english_text, target_lang) if english_text else ""
        return {"text": translated, "is_exact": is_exact, "is_phrase": is_phrase, "is_auto_translation": bool(translated), "tts_audio_url": None}

    if source_lang == "en":
        if target_lang == "om":
            out, is_exact, is_phrase = translate_multipart_text(t, "en_om")
            return {"text": out, "is_exact": is_exact, "is_phrase": is_phrase, "is_auto_translation": False, "tts_audio_url": None}

        tokens = t.split()
        if len(tokens) == 1:
            row = _get_word_by_key("en", make_search_key(_strip_edge_punct(tokens[0])))
            if row:
                translated, tts_url, _ = _get_or_generate_word_translation(
                    row[0],
                    row[1],
                    target_lang,
                    allow_tts_generation=False,
                )
                if translated:
                    return {"text": translated, "is_exact": 1, "is_phrase": 0, "is_auto_translation": True, "tts_audio_url": tts_url}

        translated = _auto_translate_from_english(t, target_lang)
        return {"text": translated, "is_exact": 0, "is_phrase": 0, "is_auto_translation": bool(translated), "tts_audio_url": None}

    # Non-base source languages always pivot through English.
    src_google = _google_lang_code(source_lang)
    source_to_english = google_translate_text_v2(t, target="en", source=src_google)
    if not source_to_english:
        return {"text": "", "is_exact": 0, "is_phrase": 0, "is_auto_translation": True, "tts_audio_url": None}

    if target_lang == "en":
        return {"text": source_to_english, "is_exact": 0, "is_phrase": 0, "is_auto_translation": True, "tts_audio_url": None}

    if target_lang == "om":
        out, is_exact, is_phrase = translate_multipart_text(source_to_english, "en_om")
        return {"text": out, "is_exact": is_exact, "is_phrase": is_phrase, "is_auto_translation": True, "tts_audio_url": None}

    out = _auto_translate_from_english(source_to_english, target_lang)
    return {"text": out, "is_exact": 0, "is_phrase": 0, "is_auto_translation": True, "tts_audio_url": None}


import re

EN_FILLERS = {"please"}
EN_ARTICLES = {"a", "an", "the"}

def key_for(text: str) -> str:
    """One key rule everywhere."""
    return make_search_key(_strip_edge_punct(normalize_text(text)))

def generate_english_alias_texts(english: str) -> list[str]:
    """
    Generate a small set of safe aliases (not too aggressive).
    You can expand later.
    """
    s = normalize_text(english)
    if not s:
        return []

    toks = s.split()
    aliases = {s}

    # Remove trailing "please"
    if toks and toks[-1].casefold() in EN_FILLERS:
        aliases.add(" ".join(toks[:-1]))

    # Remove articles (a/an/the) - lightweight
    no_articles = [t for t in toks if t.casefold() not in EN_ARTICLES]
    if no_articles and no_articles != toks:
        aliases.add(" ".join(no_articles))

    # Remove both: articles + trailing please
    toks2 = no_articles
    if toks2 and toks2[-1].casefold() in EN_FILLERS:
        aliases.add(" ".join(toks2[:-1]))

    # De-dup, keep non-empty
    out = []
    seen = set()
    for a in aliases:
        a = normalize_text(a)
        if a and a not in seen:
            seen.add(a)
            out.append(a)
    return out

def generate_oromo_alias_texts(oromo: str) -> list[str]:
    """
    Oromo aliases: keep conservative.
    For now, just the original normalized text.
    You can add Oromo-specific variants later.
    """
    s = normalize_text(oromo)
    return [s] if s else []

def suggest_phrases_from_text(text: str, direction: str, limit: int = 8, min_words: int = 2, max_words: int = 6):
    """
    Suggest approved phrases found inside the user's text using phrase_aliases.
    Returns a list of dicts:
      { "phrase_id": int, "source_text": str, "english": str, "oromo": str, "matched_ngram": str }
    direction:
      - 'en_om' means user typed English, target Oromo
      - 'om_en' means user typed Oromo, target English
    """
    parts = split_segments(text)
    if not parts:
        return []

    # Extract segments, normalize, split to words
    segments_words = []
    for seg_text, _, _ in parts:
        seg = normalize_text(seg_text)
        if not seg:
            continue
        words = seg.split()
        if words:
            segments_words.append(words)

    if not segments_words:
        return []

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    # For scoring and de-dup
    best_by_phrase_id = {}  # phrase_id -> (score, payload)

    # Decide which alias column to search
    alias_col = "english_alias_key" if direction == "en_om" else "oromo_alias_key"

    # We will do lookups for each n-gram key.
    # To keep it efficient, we can cache alias lookups in a dict
    alias_cache = {}  # key -> row

    for words in segments_words:
        n = len(words)
        for i in range(n):
            for L in range(max_words, min_words - 1, -1):
                if i + L > n:
                    continue
                ngram = " ".join(words[i:i+L])
                k = key_for(ngram)
                if not k:
                    continue

                # Cache DB lookups for speed
                if k in alias_cache:
                    row = alias_cache[k]
                else:
                    row = c.execute(f"""
                        SELECT a.phrase_id, p.english, p.oromo
                        FROM phrase_aliases a
                        JOIN phrases p ON p.id = a.phrase_id
                        WHERE p.status='approved' AND a.{alias_col}=? 
                        LIMIT 1
                    """, (k,)).fetchone()
                    alias_cache[k] = row

                if not row:
                    continue

                phrase_id, en, om = row
                # Score: prefer longer phrase, earlier in sentence
                score = (L * 1000) - i

                payload = {
                    "phrase_id": phrase_id,
                    "source_text": ngram,
                    "matched_ngram": ngram,
                    "english": en,
                    "oromo": om,
                }

                prev = best_by_phrase_id.get(phrase_id)
                if (prev is None) or (score > prev[0]):
                    best_by_phrase_id[phrase_id] = (score, payload)

    conn.close()

    # Sort by score descending
    suggestions = [v[1] for v in best_by_phrase_id.values()]
    suggestions.sort(key=lambda d: (
        -len(normalize_text(d["matched_ngram"]).split()),
        -len(normalize_text(d["matched_ngram"])),
        d["matched_ngram"].lower()
    ))

    return suggestions[:limit]
EN_DROP_WORDS = {"the", "a", "an"}         # safe to drop in Oromo output
EN_SOFT_WORDS = {"please"}                # optional: skip in fallback if not part of a matched phrase

def postprocess_segment(text_out: str, direction: str) -> str:
    s = (text_out or "").strip()
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"\s+([?.!,;:])", r"\1", s)        # no space before punctuation
    s = re.sub(r"([?.!,;:])\1+", r"\1", s)        # collapse repeated punctuation
    if direction == "en_om":
        # remove any leftover English articles that slipped through
        s = re.sub(r"\b(the|a|an)\b", "", s, flags=re.I)
        s = re.sub(r"\s+", " ", s).strip()
        s = re.sub(r"\s+([?.!,;:])", r"\1", s)
    return s
def apply_grammar_templates(output_text: str, direction: str) -> str:
    if direction != "en_om":
        return output_text

    s = normalize_text(output_text)

    # Clean common artifacts from word fallback
    s = re.sub(r"\bthe\b", "", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip()

    # Fix spacing before punctuation
    s = re.sub(r"\s+([?.!,;:])", r"\1", s)

    return s


def try_simple_sov_reorder(text: str, direction: str) -> str:
    if direction != "en_om":
        return text

    words = text.split()
    if len(words) != 3:
        return text

    subj, verb, obj = words

    # simple Oromo pronoun subjects
    if subj.casefold() in {"ani", "ati", "inni", "isheen", "nuti", "isin"}:
        return f"{subj} {obj} {verb}"

    return text
EN_OM_TEMPLATES = [
    # can you (please) show me the way
    (
        re.compile(r"^(please\s+)?can\s+you\s+show\s+me\s+the\s+way\??$", re.I),
        "Karaa natti agarsiisuu dandeessaa?"
    ),

    (
        re.compile(r"^(please\s+)?can\s+you\s+show\s+me\s+the\s+way\?$", re.I),
        "Karaa natti agarsiisuu dandeessaa?"
    ),
]



# ------------------ LEARN ------------------

def _bulk_fetch_generated_tts_urls(entry_type: str, entry_ids, text_by_key: dict, langs=None):
    if not entry_ids:
        return {}
    lang_list = tuple(langs or ("en", "am"))
    if not lang_list:
        return {}
    placeholders = ",".join("?" for _ in entry_ids)
    lang_placeholders = ",".join("?" for _ in lang_list)
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute(
        f"""
        SELECT entry_id, lang_code, text_hash, file_path
        FROM generated_tts_audio
        WHERE entry_type=?
          AND entry_id IN ({placeholders})
          AND lang_code IN ({lang_placeholders})
        ORDER BY id DESC
        """,
        (entry_type, *entry_ids, *lang_list),
    )
    rows = c.fetchall()
    conn.close()

    out = {}
    fallback = {}
    for eid, lang_code, text_hash, file_path in rows:
        key = (int(eid or 0), lang_code)
        expected_hash = text_by_key.get(key)
        if not _has_usable_audio_ref(file_path or ""):
            continue
        url = _public_audio_url(file_path)
        # Prefer exact text-hash matches when available.
        if expected_hash and expected_hash == (text_hash or ""):
            if key not in out:
                out[key] = url
            continue
        # Fallback to latest available saved audio for that entry/lang.
        if key not in fallback:
            fallback[key] = url
    for key, url in fallback.items():
        if key not in out:
            out[key] = url
    return out


def _bulk_fetch_saved_tts_by_entry_lang(entry_type: str, entry_ids, langs=None, text_by_key: dict = None):
    """
    DB-first bulk audio lookup for Learn page.
    Returns latest usable generated_tts_audio URL per (entry_id, lang_code),
    without requiring text-hash matches or search side-effects.
    """
    if not entry_ids:
        return {}
    requested = {_canonical_tts_lang_code(lc) for lc in (langs or ("en", "am", "ar", "fr", "zh-CN", "om"))}
    requested.discard("")
    if not requested:
        return {}

    placeholders = ",".join("?" for _ in entry_ids)
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute(
        f"""
        SELECT entry_id, lang_code, file_path, text_hash
        FROM generated_tts_audio
        WHERE entry_type=?
          AND entry_id IN ({placeholders})
        ORDER BY id DESC
        """,
        (entry_type, *entry_ids),
    )
    rows = c.fetchall()
    conn.close()

    expected_hash_by_key = {}
    for (entry_id, lang_code), text_value in (text_by_key or {}).items():
        canonical_lang = _canonical_tts_lang_code(lang_code or "")
        text_norm = normalize_text(text_value or "")
        if (not canonical_lang) or (not text_norm):
            continue
        expected_hash_by_key[(int(entry_id or 0), canonical_lang)] = _text_hash(text_norm)

    out = {}
    for entry_id, lang_code, file_path, text_hash in rows:
        canonical_lang = _canonical_tts_lang_code(lang_code or "")
        if canonical_lang not in requested:
            continue
        key = (int(entry_id or 0), canonical_lang)
        if key in out:
            continue
        expected_hash = expected_hash_by_key.get(key, "")
        if expected_hash and normalize_text(text_hash or "") != expected_hash:
            continue
        if not normalize_text(file_path or ""):
            continue
        if not _has_usable_audio_ref(file_path or ""):
            continue
        url = _public_audio_url(file_path or "")
        if not url:
            continue
        out[key] = url
    return out


def get_saved_extra_translations(word_id: int):
    """
    DB-only extra-language lookup for render routes.
    Never triggers provider/backfill generation.
    """
    return _get_cached_extra_translations_for_word(word_id, EXTRA_GENERATED_LANGS, log_hits=False)


def _bulk_fetch_translation_cache_tts_urls(entry_type: str, entry_ids, langs=None):
    if entry_type not in ("word", "phrase") or not entry_ids:
        return {}
    lang_list = tuple(langs or EXTRA_GENERATED_LANGS)
    if not lang_list:
        return {}

    placeholders = ",".join("?" for _ in entry_ids)
    lang_placeholders = ",".join("?" for _ in lang_list)
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    if entry_type == "word":
        c.execute(
            f"""
            SELECT word_id, lang_code, tts_audio_url
            FROM generated_translations
            WHERE word_id IN ({placeholders})
              AND lang_code IN ({lang_placeholders})
              AND tts_audio_url IS NOT NULL
              AND TRIM(tts_audio_url) != ''
            """,
            (*entry_ids, *lang_list),
        )
    else:
        c.execute(
            f"""
            SELECT phrase_id, lang_code, tts_audio_url
            FROM generated_phrase_translations
            WHERE phrase_id IN ({placeholders})
              AND lang_code IN ({lang_placeholders})
              AND tts_audio_url IS NOT NULL
              AND TRIM(tts_audio_url) != ''
            """,
            (*entry_ids, *lang_list),
        )
    rows = c.fetchall()
    conn.close()

    out = {}
    for entry_id, lang_code, tts_audio_url in rows:
        key = (int(entry_id or 0), lang_code)
        if key in out:
            continue
        normalized = _normalize_cached_tts_url(tts_audio_url or "")
        if not normalized:
            continue
        if not _has_usable_audio_ref(normalized):
            continue
        out[key] = _public_audio_url(normalized)
    return out


def _bulk_fetch_approved_oromo_audio_urls(entry_type: str, entry_ids):
    if not entry_ids:
        return {}
    placeholders = ",".join("?" for _ in entry_ids)
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute(
        f"""
        SELECT entry_id, file_path
        FROM audio
        WHERE status='approved'
          AND entry_type=?
          AND lang='oromo'
          AND entry_id IN ({placeholders})
        ORDER BY id DESC
        """,
        (entry_type, *entry_ids),
    )
    rows = c.fetchall()
    conn.close()

    out = {}
    for entry_id, file_path in rows:
        eid = int(entry_id or 0)
        if eid in out:
            continue
        if not _has_usable_audio_ref(file_path or ""):
            continue
        out[eid] = _public_audio_url(file_path)
    return out


def _load_learn_rows():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    c.execute(
        """
        SELECT id, english, oromo
        FROM phrases
        WHERE status='approved'
          AND english IS NOT NULL AND TRIM(english) != ''
        ORDER BY id DESC
        LIMIT ?
        """,
        (int(LEARN_RECENT_PHRASE_LIMIT),),
    )
    phrase_rows = c.fetchall()

    phrase_ids = [int(p[0]) for p in phrase_rows if p and p[0]]
    phrase_translations = {}
    if phrase_ids:
        id_marks = ",".join("?" for _ in phrase_ids)
        c.execute(
            f"""
            SELECT phrase_id, lang_code, translated_text
            FROM generated_phrase_translations
            WHERE phrase_id IN ({id_marks})
              AND lang_code IN (?, ?, ?, ?)
              AND translated_text IS NOT NULL
              AND TRIM(translated_text) != ''
            """,
            (*phrase_ids, "am", "ar", "fr", "zh-CN"),
        )
        for pid, lang, txt in c.fetchall():
            phrase_translations.setdefault(int(pid), {})[lang] = normalize_text(txt or "")
    conn.close()

    phrase_text_by_key = {}
    for pid, en, om in phrase_rows:
        pid_int = int(pid or 0)
        if pid_int <= 0:
            continue
        en_norm = normalize_text(en or "")
        om_norm = normalize_text(om or "")
        if en_norm:
            phrase_text_by_key[(pid_int, "en")] = en_norm
        if om_norm:
            phrase_text_by_key[(pid_int, "om")] = om_norm
    for pid, tr_map in (phrase_translations or {}).items():
        pid_int = int(pid or 0)
        for lang_code, translated_text in ((tr_map or {}).items()):
            txt = normalize_text(translated_text or "")
            if txt:
                phrase_text_by_key[(pid_int, lang_code)] = txt

    phrase_tts = _bulk_fetch_saved_tts_by_entry_lang(
        "phrase",
        phrase_ids,
        langs=("en", "am", "ar", "fr", "zh-CN", "om"),
        text_by_key=phrase_text_by_key,
    )
    phrase_oromo_audio = _bulk_fetch_approved_oromo_audio_urls("phrase", phrase_ids)

    rows = []
    phrases_with_missing_oromo_but_shown = 0

    for pid, en, om in phrase_rows:
        pid_int = int(pid)
        tr = phrase_translations.get(pid_int, {})
        om_text = normalize_text(om or "")
        audio_map = {
            "en": phrase_tts.get((pid_int, "en"), ""),
            "am": phrase_tts.get((pid_int, "am"), ""),
            "ar": phrase_tts.get((pid_int, "ar"), ""),
            "fr": phrase_tts.get((pid_int, "fr"), ""),
            "zh-CN": phrase_tts.get((pid_int, "zh-CN"), ""),
            "oromo": phrase_oromo_audio.get(pid_int, "") or phrase_tts.get((pid_int, "om"), ""),
        }
        if not om_text:
            phrases_with_missing_oromo_but_shown += 1
        rows.append({
            "entry_type": "phrase",
            "entry_id": pid_int,
            "english": normalize_text(en or ""),
            "oromo": om_text,
            "am": normalize_text(tr.get("am", "") or ""),
            "ar": normalize_text(tr.get("ar", "") or ""),
            "fr": normalize_text(tr.get("fr", "") or ""),
            "zh-CN": normalize_text(tr.get("zh-CN", "") or ""),
            "audio": audio_map,
        })

    words_loaded_raw = 0
    phrases_loaded_raw = int(len(phrase_rows))
    word_audio_rows_found = 0
    phrase_audio_rows_found = int(len(phrase_tts) + len(phrase_oromo_audio))
    audio_rows_found = int(word_audio_rows_found + phrase_audio_rows_found)
    rows_with_audio_in_db_all = 0
    rows_with_audio_in_db_words_all = 0
    rows_with_audio_in_db_phrases_all = 0
    rows_with_any_audio = 0
    rows_rendered_with_audio_url = 0
    audio_attached_count = 0
    phrase_rows_visible = []
    for r in rows:
        audio = (r or {}).get("audio") or {}
        entry_type = (r or {}).get("entry_type")
        entry_id = int((r or {}).get("entry_id") or 0)
        has_audio_db = False
        if entry_type == "phrase":
            has_audio_db = any((entry_id, lc) in phrase_tts for lc in ("en", "am", "ar", "fr", "zh-CN", "om")) or (entry_id in phrase_oromo_audio)
        if has_audio_db:
            rows_with_audio_in_db_all += 1
            if entry_type == "phrase":
                rows_with_audio_in_db_phrases_all += 1

        row_has_audio = False
        for u in audio.values():
            if normalize_text(u or ""):
                audio_attached_count += 1
                row_has_audio = True
        if row_has_audio:
            rows_with_any_audio += 1
            rows_rendered_with_audio_url += 1
        if entry_type == "phrase":
            phrase_rows_visible.append(r)
    rows = phrase_rows_visible

    newest_phrase_ids = [int((r or {}).get("entry_id") or 0) for r in rows]
    phrases_loaded_with_audio = int(len(rows))
    words_loaded_with_audio = 0
    total_rows_loaded = int(len(rows))
    learn_visible_phrase_count = total_rows_loaded

    app.logger.info(
        "/learn loader total_rows_loaded=%s phrases_loaded_with_audio=%s words_loaded_with_audio=%s words_loaded_raw=%s phrases_loaded_raw=%s audio_rows_found=%s word_audio_rows_found=%s phrase_audio_rows_found=%s rows_with_audio_in_db_all=%s rows_with_audio_in_db_words_all=%s rows_with_audio_in_db_phrases_all=%s rows_with_any_audio=%s rows_rendered_with_audio_url=%s audio_attached_count=%s phrases_with_missing_oromo_but_shown=%s learn_visible_phrase_count=%s learn_recent_limit=%s newest_phrase_ids=%s",
        total_rows_loaded,
        phrases_loaded_with_audio,
        words_loaded_with_audio,
        words_loaded_raw,
        phrases_loaded_raw,
        audio_rows_found,
        word_audio_rows_found,
        phrase_audio_rows_found,
        rows_with_audio_in_db_all,
        rows_with_audio_in_db_words_all,
        rows_with_audio_in_db_phrases_all,
        rows_with_any_audio,
        rows_rendered_with_audio_url,
        audio_attached_count,
        phrases_with_missing_oromo_but_shown,
        learn_visible_phrase_count,
        int(LEARN_RECENT_PHRASE_LIMIT),
        newest_phrase_ids[:20],
    )
    return rows


def _warmup_learn_tts_for_rows(learn_rows, max_entries: int = 0):
    summary = {
        "processed_entries": 0,
        "generated": 0,
        "cached": 0,
        "failed": 0,
        "skipped_missing_text": 0,
        "skipped_missing_voice": 0,
    }
    cap = int(max_entries or 0)
    if cap <= 0 or not learn_rows:
        return summary
    if not _get_azure_speech_key() or not _get_azure_speech_region():
        app.logger.info("Learn TTS warmup skipped: missing Azure Speech credentials.")
        return summary

    needs_om = bool(_azure_voice_for_lang("om"))
    for row in learn_rows:
        if summary["processed_entries"] >= cap:
            break
        entry_type = (row or {}).get("entry_type")
        entry_id = int((row or {}).get("entry_id") or 0)
        if entry_type not in ("word", "phrase") or entry_id <= 0:
            continue

        audio = (row or {}).get("audio") or {}
        missing_core = any(not normalize_text(audio.get(lang, "") or "") for lang in ("en", "am", "ar", "fr", "zh-CN"))
        missing_oromo = needs_om and (not normalize_text(audio.get("oromo", "") or ""))
        if not (missing_core or missing_oromo):
            continue

        row_summary = generate_tts_for_entry(entry_type, entry_id, force_regenerate=False)
        summary["processed_entries"] += 1
        for key in ("generated", "cached", "failed", "skipped_missing_text", "skipped_missing_voice"):
            summary[key] += int(row_summary.get(key, 0) or 0)

    if summary["processed_entries"] > 0:
        app.logger.info(
            "Learn TTS warmup processed=%s generated=%s cached=%s failed=%s skipped_missing_text=%s skipped_missing_voice=%s",
            summary["processed_entries"],
            summary["generated"],
            summary["cached"],
            summary["failed"],
            summary["skipped_missing_text"],
            summary["skipped_missing_voice"],
        )
    return summary


@app.route("/learn", methods=["GET"])
def learn():
    _log_db_context("/learn")
    trending = get_trending(limit=15)
    learn_rows = _load_learn_rows()
    total_rows_loaded = int(len(learn_rows or []))
    words_loaded_with_audio = 0
    phrases_loaded_with_audio = 0
    rows_rendered_with_audio_url = 0
    for r in (learn_rows or []):
        if (r or {}).get("entry_type") == "word":
            words_loaded_with_audio += 1
        elif (r or {}).get("entry_type") == "phrase":
            phrases_loaded_with_audio += 1
        a = (r or {}).get("audio") or {}
        if any(normalize_text(v or "") for v in a.values()):
            rows_rendered_with_audio_url += 1
    rows_with_any_audio = int(rows_rendered_with_audio_url)
    learn_visible_phrase_count = int(phrases_loaded_with_audio)
    newest_phrase_ids_selected = [
        int((r or {}).get("entry_id") or 0)
        for r in (learn_rows or [])
        if (r or {}).get("entry_type") == "phrase"
    ]
    learn_render_path = "table_rows" if total_rows_loaded > 0 else "table_rows_empty"
    legacy_cards_path = False
    app.logger.info(
        "/learn render template_version=%s build=%s render_path=%s legacy_cards_path=%s phrases_loaded_with_audio=%s words_loaded_with_audio=%s total_rows_loaded=%s rows_with_any_audio=%s rows_rendered_with_audio_url=%s learn_visible_phrase_count=%s learn_recent_limit=%s newest_phrase_ids_selected=%s audio_js=%s pwa_ui_js=%s sw_js=%s",
        LEARN_TEMPLATE_VERSION,
        APP_BUILD_TOKEN,
        learn_render_path,
        legacy_cards_path,
        phrases_loaded_with_audio,
        words_loaded_with_audio,
        total_rows_loaded,
        rows_with_any_audio,
        rows_rendered_with_audio_url,
        learn_visible_phrase_count,
        int(LEARN_RECENT_PHRASE_LIMIT),
        newest_phrase_ids_selected[:20],
        AUDIO_JS_VERSION,
        PWA_UI_JS_VERSION,
        SW_JS_VERSION,
    )
    learn_debug = {
        "words_loaded_with_audio": words_loaded_with_audio,
        "phrases_loaded_with_audio": phrases_loaded_with_audio,
        "words_loaded": words_loaded_with_audio,
        "phrases_loaded": phrases_loaded_with_audio,
        "total_rows_loaded": total_rows_loaded,
        "rows_loaded": total_rows_loaded,
        "rows_with_any_audio": rows_with_any_audio,
        "rows_rendered_with_audio_url": rows_rendered_with_audio_url,
        "learn_visible_phrase_count": learn_visible_phrase_count,
        "learn_recent_limit": int(LEARN_RECENT_PHRASE_LIMIT),
        "newest_phrase_ids_selected": newest_phrase_ids_selected[:20],
        "template_version": LEARN_TEMPLATE_VERSION,
        "build_token": APP_BUILD_TOKEN,
        "render_path": learn_render_path,
        "legacy_cards_path": legacy_cards_path,
        "audio_js_version": AUDIO_JS_VERSION,
        "pwa_ui_js_version": PWA_UI_JS_VERSION,
        "sw_js_version": SW_JS_VERSION,
        "sw_canonical_url": SW_CANONICAL_URL,
    }
    resp = make_response(
        render_template("learn.html", trending=trending, learn_rows=learn_rows, learn_debug=learn_debug)
    )
    resp.headers["X-Gadaa-Build"] = APP_BUILD_TOKEN
    resp.headers["X-Learn-Template-Version"] = LEARN_TEMPLATE_VERSION
    resp.headers["X-Learn-Render-Path"] = learn_render_path
    resp.headers["X-Learn-Legacy-Cards"] = "1" if legacy_cards_path else "0"
    resp.headers["X-Audio-JS-Version"] = AUDIO_JS_VERSION
    resp.headers["X-PWA-UI-JS-Version"] = PWA_UI_JS_VERSION
    resp.headers["X-SW-JS-Version"] = SW_JS_VERSION
    return resp


# ------------------ SUPPORT ------------------

@app.route("/support", methods=["GET"])
def support():
    trending = get_trending(limit=10)
    return render_template("support.html", trending=trending)


# ------------------ HOME ------------------

@app.route("/", methods=["GET", "POST"])
def home():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    result = None
    result_id = None
    suggestions = None
    audio = None

    if request.method == "POST":
        raw = request.form.get("word", "")
        word = make_search_key(raw)

        c.execute("""
            SELECT id, english, oromo
            FROM words
            WHERE status='approved' AND (english_key=? OR oromo_key=?)
        """, (word, word))
        row = c.fetchone()

        if row:
            result_id = row[0]
            result = (row[1], row[2])
            audio = get_approved_audio("word", result_id)

        if not row and word:
            suggestions = {
                "en": suggest_terms(word, "en_om"),
                "om": suggest_terms(word, "om_en")
            }

    conn.close()

    trending = get_trending(limit=15)
    approved_oromo_audio_word_ids = get_approved_oromo_audio_ids("word")

    return render_template(
        "index.html",
        result=result,
        result_id=result_id,
        audio=audio,
        suggestions=suggestions,
        trending=trending,
        approved_oromo_audio_word_ids=approved_oromo_audio_word_ids
        
    )

# ------------------ DICTIONARY ------------------

def make_phrase_slug(english_text: str) -> str:
    """
    Build a stable phrase slug from English text:
    - lowercase
    - punctuation removed safely
    - spaces/underscores -> hyphens
    """
    t = normalize_text(english_text or "").casefold()
    if not t:
        return ""
    t = unicodedata.normalize("NFKD", t)
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = re.sub(r"[^\w\s-]", " ", t)
    t = re.sub(r"[\s_]+", "-", t).strip("-")
    t = re.sub(r"-{2,}", "-", t)
    return t

@app.route("/dictionary", methods=["GET", "POST"])
def dictionary():
    _log_db_context("/dictionary")
    result = None
    result_id = None
    suggestions = None
    audio = None
    source_lang = "om"
    target_lang = "en"
    is_auto_translation = False
    tts_audio_url = None
    lookup_error = None
    other_translations = {}

    # --- GET search (?q=) ---
    q = request.args.get("q", "").strip()
    source_lang = (request.args.get("source_lang") or source_lang).strip()
    target_lang = (request.args.get("target_lang") or target_lang).strip()

    # --- POST search (fallback hvis form bruker POST) ---
    if request.method == "POST":
        q = (request.form.get("q") or request.form.get("word") or "").strip()
        source_lang = (request.form.get("source_lang") or source_lang).strip()
        target_lang = (request.form.get("target_lang") or target_lang).strip()

    if not _is_supported_lang(source_lang):
        source_lang = "om"
    if not _is_supported_lang(target_lang):
        target_lang = "en"
    if source_lang == target_lang:
        target_lang = "en" if source_lang != "en" else "om"

    if q:
        try:
            result, result_id, lookup_error, from_base = _dictionary_lookup_result(q, source_lang, target_lang)
            is_auto_translation = bool(result and result.get("is_auto_translation"))
            tts_audio_url = result.get("tts_audio_url") if result else None
            if result and result.get("auto_unavailable"):
                lookup_error = "Auto translation unavailable. Showing base Oromo-English result."

            if result_id:
                audio, audio_meta = _get_saved_audio_for_entry(
                    "word",
                    int(result_id),
                    english_text=(result or {}).get("english", ""),
                    oromo_text=(result or {}).get("oromo", ""),
                    allow_generate=False,
                    return_meta=True,
                )
                app.logger.info(
                    "/dictionary audio loader rows_loaded=%s audio_rows_found=%s audio_urls_attached=%s entry_type=%s entry_id=%s",
                    1,
                    int((audio_meta or {}).get("audio_rows_found", 0) or 0),
                    int((audio_meta or {}).get("audio_urls_attached", 0) or 0),
                    "word",
                    int(result_id or 0),
                )
                try:
                    other_translations = get_saved_extra_translations(result_id)
                except Exception as e:
                    app.logger.exception(f"/dictionary extra translations failed: {repr(e)}")
                    other_translations = {}
            else:
                app.logger.info(
                    "/dictionary audio loader rows_loaded=%s audio_rows_found=%s audio_urls_attached=%s",
                    0,
                    0,
                    0,
                )

            if (not from_base) and source_lang in ("om", "en"):
                word = make_search_key(q)
                suggestions = {
                    "en": suggest_terms(word, "en_om"),
                    "om": suggest_terms(word, "om_en")
                }
                if result is None:
                    # Trigger template "no result found" block safely.
                    result = {}
        except Exception as e:
            app.logger.exception(f"/dictionary lookup failed: {repr(e)}")
            lookup_error = "Translation is temporarily unavailable. Showing base dictionary data."
            result = {
                "source_text": q,
                "target_text": "",
                "english": "",
                "oromo": "",
                "is_auto_translation": False,
                "tts_audio_url": None
            }

    # full dictionary list (load after lookup work so on-use cache writes are
    # not competing with this route-level DB handle)
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("""
            SELECT id, english, oromo
            FROM words
            WHERE status='approved'
            ORDER BY english ASC
        """)
    all_words = c.fetchall()
    c.execute("""
            SELECT id, english, oromo
            FROM phrases
            WHERE status='approved'
            ORDER BY english ASC
        """)
    all_phrases = c.fetchall()
    phrase_slugs = {}
    for pid, en, _om in all_phrases:
        pid_int = int(pid or 0)
        if pid_int <= 0:
            continue
        phrase_slugs[pid_int] = make_phrase_slug(en or "")

    list_other_translations = {}
    list_phrase_translations = {}

    try:
        placeholders = ",".join("?" for _ in EXTRA_GENERATED_LANGS)
        c.execute(
            f"""
            SELECT gt.word_id, gt.lang_code, gt.translated_text
            FROM generated_translations gt
            JOIN words w ON w.id = gt.word_id
            WHERE w.status='approved'
              AND gt.lang_code IN ({placeholders})
              AND gt.translated_text IS NOT NULL
              AND TRIM(gt.translated_text) != ''
            """,
            EXTRA_GENERATED_LANGS,
        )
        for wid, lang_code, translated_text in c.fetchall():
            wid_int = int(wid or 0)
            txt = normalize_text(translated_text or "")
            if not wid_int or not _is_meaningful_generated_text(txt):
                continue
            row = list_other_translations.setdefault(wid_int, {})
            row[lang_code] = txt
    except Exception as e:
        app.logger.exception(f"/dictionary list extra translations failed: {repr(e)}")
        list_other_translations = {}

    try:
        placeholders = ",".join("?" for _ in EXTRA_GENERATED_LANGS)
        c.execute(
            f"""
            SELECT gt.phrase_id, gt.lang_code, gt.translated_text
            FROM generated_phrase_translations gt
            JOIN phrases p ON p.id = gt.phrase_id
            WHERE p.status='approved'
              AND gt.lang_code IN ({placeholders})
              AND gt.translated_text IS NOT NULL
              AND TRIM(gt.translated_text) != ''
            """,
            EXTRA_GENERATED_LANGS,
        )
        for pid, lang_code, translated_text in c.fetchall():
            pid_int = int(pid or 0)
            txt = normalize_text(translated_text or "")
            if not pid_int or not _is_meaningful_generated_text(txt):
                continue
            row = list_phrase_translations.setdefault(pid_int, {})
            row[lang_code] = txt
    except Exception as e:
        app.logger.exception(f"/dictionary list phrase translations failed: {repr(e)}")
        list_phrase_translations = {}

    conn.close()

    trending = get_trending(limit=15)
    approved_oromo_audio_word_ids = get_approved_oromo_audio_ids("word")

    return render_template(
        "dictionary.html",
        q=q,
        result=result,
        result_id=result_id,
        source_lang=source_lang,
        target_lang=target_lang,
        language_options=LANGUAGE_OPTIONS,
        result_is_rtl=_is_rtl_lang(target_lang),
        source_speech_lang=_speech_lang_code(source_lang),
        target_speech_lang=_speech_lang_code(target_lang),
        is_auto_translation=is_auto_translation,
        tts_audio_url=tts_audio_url,
        lookup_error=lookup_error,
        other_translations=other_translations,
        list_other_translations=list_other_translations,
        extra_generated_langs=EXTRA_GENERATED_LANGS,
        audio=audio,
        words=all_words,
        phrases=all_phrases,
        suggestions=suggestions,
        trending=trending,
        approved_oromo_audio_word_ids=approved_oromo_audio_word_ids,
        list_phrase_translations=list_phrase_translations,
        phrase_slugs=phrase_slugs,
    )

@app.route("/phrase/<slug>", methods=["GET"])
def phrase_detail(slug):
    _log_db_context("/phrase")
    incoming_slug = normalize_text(unquote(slug or "")).strip().casefold()
    if not incoming_slug:
        abort(404)

    row = None
    other_translations = {}
    audio = {}
    conn = None
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute(
            """
            SELECT id, english, oromo
            FROM phrases
            WHERE status='approved'
            ORDER BY id ASC
            """
        )
        for pid, en, om in c.fetchall():
            s = make_phrase_slug(en or "")
            if s == incoming_slug:
                row = (int(pid or 0), en or "", om or "", s)
                break

        if not row:
            conn.close()
            conn = None
            abort(404)

        pid, en, om, canonical_slug = row
        if not canonical_slug:
            conn.close()
            conn = None
            abort(404)
        if incoming_slug != canonical_slug:
            conn.close()
            conn = None
            return redirect(f"/phrase/{canonical_slug}", code=301)

        placeholders = ",".join("?" for _ in EXTRA_GENERATED_LANGS)
        c.execute(
            f"""
            SELECT lang_code, translated_text
            FROM generated_phrase_translations
            WHERE phrase_id=?
              AND lang_code IN ({placeholders})
              AND translated_text IS NOT NULL
              AND TRIM(translated_text) != ''
            """,
            (pid, *EXTRA_GENERATED_LANGS),
        )
        other_translations = {}
        for lang_code, translated_text in c.fetchall():
            txt = normalize_text(translated_text or "")
            if txt and _is_meaningful_generated_text(txt):
                other_translations[lang_code] = txt

        try:
            audio = _get_saved_audio_for_entry(
                "phrase",
                int(pid),
                english_text=normalize_text(en or ""),
                oromo_text=normalize_text(om or ""),
                allow_generate=False,
            ) or {}
        except Exception as e:
            app.logger.exception(f"/phrase audio lookup failed: {repr(e)}")
            audio = {}

        conn.close()
        conn = None
    except HTTPException:
        raise
    except Exception as e:
        if conn:
            try:
                conn.close()
            except Exception:
                pass
        app.logger.exception(f"/phrase lookup failed: {repr(e)}")
        abort(404)

    phrase = {
        "id": int(pid),
        "english": en,
        "oromo": om,
        "audio": audio,
    }

    page_title = f"{en} meaning in Oromo ({om}) | {APP_NAME}" if om else f"{en} meaning in Oromo | {APP_NAME}"
    meta_description = (
        f"{en} means {om or 'this Oromo translation'}. "
        f"Find phrase translations in Oromo, Amharic, Arabic, French, and Chinese on {APP_NAME}."
    )[:160]
    canonical_url = f"{_site_base_url()}/phrase/{canonical_slug}"

    return render_template(
        "phrase.html",
        phrase=phrase,
        other_translations=other_translations,
        language_options=LANGUAGE_OPTIONS,
        current_year=datetime.utcnow().year,
        APP_NAME=APP_NAME,
        page_title=page_title,
        meta_description=meta_description,
        canonical_url=canonical_url,
    )

@app.route("/word/<path:term>", methods=["GET"])
def word_detail(term):
    _log_db_context("/word")
    raw = normalize_text(unquote(term or ""))
    key = make_search_key(_strip_edge_punct(raw))
    if not key:
        abort(404)

    row = None
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute(
            """
            SELECT id, english, oromo
            FROM words
            WHERE status='approved' AND (english_key=? OR oromo_key=?)
            LIMIT 1
            """,
            (key, key),
        )
        row = c.fetchone()
        conn.close()
    except Exception as e:
        app.logger.exception(f"/word lookup failed: {repr(e)}")
        row = None

    if not row:
        abort(404)

    wid, en, om = row
    english_word = (en or "").strip()
    if not english_word:
        abort(404)

    canonical_path = f"/word/{quote(english_word, safe='')}"
    incoming_term = normalize_text(unquote(term or ""))
    # Canonicalize to one stable URL per base entry to reduce duplicate indexing.
    if incoming_term.casefold() != english_word.casefold():
        return redirect(canonical_path, code=301)

    audio = {}
    try:
        audio, audio_meta = _get_saved_audio_for_entry(
            "word",
            int(wid),
            english_text=normalize_text(en or ""),
            oromo_text=normalize_text(om or ""),
            allow_generate=False,
            return_meta=True,
        )
        app.logger.info(
            "/word audio loader rows_loaded=%s audio_rows_found=%s audio_urls_attached=%s entry_type=%s entry_id=%s",
            1,
            int((audio_meta or {}).get("audio_rows_found", 0) or 0),
            int((audio_meta or {}).get("audio_urls_attached", 0) or 0),
            "word",
            int(wid or 0),
        )
    except Exception as e:
        app.logger.exception(f"/word audio lookup failed: {repr(e)}")
        audio = {}

    word = {
        "id": wid,
        "english": en or "",
        "oromo": om or "",
        "en": en or "",
        "om": om or "",
        "explanation": "",
        "audio_oromo": audio.get("oromo", ""),
        "audio_english": audio.get("english", ""),
    }

    other_translations = {}
    try:
        other_translations = get_saved_extra_translations(wid) or {}
    except Exception as e:
        app.logger.exception(f"/word extra translations failed: {repr(e)}")
        other_translations = {}

    oromo_word = (om or "").strip()

    page_title = f"{english_word} meaning in Oromo"
    if oromo_word:
        page_title += f" ({oromo_word})"
    page_title += f" | Amharic, Arabic, French & Chinese | {APP_NAME}"

    meta_description = (
        f"{english_word} means {oromo_word or 'this Oromo translation'}. "
        f"Find translations in Oromo, Amharic, Arabic, French, and Chinese on {APP_NAME}."
    )[:160]

    canonical_url = f"{_site_base_url()}/word/{quote(english_word, safe='')}"

    return render_template(
        "words.html",
        word=word,
        other_translations=other_translations,
        current_year=datetime.utcnow().year,
        APP_NAME=APP_NAME,
        page_title=page_title,
        meta_description=meta_description,
        canonical_url=canonical_url,
    )
# ------------------ TRANSLATE ------------------


# Make sure these exist somewhere in your file/module:
# - DB_NAME
# - normalize_text
# - make_search_key
# - detect_direction_auto
# - _strip_edge_punct
# - translate_multipart_text
# - record_search
# - suggest_terms
# - get_trending
# - get_approved_audio
# - get_approved_oromo_audio_ids

# Tokenizer used for "single-word" detection (keeps punctuation separate)
_TOKEN_RE = re.compile(r"\s+|[^\w\s]+|[\w']+", re.UNICODE)


def build_key_candidates(s: str):
    """Same key strategy as translate_text(): robust vs punctuation/spacing mismatch."""
    s = normalize_text(s)
    cands = []

    k1 = make_search_key(s)
    if k1:
        cands.append(k1)

    s2 = _strip_edge_punct(s)
    k2 = make_search_key(s2)
    if k2 and k2 not in cands:
        cands.append(k2)

    s3 = re.sub(r"\s+", " ", s2).strip()
    k3 = make_search_key(s3)
    if k3 and k3 not in cands:
        cands.append(k3)

    return cands

def phrase_key_candidates(s: str):
    s = normalize_text(s)
    cands = []
    k1 = make_search_key(s)                       # keep punctuation if present
    if k1:
        cands.append(k1)

    s2 = _strip_edge_punct(s)                     # remove edge punctuation
    k2 = make_search_key(s2)
    if k2 and k2 not in cands:
        cands.append(k2)

    return cands


def find_exact_base_match(text: str, source_lang: str):
    if source_lang not in ("om", "en"):
        return None, None

    clean_exact = normalize_text(text)
    key_candidates = build_key_candidates(text)
    phrase_col = "oromo_key" if source_lang == "om" else "english_key"
    word_col = "oromo_key" if source_lang == "om" else "english_key"

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    matched = None
    audio = None

    if key_candidates:
        for k in key_candidates:
            pr = c.execute(
                f"SELECT id, english, oromo FROM phrases WHERE status='approved' AND {phrase_col}=? LIMIT 1",
                (k,)
            ).fetchone()
            if pr:
                matched = {"type": "phrase", "id": pr[0], "english": pr[1], "oromo": pr[2]}
                audio = _get_saved_audio_for_entry(
                    "phrase",
                    int(pr[0]),
                    english_text=normalize_text(pr[1] or ""),
                    oromo_text=normalize_text(pr[2] or ""),
                    allow_generate=False,
                )
                break

    if not matched:
        tokens = [t for t in _TOKEN_RE.findall(clean_exact) if not t.isspace()]
        word_tokens = [t for t in tokens if re.fullmatch(r"[\w']+", t)]
        if len(word_tokens) == 1 and len([t for t in tokens if re.fullmatch(r"[\w']+", t)]) == 1:
            wkey = make_search_key(word_tokens[0])
            if wkey:
                wr = c.execute(
                    f"SELECT id, english, oromo FROM words WHERE status='approved' AND {word_col}=? LIMIT 1",
                    (wkey,)
                ).fetchone()
                if wr:
                    matched = {"type": "word", "id": wr[0], "english": wr[1], "oromo": wr[2]}
                    audio = _get_saved_audio_for_entry(
                        "word",
                        int(wr[0]),
                        english_text=normalize_text(wr[1] or ""),
                        oromo_text=normalize_text(wr[2] or ""),
                        allow_generate=False,
                    )

    conn.close()
    return matched, audio


@app.route("/translate", methods=["GET", "POST"])
def translate():
    _log_db_context("/translate")
    result = None
    text = ""
    direction = "om_en"
    source_lang = "om"
    target_lang = "en"
    suggestions = None
    audio = None
    matched = None
    phrase_suggestions = None
    is_auto_translation = False
    tts_audio_url = None
    translate_error = None
    is_query_variant = (request.method == "GET" and bool(request.query_string))

    if request.method == "POST":
        text = request.form.get("text", "")
        source_lang = (request.form.get("source_lang") or "").strip()
        target_lang = (request.form.get("target_lang") or "").strip()
        legacy_direction = (request.form.get("direction") or "").strip()

        # Backward compatibility for old form payloads
        if not source_lang or not target_lang:
            if legacy_direction == "om_en":
                source_lang, target_lang = "om", "en"
            elif legacy_direction == "en_om":
                source_lang, target_lang = "en", "om"
            elif legacy_direction == "auto":
                detected = detect_direction_auto(text)
                source_lang, target_lang = ("om", "en") if detected == "om_en" else ("en", "om")

        if not _is_supported_lang(source_lang):
            source_lang = "om"
        if not _is_supported_lang(target_lang):
            target_lang = "en"
        if source_lang == target_lang:
            target_lang = "en" if source_lang != "en" else "om"

        if source_lang == "om" and target_lang == "en":
            direction = "om_en"
        elif source_lang == "en" and target_lang == "om":
            direction = "en_om"
        else:
            direction = f"{source_lang}_{target_lang}"

        if source_lang in ("om", "en"):
            matched, audio = find_exact_base_match(text, source_lang)

        # âœ… IMPORTANT: use multipart translator (handles commas + sentences correctly)
        tr = safe_translate_multilingual(text, source_lang, target_lang)
        result = tr["text"]
        is_exact = tr["is_exact"]
        is_phrase = tr["is_phrase"]
        is_auto_translation = tr["is_auto_translation"]
        tts_audio_url = tr["tts_audio_url"]

        if text and not result:
            if source_lang == "om" and target_lang not in ("om", "en"):
                base_en, _, _ = translate_multipart_text(text, "om_en")
                result = base_en
                translate_error = "Auto translation unavailable. Showing base Oromo-English result."
            elif source_lang == "en" and target_lang not in ("om", "en"):
                result = normalize_text(text)
                translate_error = "Auto translation unavailable. Showing base Oromo-English result."
            else:
                translate_error = "Translation service is temporarily unavailable. Please try again."

        record_search(text, direction, is_phrase, is_exact)

        if source_lang in ("om", "en") and target_lang in ("om", "en") and not is_exact:
            clean_exact = normalize_text(text)
            tokens = [t for t in _TOKEN_RE.findall(clean_exact) if not t.isspace()]
            word_tokens = [t for t in tokens if re.fullmatch(r"[\w']+", t)]
            if len(word_tokens) == 1:
                suggestions = suggest_terms(word_tokens[0], direction)

    trending = get_trending(limit=15)
    approved_oromo_audio_phrase_ids = get_approved_oromo_audio_ids("phrase")
    approved_oromo_audio_word_ids = get_approved_oromo_audio_ids("word")

    resp = make_response(render_template(
        "translate.html",
        result=result,
        text=text,
        direction=direction,
        source_lang=source_lang,
        target_lang=target_lang,
        language_options=LANGUAGE_OPTIONS,
        result_is_rtl=_is_rtl_lang(target_lang),
        source_speech_lang=_speech_lang_code(source_lang),
        target_speech_lang=_speech_lang_code(target_lang),
        is_auto_translation=is_auto_translation,
        tts_audio_url=tts_audio_url,
        translate_error=translate_error,
        suggestions=suggestions,
        phrase_suggestions=phrase_suggestions,
        trending=trending,
        matched=matched,
        audio=audio,
        approved_oromo_audio_word_ids=approved_oromo_audio_word_ids,
        approved_oromo_audio_phrase_ids=approved_oromo_audio_phrase_ids,
        is_query_variant=is_query_variant,
    ))
    if is_query_variant:
        resp.headers["X-Robots-Tag"] = "noindex, follow"
    return resp


# ------------------ MULTI-SENTENCE TRANSLATION ------------------

_WORDLIKE_RE = re.compile(r"[\w']+", re.UNICODE)
_TOKEN_RE = re.compile(r"\s+|[^\w\s]+|[\w']+", re.UNICODE)
_BOUNDARY_RE = re.compile(r"[.!?,;:]")

def _strip_trailing_punct(s: str) -> str:
    # remove trailing boundary punctuation only (.,!?;:)
    return re.sub(r"[.!?,;:]+$", "", (s or "").strip())


def upsert_phrase_aliases(phrase_id: int, english: str, oromo: str, source: str = "auto"):
    """
    Insert alias keys for a phrase. Safe to call multiple times.
    """
    en_aliases = generate_english_alias_texts(english)
    om_aliases = generate_oromo_alias_texts(oromo)

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    # Clear old aliases for this phrase (simple & safe)
    c.execute("DELETE FROM phrase_aliases WHERE phrase_id=?", (phrase_id,))

    # Insert new aliases
    for a in en_aliases:
        k = key_for(a)
        if k:
            c.execute("""
                INSERT OR IGNORE INTO phrase_aliases (phrase_id, english_alias_key, source)
                VALUES (?, ?, ?)
            """, (phrase_id, k, source))

    for a in om_aliases:
        k = key_for(a)
        if k:
            c.execute("""
                INSERT OR IGNORE INTO phrase_aliases (phrase_id, oromo_alias_key, source)
                VALUES (?, ?, ?)
            """, (phrase_id, k, source))

    conn.commit()
    conn.close()


def split_segments(text: str):
    """Return list of (segment_text, punctuation, trailing_ws)."""
    t = normalize_text(text)
    if not t:
        return []

    out, buf = [], []
    i, n = 0, len(t)

    while i < n:
        ch = t[i]
        if _BOUNDARY_RE.match(ch):
            seg = "".join(buf)
            buf = []

            punct = []
            while i < n and _BOUNDARY_RE.match(t[i]):
                punct.append(t[i]); i += 1

            ws = []
            while i < n and t[i].isspace():
                ws.append(t[i]); i += 1

            out.append((seg, "".join(punct), "".join(ws)))
        else:
            buf.append(ch); i += 1

    tail = "".join(buf)
    if tail.strip():
        out.append((tail, "", ""))

    return out

def _phrase_lookup(cur, direction: str, k: str):
    if direction == "om_en":
        row = cur.execute(
            "SELECT english FROM phrases WHERE status='approved' AND oromo_key=? LIMIT 1",
            (k,)
        ).fetchone()
    else:
        row = cur.execute(
            "SELECT oromo FROM phrases WHERE status='approved' AND english_key=? LIMIT 1",
            (k,)
        ).fetchone()
    return row[0] if row else None

def _word_lookup(cur, direction: str, k: str):
    if direction == "om_en":
        row = cur.execute(
            "SELECT english FROM words WHERE status='approved' AND oromo_key=? LIMIT 1",
            (k,)
        ).fetchone()
    else:
        row = cur.execute(
            "SELECT oromo FROM words WHERE status='approved' AND english_key=? LIMIT 1",
            (k,)
        ).fetchone()
    return row[0] if row else None

def translate_segment_best(segment_text: str, direction: str, cur, max_phrase_words: int = 12):
    """
    Translate one segment using longest-phrase-first over word positions.
    Preserves original spacing/punctuation inside the segment.
    """
    if not segment_text or not segment_text.strip():
        return "", 0, 0

    # Tokenize into pieces but we also need a word list for phrase scanning
    tokens = _TOKEN_RE.findall(segment_text)

    # Build mapping from "word index" to token positions
    word_positions = []  # list of (token_idx, word_text)
    for ti, tok in enumerate(tokens):
        if _WORDLIKE_RE.fullmatch(tok):
            word_positions.append((ti, tok))

    # If no words, return as-is
    if not word_positions:
        return segment_text, 0, 0

    # Convenience: list of word strings (in order)
    words = [w for _, w in word_positions]
    n_words = len(words)

    out_tokens = tokens[:]  # we'll replace words/phrases in place
    consumed_word = [False] * n_words
    any_exact = 0
    any_phrase = 0

    i = 0
    while i < n_words:
        if consumed_word[i]:
            i += 1
            continue

        # Try longest phrase starting at i
        found_translation = None
        found_len = 0

        max_len = min(max_phrase_words, n_words - i)
        for L in range(max_len, 1, -1):  # phrases of length >= 2
            phrase_text = " ".join(words[i:i+L])
            k = key_for(phrase_text)
            if not k:
                continue
            tr = _phrase_lookup(cur, direction, k)
            if tr:
                found_translation = tr
                found_len = L
                break

        if found_translation:
            any_exact = 1
            any_phrase = 1

            # Replace the first word token with translation, blank out the rest words in phrase
            first_token_idx = word_positions[i][0]
            out_tokens[first_token_idx] = found_translation

            # Blank out the remaining word tokens and any immediate punctuation attached between them
            for j in range(i, i + found_len):
                consumed_word[j] = True
                if j == i:
                    continue
                tok_idx = word_positions[j][0]
                out_tokens[tok_idx] = ""  # remove word itself

            i += found_len
            continue

        # No phrase found => translate single word
        w = words[i]
        k = key_for(w)
        tr = _word_lookup(cur, direction, k) if k else None
        if tr:
            any_exact = 1
            out_tokens[word_positions[i][0]] = tr
        # else keep original word
        consumed_word[i] = True
        i += 1

    # Join, then clean extra spaces caused by removed tokens (keep user punctuation)
    result = "".join(out_tokens)
    result = re.sub(r"\s+", " ", result).strip()
    return result, any_exact, any_phrase

def translate_multipart_text(text: str, direction: str):
    parts = split_segments(text)
    if not parts:
        return "", 0, 0

    out = []
    any_exact = 0
    any_phrase = 0

    for seg_text, punct, ws in parts:
        if seg_text:
            tr, ex, ph = translate_text(seg_text, direction)

            # âœ… IMPORTANT: avoid doubled punctuation when phrase already ends with . or ?
            # If input has punctuation, input punctuation should win.
            if ph and punct:
                tr = _strip_trailing_punct(tr)

            out.append(tr)
            any_exact |= ex
            any_phrase |= ph
        else:
            out.append(seg_text)

        out.append(punct)
        out.append(ws)

    return "".join(out), int(any_exact), int(any_phrase)

def lookup_phrase_via_alias(cur, direction: str, alias_key: str):
    if direction == "om_en":
        row = cur.execute("""
            SELECT p.english
            FROM phrase_aliases a
            JOIN phrases p ON p.id = a.phrase_id
            WHERE p.status='approved' AND a.oromo_alias_key=?
            LIMIT 1
        """, (alias_key,)).fetchone()
    else:
        row = cur.execute("""
            SELECT p.oromo
            FROM phrase_aliases a
            JOIN phrases p ON p.id = a.phrase_id
            WHERE p.status='approved' AND a.english_alias_key=?
            LIMIT 1
        """, (alias_key,)).fetchone()
    return row[0] if row else None

def lookup_word(cur, direction: str, w_key: str):
    if direction == "om_en":
        row = cur.execute("""
            SELECT english FROM words WHERE status='approved' AND oromo_key=? LIMIT 1
        """, (w_key,)).fetchone()
    else:
        row = cur.execute("""
            SELECT oromo FROM words WHERE status='approved' AND english_key=? LIMIT 1
        """, (w_key,)).fetchone()
    return row[0] if row else None

def translate_segment_longest_phrase(cur, segment_text: str, direction: str, max_phrase_words: int = 12):
    """
    Greedy longest-phrase-first over words in the segment.
    Returns (translated_segment, any_exact, any_phrase).
    """
    seg = normalize_text(segment_text)
    if not seg:
        return "", 0, 0

    # âœ… Grammar template check FIRST
    if direction == "en_om":
        for pattern, replacement in EN_OM_TEMPLATES:
            if pattern.match(seg):
                return replacement, 1, 1

    
    words = seg.split()
    out = []
    i = 0
    any_exact = 0
    any_phrase = 0

    while i < len(words):
        best_tr = None
        best_len = 0

        # Try phrases first (length >= 2)
        Lmax = min(max_phrase_words, len(words) - i)
        for L in range(Lmax, 1, -1):
            phrase_text = " ".join(words[i:i+L])
            k = key_for(phrase_text)
            if not k:
                continue
            tr = lookup_phrase_via_alias(cur, direction, k)
            if tr:
                best_tr = tr
                best_len = L
                break

        if best_tr:
            out.append(best_tr)
            any_exact = 1
            any_phrase = 1
            i += best_len
            continue

        # ---- word fallback ----
        w = words[i]
        w_cf = w.casefold()

        # Drop safe English filler words in English->Oromo fallback
        if direction == "en_om" and (w_cf in EN_DROP_WORDS or w_cf in EN_SOFT_WORDS):
            i += 1
            continue

        wk = key_for(w)
        trw = lookup_word(cur, direction, wk) if wk else None
        if trw:
            out.append(trw)
            any_exact = 1
        else:
            out.append(w)

        i += 1

    result = " ".join(out)
    result = postprocess_segment(result, direction)
    return result, int(any_exact), int(any_phrase)



# ------------------ PUBLIC SUBMISSION (WORDS) ------------------

@app.route("/submit", methods=["GET", "POST"])
def submit():
    msg = None

    if request.method == "POST":
        mode = (request.form.get("mode") or "").strip().lower()
        f = request.files.get("file")

        # ---------- FILE MODE ----------
        if mode == "file" or (f and f.filename):
            if not f or not f.filename:
                msg = "Please choose a CSV or XLSX file."
                return render_template("submit.html", msg=msg)

            import tempfile

            filename = (f.filename or "").lower().strip()

            # Save upload to temp file (streamed, avoids RAM spike)
            with tempfile.NamedTemporaryFile(delete=False, suffix=filename) as tmp:
                f.save(tmp.name)
                path = tmp.name

            try:
                if filename.endswith(".csv"):
                    pairs = parse_csv_pairs_from_path(path)
                elif filename.endswith(".xlsx"):
                    pairs = parse_xlsx_pairs_from_path(path)
                else:
                    msg = "Only .csv or .xlsx files are allowed."
                    return render_template("submit.html", msg=msg)
            except Exception as e:
                app.logger.exception(f"submit (words) file parse error: {repr(e)}")
                msg = "Could not read the file. Please check its format."
                return render_template("submit.html", msg=msg)
            finally:
                try:
                    os.remove(path)
                except OSError:
                    pass

            if not pairs:
                msg = "No rows found in the file."
                return render_template("submit.html", msg=msg)

            for en, om in pairs:
                if not en or not om:
                    msg = "Rejected: Every row must include BOTH English and Oromo."
                    return render_template("submit.html", msg=msg)

            inserted = 0
            skipped = 0
            repaired = 0

            conn = sqlite3.connect(DB_NAME)
            for en, om in pairs:
                _wid, was_inserted, was_repaired = _upsert_pending_word_base(conn, en, om)
                if was_inserted:
                    inserted += 1
                elif was_repaired:
                    repaired += 1
                else:
                    skipped += 1

            conn.commit()
            conn.close()

            msg = (
                f"Thanks! File submitted. "
                f"Added: {inserted} | Completed partial entries: {repaired} | Skipped duplicates: {skipped}. "
                f"Waiting for admin approval."
            )
            return render_template("submit.html", msg=msg)
        
        # ---------- TEXT MODE ----------
        english_raw = (request.form.get("english") or "").strip()
        oromo_raw = (request.form.get("oromo") or "").strip()
        
        english_key = make_search_key(english_raw)
        oromo_key = make_search_key(oromo_raw)

        if not english_key or not oromo_key:
            msg = "Please provide both English and Oromo."
            return render_template("submit.html", msg=msg)
        

        conn = sqlite3.connect(DB_NAME)
        _wid, was_inserted, was_repaired = _upsert_pending_word_base(conn, english_raw, oromo_raw)
        conn.commit()
        conn.close()

        if not was_inserted and not was_repaired:
            msg = "This word already exists (or is pending). Try another."
            return render_template("submit.html", msg=msg)
        if was_repaired:
            msg = "Existing partial entry was completed and is waiting for admin approval."
            return render_template("submit.html", msg=msg)

        msg = "Thank you! Your word is waiting for admin approval."
        return render_template("submit.html", msg=msg)

    return render_template("submit.html", msg=msg)

        


# ------------------ PUBLIC SUBMISSION (PHRASES) ------------------

@app.route("/submit_phrase", methods=["GET", "POST"])
def submit_phrase():
    msg = None

    if request.method == "POST":
        mode = (request.form.get("mode") or "").strip().lower()
        f = request.files.get("file")

        # ---------- FILE MODE ----------
        if mode == "file" or (f and f.filename):
            if not f or not f.filename:
                msg = "Please choose a CSV or XLSX file."
                return render_template("submit_phrase.html", msg=msg)

            import tempfile

            filename = (f.filename or "").lower().strip()

            # Save upload to temp file (no RAM spike)
            with tempfile.NamedTemporaryFile(delete=False, suffix=filename) as tmp:
                f.save(tmp.name)
                path = tmp.name

            try:
                if filename.endswith(".csv"):
                    pairs = parse_csv_pairs_from_path(path)
                elif filename.endswith(".xlsx"):
                    pairs = parse_xlsx_pairs_from_path(path)
                else:
                    msg = "Only .csv or .xlsx files are allowed."
                    return render_template("submit_phrase.html", msg=msg)
            except Exception as e:
                app.logger.exception(f"submit_phrase file parse error: {repr(e)}")
                msg = "Could not read the file. Please check its format."
                return render_template("submit_phrase.html", msg=msg)
            finally:
                try:
                    os.remove(path)
                except OSError:
                    pass

            if not pairs:
                msg = "No rows found in the file."
                return render_template("submit_phrase.html", msg=msg)

            for en, om in pairs:
                if not en or not om:
                    msg = "Rejected: Every row must include BOTH English and Oromo."
                    return render_template("submit_phrase.html", msg=msg)

            inserted = 0
            skipped = 0

            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()

            for en, om in pairs:
                c.execute(
                    "SELECT 1 FROM phrases WHERE english=? OR oromo=? LIMIT 1",
                    (en, om),
                )
                if c.fetchone():
                    skipped += 1
                    continue

                c.execute(
                    "INSERT INTO phrases (english, oromo, status) VALUES (?, ?, 'pending')",
                    (en, om),
                )
                inserted += 1

            conn.commit()
            conn.close()

            msg = (
                f"Thanks! Phrase file submitted. "
                f"Added: {inserted} | Skipped duplicates: {skipped}. "
                f"Waiting for admin approval."
            )
            return render_template("submit_phrase.html", msg=msg)

        # ---------- TEXT MODE ----------
        english = normalize_text(request.form.get("english", ""))
        oromo = normalize_text(request.form.get("oromo", ""))

        if not english or not oromo:
            msg = "Please provide both English and Oromo phrase."
            return render_template("submit_phrase.html", msg=msg)

        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()

        c.execute(
            "SELECT 1 FROM phrases WHERE english=? OR oromo=?",
            (english, oromo),
        )
        if c.fetchone():
            conn.close()
            msg = "This phrase already exists (or is pending). Try another."
            return render_template("submit_phrase.html", msg=msg)

        c.execute(
            "INSERT INTO phrases (english, oromo, status) VALUES (?, ?, 'pending')",
            (english, oromo),
        )
        conn.commit()
        conn.close()

        msg = "Thank you! Your phrase is waiting for admin approval."
        return render_template("submit_phrase.html", msg=msg)

    return render_template("submit_phrase.html", msg=msg)


# ------------------ LEGACY: COMMUNITY FILE SUBMISSION ------------------

@app.route("/submit_file", methods=["GET", "POST"])
def submit_file():
    msg = None

    if request.method == "POST":
        f = request.files.get("file")
        if not f or not f.filename:
            msg = "Please choose a file."
            return render_template("submit_file.html", msg=msg)

        import tempfile

        filename = (f.filename or "").lower().strip()

        # Save upload to temp file (prevents RAM spike)
        with tempfile.NamedTemporaryFile(delete=False, suffix=filename) as tmp:
            f.save(tmp.name)
            path = tmp.name

        try:
            if filename.endswith(".csv"):
                pairs = parse_csv_pairs_from_path(path)
            elif filename.endswith(".xlsx"):
                pairs = parse_xlsx_pairs_from_path(path)
            else:
                msg = "Only .csv or .xlsx files are allowed."
                return render_template("submit_file.html", msg=msg)
        except Exception as e:
            app.logger.exception(f"submit_file parse error: {repr(e)}")
            msg = "Could not read the file. Please check its format."
            return render_template("submit_file.html", msg=msg)
        finally:
            try:
                os.remove(path)
            except OSError:
                pass

        if not pairs:
            msg = "No rows found in the file."
            return render_template("submit_file.html", msg=msg)

        for en, om in pairs:
            if not en or not om:
                msg = "Rejected: Every row must include BOTH English and Oromo."
                return render_template("submit_file.html", msg=msg)

        inserted = 0
        skipped = 0
        repaired = 0

        conn = sqlite3.connect(DB_NAME)
        for en, om in pairs:
            _wid, was_inserted, was_repaired = _upsert_pending_word_base(conn, en, om)
            if was_inserted:
                inserted += 1
            elif was_repaired:
                repaired += 1
            else:
                skipped += 1

        conn.commit()
        conn.close()

        msg = (
            f"Thanks! File submitted. "
            f"Added: {inserted} | Completed partial entries: {repaired} | Skipped duplicates: {skipped}. "
            f"Waiting for admin approval."
        )
        return render_template("submit_file.html", msg=msg)

    return render_template("submit_file.html", msg=msg)

# ------------------ RECORDER LOGIN + DASHBOARD ------------------
# Requires env var: RECORDER_PASSWORD

@app.route("/recorder", methods=["GET", "POST"])
def recorder_login():
    msg = None
    if request.method == "POST":
        pw = (request.form.get("password") or "").strip()
        correct = os.environ.get("RECORDER_PASSWORD", "").strip()
        if correct and pw == correct:
            session["recorder"] = 1
            return redirect("/recorder/dashboard")
        msg = "Invalid recorder password."
    return render_template("recorder_login.html", msg=msg)


@app.route("/recorder/logout")
def recorder_logout():
    session.pop("recorder", None)
    return redirect("/")


@app.route("/recorder/dashboard")
def recorder_dashboard():
    if not require_recorder():
        return redirect("/recorder")

    q = normalize_text(request.args.get("q", "") or "")

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    if q:
        like = "%" + q + "%"
        c.execute("""
            SELECT id, english, oromo
            FROM words
            WHERE status='approved' AND (english LIKE ? OR oromo LIKE ?)
            ORDER BY english ASC
            LIMIT 200
        """, (like, like))
        words = c.fetchall()

        c.execute("""
            SELECT id, english, oromo
            FROM phrases
            WHERE status='approved' AND (english LIKE ? OR oromo LIKE ?)
            ORDER BY id DESC
            LIMIT 200
        """, (like, like))
        phrases = c.fetchall()
    else:
        c.execute("""
            SELECT id, english, oromo
            FROM words
            WHERE status='approved'
            ORDER BY english ASC
            LIMIT 100
        """)
        words = c.fetchall()

        c.execute("""
            SELECT id, english, oromo
            FROM phrases
            WHERE status='approved'
            ORDER BY id DESC
            LIMIT 100
        """)
        phrases = c.fetchall()

    conn.close()

    approved_word_ids = get_approved_oromo_audio_ids("word")
    approved_phrase_ids = get_approved_oromo_audio_ids("phrase")

    return render_template(
        "recorder_dashboard.html",
        q=q,
        words=words,
        phrases=phrases,
        approved_word_ids=approved_word_ids,
        approved_phrase_ids=approved_phrase_ids
    )


@app.route("/recorder/entry/<entry_type>/<int:entry_id>")
def recorder_entry(entry_type, entry_id):
    if not require_recorder():
        return redirect("/recorder")

    entry_type = (entry_type or "").strip().lower()
    if entry_type not in ("word", "phrase"):
        abort(400)

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    if entry_type == "word":
        c.execute(
            "SELECT id, english, oromo FROM words WHERE id=? AND status='approved'",
            (entry_id,)
        )
    else:
        c.execute(
            "SELECT id, english, oromo FROM phrases WHERE id=? AND status='approved'",
            (entry_id,)
        )

    row = c.fetchone()
    conn.close()

    if not row:
        abort(404)

    audio = get_approved_audio(entry_type, entry_id)
    return render_template(
        "recorder_entry.html",
        entry_type=entry_type,
        entry_id=entry_id,
        english=row[1],
        oromo=row[2],
        audio=audio
    )


# ------------------ RECORDER API: GET CURRENT AUDIO + DELETE ------------------

from werkzeug.exceptions import RequestEntityTooLarge


@app.errorhandler(RequestEntityTooLarge)
def handle_413(e):
    # JSON for API endpoints
    if request.path.startswith("/api/") or request.path.startswith("/recorder/api/"):
        return jsonify({"ok": False, "error": "File too large. Please upload a smaller file."}), 413

    # HTML for normal pages
    msg = "File too large. Please upload a smaller file (or increase MAX_UPLOAD_MB)."
    if request.path.startswith("/submit_phrase"):
        return render_template("submit_phrase.html", msg=msg), 413
    if request.path.startswith("/submit_file"):
        return render_template("submit_file.html", msg=msg), 413
    if request.path.startswith("/submit"):
        return render_template("submit.html", msg=msg), 413

    return "File too large.", 413


@app.route("/recorder/api/audio", methods=["GET"])
def recorder_api_audio_get():
    """
    Recorder-only:
    GET /recorder/api/audio?entry_type=word&entry_id=123
    Returns approved audio urls for that entry (dict)
    """
    if not require_recorder():
        return jsonify({"ok": False, "error": "Recorder login required"}), 401

    entry_type = (request.args.get("entry_type") or "").strip().lower()
    entry_id_raw = (request.args.get("entry_id") or "").strip()

    if entry_type not in ("word", "phrase"):
        return jsonify({"ok": False, "error": "Invalid entry_type"}), 400
    if not entry_id_raw.isdigit():
        return jsonify({"ok": False, "error": "Invalid entry_id"}), 400

    entry_id = int(entry_id_raw)
    audio = get_approved_audio(entry_type, entry_id)
    return jsonify({"ok": True, "audio": audio})


@app.route("/recorder/api/delete-audio", methods=["POST"])
def recorder_api_delete_audio():
    """
    Recorder-only:
    POST { entry_type, entry_id, lang='oromo' }
    Deletes APPROVED Oromo audio (and file) for that entry.
    """
    if not require_recorder():
        return jsonify({"ok": False, "error": "Recorder login required"}), 401

    entry_type = (request.form.get("entry_type") or "").strip().lower()
    entry_id_raw = (request.form.get("entry_id") or "").strip()
    lang = (request.form.get("lang") or "oromo").strip().lower()

    if entry_type not in ("word", "phrase"):
        return jsonify({"ok": False, "error": "Invalid entry_type"}), 400
    if not entry_id_raw.isdigit():
        return jsonify({"ok": False, "error": "Invalid entry_id"}), 400
    if lang != "oromo":
        return jsonify({"ok": False, "error": "Only Oromo audio is allowed."}), 400

    entry_id = int(entry_id_raw)

    deleted = delete_audio_for_entry_lang(entry_type, entry_id, lang, statuses=("approved",))
    return jsonify({"ok": True, "deleted": deleted})


@app.route("/recorder/api/submit-audio", methods=["POST"])
def recorder_api_submit_audio():
    if not require_recorder():
        return jsonify({"ok": False, "error": "Recorder login required"}), 401

    entry_type = (request.form.get("entry_type") or "").strip().lower()
    entry_id_raw = (request.form.get("entry_id") or "").strip()
    lang = (request.form.get("lang") or "oromo").strip().lower()

    if entry_type not in ("word", "phrase"):
        return jsonify({"ok": False, "error": "Invalid entry_type"}), 400
    if lang not in ("oromo", "english"):
        return jsonify({"ok": False, "error": "Invalid lang"}), 400
    if not entry_id_raw.isdigit():
        return jsonify({"ok": False, "error": "Invalid entry_id"}), 400

    entry_id = int(entry_id_raw)

    f = request.files.get("audio")
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "Missing audio file"}), 400
    if not allowed_audio(f.filename):
        return jsonify({"ok": False, "error": "Allowed audio: mp3, wav, m4a, webm, ogg"}), 400

    # entry must exist + approved
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    if entry_type == "word":
        c.execute("SELECT id FROM words WHERE id=? AND status='approved'", (entry_id,))
    else:
        c.execute("SELECT id FROM phrases WHERE id=? AND status='approved'", (entry_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "Entry not found or not approved"}), 404

    # âœ… Recorder uploads: replace existing audio (approved + pending) for this entry/lang
    conn.close()
    delete_audio_for_entry_lang(entry_type, entry_id, lang, statuses=("approved", "pending"))

    original = secure_filename(f.filename)
    ext = original.rsplit(".", 1)[1].lower()
    new_name = f"{entry_type}_{entry_id}_{lang}_{uuid4().hex}.{ext}"
    save_path = os.path.join(UPLOAD_FOLDER, new_name)
    f.save(save_path)

    rel_path = f"uploads/{new_name}"

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    # âœ… auto-approve for recorder
    c.execute("""
        INSERT INTO audio (entry_type, entry_id, lang, file_path, status)
        VALUES (?, ?, ?, ?, 'approved')
    """, (entry_type, entry_id, lang, rel_path))
    conn.commit()
    conn.close()

    url = _public_audio_url(rel_path)
    return jsonify({"ok": True, "message": "Saved âœ… Published now.", "url": url})


# ------------------ API AUDIO SUBMISSION (PUBLIC + RECORDER MODE) ------------------

def _handle_audio_submission(is_recorder: bool):
    entry_type = (request.form.get("entry_type") or "").strip().lower()
    entry_id_raw = (request.form.get("entry_id") or "").strip()
    lang = (request.form.get("lang") or "oromo").strip().lower()

    if entry_type not in ("word", "phrase"):
        return jsonify({"ok": False, "error": "Invalid entry_type"}), 400
    if not entry_id_raw.isdigit():
        return jsonify({"ok": False, "error": "Invalid entry_id"}), 400
    if lang != "oromo":
        return jsonify({"ok": False, "error": "Only Oromo audio is allowed."}), 400

    entry_id = int(entry_id_raw)

    f = request.files.get("audio")
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "Missing audio file"}), 400

    original = secure_filename(f.filename)
    if "." not in original:
        return jsonify({"ok": False, "error": "Audio file must have an extension (webm/mp3/wav/m4a/ogg)."}), 400
    if not allowed_audio(original):
        return jsonify({"ok": False, "error": "Allowed audio: mp3, wav, m4a, webm, ogg"}), 400

    ext = original.rsplit(".", 1)[1].lower()

    # âœ… connect with timeout (prevents â€œstuck foreverâ€ on DB lock)
    conn = sqlite3.connect(DB_NAME, timeout=30)
    c = conn.cursor()
    try:
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")

        # Entry must exist + be approved
        if entry_type == "word":
            c.execute("SELECT id FROM words WHERE id=? AND status='approved'", (entry_id,))
        else:
            c.execute("SELECT id FROM phrases WHERE id=? AND status='approved'", (entry_id,))
        row = c.fetchone()
        if not row:
            return jsonify({"ok": False, "error": "Entry not found or not approved"}), 404

        if is_recorder:
            # âœ… IMPORTANT: do deletes BEFORE saving file + inserting
            conn.close()
            delete_audio_for_entry_lang(entry_type, entry_id, lang, statuses=("approved", "pending"))

            conn = sqlite3.connect(DB_NAME, timeout=30)
            c = conn.cursor()
            conn.execute("PRAGMA journal_mode=WAL;")
            conn.execute("PRAGMA synchronous=NORMAL;")
        else:
            # Public users: block if approved exists
            c.execute("""
                SELECT 1 FROM audio
                WHERE entry_type=? AND entry_id=? AND lang=? AND status='approved'
                LIMIT 1
            """, (entry_type, entry_id, lang))
            if c.fetchone():
                return jsonify({"ok": False, "error": "This entry already has approved audio."}), 409

        # Save file
        new_name = f"{entry_type}_{entry_id}_{lang}_{uuid4().hex}.{ext}"
        save_path = os.path.join(UPLOAD_FOLDER, new_name)
        f.save(save_path)

        rel_path = f"uploads/{new_name}"
        status = "approved" if is_recorder else "pending"

        c.execute("""
            INSERT INTO audio (entry_type, entry_id, lang, file_path, status)
            VALUES (?, ?, ?, ?, ?)
        """, (entry_type, entry_id, lang, rel_path, status))

        conn.commit()

        return jsonify({
            "ok": True,
            "message": ("Saved âœ… Published now." if is_recorder else "Oromo audio submitted for admin approval."),
            "status": status,
            "url": _public_audio_url(rel_path)
        })

    except sqlite3.OperationalError as e:
        return jsonify({"ok": False, "error": f"Database error: {str(e)}"}), 500

    finally:
        try:
            conn.close()
        except Exception:
            pass


# ------------------ COMMUNITY AUDIO UPLOAD PAGE (OROMO ONLY) ------------------

@app.route("/upload_audio/<entry_type>/<int:entry_id>/<lang>", methods=["GET", "POST"])
def upload_audio(entry_type, entry_id, lang):
    """
    Manual file upload page (public).
    âœ… Oromo ONLY
    âœ… Allow unlimited pending submissions
    âœ… Block only if an APPROVED Oromo audio already exists for this entry
    """
    entry_type = (entry_type or "").strip().lower()
    lang = (lang or "").strip().lower()

    if entry_type not in ("word", "phrase"):
        return "Invalid entry type", 400

    if lang != "oromo":
        return "Only Oromo audio is allowed.", 400

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    if entry_type == "word":
        c.execute("SELECT id, english, oromo FROM words WHERE id=? AND status='approved'", (entry_id,))
    else:
        c.execute("SELECT id, english, oromo FROM phrases WHERE id=? AND status='approved'", (entry_id,))

    row = c.fetchone()
    if not row:
        conn.close()
        return "Entry not found or not approved.", 404

    c.execute("""
        SELECT 1 FROM audio
        WHERE entry_type=? AND entry_id=? AND lang=? AND status='approved'
        LIMIT 1
    """, (entry_type, entry_id, lang))

    already_approved = c.fetchone() is not None
    if already_approved:
        conn.close()
        return "Audio already approved for this entry.", 409

    if request.method == "POST":
        f = request.files.get("audio")
        if not f or not f.filename:
            conn.close()
            return "Please choose an audio file.", 400

        original = secure_filename(f.filename)
        if "." not in original:
            conn.close()
            return "Audio file must have an extension (webm/mp3/wav/m4a/ogg).", 400

        if not allowed_audio(original):
            conn.close()
            return "Allowed audio: mp3, wav, m4a, webm, ogg", 400

        ext = original.rsplit(".", 1)[1].lower()
        new_name = f"{entry_type}_{entry_id}_{lang}_{uuid4().hex}.{ext}"

        save_path = os.path.join(UPLOAD_FOLDER, new_name)
        f.save(save_path)

        rel_path = f"uploads/{new_name}"

        c.execute("""
            INSERT INTO audio (entry_type, entry_id, lang, file_path, status)
            VALUES (?, ?, ?, ?, 'pending')
        """, (entry_type, entry_id, lang, rel_path))

        conn.commit()
        conn.close()

        return "Thanks! Oromo audio submitted for admin approval."

    conn.close()
    return render_template(
        "upload_audio.html",
        entry_type=entry_type,
        entry_id=entry_id,
        lang=lang,
        english=row[1],
        oromo=row[2]
    )


# ------------------ ADMIN LOGIN ------------------

@app.route("/admin", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("SELECT id, password FROM admin WHERE email=?", (email,))
        admin_row = c.fetchone()
        conn.close()

        if admin_row and check_password_hash(admin_row[1], password):
            session["admin"] = admin_row[0]
            return redirect("/dashboard")

        return "Invalid login"

    return render_template("admin_login.html")


# ------------------ ADMIN MANAGEMENT UNLOCK ------------------

@app.route("/admin/manage/unlock", methods=["GET", "POST"])
def admin_manage_unlock():
    if not require_admin():
        return redirect("/admin")

    msg = None
    real_pw = (os.environ.get("ADMIN_MANAGE_PASSWORD") or "").strip()

    if request.method == "POST":
        entered = (request.form.get("manage_password") or "").strip()

        if not real_pw:
            msg = "ADMIN_MANAGE_PASSWORD is not set on the server."
        elif entered == real_pw:
            session["manage_unlocked"] = True
            session.permanent = False
            return redirect("/admin/manage")
        else:
            msg = "Wrong management password."

    return render_template("admin_manage_unlock.html", msg=msg)


# ------------------ ADMIN DASHBOARD ------------------

@app.route("/dashboard")
def dashboard():
    if not require_admin():
        return redirect("/admin")

    def _latest_audio_job_result(entry_type: str):
        job_type = f"admin_audio_regen_{entry_type}"
        conn2 = None
        try:
            conn2 = sqlite3.connect(DB_NAME)
            c2 = conn2.cursor()
            c2.execute(
                """
                SELECT id, status, result_json, last_error, created_at, finished_at
                FROM post_import_jobs
                WHERE job_type=?
                ORDER BY id DESC
                LIMIT 1
                """,
                (job_type,),
            )
            row = c2.fetchone()
            if not row:
                return {}
            result_payload = {}
            try:
                result_payload = json.loads((row[2] or "{}"))
            except Exception:
                result_payload = {}
            return {
                "job_id": int(row[0] or 0),
                "status": normalize_text(row[1] or ""),
                "summary": result_payload or {},
                "error": normalize_text(row[3] or ""),
                "created_at": normalize_text(row[4] or ""),
                "finished_at": normalize_text(row[5] or ""),
            }
        except Exception:
            app.logger.exception("dashboard_latest_audio_job_failed entry_type=%s", entry_type)
            return {}
        finally:
            if conn2 is not None:
                try:
                    conn2.close()
                except Exception:
                    pass

    def _dashboard_missing_audio_from_latest(entry_type: str, latest_job: dict):
        summary = (latest_job or {}).get("summary", {}) if isinstance(latest_job, dict) else {}
        return {
            "entry_type": entry_type,
            "scanned_entries": int((summary or {}).get("scanned_entries", 0) or 0),
            "missing_entries": int((summary or {}).get("target_entries", 0) or 0),
            "audio_attempted": int((summary or {}).get("audio_attempted", 0) or 0),
            "audio_skipped_existing": int((summary or {}).get("audio_skipped_existing", 0) or 0),
            "audio_missing_voice": int((summary or {}).get("audio_missing_voice", 0) or 0),
        }

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    admin_msg = normalize_text(request.args.get("msg") or "")

    c.execute("SELECT id, english, oromo FROM words WHERE status='pending' ORDER BY id DESC LIMIT 100")
    pending_words = c.fetchall()

    c.execute("SELECT id, english, oromo FROM phrases WHERE status='pending' ORDER BY id DESC LIMIT 100")
    pending_phrases = c.fetchall()

    c.execute("""
        SELECT id, entry_type, entry_id, lang, file_path
        FROM audio
        WHERE status='pending'
        ORDER BY id DESC
        LIMIT 200
    """)
    pending_audio = c.fetchall()

    pending_word_ids = sorted({int(r[2] or 0) for r in (pending_audio or []) if normalize_text(r[1] or "") == "word" and int(r[2] or 0) > 0})
    pending_phrase_ids = sorted({int(r[2] or 0) for r in (pending_audio or []) if normalize_text(r[1] or "") == "phrase" and int(r[2] or 0) > 0})

    words_lookup = {}
    if pending_word_ids:
        marks = ",".join("?" for _ in pending_word_ids)
        c.execute(
            f"SELECT id, english, oromo FROM words WHERE status='approved' AND id IN ({marks})",
            tuple(pending_word_ids),
        )
        words_lookup = {int(row[0]): (row[1], row[2]) for row in c.fetchall()}

    phrases_lookup = {}
    if pending_phrase_ids:
        marks = ",".join("?" for _ in pending_phrase_ids)
        c.execute(
            f"SELECT id, english, oromo FROM phrases WHERE status='approved' AND id IN ({marks})",
            tuple(pending_phrase_ids),
        )
        phrases_lookup = {int(row[0]): (row[1], row[2]) for row in c.fetchall()}

    conn.close()

    latest_phrase_audio_job = _latest_audio_job_result("phrase")
    latest_word_audio_job = _latest_audio_job_result("word")
    phrase_audio_missing = _dashboard_missing_audio_from_latest("phrase", latest_phrase_audio_job)
    word_audio_missing = _dashboard_missing_audio_from_latest("word", latest_word_audio_job)

    return render_template(
        "admin_dashboard.html",
        pending=pending_words,
        pending_phrases=pending_phrases,
        pending_audio=pending_audio,
        words_lookup=words_lookup,
        phrases_lookup=phrases_lookup,
        admin_msg=admin_msg,
        phrase_audio_missing=phrase_audio_missing,
        word_audio_missing=word_audio_missing,
        latest_phrase_audio_job=latest_phrase_audio_job,
        latest_word_audio_job=latest_word_audio_job,
    )


@app.route("/admin/debug/db-counts", methods=["GET"])
def admin_debug_db_counts():
    if not require_admin():
        return redirect("/admin")

    try:
        counts = _words_table_counts()
        payload = {
            "db_path": DB_NAME,
            **counts,
        }
        app.logger.info(f"admin_debug_db_counts db_path={DB_NAME} counts={counts}")
        return jsonify(payload)
    except Exception as e:
        app.logger.exception(f"admin_debug_db_counts failed: {repr(e)}")
        return jsonify({
            "error": "Could not read DB counts.",
            "db_path": DB_NAME
        }), 500


def _queue_admin_audio_regen_response(entry_type: str):
    if not require_admin():
        return (jsonify({"ok": False, "error": "Unauthorized"}), 401) if request.is_json else redirect("/admin")

    payload = request.get_json(silent=True) if request.is_json else {}
    def _norm_any(v) -> str:
        return normalize_text("" if v is None else str(v))

    limit_raw = _norm_any(
        (payload or {}).get("limit")
        if request.is_json
        else (request.form.get("limit") or request.args.get("limit") or "")
    )
    chunk_raw = _norm_any(
        (payload or {}).get("chunk_size")
        if request.is_json
        else (request.form.get("chunk_size") or request.args.get("chunk_size") or "")
    )
    limit = int(limit_raw) if (limit_raw and limit_raw.isdigit()) else 0
    limit = max(0, min(limit, 100000))
    chunk_size = int(chunk_raw) if (chunk_raw and chunk_raw.isdigit()) else IMPORT_BATCH_SIZE
    chunk_size = max(1, min(int(chunk_size or IMPORT_BATCH_SIZE), 1000))

    queued_ok, job_id = trigger_admin_audio_regen_async(entry_type, limit=limit, chunk_size=chunk_size)
    queued_jobs = 1 if queued_ok else 0
    out = {
        "ok": bool(queued_ok),
        "entry_type": normalize_text(entry_type or ""),
        "queued_jobs": int(queued_jobs),
        "job_id": int(job_id or 0),
        "limit": int(limit or 0),
        "chunk_size": int(chunk_size or IMPORT_BATCH_SIZE),
        "scanned_entries": 0,
        "missing_entries": 0,
        "audio_attempted": 0,
        "audio_skipped_existing": 0,
        "audio_missing_voice": 0,
        "audio_generated": 0,
        "audio_failed": 0,
        "db_path": DB_NAME,
        "upload_folder": UPLOAD_FOLDER,
        "queue_job_started": bool(queued_ok),
    }
    app.logger.info("admin_audio_regen_queue_request %s", out)
    if request.is_json:
        return jsonify(out), (200 if queued_ok else 500)
    if queued_ok:
        # Job stats are unknown at queue-time; show only queue confirmation.
        msg = (
            f"{str(entry_type).capitalize()} audio job queued. "
            f"Results will appear after completion. "
            f"(Job #{int(job_id or 0)})"
        )
    else:
        msg = f"Failed to queue missing {entry_type} audio regeneration job."
    return redirect(f"/dashboard?msg={quote(msg, safe='')}")


@app.route("/admin/queue-regenerate-missing-phrase-audio", methods=["POST"])
def admin_queue_regenerate_missing_phrase_audio():
    return _queue_admin_audio_regen_response("phrase")


@app.route("/admin/queue-regenerate-missing-word-audio", methods=["POST"])
def admin_queue_regenerate_missing_word_audio():
    return _queue_admin_audio_regen_response("word")


def _safe_admin_next_url(raw_next: str) -> str:
    nxt = normalize_text(raw_next or "")
    if nxt.startswith("/") and not nxt.startswith("//"):
        return nxt
    return "/admin/manage"


@app.route("/admin/cancel-job/<int:job_id>", methods=["POST"])
def admin_cancel_job(job_id: int):
    if not require_admin():
        return redirect("/admin")

    next_url = _safe_admin_next_url(request.form.get("next") or request.args.get("next") or "/admin/manage")
    jid = int(job_id or 0)
    if jid <= 0:
        return redirect(f"{next_url}?msg={quote('Invalid job id.', safe='')}")

    conn = None
    try:
        conn = sqlite3.connect(DB_NAME, timeout=30)
        c = conn.cursor()
        c.execute("BEGIN IMMEDIATE")
        c.execute(
            """
            SELECT status, cancel_requested, result_json, job_type
            FROM post_import_jobs
            WHERE id=?
            LIMIT 1
            """,
            (jid,),
        )
        row = c.fetchone()
        if not row:
            conn.commit()
            return redirect(f"{next_url}?msg={quote(f'Job #{jid} was not found.', safe='')}")

        status = normalize_text((row or ["", 0, "{}", ""])[0] or "").lower()
        cancel_requested = bool(int((row or ["", 0, "{}", ""])[1] or 0))
        result_payload = {}
        try:
            result_payload = json.loads((row or ["", 0, "{}", ""])[2] or "{}")
        except Exception:
            result_payload = {}
        job_type = normalize_text((row or ["", 0, "{}", ""])[3] or "")

        if status in ("done", "failed", "canceled"):
            conn.commit()
            return redirect(f"{next_url}?msg={quote(f'Job #{jid} is already {status}.', safe='')}")

        result_payload["job_id"] = jid
        result_payload["job_type"] = job_type
        result_payload["canceled"] = True
        result_payload["ok"] = False

        if status == "pending":
            c.execute(
                """
                UPDATE post_import_jobs
                SET status='canceled',
                    cancel_requested=1,
                    finished_at=CURRENT_TIMESTAMP,
                    last_error='cancel_requested_before_start',
                    result_json=?
                WHERE id=?
                """,
                (json.dumps(result_payload, separators=(",", ":")), jid),
            )
            conn.commit()
            return redirect(f"{next_url}?msg={quote(f'Job #{jid} canceled before start.', safe='')}")

        c.execute(
            """
            UPDATE post_import_jobs
            SET cancel_requested=1
            WHERE id=?
            """,
            (jid,),
        )
        conn.commit()
        if cancel_requested:
            msg = f"Stop already requested for Job #{jid}. Waiting for worker to stop."
        else:
            msg = f"Stop requested for Job #{jid}. Worker will stop at next entry checkpoint."
        return redirect(f"{next_url}?msg={quote(msg, safe='')}")
    except Exception:
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
        app.logger.exception("Failed to request cancel for job id=%s", jid)
        return redirect(f"{next_url}?msg={quote(f'Failed to stop Job #{jid}.', safe='')}")
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


# ------------------ ADMIN MANAGEMENT ------------------

def _coerce_int(value, default: int = 0, minimum: int = 0) -> int:
    try:
        out = int((value or default))
    except Exception:
        out = int(default)
    return max(int(minimum), out)


def _coerce_bool(value) -> bool:
    raw = normalize_text(str(value or "")).strip().lower()
    return raw in {"1", "true", "yes", "on"}


def import_missing_phrase_audio_from_source(
    limit: int = 100,
    offset: int = 0,
    entry_id: int = 0,
    dry_run: bool = False,
    cancel_check=None,
):
    summary = {
        "ok": True,
        "source_base_url": AUDIO_SOURCE_BASE_URL,
        "db_path": DB_NAME,
        "upload_folder": UPLOAD_FOLDER,
        "entry_type": "phrase",
        "limit": int(limit or 0),
        "offset": int(offset or 0),
        "entry_id": int(entry_id or 0),
        "dry_run": bool(dry_run),
        "scanned": 0,
        "downloaded": 0,
        "would_download": 0,
        "skipped_existing": 0,
        "skipped_no_file_path": 0,
        "failed": 0,
        "sample_failures": [],
    }
    if not AUDIO_SOURCE_BASE_URL:
        summary["ok"] = False
        summary["error"] = "AUDIO_SOURCE_BASE_URL is not configured."
        return summary

    safe_limit = _coerce_int(limit, default=100, minimum=1)
    safe_limit = min(safe_limit, 2000)
    safe_offset = _coerce_int(offset, default=0, minimum=0)
    safe_entry_id = _coerce_int(entry_id, default=0, minimum=0)

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    sql = """
        SELECT id, entry_id, lang_code, file_path
        FROM generated_tts_audio
        WHERE entry_type='phrase'
    """
    params = []
    if safe_entry_id > 0:
        sql += " AND entry_id=?"
        params.append(safe_entry_id)
    sql += " ORDER BY id ASC LIMIT ? OFFSET ?"
    params.extend([safe_limit, safe_offset])
    c.execute(sql, tuple(params))
    rows = c.fetchall() or []
    conn.close()

    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    app.logger.info(
        "admin_phrase_audio_import start source_base_url=%s upload_folder=%s db_path=%s limit=%s offset=%s entry_id=%s dry_run=%s selected_rows=%s",
        AUDIO_SOURCE_BASE_URL,
        UPLOAD_FOLDER,
        DB_NAME,
        safe_limit,
        safe_offset,
        safe_entry_id,
        bool(dry_run),
        len(rows),
    )

    sess = requests.Session()
    for row_id, phrase_id, lang_code, file_path in rows:
        if callable(cancel_check) and bool(cancel_check()):
            summary["canceled"] = True
            break
        summary["scanned"] += 1
        fp = normalize_text(file_path or "")
        if not fp:
            summary["skipped_no_file_path"] += 1
            continue

        local_ref = _canonical_local_audio_ref(fp)
        name = os.path.basename(local_ref or "")
        if not name:
            summary["skipped_no_file_path"] += 1
            continue

        dst_abs = os.path.join(UPLOAD_FOLDER, name)
        static_abs = os.path.join(STATIC_UPLOADS_FOLDER, name)
        if os.path.isfile(dst_abs) or os.path.isfile(static_abs):
            summary["skipped_existing"] += 1
            continue

        source_url = f"{AUDIO_SOURCE_BASE_URL}/{local_ref.lstrip('/')}"
        if dry_run:
            summary["would_download"] += 1
            continue

        try:
            with sess.get(source_url, timeout=(8, 60), stream=True) as resp:
                if int(resp.status_code or 0) != 200:
                    raise RuntimeError(f"http_status={resp.status_code}")
                content_type = normalize_text(resp.headers.get("Content-Type", "")).lower()
                if (not content_type.startswith("audio/")) and (not name.lower().endswith(".mp3")):
                    raise RuntimeError(f"invalid_content_type={content_type or 'missing'}")

                tmp_abs = dst_abs + ".part"
                written = 0
                try:
                    with open(tmp_abs, "wb") as f:
                        for chunk in resp.iter_content(chunk_size=65536):
                            if not chunk:
                                continue
                            f.write(chunk)
                            written += len(chunk)

                    if written <= 0:
                        raise RuntimeError("empty_download")
                    if os.path.isfile(dst_abs):
                        summary["skipped_existing"] += 1
                        try:
                            os.remove(tmp_abs)
                        except Exception:
                            pass
                        continue
                    os.replace(tmp_abs, dst_abs)
                    summary["downloaded"] += 1
                finally:
                    if os.path.isfile(tmp_abs):
                        try:
                            os.remove(tmp_abs)
                        except Exception:
                            pass
        except Exception as e:
            summary["failed"] += 1
            if len(summary["sample_failures"]) < 10:
                summary["sample_failures"].append(
                    {
                        "row_id": int(row_id or 0),
                        "phrase_id": int(phrase_id or 0),
                        "lang_code": normalize_text(lang_code or ""),
                        "file_path": fp,
                        "source_url": source_url,
                        "error": repr(e),
                    }
                )
            app.logger.warning(
                "admin_phrase_audio_import failed row_id=%s phrase_id=%s lang=%s file_path=%s source_url=%s error=%s",
                row_id,
                phrase_id,
                lang_code,
                fp,
                source_url,
                repr(e),
            )

    app.logger.info(
        "admin_phrase_audio_import done scanned=%s downloaded=%s would_download=%s skipped_existing=%s skipped_no_file_path=%s failed=%s",
        summary["scanned"],
        summary["downloaded"],
        summary["would_download"],
        summary["skipped_existing"],
        summary["skipped_no_file_path"],
        summary["failed"],
    )
    return summary


@app.route("/admin/import-missing-phrase-audio", methods=["POST"])
def admin_import_missing_phrase_audio():
    if not require_admin():
        return jsonify({"ok": False, "error": "Unauthorized"}), 401

    payload = request.get_json(silent=True) if request.is_json else None
    src = payload if isinstance(payload, dict) else request.form
    limit = _coerce_int((src or {}).get("limit"), default=100, minimum=1)
    offset = _coerce_int((src or {}).get("offset"), default=0, minimum=0)
    entry_id = _coerce_int((src or {}).get("entry_id"), default=0, minimum=0)
    dry_run = _coerce_bool((src or {}).get("dry_run"))

    queued_ok, job_id = trigger_admin_import_missing_phrase_audio_async(
        limit=limit,
        offset=offset,
        entry_id=entry_id,
        dry_run=dry_run,
    )
    payload = {
        "ok": bool(queued_ok),
        "queued_jobs": 1 if queued_ok else 0,
        "job_id": int(job_id or 0),
        "entry_type": "phrase",
        "limit": int(limit or 0),
        "offset": int(offset or 0),
        "entry_id": int(entry_id or 0),
        "dry_run": bool(dry_run),
        "queue_job_started": bool(queued_ok),
    }
    if request.is_json:
        return jsonify(payload), (200 if queued_ok else 500)
    if queued_ok:
        msg = f"Phrase audio source import queued (Job #{int(job_id or 0)})."
    else:
        msg = "Failed to queue phrase audio source import."
    return redirect(f"/dashboard?msg={quote(msg, safe='')}")

@app.route("/admin/manage", methods=["GET", "POST"])
def admin_manage():
    if not require_admin():
        return redirect("/admin")

    if not session.get("manage_unlocked"):
        return redirect("/admin/manage/unlock")

    msg = None

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        if action in ("add_admin", "delete_admin"):
            msg = "Admin account changes are disabled here. Use password management only."

        elif action == "update_word":
            wid_raw = (request.form.get("word_id") or "").strip()
            en = normalize_text(request.form.get("english") or "")
            om = normalize_text(request.form.get("oromo") or "")
            if not wid_raw.isdigit():
                msg = "Invalid word id."
            elif not en or not om:
                msg = "Both English and Oromo are required."
            else:
                wid = int(wid_raw)
                conn = sqlite3.connect(DB_NAME)
                c = conn.cursor()

                c.execute("SELECT 1 FROM words WHERE id=? AND status='approved'", (wid,))
                if not c.fetchone():
                    msg = "Word not found (or not approved)."
                else:
                    c.execute("""
                        SELECT 1 FROM words
                        WHERE id != ? AND (english=? OR oromo=?)
                        LIMIT 1
                    """, (wid, en, om))
                    if c.fetchone():
                        msg = "Duplicate conflict: another word already uses that English or Oromo."
                    else:
                        en_key = make_search_key(_strip_edge_punct(en))
                        om_key = make_search_key(_strip_edge_punct(om))
                        c.execute(
                            "UPDATE words SET english=?, oromo=?, english_key=?, oromo_key=? WHERE id=?",
                            (en, om, en_key, om_key, wid)
                        )
                        conn.commit()
                        clear_generated_translations_for_word(wid)
                        msg = "Word updated."
                conn.close()

        elif action == "delete_word":
            wid_raw = (request.form.get("word_id") or "").strip()
            if not wid_raw.isdigit():
                msg = "Invalid word id."
            else:
                wid = int(wid_raw)
                delete_audio_for_entry("word", wid)

                conn = sqlite3.connect(DB_NAME)
                c = conn.cursor()
                c.execute("DELETE FROM words WHERE id=? AND status='approved'", (wid,))
                conn.commit()
                conn.close()
                clear_generated_translations_for_word(wid)
                msg = "Word deleted permanently."

        elif action == "update_phrase":
            pid_raw = (request.form.get("phrase_id") or "").strip()
            en = normalize_text(request.form.get("english") or "")
            om = normalize_text(request.form.get("oromo") or "")
            if not pid_raw.isdigit():
                msg = "Invalid phrase id."
            elif not en or not om:
                msg = "Both English and Oromo are required."
            else:
                pid = int(pid_raw)
                conn = sqlite3.connect(DB_NAME)
                c = conn.cursor()

                c.execute("SELECT 1 FROM phrases WHERE id=? AND status='approved'", (pid,))
                if not c.fetchone():
                    msg = "Phrase not found (or not approved)."
                else:
                    c.execute("""
                        SELECT 1 FROM phrases
                        WHERE id != ? AND (english=? OR oromo=?)
                        LIMIT 1
                    """, (pid, en, om))
                    if c.fetchone():
                        msg = "Duplicate conflict: another phrase already uses that English or Oromo."
                    else:
                        en_key = make_search_key(_strip_edge_punct(en))
                        om_key = make_search_key(_strip_edge_punct(om))
                        c.execute(
                            "UPDATE phrases SET english=?, oromo=?, english_key=?, oromo_key=? WHERE id=?",
                            (en, om, en_key, om_key, pid)
                        )
                        conn.commit()
                        upsert_phrase_aliases(pid, en, om, source="admin_manage")
                        msg = "Phrase updated."
                conn.close()

        elif action == "delete_phrase":
            pid_raw = (request.form.get("phrase_id") or "").strip()
            if not pid_raw.isdigit():
                msg = "Invalid phrase id."
            else:
                pid = int(pid_raw)
                delete_audio_for_entry("phrase", pid)

                conn = sqlite3.connect(DB_NAME)
                c = conn.cursor()
                c.execute("DELETE FROM phrases WHERE id=? AND status='approved'", (pid,))
                conn.commit()
                conn.close()
                msg = "Phrase deleted permanently."

        else:
            msg = "Unknown action."

    word_q = (request.args.get("word_q") or "").strip()
    phrase_q = (request.args.get("phrase_q") or "").strip()

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute(
        """
        SELECT id, job_type, status, created_at, started_at, cancel_requested
        FROM post_import_jobs
        WHERE status='running'
        ORDER BY id DESC
        LIMIT 20
        """
    )
    running_jobs = [
        {
            "id": int((r or [0, "", "", "", "", 0])[0] or 0),
            "job_type": normalize_text((r or [0, "", "", "", "", 0])[1] or ""),
            "status": normalize_text((r or [0, "", "", "", "", 0])[2] or ""),
            "created_at": normalize_text((r or [0, "", "", "", "", 0])[3] or ""),
            "started_at": normalize_text((r or [0, "", "", "", "", 0])[4] or ""),
            "cancel_requested": bool(int((r or [0, "", "", "", "", 0])[5] or 0)),
        }
        for r in (c.fetchall() or [])
    ]

    if word_q:
        q = "%" + normalize_text(word_q) + "%"
        c.execute("""
            SELECT id, english, oromo
            FROM words
            WHERE status='approved' AND (english LIKE ? OR oromo LIKE ?)
            ORDER BY english ASC
            LIMIT 200
        """, (q, q))
    else:
        c.execute("""
            SELECT id, english, oromo
            FROM words
            WHERE status='approved'
            ORDER BY id DESC
            LIMIT 50
        """)
    approved_words = c.fetchall()
    generated_translations_by_word = {}

    try:
        word_ids = [int(w[0]) for w in approved_words if w and w[0]]
        if word_ids:
            word_placeholders = ",".join("?" for _ in word_ids)
            lang_placeholders = ",".join("?" for _ in EXTRA_GENERATED_LANGS)
            c.execute(
                f"""
                SELECT word_id, lang_code, translated_text
                FROM generated_translations
                WHERE word_id IN ({word_placeholders})
                  AND lang_code IN ({lang_placeholders})
                  AND translated_text IS NOT NULL
                  AND TRIM(translated_text) != ''
                """,
                (*word_ids, *EXTRA_GENERATED_LANGS),
            )
            for wid, lang, txt in c.fetchall():
                wid_int = int(wid or 0)
                text_norm = normalize_text(txt or "")
                if not wid_int or not _is_meaningful_generated_text(text_norm):
                    continue
                generated_translations_by_word.setdefault(wid_int, {})[lang] = text_norm
    except Exception as e:
        app.logger.exception(f"admin_manage generated translation read failed: {repr(e)}")
        generated_translations_by_word = {}

    if phrase_q:
        q = "%" + normalize_text(phrase_q) + "%"
        c.execute("""
            SELECT id, english, oromo
            FROM phrases
            WHERE status='approved' AND (english LIKE ? OR oromo LIKE ?)
            ORDER BY id DESC
            LIMIT 200
        """, (q, q))
    else:
        c.execute("""
            SELECT id, english, oromo
            FROM phrases
            WHERE status='approved'
            ORDER BY id DESC
            LIMIT 50
        """)
    approved_phrases = c.fetchall()

    conn.close()

    return render_template(
        "admin_manage.html",
        msg=msg,
        approved_words=approved_words,
        generated_translations_by_word=generated_translations_by_word,
        extra_generated_langs=EXTRA_GENERATED_LANGS,
        language_options=LANGUAGE_OPTIONS,
        approved_phrases=approved_phrases,
        word_q=word_q,
        phrase_q=phrase_q,
        running_jobs=running_jobs,
    )

# ------------------ CHANGE PASSWORD ------------------

@app.route("/admin/change_password", methods=["GET", "POST"])
def admin_change_password():
    if not require_admin():
        return redirect("/admin")

    msg = None
    admin_id = session.get("admin")

    if request.method == "POST":
        current_pw = (request.form.get("current_password") or "").strip()
        new_pw = (request.form.get("new_password") or "").strip()
        new_pw2 = (request.form.get("new_password2") or "").strip()

        if len(new_pw) < 6:
            msg = "New password must be at least 6 characters."
        elif new_pw != new_pw2:
            msg = "New passwords do not match."
        else:
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            c.execute("SELECT password FROM admin WHERE id=?", (admin_id,))
            row = c.fetchone()

            if not row or not check_password_hash(row[0], current_pw):
                msg = "Current password is incorrect."
            else:
                c.execute(
                    "UPDATE admin SET password=? WHERE id=?",
                    (generate_password_hash(new_pw), admin_id)
                )
                conn.commit()
                msg = "Password updated."
            conn.close()

    return render_template("admin_change_password.html", msg=msg)


# ------------------ ADMIN IMPORT (ENGLISH-ONLY -> GOOGLE) ------------------

def _find_word_by_english(conn, english_word: str):
    c = conn.cursor()
    norm = normalize_text(english_word or "")
    key = make_search_key(_strip_edge_punct(norm))
    if key:
        c.execute(
            "SELECT id, english, oromo FROM words WHERE english_key=? OR english=? LIMIT 1",
            (key, norm),
        )
    else:
        c.execute("SELECT id, english, oromo FROM words WHERE english=? LIMIT 1", (norm,))
    return c.fetchone()

def _insert_admin_import_word_english_only(
    conn,
    english_norm: str,
    english_key: str,
    oromo_text: str,
    status: str = "approved",
):
    """
    Admin import insert path: English identity only.
    Returns: (word_id, inserted_new, reason)
    reason in {"inserted", "updated_existing", "existing_english", "invalid"}
    """
    en = english_norm or ""
    en_key = english_key or ""
    om = normalize_text(oromo_text or "")
    om_key = make_search_key(_strip_edge_punct(om)) if om else ""

    # English is identity for import. Oromo can be empty and corrected later.
    if not en or not en_key:
        return None, False, "invalid"

    c = conn.cursor()
    c.execute(
        "SELECT id, oromo FROM words WHERE english_key=? OR english=? LIMIT 1",
        (en_key, en),
    )
    row = c.fetchone()
    if row:
        existing_id = int((row or [0, ""])[0] or 0)
        existing_oromo = normalize_text((row or [0, ""])[1] or "")
        if existing_id and om and _is_missing_value(existing_oromo):
            c.execute(
                """
                UPDATE words
                SET oromo=?, oromo_key=?
                WHERE id=?
                """,
                (om, om_key, existing_id),
            )
            return existing_id, False, "updated_existing"
        return existing_id, False, "existing_english"

    safe_status = status if status in ("pending", "approved") else "approved"
    c.execute(
        "INSERT INTO words (english, oromo, english_key, oromo_key, status) VALUES (?, ?, ?, ?, ?)",
        (en, om, en_key, om_key, safe_status),
    )
    return c.lastrowid, True, "inserted"


def _find_phrase_by_english(conn, english_phrase: str):
    c = conn.cursor()
    norm = normalize_text(english_phrase or "")
    key = make_search_key(_strip_edge_punct(norm))
    if key:
        c.execute(
            "SELECT id, english, oromo FROM phrases WHERE english_key=? OR english=? LIMIT 1",
            (key, norm),
        )
    else:
        c.execute("SELECT id, english, oromo FROM phrases WHERE english=? LIMIT 1", (norm,))
    return c.fetchone()


def _insert_admin_import_phrase_english_only(
    conn,
    english_norm: str,
    english_key: str,
    oromo_text: str,
    status: str = "approved",
):
    """
    Admin import insert path for phrases.
    Returns: (phrase_id, inserted_new, reason)
    reason in {"inserted", "updated_existing", "existing_english", "invalid"}
    """
    en = english_norm or ""
    en_key = english_key or ""
    om = normalize_text(oromo_text or "")
    om_key = make_search_key(_strip_edge_punct(om)) if om else MISSING_OROMO_KEY_SENTINEL

    if not en or not en_key:
        return None, False, "invalid"

    c = conn.cursor()
    c.execute(
        "SELECT id, oromo FROM phrases WHERE english_key=? OR english=? LIMIT 1",
        (en_key, en),
    )
    row = c.fetchone()
    if row:
        existing_id = int((row or [0, ""])[0] or 0)
        existing_oromo = normalize_text((row or [0, ""])[1] or "")
        if existing_id and om and _is_missing_value(existing_oromo):
            c.execute(
                """
                UPDATE phrases
                SET oromo=?, oromo_key=?
                WHERE id=?
                """,
                (om, om_key, existing_id),
            )
            return existing_id, False, "updated_existing"
        return existing_id, False, "existing_english"

    safe_status = status if status in ("pending", "approved") else "approved"
    c.execute(
        "INSERT INTO phrases (english, oromo, english_key, oromo_key, status) VALUES (?, ?, ?, ?, ?)",
        (en, om, en_key, om_key, safe_status),
    )
    return c.lastrowid, True, "inserted"


def _get_word_english_by_id(word_id: int) -> str:
    if not word_id:
        return ""
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute(
            "SELECT english FROM words WHERE id=? AND status='approved' LIMIT 1",
            (word_id,),
        )
        row = c.fetchone()
        conn.close()
        return normalize_text((row or [""])[0] or "")
    except Exception:
        return ""


def _count_missing_generated_langs(conn, word_id: int) -> int:
    if not word_id:
        return len(EXTRA_GENERATED_LANGS)
    try:
        c = conn.cursor()
        placeholders = ",".join("?" for _ in EXTRA_GENERATED_LANGS)
        c.execute(
            f"""
            SELECT lang_code, translated_text
            FROM generated_translations
            WHERE word_id=?
              AND lang_code IN ({placeholders})
              AND translated_text IS NOT NULL
              AND TRIM(translated_text) != ''
            """,
            (word_id, *EXTRA_GENERATED_LANGS),
        )
        have_langs = set()
        for lang_code, translated_text in c.fetchall():
            if _is_meaningful_generated_text(translated_text):
                have_langs.add(lang_code)
        have_count = len(have_langs)
        missing = len(EXTRA_GENERATED_LANGS) - have_count
        return missing if missing > 0 else 0
    except Exception:
        return len(EXTRA_GENERATED_LANGS)


def _count_missing_generated_phrase_langs(conn, phrase_id: int) -> int:
    if not phrase_id:
        return len(EXTRA_GENERATED_LANGS)
    try:
        c = conn.cursor()
        placeholders = ",".join("?" for _ in EXTRA_GENERATED_LANGS)
        c.execute(
            f"""
            SELECT lang_code, translated_text
            FROM generated_phrase_translations
            WHERE phrase_id=?
              AND lang_code IN ({placeholders})
              AND translated_text IS NOT NULL
              AND TRIM(translated_text) != ''
            """,
            (phrase_id, *EXTRA_GENERATED_LANGS),
        )
        have_langs = set()
        for lang_code, translated_text in c.fetchall():
            if _is_meaningful_generated_text(translated_text):
                have_langs.add(lang_code)
        have_count = len(have_langs)
        missing = len(EXTRA_GENERATED_LANGS) - have_count
        return missing if missing > 0 else 0
    except Exception:
        return len(EXTRA_GENERATED_LANGS)


def _merge_generated_stats(acc: dict, stats: dict):
    for lang in EXTRA_GENERATED_LANGS:
        src = stats.get(lang, {}) if isinstance(stats, dict) else {}
        dst = acc.setdefault(lang, {
            "items_seen": 0,
            "already_cached": 0,
            "invalid_cached_treated_missing": 0,
            "missing_before": 0,
            "saved": 0,
            "empty_results": 0,
            "batch_mismatch_fallback": 0,
            "provider_errors": 0,
        })
        for k in (
            "items_seen",
            "already_cached",
            "invalid_cached_treated_missing",
            "missing_before",
            "saved",
            "empty_results",
            "batch_mismatch_fallback",
            "provider_errors",
        ):
            dst[k] = int(dst.get(k, 0) or 0) + int(src.get(k, 0) or 0)
    return acc


@app.route("/admin/repair-generated", methods=["GET", "POST"])
def admin_repair_generated():
    if not require_admin():
        return redirect("/admin")

    default_max_words = 5000
    max_words = default_max_words
    summary = {}
    msg = ""
    has_run = False
    conn = None

    try:
        if request.method == "POST":
            has_run = True
            max_words_raw = normalize_text(request.form.get("max_words") or "")
            if max_words_raw.isdigit():
                max_words = max(1, min(int(max_words_raw), 50000))
            else:
                max_words = default_max_words
            queued_ok, job_id = trigger_admin_repair_generated_async(
                "word",
                max_items=max_words,
                chunk_size=IMPORT_BATCH_SIZE,
            )
            summary = {
                "queued_jobs": 1 if queued_ok else 0,
                "job_id": int(job_id or 0),
                "max_words": int(max_words or default_max_words),
            }
            if queued_ok:
                msg = f"Word language repair queued (Job #{int(job_id or 0)})."
            else:
                msg = "Failed to queue word language repair."
    except Exception as e:
        app.logger.exception(f"admin_repair_generated failed: {repr(e)}")
        msg = "Repair failed safely due to an internal error. Please try again."
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass

    try:
        safe_summary = summary or {}
        safe_stats = (safe_summary.get("stats_by_lang", {}) if isinstance(safe_summary, dict) else {}) or {}
        safe_language_options = LANGUAGE_OPTIONS if isinstance(LANGUAGE_OPTIONS, dict) else {}
        safe_extra_langs = EXTRA_GENERATED_LANGS if EXTRA_GENERATED_LANGS else tuple()
        return render_template(
            "admin_repair_generated.html",
            msg=msg,
            message=(msg or ""),
            summary=safe_summary,
            stats=safe_stats,
            has_run=has_run,
            max_words=max_words,
            extra_generated_langs=safe_extra_langs,
            language_options=safe_language_options,
        )
    except Exception as e:
        app.logger.exception(f"admin_repair_generated render failed: {repr(e)}")
        return (
            "<h3>Repair Missing Languages is temporarily unavailable.</h3>"
            "<p>Please return to dashboard and try again.</p>"
        ), 200


@app.route("/admin/repair-generated-phrases", methods=["GET", "POST"])
def admin_repair_generated_phrases():
    if not require_admin():
        return redirect("/admin")

    default_max_phrases = 5000
    max_phrases = default_max_phrases
    summary = {}
    msg = ""
    has_run = False
    conn = None

    try:
        _log_runtime_repair_context("admin_repair_generated_phrases")
        if request.method == "POST":
            has_run = True
            max_phrases_raw = normalize_text(request.form.get("max_phrases") or "")
            if max_phrases_raw.isdigit():
                max_phrases = max(1, min(int(max_phrases_raw), 50000))
            else:
                max_phrases = default_max_phrases
            queued_ok, job_id = trigger_admin_repair_generated_async(
                "phrase",
                max_items=max_phrases,
                chunk_size=IMPORT_BATCH_SIZE,
            )
            summary = {
                "queued_jobs": 1 if queued_ok else 0,
                "job_id": int(job_id or 0),
                "max_phrases": int(max_phrases or default_max_phrases),
            }
            if queued_ok:
                msg = f"Phrase language repair queued (Job #{int(job_id or 0)})."
            else:
                msg = "Failed to queue phrase language repair."
    except Exception as e:
        app.logger.exception(f"admin_repair_generated_phrases failed: {repr(e)}")
        msg = "Phrase repair failed safely due to an internal error. Please try again."
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass

    try:
        safe_summary = summary or {}
        safe_stats = (safe_summary.get("stats_by_lang", {}) if isinstance(safe_summary, dict) else {}) or {}
        safe_language_options = LANGUAGE_OPTIONS if isinstance(LANGUAGE_OPTIONS, dict) else {}
        safe_extra_langs = EXTRA_GENERATED_LANGS if EXTRA_GENERATED_LANGS else tuple()
        return render_template(
            "admin_repair_generated_phrases.html",
            msg=msg,
            message=(msg or ""),
            summary=safe_summary,
            stats=safe_stats,
            has_run=has_run,
            max_phrases=max_phrases,
            extra_generated_langs=safe_extra_langs,
            language_options=safe_language_options,
        )
    except Exception as e:
        app.logger.exception(f"admin_repair_generated_phrases render failed: {repr(e)}")
        return (
            "<h3>Repair Missing Phrase Languages is temporarily unavailable.</h3>"
            "<p>Please return to dashboard and try again.</p>"
        ), 200


def run_repair_missing_audio(
    limit: int = 0,
    generate_missing: bool = True,
    source_dirs=None,
    cancel_check=None,
):
    _log_runtime_repair_context("run_repair_missing_audio")
    words = _fetch_approved_word_items(limit=limit)
    phrases = _fetch_approved_phrase_items(limit=limit)
    linkage = run_backfill_existing_audio_linkage(
        limit=0,
        dry_run=False,
        source_dirs=source_dirs,
        promote_to_uploads=True,
        cancel_check=cancel_check,
    )
    if callable(cancel_check) and bool(cancel_check()):
        return {
            "words_scanned": len(words),
            "phrases_scanned": len(phrases),
            "items_scanned": int(len(words) + len(phrases)),
            "audio_reused": 0,
            "audio_generated": 0,
            "missing_text": 0,
            "missing_voice_config": 0,
            "failures": 0,
            "processed_items": 0,
            "linkage": linkage,
            "per_language_counts": {},
            "generation_performed": False,
            "canceled": True,
        }
    tts = {
        "processed_items": 0,
        "generated": 0,
        "cached": 0,
        "failed": 0,
        "skipped_missing_text": 0,
        "skipped_missing_voice": 0,
    }
    if generate_missing:
        try:
            tts = run_tts_backfill(
                entry_type="all",
                entry_id=0,
                force_regenerate=False,
                limit=int(limit or 0),
                cancel_check=cancel_check,
            )
            if callable(cancel_check) and bool(cancel_check()):
                tts["canceled"] = True
        except Exception:
            app.logger.exception("run_repair_missing_audio full generation failed")
            tts["failed"] = int(tts.get("failed", 0) or 0) + 1

    per_language = {}
    conn = None
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute(
            """
            SELECT entry_type, lang_code, COUNT(*) AS n
            FROM generated_tts_audio
            GROUP BY entry_type, lang_code
            ORDER BY entry_type, lang_code
            """
        )
        for entry_type, lang_code, n in c.fetchall():
            key = f"{entry_type}:{lang_code}"
            per_language[key] = int(n or 0)
    except Exception:
        app.logger.exception("run_repair_missing_audio failed while counting per-language rows")
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass

    out = {
        "words_scanned": len(words),
        "phrases_scanned": len(phrases),
        "items_scanned": int(len(words) + len(phrases)),
        "audio_reused": int(tts.get("cached", 0) or 0),
        "audio_generated": int(tts.get("generated", 0) or 0),
        "missing_text": int(tts.get("skipped_missing_text", 0) or 0),
        "missing_voice_config": int(tts.get("skipped_missing_voice", 0) or 0),
        "failures": int(tts.get("failed", 0) or 0),
        "processed_items": int(tts.get("processed_items", 0) or 0),
        "linkage": linkage,
        "per_language_counts": per_language,
        "generation_performed": bool(generate_missing),
        "canceled": bool(tts.get("canceled")),
    }
    if int(out.get("failures", 0) or 0) > 0:
        app.logger.error(
            "run_repair_missing_audio failures detected items_scanned=%s processed_items=%s generated=%s cached=%s failures=%s",
            out.get("items_scanned", 0),
            out.get("processed_items", 0),
            out.get("audio_generated", 0),
            out.get("audio_reused", 0),
            out.get("failures", 0),
        )
    app.logger.info("run_repair_missing_audio summary: %s", out)
    return out


@app.route("/admin/repair-missing-audio", methods=["GET", "POST"])
def admin_repair_missing_audio():
    if not require_admin():
        return redirect("/admin")

    default_limit = 5000
    limit = default_limit
    summary = {}
    msg = ""
    has_run = False
    safe_source_dirs = [STATIC_UPLOADS_FOLDER]
    selected_action = "import_only"

    try:
        _log_runtime_repair_context("admin_repair_missing_audio")
        if request.method == "POST":
            has_run = True
            limit_raw = normalize_text(request.form.get("max_items") or "")
            if limit_raw.isdigit():
                limit = max(1, min(int(limit_raw), 50000))
            else:
                limit = default_limit

            selected_action = normalize_text(request.form.get("repair_action") or "import_only")
            queued_ok, job_id = trigger_admin_repair_missing_audio_async(
                selected_action,
                limit=limit,
            )
            summary = {
                "mode": ("full_repair" if selected_action == "full_repair" else "import_only"),
                "queued_jobs": 1 if queued_ok else 0,
                "job_id": int(job_id or 0),
                "max_items": int(limit or default_limit),
            }
            if queued_ok:
                msg = f"Audio repair job queued (mode={summary['mode']}, Job #{int(job_id or 0)})."
            else:
                msg = "Failed to queue audio repair job."
            app.logger.info("admin repair missing audio queued mode=%s data=%s", selected_action, summary)
    except Exception as e:
        app.logger.exception(f"admin_repair_missing_audio failed: {repr(e)}")
        msg = "Audio repair failed safely due to an internal error. Please try again."

    try:
        return render_template(
            "admin_repair_audio.html",
            msg=msg,
            message=(msg or ""),
            summary=(summary or {}),
            has_run=has_run,
            max_items=limit,
            extra_generated_langs=EXTRA_GENERATED_LANGS if EXTRA_GENERATED_LANGS else tuple(),
            language_options=LANGUAGE_OPTIONS if isinstance(LANGUAGE_OPTIONS, dict) else {},
            upload_folder=UPLOAD_FOLDER,
            is_render_disk=IS_RENDER_DISK,
            recommend_generate_cmd=f"flask repair-missing-audio --limit {int(limit or default_limit)}",
            recommend_import_cmd="flask import-existing-audio",
            selected_action=selected_action,
        )
    except Exception as e:
        app.logger.exception(f"admin_repair_missing_audio render failed: {repr(e)}")
        return (
            "<h3>Repair Missing Audio is temporarily unavailable.</h3>"
            "<p>Please return to dashboard and try again.</p>"
        ), 200


@app.route("/admin/import", methods=["GET", "POST"])
def admin_import():
    if not require_admin():
        return redirect("/admin")

    msg = None
    conn = None
    req_started_at_iso = None
    req_started_perf = None

    try:
        if request.method == "POST":
            req_started_at_iso = datetime.utcnow().isoformat() + "Z"
            req_started_perf = time.perf_counter()

            def _perf_log(phase: str, **fields):
                payload = {
                    "phase": phase,
                    "ts": datetime.utcnow().isoformat() + "Z",
                    "elapsed_s": round(time.perf_counter() - req_started_perf, 3) if req_started_perf else None,
                }
                payload.update(fields or {})
                app.logger.info("admin_import perf: %s", payload)
                print(f"admin_import perf: {payload}")

            _perf_log(
                "request_start",
                started_at=req_started_at_iso,
                is_json=bool(request.is_json),
                runtime=APP_RUNTIME,
                db_path=app.DB_NAME,
                upload_folder=UPLOAD_FOLDER,
                persistent_storage=bool(IS_PERSISTENT_STORAGE),
            )
            raw_items = []

            if request.is_json:
                data = request.get_json(silent=True) or {}
                incoming_words = data.get("words", [])
                incoming_phrases = data.get("phrases", [])
                if not isinstance(incoming_words, list):
                    return jsonify({"error": "JSON must include 'words' as a list"}), 400
                if incoming_phrases is not None and not isinstance(incoming_phrases, list):
                    return jsonify({"error": "JSON 'phrases' must be a list"}), 400
                for item in (incoming_words or []):
                    if isinstance(item, dict):
                        en_raw = item.get("english", "") or item.get("text", "") or item.get("word", "")
                        om_raw = item.get("oromo", "")
                    else:
                        en_raw = item
                        om_raw = ""
                    raw_items.append(("word", str(en_raw or ""), str(om_raw or "")))
                for item in (incoming_phrases or []):
                    if isinstance(item, dict):
                        en_raw = item.get("english", "") or item.get("phrase", "") or item.get("text", "")
                        om_raw = item.get("oromo", "")
                    else:
                        en_raw = item
                        om_raw = ""
                    raw_items.append(("phrase", str(en_raw or ""), str(om_raw or "")))
            else:
                if not request.files or len(request.files) == 0:
                    msg = "No upload detected. Please choose a TXT / CSV / XLSX file."
                    return render_template("admin_import.html", msg=msg)

                f = request.files.get("file") or request.files.get("txt_file")
                if not f or not f.filename:
                    msg = "Please upload a TXT / CSV / XLSX file."
                    return render_template("admin_import.html", msg=msg)

                filename = (f.filename or "").lower().strip()
                file_bytes = f.read()
                if file_bytes is None or len(file_bytes) == 0:
                    msg = "Uploaded file is empty or unreadable."
                    return render_template("admin_import.html", msg=msg)

                try:
                    if filename.endswith(".txt"):
                        raw_rows = [("", str(x or ""), "") for x in parse_txt_english_rows(file_bytes)]
                    elif filename.endswith(".csv"):
                        raw_rows = parse_csv_admin_import_rows(file_bytes)
                    elif filename.endswith(".xlsx"):
                        raw_rows = parse_xlsx_admin_import_rows(file_bytes)
                    else:
                        msg = "Only .txt, .csv, .xlsx files are supported."
                        return render_template("admin_import.html", msg=msg)
                except Exception as e:
                    app.logger.exception(f"admin_import parse error: {repr(e)}")
                    msg = "Could not read the file. Please check its format."
                    return render_template("admin_import.html", msg=msg)

                raw_items = [(str(t or ""), str(en or ""), str(om or "")) for t, en, om in (raw_rows or [])]
                _perf_log(
                    "file_parsed",
                    filename=filename,
                    bytes_read=len(file_bytes or b""),
                    parsed_rows=len(raw_items),
                )

            rows_read = len(raw_items)
            app.logger.info(
                "admin_import diagnostics: rows_read=%s first10_raw=%s",
                rows_read,
                [x[1] for x in raw_items[:10]],
            )
            print(f"admin_import: rows_read={rows_read}")
            _perf_log("rows_loaded", rows_read=rows_read)

            empty_rows = 0
            duplicate_rows = 0
            over_limit_rows = 0
            seen_keys = set()
            unique_items = []
            unique_preview = []
            key_to_item_index = {}

            norm_loop_started = time.perf_counter()
            for row_idx, (source_type, raw_en, raw_om) in enumerate(raw_items, start=1):
                raw_orig = str(raw_en or "")
                w = normalize_text(raw_en or "")
                om_in = normalize_text(raw_om or "")
                k = make_search_key(_strip_edge_punct(w))
                if not w or not k:
                    empty_rows += 1
                    continue
                entry_type = source_type if source_type in ("word", "phrase") else ("phrase" if (om_in or len(w.split()) > 1) else "word")
                dedup_key = (entry_type, k)
                if dedup_key in seen_keys:
                    existing_idx = key_to_item_index.get(dedup_key)
                    if (
                        existing_idx is not None
                        and entry_type == "phrase"
                        and om_in
                        and (not normalize_text((unique_items[existing_idx][3] if len(unique_items[existing_idx]) > 3 else "") or ""))
                    ):
                        prev = unique_items[existing_idx]
                        unique_items[existing_idx] = (prev[0], prev[1], prev[2], om_in)
                    duplicate_rows += 1
                    continue
                seen_keys.add(dedup_key)
                key_to_item_index[dedup_key] = len(unique_items)
                unique_items.append((entry_type, w, k, om_in))
                if len(unique_preview) < 10:
                    unique_preview.append((entry_type, raw_orig, w, k, om_in))
                if row_idx % 100 == 0:
                    elapsed = max(0.001, (time.perf_counter() - norm_loop_started))
                    _perf_log(
                        "normalize_dedup_progress",
                        rows_processed=row_idx,
                        rows_per_sec=round(row_idx / elapsed, 2),
                        unique_so_far=len(unique_items),
                        duplicate_rows=duplicate_rows,
                        empty_rows=empty_rows,
                    )

            app.logger.info(
                "admin_import diagnostics: unique_count=%s first10_unique_norm=%s",
                len(unique_items),
                unique_preview,
            )
            _perf_log(
                "normalize_dedup_done",
                unique_count=len(unique_items),
                duplicate_rows=duplicate_rows,
                empty_rows=empty_rows,
            )

            if not unique_items:
                summary = (
                    f"No valid rows to import. Empty rows: {empty_rows} | "
                    f"Duplicates: {duplicate_rows}."
                )
                if request.is_json:
                    return jsonify({
                        "rows_read": rows_read,
                        "words_inserted": 0,
                        "phrases_inserted": 0,
                        "skipped": 0,
                        "errors": 0,
                        "imported_with_oromo": 0,
                        "imported_missing_oromo": 0,
                        "phrase_translations_saved": 0,
                        "phrase_audio_generated_by_language": {lang: 0 for lang in LEARN_TTS_LANGS},
                        "empty_rows": empty_rows,
                        "duplicate_rows": duplicate_rows,
                        "message": summary,
                    }), 400
                msg = summary
                return render_template("admin_import.html", msg=msg)

            inserted_words = 0
            inserted_phrases = 0
            updated_words = 0
            updated_phrases = 0
            skipped_existing_precheck = 0
            skipped_existing_during_insert = 0
            failed = 0
            failed_base_inserts = 0
            imported_with_oromo_words = 0
            imported_with_oromo_phrases = 0
            imported_missing_oromo_words = 0
            imported_missing_oromo_phrases = 0
            affected_word_items = []
            affected_phrase_items = []

            conn = sqlite3.connect(app.DB_NAME)
            app.logger.info(
                "admin_import diagnostics: db_path=%s abs_db_path=%s",
                app.DB_NAME,
                os.path.abspath(app.DB_NAME),
            )
            _perf_log("db_opened", db_path=app.DB_NAME)
            # Keep request path fast: do one base upsert pass and move enrichment to background worker.
            items_for_base_insert = list(unique_items[:IMPORT_MAX_WORDS])
            over_limit_rows = max(0, len(unique_items) - len(items_for_base_insert))
            app.logger.info(
                "admin_import diagnostics: pre_translation_summary valid_unique=%s attempted_new=%s imported_so_far=%s skipped_precheck=%s failed_so_far=%s over_limit=%s empty=%s dup_rows=%s",
                len(unique_items),
                len(items_for_base_insert),
                (inserted_words + inserted_phrases),
                skipped_existing_precheck,
                failed,
                over_limit_rows,
                empty_rows,
                duplicate_rows,
            )

            insert_started = time.perf_counter()
            insert_rows_processed = 0
            for entry_type, en, en_key, om_in in items_for_base_insert:
                try:
                    om_text = normalize_text(om_in or "")
                    if entry_type == "word":
                        if om_text:
                            imported_with_oromo_words += 1
                        else:
                            imported_missing_oromo_words += 1
                    else:
                        if om_text:
                            imported_with_oromo_phrases += 1
                        else:
                            imported_missing_oromo_phrases += 1

                    if entry_type == "word":
                        eid, was_inserted, insert_reason = _insert_admin_import_word_english_only(
                            conn,
                            english_norm=en,
                            english_key=en_key,
                            oromo_text=om_text,
                            status="approved",
                        )
                    else:
                        eid, was_inserted, insert_reason = _insert_admin_import_phrase_english_only(
                            conn,
                            english_norm=en,
                            english_key=en_key,
                            oromo_text=om_text,
                            status="approved",
                        )

                    if not eid:
                        failed += 1
                        failed_base_inserts += 1
                        continue

                    if was_inserted:
                        if entry_type == "word":
                            inserted_words += 1
                            affected_word_items.append((int(eid), en))
                        else:
                            inserted_phrases += 1
                            affected_phrase_items.append((int(eid), en))
                    elif insert_reason == "updated_existing":
                        if entry_type == "word":
                            updated_words += 1
                            affected_word_items.append((int(eid), en))
                        else:
                            updated_phrases += 1
                            affected_phrase_items.append((int(eid), en))
                    elif insert_reason == "existing_english":
                        skipped_existing_during_insert += 1
                        if entry_type == "word":
                            affected_word_items.append((int(eid), en))
                        else:
                            affected_phrase_items.append((int(eid), en))
                    else:
                        failed += 1
                        failed_base_inserts += 1
                        continue

                    insert_rows_processed += 1
                    if insert_rows_processed % 100 == 0:
                        elapsed = max(0.001, (time.perf_counter() - insert_started))
                        _perf_log(
                            "insert_progress",
                            rows_processed=insert_rows_processed,
                            rows_per_sec=round(insert_rows_processed / elapsed, 2),
                            inserted_words=inserted_words,
                            inserted_phrases=inserted_phrases,
                            failed=failed,
                        )
                except Exception as row_err:
                    app.logger.exception(f"admin_import row insert failed type={entry_type} en={repr(en)}: {repr(row_err)}")
                    failed += 1
                    failed_base_inserts += 1
                    continue

            _perf_log(
                "insert_done",
                rows_processed=insert_rows_processed,
                inserted_words=inserted_words,
                inserted_phrases=inserted_phrases,
                failed=failed,
            )

            commit_t0 = time.perf_counter()
            conn.commit()
            _perf_log("db_commit_done", commit_elapsed_s=round(time.perf_counter() - commit_t0, 3))
            conn.close()
            conn = None

            affected_words = list(affected_word_items or [])
            affected_phrases = list(affected_phrase_items or [])
            word_ids = [int(w[0]) for w in affected_words if w and int(w[0] or 0) > 0]
            phrase_ids = [int(p[0]) for p in affected_phrases if p and int(p[0] or 0) > 0]
            import_stage_summary = {
                "word": {
                    "inserted": int(inserted_words),
                    "updated": int(updated_words),
                    "imported_with_oromo": int(imported_with_oromo_words),
                    "imported_missing_oromo": int(imported_missing_oromo_words),
                },
                "phrase": {
                    "inserted": int(inserted_phrases),
                    "updated": int(updated_phrases),
                    "imported_with_oromo": int(imported_with_oromo_phrases),
                    "imported_missing_oromo": int(imported_missing_oromo_phrases),
                },
            }
            pipeline_queued = trigger_post_import_pipeline_async(
                word_ids,
                phrase_ids,
                chunk_size=IMPORT_BATCH_SIZE,
                import_summary=import_stage_summary,
            )
            _perf_log(
                "post_import_pipeline_queued",
                queued=bool(pipeline_queued),
                affected_word_items=len(affected_word_items),
                affected_phrase_items=len(affected_phrase_items),
            )

            skipped_existing_total = skipped_existing_precheck + skipped_existing_during_insert
            rows_inserted_total = inserted_words + inserted_phrases
            skipped_total = skipped_existing_total + duplicate_rows + empty_rows + over_limit_rows
            errors_total = failed

            app.logger.info(
                "admin_import summary: rows_read=%s words_inserted=%s phrases_inserted=%s skipped=%s errors=%s db_path=%s",
                rows_read, inserted_words, inserted_phrases, skipped_total, errors_total, app.DB_NAME
            )
            app.logger.info(
                "admin_import phrase_summary imported_with_oromo=%s imported_missing_oromo=%s phrase_translations_saved=%s phrase_audio_generated_by_language=%s unresolved_missing_oromo_ids=%s",
                imported_with_oromo_phrases,
                imported_missing_oromo_phrases,
                0,
                {lang: 0 for lang in LEARN_TTS_LANGS},
                [],
            )
            print(f"admin_import: rows_inserted_total={rows_inserted_total}")
            print(f"admin_import: errors_total={errors_total}")
            app.logger.info("Import completed. Post-import pipeline queued=%s", bool(pipeline_queued))
            print(f"Import completed. Post-import pipeline queued={bool(pipeline_queued)}")
            _perf_log(
                "request_complete",
                rows_read=rows_read,
                rows_inserted_total=rows_inserted_total,
                words_inserted=inserted_words,
                phrases_inserted=inserted_phrases,
                skipped=skipped_total,
                errors=errors_total,
                external_calls={
                    "google_calls_used": 0,
                    "translation_warmup_words": 0,
                    "translation_warmup_phrases": 0,
                    "tts_enabled": False,
                },
            )

            msg = (
                f"Import done. Rows read: {rows_read} | "
                f"Words inserted: {inserted_words} | "
                f"Words updated: {updated_words} | "
                f"Phrases inserted: {inserted_phrases} | "
                f"Phrases updated: {updated_phrases} | "
                f"Skipped: {skipped_total} | "
                f"Errors: {errors_total}."
            )

            if request.is_json:
                return jsonify({
                    "rows_read": rows_read,
                    "words_inserted": inserted_words,
                    "phrases_inserted": inserted_phrases,
                    "rows_inserted_total": rows_inserted_total,
                    "skipped": skipped_total,
                    "skipped_existing_precheck": skipped_existing_precheck,
                    "skipped_existing_during_insert": skipped_existing_during_insert,
                    "skipped_existing_total": skipped_existing_total,
                    "failed": failed,
                    "failed_base_inserts": failed_base_inserts,
                    "word_updated": updated_words,
                    "phrase_updated": updated_phrases,
                    "imported_with_oromo_words": imported_with_oromo_words,
                    "imported_missing_oromo_words": imported_missing_oromo_words,
                    "imported_with_oromo_phrases": imported_with_oromo_phrases,
                    "imported_missing_oromo_phrases": imported_missing_oromo_phrases,
                    "empty_rows": empty_rows,
                    "duplicate_rows": duplicate_rows,
                    "ignored_due_limit": over_limit_rows,
                    "phrase_translations_saved": 0,
                    "phrase_audio_generated_by_language": {lang: 0 for lang in LEARN_TTS_LANGS},
                    "unresolved_phrase_oromo_ids": [],
                    "updated_missing_word_translations": 0,
                    "updated_missing_phrase_translations": 0,
                    "cached_generated_word_translations": 0,
                    "cached_generated_phrase_translations": 0,
                    "generated_word_by_lang": {},
                    "generated_phrase_by_lang": {},
                    "tts_summary": {},
                    "post_import_pipeline_queued": bool(pipeline_queued),
                    "post_import_pipeline_import_summary": import_stage_summary,
                    "google_calls_used": 0,
                    "batch_size": IMPORT_BATCH_SIZE,
                    "max_items": IMPORT_MAX_WORDS,
                    "db_path": app.DB_NAME,
                    "message": msg
                })

    except Exception as e:
        app.logger.exception(f"admin_import failed: {repr(e)}")
        safe_msg = "Import failed safely due to an internal error. Please check file format and try again."
        if request.method == "POST" and request.is_json:
            return jsonify({"error": safe_msg}), 500
        msg = safe_msg
    finally:
        if req_started_perf:
            ended_iso = datetime.utcnow().isoformat() + "Z"
            duration_s = (time.perf_counter() - req_started_perf)
            app.logger.info(
                "admin_import request_end: started_at=%s ended_at=%s duration_s=%.3f",
                req_started_at_iso,
                ended_iso,
                duration_s,
            )
            print(
                "admin_import request_end: "
                f"started_at={req_started_at_iso} "
                f"ended_at={ended_iso} "
                f"duration_s={round(duration_s, 3)}"
            )
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass

    try:
        return render_template("admin_import.html", msg=msg)
    except Exception as e:
        app.logger.exception(f"admin_import render failed: {repr(e)}")
        # Final guardrail: GET/POST should fail safely without a hard 500.
        return (
            "<h3>Admin import is temporarily unavailable.</h3>"
            "<p>Please return to dashboard and try again.</p>"
        ), 200


# ------------------ APPROVE / REJECT WORDS ------------------

@app.route("/approve/<int:word_id>")
def approve(word_id):
    if not require_admin():
        return redirect("/admin")

    wid = int(word_id or 0)
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("UPDATE words SET status='approved' WHERE id=?", (wid,))
    c.execute("SELECT english FROM words WHERE id=? LIMIT 1", (wid,))
    row = c.fetchone()
    conn.commit()
    conn.close()

    en = normalize_text((row or [""])[0] or "")
    if wid and en:
        try:
            trigger_post_import_pipeline_async([wid], [], chunk_size=1)
        except Exception as e:
            app.logger.exception(f"approve word queue failed word_id={wid}: {repr(e)}")
    return redirect("/dashboard")


@app.route("/reject/<int:word_id>")
def reject(word_id):
    if not require_admin():
        return redirect("/admin")

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("DELETE FROM words WHERE id=? AND status='pending'", (word_id,))
    conn.commit()
    conn.close()
    return redirect("/dashboard")


# ------------------ APPROVE / REJECT PHRASES ------------------

@app.route("/approve_phrase/<int:phrase_id>")
def approve_phrase(phrase_id):
    if not require_admin():
        return redirect("/admin")

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("UPDATE phrases SET status='approved' WHERE id=?", (phrase_id,))
    conn.commit()
    conn.close()
    try:
        pid = int(phrase_id or 0)
        if pid > 0:
            trigger_post_import_pipeline_async([], [pid], chunk_size=1)
    except Exception as e:
        app.logger.exception(f"approve phrase queue failed phrase_id={phrase_id}: {repr(e)}")
    return redirect("/dashboard")


@app.route("/reject_phrase/<int:phrase_id>")
def reject_phrase(phrase_id):
    if not require_admin():
        return redirect("/admin")

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("DELETE FROM phrases WHERE id=? AND status='pending'", (phrase_id,))
    conn.commit()
    conn.close()
    return redirect("/dashboard")


# ------------------ APPROVE / REJECT AUDIO ------------------

@app.route("/approve_audio/<int:audio_id>")
def approve_audio(audio_id):
    if not require_admin():
        return redirect("/admin")

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("UPDATE audio SET status='approved' WHERE id=?", (audio_id,))
    conn.commit()
    conn.close()
    return redirect("/dashboard")


@app.route("/reject_audio/<int:audio_id>")
def reject_audio(audio_id):
    if not require_admin():
        return redirect("/admin")

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT file_path FROM audio WHERE id=? AND status='pending'", (audio_id,))
    row = c.fetchone()
    c.execute("DELETE FROM audio WHERE id=? AND status='pending'", (audio_id,))
    conn.commit()
    conn.close()

    if row and row[0]:
        abs_path = _audio_abs_path(row[0])
        if abs_path and os.path.isfile(abs_path):
            try:
                os.remove(abs_path)
            except Exception:
                app.logger.exception(f"Could not delete pending audio file: {abs_path}")

    return redirect("/dashboard")

# ------------------ LOGOUT ------------------

@app.route("/logout")
def logout():
    session.pop("admin", None)
    return redirect("/")


# ------------------ CREATE FIRST ADMIN (RUN ONCE) ------------------

@app.route("/create_admin")
def create_admin():
    if os.environ.get("ENABLE_CREATE_ADMIN") != "1":
        return "Disabled."

    email = "jewargure1@gmail.com"
    password = generate_password_hash("admin123")

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    c.execute("SELECT 1 FROM admin WHERE email=?", (email,))
    if not c.fetchone():
        c.execute("INSERT INTO admin (email, password) VALUES (?, ?)", (email, password))
        conn.commit()

    conn.close()
    return "Admin created (or already exists). You can now login."


# ------------------ GADAA AI (FREE DEMO - NO PAID API) ------------------
# URL: /gadaa-ai
# - bilingual Oromo + English
# - rule-based tutor using your own SQLite dictionary
# - promo-friendly, no card required
# - later: can swap backend to paid AI without changing UI

from collections import defaultdict
import time

# Simple in-memory rate limit (per IP). Resets on restart (OK for demo).
_AI_LIMIT_WINDOW_SEC = 60
_AI_LIMIT_MAX_REQ = 20
_ai_hits = defaultdict(list)  # ip -> [timestamps]


def _client_ip():
    # Works behind proxies with ProxyFix
    return (
        request.headers.get("X-Forwarded-For", "").split(",")[0].strip()
        or request.remote_addr
        or "unknown"
    )


def _rate_limit_ok():
    ip = _client_ip()
    now = time.time()
    hits = _ai_hits[ip]
    # keep only last window
    hits = [t for t in hits if now - t < _AI_LIMIT_WINDOW_SEC]
    if len(hits) >= _AI_LIMIT_MAX_REQ:
        _ai_hits[ip] = hits
        return False, ip, len(hits)
    hits.append(now)
    _ai_hits[ip] = hits
    return True, ip, len(hits)


def _db_lookup_word_or_phrase(q: str):
    """
    Try exact match in phrases then words (approved only).
    Returns dict: {"type": "phrase"/"word", "id": int, "english": str, "oromo": str} or None
    """
    clean = normalize_text(q)
    if not clean:
        return None

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    # phrases exact match
    c.execute(
        "SELECT id, english, oromo FROM phrases "
        "WHERE status='approved' AND (english=? OR oromo=?) LIMIT 1",
        (clean, clean),
    )
    r = c.fetchone()
    if r:
        conn.close()
        return {"type": "phrase", "id": r[0], "english": r[1], "oromo": r[2]}

    # word exact match (single token)
    if len(clean.split()) == 1:
        c.execute(
            "SELECT id, english, oromo FROM words "
            "WHERE status='approved' AND (english=? OR oromo=?) LIMIT 1",
            (clean, clean),
        )
        r = c.fetchone()
        if r:
            conn.close()
            return {"type": "word", "id": r[0], "english": r[1], "oromo": r[2]}

    conn.close()
    return None


def _db_suggest(clean: str, limit=6):
    """
    Suggestions from both directions.
    """
    s_en = suggest_terms(clean, "en_om", limit=limit)
    s_om = suggest_terms(clean, "om_en", limit=limit)

    # flatten unique
    items = []
    for k in ("closest", "prefix", "partial"):
        items += s_en.get(k, [])
        items += s_om.get(k, [])

    out = []
    seen = set()
    for x in items:
        if x and x not in seen:
            seen.add(x)
            out.append(x)
        if len(out) >= limit:
            break
    return out


def _make_lesson_card(entry: dict):
    """
    Build a small â€œteacher styleâ€ response.
    """
    en = entry["english"]
    om = entry["oromo"]

    # very lightweight â€œlessonâ€
    if entry["type"] == "word":
        examples = [
            f"Example (EN): I use **{en}** in a sentence.",
            f"Example (OM): Ani jecha **{om}** keessatti fayyadama.",
        ]
    else:
        examples = [
            f"Example (EN): **{en}**",
            f"Example (OM): **{om}**",
        ]

    quiz = [
        f"Quick quiz: What is the Oromo for **{en}**?",
        f"Answer: **{om}**",
    ]

    return {
        "title": f"ðŸ“˜ {entry['type'].capitalize()} lesson",
        "english": en,
        "oromo": om,
        "examples": examples,
        "quiz": quiz,
        "audio": get_approved_audio(entry["type"], entry["id"]),  # shows Oromo if exists
    }


@app.route("/gadaa-ai", methods=["GET"])
def gadaa_ai_page():
    trending = get_trending(limit=12)
    return render_template("gadaa_ai.html", trending=trending)


@app.route("/api/gadaa-ai", methods=["POST"])
def gadaa_ai_api():
    ok, ip, count = _rate_limit_ok()
    if not ok:
        return jsonify({
            "ok": False,
            "error": "Too many requests. Please wait 1 minute and try again."
        }), 429

    data = request.get_json(silent=True) or {}
    msg = (data.get("message") or "").strip()
    if not msg:
        return jsonify({"ok": False, "error": "Empty message"}), 400

    clean = normalize_text(msg)

    # basic commands
    if clean in ("help", "what can you do", "menu"):
        return jsonify({
            "ok": True,
            "reply": {
                "type": "text",
                "text": (
                    "Hi! Iâ€™m **Gadaa AI (free demo)**.\n\n"
                    "Try:\n"
                    "- Type a word/phrase in Oromo or English (exact match)\n"
                    "- Ask: `quiz me` or `lesson`\n"
                    "- Ask: `suggest <word>`\n\n"
                    "Note: This demo uses your dictionary database (no paid AI yet)."
                )
            }
        })

    if clean.startswith("suggest "):
        term = clean.replace("suggest ", "", 1).strip()
        sug = _db_suggest(term, limit=8) if term else []
        if not sug:
            text = "No suggestions found."
        else:
            text = "Suggestions: " + ", ".join(sug)
        return jsonify({"ok": True, "reply": {"type": "text", "text": text}})

    # "quiz me" -> use a trending query if exists, else generic
    if clean in ("quiz me", "quiz", "test me"):
        trending = get_trending(limit=1)
        pick = trending[0][0] if trending else "hello"
        entry = _db_lookup_word_or_phrase(pick) or _db_lookup_word_or_phrase("hello")
        if entry:
            card = _make_lesson_card(entry)
            return jsonify({"ok": True, "reply": {"type": "card", "card": card}})
        return jsonify({
            "ok": True,
            "reply": {"type": "text", "text": "I need more approved words/phrases to quiz you."}
        })

    # Normal: try dictionary lookup
    entry = _db_lookup_word_or_phrase(msg)
    if entry:
        card = _make_lesson_card(entry)
        return jsonify({"ok": True, "reply": {"type": "card", "card": card}})

    # Not found: suggestions
    sug = _db_suggest(clean, limit=8) if clean else []
    if sug:
        return jsonify({
            "ok": True,
            "reply": {
                "type": "text",
                "text": "I couldn't find an exact match. Try one of these: " + ", ".join(sug)
            }
        })

    return jsonify({
        "ok": True,
        "reply": {
            "type": "text",
            "text": (
                "I couldn't find that in the dictionary yet.\n"
                "Tip: try a simpler word, or submit it via **Submit Word / Submit Phrase**."
            )
        }
    })


@app.cli.command("backfill-translations")
@click.option("--entry-type", type=click.Choice(["all", "word", "phrase"]), default="all", show_default=True)
@click.option("--entry-id", type=int, default=0, show_default=True)
@click.option("--overwrite-existing", is_flag=True, default=False, help="Overwrite existing generated translations.")
@click.option("--limit", type=int, default=0, show_default=True, help="Limit approved records per entry type.")
def cli_backfill_translations(entry_type, entry_id, overwrite_existing, limit):
    """
    Backfill generated translations for approved words/phrases.
    Uses English as pivot and writes to DB cache tables.
    """
    summary = run_translation_backfill(
        entry_type=entry_type,
        entry_id=int(entry_id or 0),
        overwrite_existing=bool(overwrite_existing),
        limit=int(limit or 0),
    )
    click.echo("Translation backfill completed.")
    click.echo(
        f"words_saved={summary.get('words_saved', 0)} "
        f"phrases_saved={summary.get('phrases_saved', 0)}"
    )


@app.cli.command("audit-live-word-audio")
@click.option("--word", "word_text", required=True, help="English base word to audit, e.g. about")
@click.option("--base-url", default="https://gadaadictionary.com", show_default=True, help="Live site base URL.")
@click.option("--source-lang", default="en", show_default=True)
@click.option("--target-lang", default="om", show_default=True)
def cli_audit_live_word_audio(word_text, base_url, source_lang, target_lang):
    """
    Proof-based audit for one word:
    1) local DB audio rows
    2) rendered live HTML data-audio URLs
    3) HTTP status for each rendered audio URL
    """
    base = (base_url or "").strip().rstrip("/")
    w = normalize_text(word_text or "")
    if not base or (not w):
        raise click.UsageError("--word and --base-url are required.")

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute(
        """
        SELECT w.id, w.english, w.oromo, a.file_path, a.status
        FROM words w
        LEFT JOIN audio a
          ON a.entry_type='word' AND a.entry_id=w.id AND a.lang='oromo' AND a.status='approved'
        WHERE w.status='approved' AND lower(trim(w.english))=lower(?)
        ORDER BY a.id DESC
        LIMIT 1
        """,
        (w,),
    )
    row = c.fetchone()
    conn.close()
    if not row:
        click.echo(f"NO_LOCAL_DB_MATCH word={w}")
        return

    wid, en, om, db_audio_path, db_audio_status = row
    click.echo(
        f"LOCAL_DB word_id={int(wid)} english={normalize_text(en or '')} "
        f"oromo={normalize_text(om or '')} file_path={db_audio_path or ''} status={db_audio_status or ''}"
    )

    req = requests.Session()
    req.trust_env = False
    page_url = (
        f"{base}/dictionary?q={quote(normalize_text(en or ''), safe='')}"
        f"&source_lang={quote(source_lang or 'en', safe='')}"
        f"&target_lang={quote(target_lang or 'om', safe='')}"
    )
    try:
        page = req.get(page_url, timeout=25)
    except Exception as e:
        click.echo(f"LIVE_PAGE_ERROR url={page_url} error={repr(e)}")
        return

    click.echo(f"LIVE_PAGE status={page.status_code} url={page.url}")
    html = page.text or ""
    audio_urls = re.findall(r'data-audio=\"([^\"]+)\"', html)
    click.echo(
        f"LIVE_RENDER data_audio_count={len(audio_urls)} "
        f"contains_uploads={'/uploads/' in html} "
        f"contains_static_uploads={'/static/uploads/' in html} "
        f"contains_blob={'blob.core.windows.net' in html}"
    )

    if not audio_urls:
        return

    for idx, audio_ref in enumerate(audio_urls, start=1):
        full = audio_ref if audio_ref.startswith("http") else f"{base}{audio_ref}"
        try:
            rr = req.get(full, timeout=25, allow_redirects=True)
            click.echo(
                f"LIVE_AUDIO_{idx} ref={audio_ref} final={rr.url} status={rr.status_code} "
                f"content_type={(rr.headers.get('content-type') or '').strip()}"
            )
        except Exception as e:
            click.echo(f"LIVE_AUDIO_{idx} ref={audio_ref} error={repr(e)}")


@app.cli.command("audit-live-learn-audio")
@click.option("--word", "word_text", default="", help="Backward-compatible alias for --text (word mode).")
@click.option("--text", "target_text", default="", help="Target word/phrase text expected on /learn.")
@click.option(
    "--entry-type",
    type=click.Choice(["word", "phrase"]),
    default="word",
    show_default=True,
    help="Expected Learn row type.",
)
@click.option("--base-url", default="https://gadaadictionary.com", show_default=True, help="Live site base URL.")
def cli_audit_live_learn_audio(word_text, target_text, entry_type, base_url):
    """
    Verify /learn rendered HTML contains a row for the target word/phrase and data-audio URLs.
    Supports words and phrases without triggering generation/repair.
    Does not trigger generation/repair.
    """
    base = (base_url or "").strip().rstrip("/")
    query_text = normalize_text(target_text or word_text or "")
    if not base or (not query_text):
        raise click.UsageError("--text (or --word) and --base-url are required.")

    sess = requests.Session()
    sess.trust_env = False
    page_url = f"{base}/learn"
    try:
        page = sess.get(page_url, timeout=25)
    except Exception as e:
        click.echo(f"LIVE_LEARN_ERROR url={page_url} error={repr(e)}")
        return

    html = page.text or ""
    click.echo(f"LIVE_LEARN status={page.status_code} url={page.url}")
    click.echo(
        f"LIVE_LEARN_RENDER contains_data_audio={'data-audio=' in html} "
        f"contains_uploads_tts={'/uploads/tts_' in html} "
        f"contains_text={query_text.lower() in html.lower()}"
    )

    # Narrow to the first table row containing the target word/phrase, then capture data-audio refs.
    row_re = re.compile(r"<tr[\s\S]*?</tr>", re.IGNORECASE)
    target_row = ""
    type_marker = f'data-entry-type="{entry_type}"'
    for m in row_re.finditer(html):
        row_html = m.group(0) or ""
        if query_text.lower() not in row_html.lower():
            continue
        if type_marker and (type_marker not in row_html):
            continue
        target_row = row_html
        break

    if not target_row:
        click.echo(f"LIVE_LEARN_ROW missing text={query_text} entry_type={entry_type}")
        return

    row_type_match = re.search(r'data-entry-type="([^"]+)"', target_row, re.IGNORECASE)
    row_id_match = re.search(r'data-entry-id="([^"]+)"', target_row, re.IGNORECASE)
    row_type = normalize_text((row_type_match.group(1) if row_type_match else "") or "")
    row_id = normalize_text((row_id_match.group(1) if row_id_match else "") or "")

    refs = re.findall(r'data-audio=\"([^\"]+)\"', target_row)
    click.echo(
        f"LIVE_LEARN_ROW_META text={query_text} entry_type={row_type or entry_type} entry_id={row_id or 'unknown'}"
    )
    click.echo(f"LIVE_LEARN_ROW text={query_text} entry_type={entry_type} data_audio_count={len(refs)}")
    if refs:
        click.echo("LIVE_LEARN_AUDIO_URLS " + " | ".join(refs))
    if not refs:
        return

    uploads_urls = 0
    uploads_200 = 0
    for idx, ref in enumerate(refs, start=1):
        full = ref if ref.startswith("http") else f"{base}{ref}"
        try:
            rr = sess.get(full, timeout=25, allow_redirects=True)
            is_upload_ref = ref.startswith("/uploads/") or ("/uploads/" in ref)
            if is_upload_ref:
                uploads_urls += 1
                if rr.status_code == 200:
                    uploads_200 += 1
            click.echo(
                f"LIVE_LEARN_AUDIO_{idx} ref={ref} final={rr.url} status={rr.status_code} "
                f"content_type={(rr.headers.get('content-type') or '').strip()}"
            )
        except Exception as e:
            click.echo(f"LIVE_LEARN_AUDIO_{idx} ref={ref} error={repr(e)}")
    click.echo(
        f"LIVE_LEARN_UPLOADS_CHECK entry_type={entry_type} text={query_text} "
        f"uploads_urls={uploads_urls} uploads_200={uploads_200} all_uploads_200={(uploads_urls > 0 and uploads_urls == uploads_200)}"
    )


@app.cli.command("audit-learn-delivery")
@click.option("--base-url", default="https://gadaadictionary.com", show_default=True, help="Target site base URL.")
@click.option("--path", default="/learn", show_default=True, help="Path to inspect.")
def cli_audit_learn_delivery(base_url, path):
    """
    Inspect rendered Learn HTML/headers/assets to detect template or cache/version drift.
    Read-only diagnostics; does not trigger generation or repair.
    """
    base = (base_url or "").strip().rstrip("/")
    rel = "/" + (path or "/learn").lstrip("/")
    url = f"{base}{rel}"
    sess = requests.Session()
    sess.trust_env = False

    try:
        resp = sess.get(url, timeout=25)
    except Exception as e:
        click.echo(f"LEARN_DELIVERY_ERROR url={url} error={repr(e)}")
        return

    html = resp.text or ""
    marker_match = re.search(r'<!--\s*learn_debug\s+([^>]*)-->', html, re.IGNORECASE)
    marker = normalize_text(marker_match.group(1) if marker_match else "")
    title_match = re.search(r"<title>(.*?)</title>", html, re.IGNORECASE | re.DOTALL)
    page_title = normalize_text(title_match.group(1) if title_match else "")

    click.echo(f"LEARN_DELIVERY status={resp.status_code} url={resp.url}")
    click.echo(
        "LEARN_DELIVERY_HEADERS "
        f"x_gadaa_build={resp.headers.get('X-Gadaa-Build', '')} "
        f"x_learn_template={resp.headers.get('X-Learn-Template-Version', '')} "
        f"x_learn_render_path={resp.headers.get('X-Learn-Render-Path', '')} "
        f"x_learn_legacy_cards={resp.headers.get('X-Learn-Legacy-Cards', '')} "
        f"cache_control={resp.headers.get('Cache-Control', '')}"
    )
    click.echo(
        "LEARN_DELIVERY_HTML "
        f"title={page_title or 'missing'} "
        f"has_debug_marker={bool(marker)} "
        f"has_data_audio={'data-audio=' in html} "
        f"has_table_rows={'Multilingual Study Table' in html} "
        f"has_legacy_cards={'Quick Practice Cards' in html} "
        f"contains_no_words_text={'No approved words loaded here yet' in html}"
    )
    if marker:
        click.echo(f"LEARN_DELIVERY_MARKER {marker}")

    script_srcs = re.findall(r'<script[^>]+src=\"([^\"]+)\"', html, re.IGNORECASE)
    audio_src = ""
    pwa_src = ""
    for src in script_srcs:
        if "audio.js" in src and (not audio_src):
            audio_src = src
        if "pwa-ui.js" in src and (not pwa_src):
            pwa_src = src
    click.echo(
        "LEARN_DELIVERY_ASSETS "
        f"audio_js={audio_src or 'missing'} "
        f"pwa_ui_js={pwa_src or 'missing'} "
        f"script_count={len(script_srcs)}"
    )

    sw_urls = [f"{base}/service-worker.js", f"{base}/sw.js", f"{base}/static/service-worker.js", f"{base}/static/sw.js"]
    for sw_url in sw_urls:
        try:
            sw_resp = sess.get(sw_url, timeout=25)
            body = sw_resp.text or ""
            click.echo(
                "LEARN_DELIVERY_SW "
                f"url={sw_url} status={sw_resp.status_code} "
                f"cache_control={sw_resp.headers.get('Cache-Control', '')} "
                f"service_worker_allowed={sw_resp.headers.get('Service-Worker-Allowed', '')} "
                f"contains_cache_version={'CACHE_VERSION' in body} "
                f"contains_nav_fetch={'mode === \"navigate\"' in body}"
            )
        except Exception as e:
            click.echo(f"LEARN_DELIVERY_SW url={sw_url} error={repr(e)}")


@app.cli.command("diagnose-data-model")
def cli_diagnose_data_model():
    """
    Print DB/storage diagnostics for production data-model verification.
    """
    _log_db_context("cli:diagnose-data-model")
    d = _collect_db_diagnostics()
    click.echo("Data-model diagnostics:")
    click.echo(f"RUNTIME={d.get('runtime', '')}")
    click.echo(f"BASE_DIR={d.get('base_dir', '')}")
    click.echo(f"PERSISTENT_DATA_CONFIGURED={d.get('persistent_data_configured', False)}")
    click.echo(f"PERSISTENT_STORAGE_ACTIVE={d.get('is_persistent_storage', False)}")
    click.echo(f"DB_PATH={d.get('db_path', '')}")
    click.echo(f"DB_ABS={d.get('db_abs', '')}")
    click.echo(f"DB_EXISTS={d.get('db_exists', False)}")
    click.echo(f"UPLOAD_FOLDER={d.get('upload_folder', '')}")
    click.echo(f"VOICE_MAP={d.get('voice_map', {})}")
    click.echo(f"REQUIRE_EXPLICIT_DB_PATH={d.get('require_explicit_db_path', False)}")
    click.echo(f"REQUIRE_BLOB_FOR_GENERATED_TTS={d.get('require_blob_for_generated_tts', False)}")
    click.echo(f"AZURE_BLOB_ENABLED={d.get('azure_blob_enabled', False)}")
    click.echo(
        "ROW_COUNTS "
        f"words={d.get('tables', {}).get('words', -1)} "
        f"phrases={d.get('tables', {}).get('phrases', -1)} "
        f"generated_translations={d.get('tables', {}).get('generated_translations', -1)} "
        f"generated_phrase_translations={d.get('tables', {}).get('generated_phrase_translations', -1)} "
        f"generated_tts_audio={d.get('tables', {}).get('generated_tts_audio', -1)} "
        f"audio={d.get('tables', {}).get('audio', -1)} "
        f"post_import_jobs={d.get('tables', {}).get('post_import_jobs', -1)}"
    )
    click.echo(
        "TTS_STORAGE "
        f"blob_url_rows={d.get('generated_tts_storage', {}).get('blob_url_rows', 0)} "
        f"local_path_rows={d.get('generated_tts_storage', {}).get('local_path_rows', 0)} "
        f"empty_rows={d.get('generated_tts_storage', {}).get('empty_rows', 0)}"
    )


@app.cli.command("db-diagnostics")
def cli_db_diagnostics():
    """
    Print core DB drift diagnostics for local vs production verification.
    """
    _log_db_context("cli:db-diagnostics")
    d = _collect_db_diagnostics()
    words_counts = _words_table_counts()
    click.echo(f"DB_PATH={d.get('db_path', '')}")
    click.echo(f"DB_ABS={d.get('db_abs', '')}")
    click.echo(f"UPLOAD_FOLDER={d.get('upload_folder', '')}")
    click.echo(f"RUNTIME={d.get('runtime', '')}")
    click.echo(f"PERSISTENT_STORAGE_ACTIVE={d.get('is_persistent_storage', False)}")
    click.echo(f"VOICE_MAP={d.get('voice_map', {})}")
    click.echo(f"RENDER_DISK_ACTIVE={d.get('is_render_disk', False)}")
    click.echo(
        f"words_total={d.get('tables', {}).get('words', -1)} "
        f"approved_words={words_counts.get('approved_rows', 0)} "
        f"phrases_total={d.get('tables', {}).get('phrases', -1)} "
        f"generated_translations_total={d.get('tables', {}).get('generated_translations', -1)} "
        f"generated_tts_audio_total={d.get('tables', {}).get('generated_tts_audio', -1)} "
        f"audio_total={d.get('tables', {}).get('audio', -1)} "
        f"post_import_jobs_total={d.get('tables', {}).get('post_import_jobs', -1)}"
    )


@app.cli.command("audit-phrase-translations")
def cli_audit_phrase_translations():
    """
    Print phrase translation coverage for am/ar/fr/zh-CN.
    """
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute(
        """
        SELECT COUNT(*)
        FROM phrases
        WHERE status='approved' AND english IS NOT NULL AND TRIM(english) != ''
        """
    )
    approved_total = int((c.fetchone() or [0])[0] or 0)
    c.execute(
        """
        SELECT COUNT(*)
        FROM (
            SELECT p.id
            FROM phrases p
            LEFT JOIN generated_phrase_translations g_am
              ON g_am.phrase_id = p.id
             AND g_am.lang_code='am'
             AND g_am.translated_text IS NOT NULL
             AND TRIM(g_am.translated_text) != ''
            LEFT JOIN generated_phrase_translations g_ar
              ON g_ar.phrase_id = p.id
             AND g_ar.lang_code='ar'
             AND g_ar.translated_text IS NOT NULL
             AND TRIM(g_ar.translated_text) != ''
            LEFT JOIN generated_phrase_translations g_fr
              ON g_fr.phrase_id = p.id
             AND g_fr.lang_code='fr'
             AND g_fr.translated_text IS NOT NULL
             AND TRIM(g_fr.translated_text) != ''
            LEFT JOIN generated_phrase_translations g_zh
              ON g_zh.phrase_id = p.id
             AND g_zh.lang_code='zh-CN'
             AND g_zh.translated_text IS NOT NULL
             AND TRIM(g_zh.translated_text) != ''
            WHERE p.status='approved'
              AND p.english IS NOT NULL
              AND TRIM(p.english) != ''
              AND (
                g_am.id IS NULL OR
                g_ar.id IS NULL OR
                g_fr.id IS NULL OR
                g_zh.id IS NULL
              )
        )
        """
    )
    missing_any = int((c.fetchone() or [0])[0] or 0)
    c.execute("SELECT COUNT(*) FROM generated_phrase_translations")
    total_generated_rows = int((c.fetchone() or [0])[0] or 0)
    conn.close()

    click.echo(
        f"approved_phrases_total={approved_total} "
        f"phrases_missing_one_or_more={missing_any} "
        f"generated_phrase_translations_total={total_generated_rows}"
    )


@app.cli.command("backfill-audio-linkage")
@click.option("--limit", type=int, default=0, show_default=True, help="Limit scanned tts files (0 = all).")
@click.option("--dry-run", is_flag=True, default=False, help="Preview only; do not write DB updates.")
@click.option(
    "--source-dir",
    multiple=True,
    help="Extra folder(s) to scan for existing tts_*.mp3 before promotion to upload folder.",
)
def cli_backfill_audio_linkage(limit, dry_run, source_dir):
    """
    Link existing persisted TTS files/URLs into DB caches without regeneration.
    """
    _log_db_context("cli:backfill-audio-linkage")
    click.echo(f"DB_PATH={DB_NAME}")
    click.echo(f"DB_ABS={os.path.abspath(DB_NAME)}")
    summary = run_backfill_existing_audio_linkage(
        limit=int(limit or 0),
        dry_run=bool(dry_run),
        source_dirs=list(source_dir or ()),
        promote_to_uploads=True,
    )
    click.echo("Audio linkage backfill completed.")
    click.echo(
        f"files_scanned={summary.get('files_scanned', 0)} "
        f"files_promoted={summary.get('files_promoted', 0)} "
        f"files_already_in_uploads={summary.get('files_already_in_uploads', 0)} "
        f"files_promotion_failed={summary.get('files_promotion_failed', 0)} "
        f"rows_linked={summary.get('rows_linked', 0)} "
        f"rows_already_present={summary.get('rows_already_present', 0)} "
        f"rows_skipped_missing_text={summary.get('rows_skipped_missing_text', 0)} "
        f"rows_skipped_hash_mismatch={summary.get('rows_skipped_hash_mismatch', 0)} "
        f"rows_skipped_missing_file={summary.get('rows_skipped_missing_file', 0)} "
        f"cache_rows_scanned={summary.get('cache_rows_scanned', 0)} "
        f"cache_rows_linked={summary.get('cache_rows_linked', 0)} "
        f"dry_run={summary.get('dry_run', False)}"
    )


@app.cli.command("import-existing-audio")
@click.option("--limit", type=int, default=0, show_default=True, help="Limit scanned tts files (0 = all).")
@click.option("--dry-run", is_flag=True, default=False, help="Preview only; do not write DB updates.")
@click.option(
    "--source-dir",
    multiple=True,
    help="Source folder(s) containing local tts_*.mp3 files to import/promote.",
)
def cli_import_existing_audio(limit, dry_run, source_dir):
    """
    Import/link existing local TTS files into persistent uploads + generated_tts_audio.
    This command never regenerates audio and never calls Azure.
    """
    _log_db_context("cli:import-existing-audio")
    click.echo(f"DB_PATH={DB_NAME}")
    click.echo(f"DB_ABS={os.path.abspath(DB_NAME)}")
    click.echo(f"UPLOAD_FOLDER={UPLOAD_FOLDER}")
    click.echo(f"RENDER_DISK_ACTIVE={IS_RENDER_DISK}")
    src_dirs = list(source_dir or ()) or [STATIC_UPLOADS_FOLDER]
    summary = run_backfill_existing_audio_linkage(
        limit=int(limit or 0),
        dry_run=bool(dry_run),
        source_dirs=src_dirs,
        promote_to_uploads=True,
    )
    click.echo("Existing audio import completed.")
    click.echo(
        f"files_scanned={summary.get('files_scanned', 0)} "
        f"files_promoted={summary.get('files_promoted', 0)} "
        f"files_already_in_uploads={summary.get('files_already_in_uploads', 0)} "
        f"files_promotion_failed={summary.get('files_promotion_failed', 0)} "
        f"rows_linked={summary.get('rows_linked', 0)} "
        f"rows_already_present={summary.get('rows_already_present', 0)} "
        f"rows_skipped_missing_text={summary.get('rows_skipped_missing_text', 0)} "
        f"rows_skipped_hash_mismatch={summary.get('rows_skipped_hash_mismatch', 0)} "
        f"rows_skipped_missing_file={summary.get('rows_skipped_missing_file', 0)} "
        f"dry_run={summary.get('dry_run', False)}"
    )


@app.cli.command("repair-missing-audio")
@click.option("--limit", type=int, default=0, show_default=True, help="Limit approved words/phrases scanned (0 = all).")
def cli_repair_missing_audio(limit):
    """
    Cache-first audio repair:
    1) link existing local files
    2) generate only missing audio with current Azure key
    """
    _log_db_context("cli:repair-missing-audio")
    click.echo(f"DB_PATH={DB_NAME}")
    click.echo(f"DB_ABS={os.path.abspath(DB_NAME)}")
    click.echo(f"UPLOAD_FOLDER={UPLOAD_FOLDER}")
    click.echo(f"RENDER_DISK_ACTIVE={IS_RENDER_DISK}")
    summary = run_repair_missing_audio(limit=int(limit or 0), generate_missing=True, source_dirs=[STATIC_UPLOADS_FOLDER])
    linkage = summary.get("linkage", {}) or {}
    click.echo("Repair missing audio completed.")
    click.echo(
        f"items_scanned={summary.get('items_scanned', 0)} "
        f"audio_reused={summary.get('audio_reused', 0)} "
        f"audio_generated={summary.get('audio_generated', 0)} "
        f"missing_text={summary.get('missing_text', 0)} "
        f"missing_voice_config={summary.get('missing_voice_config', 0)} "
        f"failures={summary.get('failures', 0)} "
        f"linkage_files_scanned={linkage.get('files_scanned', 0)} "
        f"linkage_files_promoted={linkage.get('files_promoted', 0)} "
        f"linkage_rows_linked={linkage.get('rows_linked', 0)} "
        f"linkage_rows_already_present={linkage.get('rows_already_present', 0)}"
    )


@app.cli.command("backfill-phrase-translations")
@click.option("--limit", type=int, default=0, show_default=True, help="Limit phrases scanned (0 = all missing).")
@click.option("--chunk-size", type=int, default=IMPORT_BATCH_SIZE, show_default=True, help="Batch size for translation requests.")
def cli_backfill_phrase_translations(limit, chunk_size):
    """
    Backfill only missing generated translations for phrases.
    Generates am/ar/fr/zh-CN from phrase English text and persists to DB cache.
    """
    summary = run_phrase_translation_backfill(
        limit=int(limit or 0),
        chunk_size=int(chunk_size or IMPORT_BATCH_SIZE),
    )
    click.echo("Phrase translation backfill completed.")
    click.echo(
        f"phrases_processed={summary.get('phrases_processed', 0)} "
        f"translations_generated={summary.get('translations_generated', 0)} "
        f"translations_cached={summary.get('translations_cached', 0)} "
        f"failures={summary.get('failures', 0)} "
        f"skipped_missing_text={summary.get('skipped_missing_text', 0)}"
    )


def _canonical_phrase_sync_lang(lang_code: str) -> str:
    k = normalize_text(lang_code or "").replace("_", "-").casefold()
    if k in ("am",):
        return "am"
    if k in ("ar",):
        return "ar"
    if k in ("fr",):
        return "fr"
    if k in ("zh-cn",):
        return "zh-CN"
    return ""


def run_sync_missing_phrase_translations(source_db_path: str, dry_run: bool = False):
    """
    Sync only missing generated_phrase_translations rows from source DB into current DB.
    Safe behavior:
      - never deletes
      - never overwrites existing non-empty target rows
      - no external API calls
    """
    summary = {
        "source_rows_scanned": 0,
        "production_rows_already_present": 0,
        "rows_inserted": 0,
        "rows_skipped_no_phrase_match": 0,
        "rows_skipped_ambiguous_phrase_match": 0,
        "rows_skipped_unsupported_lang_code": 0,
        "rows_skipped_missing_text": 0,
        "rows_updated_empty_existing": 0,
        "dry_run": bool(dry_run),
    }

    source_db = normalize_text(source_db_path or "")
    if not source_db:
        raise click.UsageError("--source-db is required.")
    if not os.path.isfile(source_db):
        raise click.UsageError(f"Source DB not found: {source_db}")

    src_conn = sqlite3.connect(source_db)
    src_cur = src_conn.cursor()
    dst_conn = sqlite3.connect(DB_NAME)
    dst_cur = dst_conn.cursor()
    try:
        # Source phrase identity map
        src_cur.execute(
            """
            SELECT id, english, oromo
            FROM phrases
            """
        )
        src_phrase_by_id = {
            int(pid or 0): (normalize_text(en or ""), normalize_text(om or ""))
            for pid, en, om in src_cur.fetchall()
            if int(pid or 0) > 0
        }

        # Destination phrase identity maps
        dst_cur.execute(
            """
            SELECT id, english, oromo
            FROM phrases
            """
        )
        dst_phrase_by_id = {}
        dst_phrase_by_text = {}
        for pid, en, om in dst_cur.fetchall():
            pid_int = int(pid or 0)
            if pid_int <= 0:
                continue
            key = (normalize_text(en or ""), normalize_text(om or ""))
            dst_phrase_by_id[pid_int] = key
            dst_phrase_by_text.setdefault(key, []).append(pid_int)

        # Existing destination translation rows
        dst_cur.execute(
            """
            SELECT phrase_id, lang_code, translated_text
            FROM generated_phrase_translations
            """
        )
        dst_existing = {}
        for pid, lang, txt in dst_cur.fetchall():
            pid_int = int(pid or 0)
            c_lang = _canonical_phrase_sync_lang(lang or "")
            if pid_int <= 0 or not c_lang:
                continue
            key = (pid_int, c_lang)
            state = dst_existing.get(key, {"row_exists": False, "non_empty": False})
            state["row_exists"] = True
            if normalize_text(txt or ""):
                state["non_empty"] = True
            dst_existing[key] = state

        # Source translation rows
        src_cur.execute(
            """
            SELECT phrase_id, lang_code, translated_text
            FROM generated_phrase_translations
            WHERE translated_text IS NOT NULL
              AND TRIM(translated_text) != ''
            ORDER BY phrase_id ASC
            """
        )
        src_rows = src_cur.fetchall()

        for src_phrase_id, src_lang, src_text in src_rows:
            summary["source_rows_scanned"] += 1
            pid_src = int(src_phrase_id or 0)
            c_lang = _canonical_phrase_sync_lang(src_lang or "")
            if not c_lang:
                summary["rows_skipped_unsupported_lang_code"] += 1
                continue
            translated = normalize_text(src_text or "")
            if not translated:
                summary["rows_skipped_missing_text"] += 1
                continue

            src_phrase_key = src_phrase_by_id.get(pid_src)
            target_phrase_id = 0

            # Preferred: same phrase_id only when phrase texts align.
            if pid_src in dst_phrase_by_id and src_phrase_key and (dst_phrase_by_id.get(pid_src) == src_phrase_key):
                target_phrase_id = pid_src
            else:
                # Fallback: exact English+Oromo text match, only when unambiguous.
                if not src_phrase_key or (not src_phrase_key[0]) or (not src_phrase_key[1]):
                    summary["rows_skipped_no_phrase_match"] += 1
                    continue
                candidates = dst_phrase_by_text.get(src_phrase_key, [])
                if len(candidates) == 1:
                    target_phrase_id = int(candidates[0] or 0)
                elif len(candidates) == 0:
                    summary["rows_skipped_no_phrase_match"] += 1
                    continue
                else:
                    summary["rows_skipped_ambiguous_phrase_match"] += 1
                    continue

            if target_phrase_id <= 0:
                summary["rows_skipped_no_phrase_match"] += 1
                continue

            key = (target_phrase_id, c_lang)
            state = dst_existing.get(key, {"row_exists": False, "non_empty": False})
            if state.get("non_empty"):
                summary["production_rows_already_present"] += 1
                continue

            if dry_run:
                if state.get("row_exists"):
                    summary["rows_updated_empty_existing"] += 1
                else:
                    summary["rows_inserted"] += 1
                # Simulate destination state for subsequent duplicates.
                dst_existing[key] = {"row_exists": True, "non_empty": True}
                continue

            if state.get("row_exists"):
                dst_cur.execute(
                    """
                    UPDATE generated_phrase_translations
                    SET translated_text=?, provider='sync_local_db', updated_at=CURRENT_TIMESTAMP
                    WHERE phrase_id=? AND lang_code=?
                      AND (translated_text IS NULL OR TRIM(translated_text) = '')
                    """,
                    (translated, int(target_phrase_id), c_lang),
                )
                if int(dst_cur.rowcount or 0) > 0:
                    summary["rows_updated_empty_existing"] += 1
                    dst_existing[key] = {"row_exists": True, "non_empty": True}
                else:
                    # Another process may have filled it between preload and now.
                    summary["production_rows_already_present"] += 1
                continue

            dst_cur.execute(
                """
                INSERT INTO generated_phrase_translations
                (phrase_id, lang_code, translated_text, provider, updated_at)
                VALUES (?, ?, ?, 'sync_local_db', CURRENT_TIMESTAMP)
                """,
                (int(target_phrase_id), c_lang, translated),
            )
            summary["rows_inserted"] += 1
            dst_existing[key] = {"row_exists": True, "non_empty": True}

        if not dry_run:
            dst_conn.commit()
    finally:
        try:
            src_conn.close()
        except Exception:
            pass
        try:
            dst_conn.close()
        except Exception:
            pass

    return summary


@app.cli.command("sync-missing-phrase-translations")
@click.option("--source-db", required=True, help="Source SQLite DB path (local export) to sync from.")
@click.option("--dry-run", is_flag=True, default=False, help="Preview only; do not write changes.")
def cli_sync_missing_phrase_translations(source_db, dry_run):
    """
    Sync missing generated_phrase_translations rows from source DB into current DB.
    Conservative matching:
      1) phrase_id + exact phrase text alignment
      2) fallback exact English+Oromo match when unique
    """
    _log_db_context("cli:sync-missing-phrase-translations")
    click.echo(f"DB_PATH={DB_NAME}")
    click.echo(f"DB_ABS={os.path.abspath(DB_NAME)}")
    click.echo(f"SOURCE_DB={source_db}")
    summary = run_sync_missing_phrase_translations(source_db_path=source_db, dry_run=bool(dry_run))
    click.echo("Phrase translation sync completed.")
    click.echo(
        f"source_rows_scanned={summary.get('source_rows_scanned', 0)} "
        f"production_rows_already_present={summary.get('production_rows_already_present', 0)} "
        f"rows_inserted={summary.get('rows_inserted', 0)} "
        f"rows_updated_empty_existing={summary.get('rows_updated_empty_existing', 0)} "
        f"rows_skipped_no_phrase_match={summary.get('rows_skipped_no_phrase_match', 0)} "
        f"rows_skipped_ambiguous_phrase_match={summary.get('rows_skipped_ambiguous_phrase_match', 0)} "
        f"rows_skipped_unsupported_lang_code={summary.get('rows_skipped_unsupported_lang_code', 0)} "
        f"rows_skipped_missing_text={summary.get('rows_skipped_missing_text', 0)} "
        f"dry_run={summary.get('dry_run', False)}"
    )


def _load_entry_identity_maps(cur, table_name: str):
    cur.execute(f"SELECT id, english, oromo FROM {table_name}")
    by_id = {}
    by_text = {}
    for row_id, en, om in cur.fetchall():
        rid = int(row_id or 0)
        if rid <= 0:
            continue
        key = (normalize_text(en or ""), normalize_text(om or ""))
        by_id[rid] = key
        by_text.setdefault(key, []).append(rid)
    return {"by_id": by_id, "by_text": by_text}


def _resolve_audio_sync_target_entry_id(entry_type: str, src_entry_id: int, src_maps: dict, dst_maps: dict):
    src = (src_maps or {}).get(entry_type, {}) or {}
    dst = (dst_maps or {}).get(entry_type, {}) or {}
    src_by_id = src.get("by_id", {}) or {}
    dst_by_id = dst.get("by_id", {}) or {}
    dst_by_text = dst.get("by_text", {}) or {}

    src_id = int(src_entry_id or 0)
    if src_id <= 0:
        return 0, "no_match"
    src_key = src_by_id.get(src_id)
    if not src_key:
        return 0, "no_match"

    if (src_id in dst_by_id) and (dst_by_id.get(src_id) == src_key):
        return src_id, "id_match"

    en_text, om_text = src_key
    if not en_text or not om_text:
        return 0, "no_match"

    candidates = dst_by_text.get(src_key, []) or []
    if len(candidates) == 1:
        return int(candidates[0] or 0), "text_match"
    if len(candidates) > 1:
        return 0, "ambiguous"
    return 0, "no_match"


def _resolve_source_audio_abs_path(source_db_path: str, file_path: str) -> str:
    fp = normalize_text(file_path or "").replace("\\", "/")
    if not fp or _is_remote_audio_ref(fp):
        return ""

    source_db_abs = os.path.abspath(source_db_path)
    source_root = os.path.dirname(source_db_abs)
    name = os.path.basename(fp)
    candidates = []

    if os.path.isabs(fp):
        candidates.append(fp)
    else:
        candidates.append(os.path.join(source_root, fp))
    if name:
        candidates.extend(
            [
                os.path.join(source_root, "uploads", name),
                os.path.join(source_root, "static", "uploads", name),
                os.path.join(source_root, name),
            ]
        )

    seen = set()
    for path in candidates:
        abs_path = os.path.abspath(path)
        if abs_path in seen:
            continue
        seen.add(abs_path)
        if os.path.isfile(abs_path):
            return abs_path
    return ""


def _canonical_audio_sync_lang(lang_code: str) -> str:
    canonical = _canonical_tts_lang_code(lang_code or "")
    if canonical in ("en", "am", "ar", "fr", "zh-CN", "om"):
        return canonical
    return ""


def _collect_phrase_audio_coverage(conn):
    c = conn.cursor()
    c.execute(
        """
        SELECT id
        FROM phrases
        WHERE status='approved'
          AND english IS NOT NULL AND TRIM(english) != ''
          AND oromo IS NOT NULL AND TRIM(oromo) != ''
        """
    )
    phrase_ids = {int(r[0] or 0) for r in c.fetchall() if int(r[0] or 0) > 0}
    total = int(len(phrase_ids))

    tts_sets = {lc: set() for lc in ("en", "am", "ar", "fr", "zh-CN", "om")}
    if phrase_ids:
        placeholders = ",".join(["?"] * len(phrase_ids))
        c.execute(
            f"""
            SELECT entry_id, lang_code, file_path
            FROM generated_tts_audio
            WHERE entry_type='phrase'
              AND entry_id IN ({placeholders})
            """,
            tuple(sorted(phrase_ids)),
        )
        for entry_id, lang_code, file_path in c.fetchall():
            pid = int(entry_id or 0)
            if pid not in phrase_ids:
                continue
            lc = _canonical_audio_sync_lang(lang_code or "")
            if lc not in tts_sets:
                continue
            if _has_usable_audio_ref(file_path or ""):
                tts_sets[lc].add(pid)

    approved_oromo = set()
    if phrase_ids:
        placeholders = ",".join(["?"] * len(phrase_ids))
        c.execute(
            f"""
            SELECT entry_id, file_path
            FROM audio
            WHERE status='approved'
              AND entry_type='phrase'
              AND lower(lang)='oromo'
              AND entry_id IN ({placeholders})
            """,
            tuple(sorted(phrase_ids)),
        )
        for entry_id, file_path in c.fetchall():
            pid = int(entry_id or 0)
            if (pid in phrase_ids) and _has_usable_audio_ref(file_path or ""):
                approved_oromo.add(pid)

    oromo_union = approved_oromo.union(tts_sets.get("om", set()))
    coverage = {
        "total_phrases": total,
        "with_audio": {
            "en": len(tts_sets.get("en", set())),
            "am": len(tts_sets.get("am", set())),
            "ar": len(tts_sets.get("ar", set())),
            "fr": len(tts_sets.get("fr", set())),
            "zh-CN": len(tts_sets.get("zh-CN", set())),
            "oromo": len(oromo_union),
        },
    }
    coverage["missing"] = {
        lc: max(0, total - int(coverage["with_audio"].get(lc, 0)))
        for lc in ("en", "am", "ar", "fr", "zh-CN", "oromo")
    }
    return coverage


def run_sync_missing_audio_from_db(source_db_path: str, dry_run: bool = False):
    """
    Sync existing local audio metadata into current DB without generation.
    Includes:
      - generated_tts_audio (en/am/ar/fr/zh-CN/om)
      - approved Oromo audio linkage from audio table (lang='oromo')
    """
    summary = {
        "dry_run": bool(dry_run),
        "source_generated_rows_scanned": 0,
        "source_approved_oromo_rows_scanned": 0,
        "rows_inserted_from_localhost_audio": 0,
        "rows_inserted_generated_tts_audio": 0,
        "rows_inserted_approved_oromo_audio": 0,
        "production_rows_already_present": 0,
        "rows_skipped_production_has_usable_audio": 0,
        "rows_skipped_source_file_missing": 0,
        "rows_skipped_no_destination_match": 0,
        "rows_skipped_ambiguous_match": 0,
        "rows_skipped_empty_file_path": 0,
        "rows_skipped_unsupported_lang_code": 0,
        "rows_skipped_unknown_entry_type": 0,
        "files_copied_to_uploads": 0,
        "files_already_present_in_uploads": 0,
        "rows_still_missing_after_sync": 0,
        "phrase_sync_inserted_en": 0,
        "phrase_sync_inserted_am": 0,
        "phrase_sync_inserted_ar": 0,
        "phrase_sync_inserted_fr": 0,
        "phrase_sync_inserted_zh-CN": 0,
        "phrase_sync_inserted_oromo": 0,
    }

    source_db = normalize_text(source_db_path or "")
    if not source_db:
        raise click.UsageError("--source-db is required.")
    if not os.path.isfile(source_db):
        raise click.UsageError(f"Source DB not found: {source_db}")

    src_conn = sqlite3.connect(source_db)
    src_cur = src_conn.cursor()
    dst_conn = sqlite3.connect(DB_NAME)
    dst_cur = dst_conn.cursor()
    try:
        src_maps = {
            "word": _load_entry_identity_maps(src_cur, "words"),
            "phrase": _load_entry_identity_maps(src_cur, "phrases"),
        }
        dst_maps = {
            "word": _load_entry_identity_maps(dst_cur, "words"),
            "phrase": _load_entry_identity_maps(dst_cur, "phrases"),
        }

        dst_existing_tts_initial = set()
        dst_existing_tts_usable = {}
        dst_cur.execute(
            """
            SELECT entry_type, entry_id, lang_code, file_path
            FROM generated_tts_audio
            """
        )
        for entry_type, entry_id, lang_code, file_path in dst_cur.fetchall():
            et = normalize_text(entry_type or "").lower()
            if et not in ("word", "phrase"):
                continue
            lc = _canonical_audio_sync_lang(lang_code or "")
            if not lc:
                continue
            eid = int(entry_id or 0)
            if eid <= 0:
                continue
            key = (et, eid, lc)
            usable = _has_usable_audio_ref(file_path or "")
            if usable:
                dst_existing_tts_initial.add(key)
            dst_existing_tts_usable[key] = bool(dst_existing_tts_usable.get(key, False) or usable)

        dst_existing_oromo_initial = set()
        dst_existing_oromo_usable = {}
        dst_cur.execute(
            """
            SELECT entry_type, entry_id, file_path
            FROM audio
            WHERE status='approved'
              AND lower(lang)='oromo'
            """
        )
        for entry_type, entry_id, file_path in dst_cur.fetchall():
            et = normalize_text(entry_type or "").lower()
            if et not in ("word", "phrase"):
                continue
            eid = int(entry_id or 0)
            if eid <= 0:
                continue
            key = (et, eid, "oromo")
            usable = _has_usable_audio_ref(file_path or "")
            if usable:
                dst_existing_oromo_initial.add(key)
            dst_existing_oromo_usable[key] = bool(dst_existing_oromo_usable.get(key, False) or usable)

        desired_keys_tts = set()
        desired_keys_oromo = set()

        src_cur.execute(
            """
            SELECT entry_type, entry_id, lang_code, text_value, text_hash, voice_provider, voice_name, file_path
            FROM generated_tts_audio
            WHERE file_path IS NOT NULL
              AND TRIM(file_path) != ''
            ORDER BY id ASC
            """
        )
        for entry_type, entry_id, lang_code, text_value, text_hash, voice_provider, voice_name, file_path in src_cur.fetchall():
            summary["source_generated_rows_scanned"] += 1
            et = normalize_text(entry_type or "").lower()
            if et not in ("word", "phrase"):
                summary["rows_skipped_unknown_entry_type"] += 1
                continue
            lc = _canonical_audio_sync_lang(lang_code or "")
            if not lc:
                summary["rows_skipped_unsupported_lang_code"] += 1
                continue
            target_id, reason = _resolve_audio_sync_target_entry_id(et, int(entry_id or 0), src_maps, dst_maps)
            if target_id <= 0:
                if reason == "ambiguous":
                    summary["rows_skipped_ambiguous_match"] += 1
                else:
                    summary["rows_skipped_no_destination_match"] += 1
                continue

            key = (et, int(target_id), lc)
            desired_keys_tts.add(key)
            if dst_existing_tts_usable.get(key, False):
                summary["rows_skipped_production_has_usable_audio"] += 1
                if key in dst_existing_tts_initial:
                    summary["production_rows_already_present"] += 1
                continue

            src_fp = normalize_text(file_path or "")
            if not src_fp:
                summary["rows_skipped_empty_file_path"] += 1
                continue

            stored_ref = src_fp
            if not _is_remote_audio_ref(src_fp):
                src_abs = _resolve_source_audio_abs_path(source_db, src_fp)
                if not src_abs:
                    summary["rows_skipped_source_file_missing"] += 1
                    continue
                name = os.path.basename(src_abs)
                if not name:
                    summary["rows_skipped_source_file_missing"] += 1
                    continue
                dst_abs = os.path.join(UPLOAD_FOLDER, name)
                if os.path.isfile(dst_abs):
                    summary["files_already_present_in_uploads"] += 1
                else:
                    summary["files_copied_to_uploads"] += 1
                    if not dry_run:
                        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
                        shutil.copy2(src_abs, dst_abs)
                stored_ref = f"uploads/{name}"

            text_val = normalize_text(text_value or "")
            text_h = normalize_text(text_hash or "")
            if (not text_h) and text_val:
                text_h = hashlib.md5(text_val.encode("utf-8")).hexdigest()
            if not text_h:
                text_h = hashlib.md5(f"{et}|{target_id}|{lc}|{os.path.basename(stored_ref)}".encode("utf-8")).hexdigest()

            if not dry_run:
                dst_cur.execute(
                    """
                    INSERT INTO generated_tts_audio
                    (entry_type, entry_id, lang_code, text_value, text_hash, voice_provider, voice_name, file_path)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        et,
                        int(target_id),
                        lc,
                        text_val or f"sync_local_db:{et}:{target_id}:{lc}",
                        text_h,
                        normalize_text(voice_provider or "") or "sync_local_db",
                        normalize_text(voice_name or ""),
                        stored_ref,
                    ),
                )
            dst_existing_tts_usable[key] = True
            summary["rows_inserted_generated_tts_audio"] += 1
            summary["rows_inserted_from_localhost_audio"] += 1
            if et == "phrase":
                summary[f"phrase_sync_inserted_{lc}"] = int(summary.get(f"phrase_sync_inserted_{lc}", 0)) + 1

        src_cur.execute(
            """
            SELECT entry_type, entry_id, file_path
            FROM audio
            WHERE status='approved'
              AND lower(lang)='oromo'
              AND file_path IS NOT NULL
              AND TRIM(file_path) != ''
            ORDER BY id ASC
            """
        )
        for entry_type, entry_id, file_path in src_cur.fetchall():
            summary["source_approved_oromo_rows_scanned"] += 1
            et = normalize_text(entry_type or "").lower()
            if et not in ("word", "phrase"):
                summary["rows_skipped_unknown_entry_type"] += 1
                continue
            target_id, reason = _resolve_audio_sync_target_entry_id(et, int(entry_id or 0), src_maps, dst_maps)
            if target_id <= 0:
                if reason == "ambiguous":
                    summary["rows_skipped_ambiguous_match"] += 1
                else:
                    summary["rows_skipped_no_destination_match"] += 1
                continue

            key = (et, int(target_id), "oromo")
            desired_keys_oromo.add(key)
            if dst_existing_oromo_usable.get(key, False):
                summary["rows_skipped_production_has_usable_audio"] += 1
                if key in dst_existing_oromo_initial:
                    summary["production_rows_already_present"] += 1
                continue

            src_fp = normalize_text(file_path or "")
            if not src_fp:
                summary["rows_skipped_empty_file_path"] += 1
                continue

            stored_ref = src_fp
            if not _is_remote_audio_ref(src_fp):
                src_abs = _resolve_source_audio_abs_path(source_db, src_fp)
                if not src_abs:
                    summary["rows_skipped_source_file_missing"] += 1
                    continue
                name = os.path.basename(src_abs)
                if not name:
                    summary["rows_skipped_source_file_missing"] += 1
                    continue
                dst_abs = os.path.join(UPLOAD_FOLDER, name)
                if os.path.isfile(dst_abs):
                    summary["files_already_present_in_uploads"] += 1
                else:
                    summary["files_copied_to_uploads"] += 1
                    if not dry_run:
                        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
                        shutil.copy2(src_abs, dst_abs)
                stored_ref = f"uploads/{name}"

            if not dry_run:
                dst_cur.execute(
                    """
                    INSERT INTO audio (entry_type, entry_id, lang, file_path, status)
                    VALUES (?, ?, 'oromo', ?, 'approved')
                    """,
                    (et, int(target_id), stored_ref),
                )
            dst_existing_oromo_usable[key] = True
            summary["rows_inserted_approved_oromo_audio"] += 1
            summary["rows_inserted_from_localhost_audio"] += 1
            if et == "phrase":
                summary["phrase_sync_inserted_oromo"] = int(summary.get("phrase_sync_inserted_oromo", 0)) + 1

        summary["rows_still_missing_after_sync"] = (
            sum(1 for k in desired_keys_tts if not dst_existing_tts_usable.get(k, False))
            + sum(1 for k in desired_keys_oromo if not dst_existing_oromo_usable.get(k, False))
        )

        if not dry_run:
            dst_conn.commit()

        phrase_cov = _collect_phrase_audio_coverage(dst_conn)
        summary["phrase_total_approved"] = int(phrase_cov.get("total_phrases", 0))
        for lc in ("en", "am", "ar", "fr", "zh-CN", "oromo"):
            summary[f"phrase_with_audio_{lc}"] = int((phrase_cov.get("with_audio", {}) or {}).get(lc, 0))
            summary[f"phrase_missing_audio_{lc}"] = int((phrase_cov.get("missing", {}) or {}).get(lc, 0))
    finally:
        try:
            src_conn.close()
        except Exception:
            pass
        try:
            dst_conn.close()
        except Exception:
            pass

    return summary


@app.cli.command("sync-missing-audio-from-db")
@click.option("--source-db", required=True, help="Source SQLite DB path (localhost export) to sync audio from.")
@click.option("--dry-run", is_flag=True, default=False, help="Preview only; do not write changes.")
def cli_sync_missing_audio_from_db(source_db, dry_run):
    """
    CLI-only audio sync from source DB to current DB.
    No generation, no external API calls, no deletes.
    """
    _log_db_context("cli:sync-missing-audio-from-db")
    click.echo(f"DB_PATH={DB_NAME}")
    click.echo(f"DB_ABS={os.path.abspath(DB_NAME)}")
    click.echo(f"SOURCE_DB={source_db}")
    summary = run_sync_missing_audio_from_db(source_db_path=source_db, dry_run=bool(dry_run))
    click.echo("Audio sync from DB completed.")
    click.echo(
        f"source_generated_rows_scanned={summary.get('source_generated_rows_scanned', 0)} "
        f"source_approved_oromo_rows_scanned={summary.get('source_approved_oromo_rows_scanned', 0)} "
        f"rows_inserted_from_localhost_audio={summary.get('rows_inserted_from_localhost_audio', 0)} "
        f"rows_inserted_generated_tts_audio={summary.get('rows_inserted_generated_tts_audio', 0)} "
        f"rows_inserted_approved_oromo_audio={summary.get('rows_inserted_approved_oromo_audio', 0)} "
        f"production_rows_already_present={summary.get('production_rows_already_present', 0)} "
        f"rows_skipped_production_has_usable_audio={summary.get('rows_skipped_production_has_usable_audio', 0)} "
        f"rows_skipped_source_file_missing={summary.get('rows_skipped_source_file_missing', 0)} "
        f"rows_skipped_no_destination_match={summary.get('rows_skipped_no_destination_match', 0)} "
        f"rows_skipped_ambiguous_match={summary.get('rows_skipped_ambiguous_match', 0)} "
        f"rows_skipped_empty_file_path={summary.get('rows_skipped_empty_file_path', 0)} "
        f"rows_skipped_unsupported_lang_code={summary.get('rows_skipped_unsupported_lang_code', 0)} "
        f"files_copied_to_uploads={summary.get('files_copied_to_uploads', 0)} "
        f"files_already_present_in_uploads={summary.get('files_already_present_in_uploads', 0)} "
        f"rows_still_missing_after_sync={summary.get('rows_still_missing_after_sync', 0)} "
        f"dry_run={summary.get('dry_run', False)}"
    )
    click.echo(
        f"PHRASE_AUDIO_SYNC inserted_en={summary.get('phrase_sync_inserted_en', 0)} "
        f"inserted_am={summary.get('phrase_sync_inserted_am', 0)} "
        f"inserted_ar={summary.get('phrase_sync_inserted_ar', 0)} "
        f"inserted_fr={summary.get('phrase_sync_inserted_fr', 0)} "
        f"inserted_zh-CN={summary.get('phrase_sync_inserted_zh-CN', 0)} "
        f"inserted_oromo={summary.get('phrase_sync_inserted_oromo', 0)}"
    )
    click.echo(
        f"PHRASE_AUDIO_COVERAGE total={summary.get('phrase_total_approved', 0)} "
        f"with_en={summary.get('phrase_with_audio_en', 0)} missing_en={summary.get('phrase_missing_audio_en', 0)} "
        f"with_am={summary.get('phrase_with_audio_am', 0)} missing_am={summary.get('phrase_missing_audio_am', 0)} "
        f"with_ar={summary.get('phrase_with_audio_ar', 0)} missing_ar={summary.get('phrase_missing_audio_ar', 0)} "
        f"with_fr={summary.get('phrase_with_audio_fr', 0)} missing_fr={summary.get('phrase_missing_audio_fr', 0)} "
        f"with_zh-CN={summary.get('phrase_with_audio_zh-CN', 0)} missing_zh-CN={summary.get('phrase_missing_audio_zh-CN', 0)} "
        f"with_oromo={summary.get('phrase_with_audio_oromo', 0)} missing_oromo={summary.get('phrase_missing_audio_oromo', 0)}"
    )
    click.echo(
        "NOTE: localhost audio sync only imports audio that already exists in source DB/files; "
        "it cannot fill non-English phrase audio if localhost never had translated phrase audio."
    )


@app.cli.command("backfill-word-audio")
@click.option("--limit", type=int, default=0, show_default=True, help="Limit approved words (0 = all).")
@click.option("--chunk-size", type=int, default=150, show_default=True, help="Batch size for translation + TTS.")
@click.option("--force-regenerate", is_flag=True, default=False, help="Regenerate audio even when cache exists.")
def cli_backfill_word_audio(limit, chunk_size, force_regenerate):
    """
    Backfill full persisted audio assets for approved WORDS only.
    Pipeline:
      1) Ensure generated translations for am/ar/fr/zh-CN (DB cache)
      2) Ensure Azure TTS cache for en/am/ar/fr/zh-CN (+ optional om)
    """
    _log_db_context("cli:backfill-word-audio")
    click.echo(f"DB_PATH={DB_NAME}")
    click.echo(f"DB_ABS={os.path.abspath(DB_NAME)}")
    summary = run_word_audio_backfill(
        limit=int(limit or 0),
        chunk_size=int(chunk_size or 150),
        force_regenerate=bool(force_regenerate),
    )
    tts = summary.get("tts", {}) or {}
    click.echo("Word audio backfill completed.")
    click.echo(
        f"words_total={summary.get('words_total', 0)} "
        f"translation_saved={summary.get('translation_saved', 0)} "
        f"tts_words_seen={tts.get('words_seen', 0)} "
        f"tts_generated={tts.get('generated', 0)} "
        f"tts_cached={tts.get('cached', 0)} "
        f"tts_failed={tts.get('failed', 0)} "
        f"tts_skipped_missing_text={tts.get('skipped_missing_text', 0)} "
        f"tts_skipped_missing_voice={tts.get('skipped_missing_voice', 0)}"
    )


@app.cli.command("backfill-tts")
@click.option("--entry-type", type=click.Choice(["all", "word", "phrase"]), default="all", show_default=True)
@click.option("--entry-id", type=int, default=0, show_default=True)
@click.option("--force-regenerate", is_flag=True, default=False, help="Regenerate audio even when cache exists.")
@click.option("--limit", type=int, default=0, show_default=True, help="Limit approved records per entry type.")
def cli_backfill_tts(entry_type, entry_id, force_regenerate, limit):
    """
    Backfill Azure-generated TTS audio for approved words/phrases.
    Generates Learn-table languages from saved text:
    en, am, ar, fr, zh-CN, and optional om when AZURE_VOICE_OM is configured.
    """
    summary = run_tts_backfill(
        entry_type=entry_type,
        entry_id=int(entry_id or 0),
        force_regenerate=bool(force_regenerate),
        limit=int(limit or 0),
    )
    click.echo("TTS backfill completed.")
    click.echo(
        f"processed_items={summary.get('processed_items', 0)} "
        f"generated={summary.get('generated', 0)} "
        f"cached={summary.get('cached', 0)} "
        f"failed={summary.get('failed', 0)} "
        f"skipped_missing_text={summary.get('skipped_missing_text', 0)} "
        f"skipped_missing_voice={summary.get('skipped_missing_voice', 0)}"
    )


@app.cli.command("migrate-tts-to-blob")
@click.option("--limit", type=int, default=0, show_default=True, help="Limit rows scanned (0 = all).")
@click.option("--chunk-size", type=int, default=100, show_default=True, help="Rows processed per chunk.")
@click.option("--dry-run", is_flag=True, default=False, help="Scan and report migration candidates without writing.")
def cli_migrate_tts_to_blob(limit, chunk_size, dry_run):
    """
    Migrate existing generated_tts_audio local file_path rows to Azure Blob URLs.
    Storage migration only; does not regenerate audio.
    """
    _log_db_context("cli:migrate-tts-to-blob")
    click.echo(f"DB_PATH={DB_NAME}")
    click.echo(f"DB_ABS={os.path.abspath(DB_NAME)}")

    summary = run_generated_tts_blob_migration(
        limit=int(limit or 0),
        chunk_size=int(chunk_size or 100),
        dry_run=bool(dry_run),
    )
    click.echo("TTS blob migration completed.")
    click.echo(
        f"rows_scanned={summary.get('rows_scanned', 0)} "
        f"rows_migration_candidates={summary.get('rows_migration_candidates', 0)} "
        f"rows_migrated={summary.get('rows_migrated', 0)} "
        f"rows_missing_file={summary.get('rows_missing_file', 0)} "
        f"rows_already_blob_backed={summary.get('rows_already_blob_backed', 0)} "
        f"failures={summary.get('failures', 0)} "
        f"dry_run={summary.get('dry_run', False)}"
    )

@app.after_request
def no_cache_html(resp):
    if resp.mimetype == "text/html":
        resp.headers["Cache-Control"] = "no-store, max-age=0"
        resp.headers["Pragma"] = "no-cache"
    return resp



# ------------------ RUN / MIGRATE ------------------

if __name__ == "__main__":
    import sys

    if len(sys.argv) > 1 and sys.argv[1] == "migrate":
        ensure_key_columns()
        backfill_keys()
        ensure_key_indexes()
        print("DB migration done")
        sys.exit(0)   # âœ… VERY IMPORTANT â†’ stop here

    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)




